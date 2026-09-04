"""OfficeTools: Excel and PDF generation/manipulation."""

from __future__ import annotations

import asyncio
import hashlib
import json
import os
import stat
from collections.abc import Iterator
from contextlib import ExitStack, contextmanager
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

from js.config import ToolLimits
from js.echo.turn_context import current_runtime_context
from js.security.guard import BehaviorGuard, SecurityDecisionType
from js.tools.files import FileTools
from js.tools.office.csv_utils import (
    _CSV_FIELD_SIZE_LOCK,
    _csv_file_fingerprint,
    _validate_csv_encoding,
)
from js.tools.office.work_runtime import (
    _escape_formula_rows,
    _escape_formula_text,
    _is_work_runtime,
    _normalize_work_cell_values,
    _publish_work_artifact,
    _validate_work_xlsx,
    _write_work_excel_cell,
)
from js.tools.registry import ToolParam, ToolResult, ToolSpec

_OFFICE_EXTRA_MSG = "Install js-agent[office] to use Excel tools."
_PDF_EXTRA_MSG = "Install js-agent[pdf] to use PDF generation tools."


def _cancel_requested(token: Any) -> bool:
    if token is None:
        return False
    for attribute in ("is_set", "cancelled", "is_cancelled"):
        value = getattr(token, attribute, None)
        if callable(value):
            try:
                value = value()
            except Exception:
                continue
        if value is True:
            return True
    return False


class OfficeTools:
    """Tools for Excel and PDF document operations."""

    def __init__(self, workspace: Path, limits: ToolLimits, guard: BehaviorGuard) -> None:
        self.workspace = workspace.resolve()
        self.limits = limits
        self.guard = guard
        self._files = FileTools(workspace, limits, guard)

    def _save_bytes(self, path: str, payload: bytes) -> Path:
        return self._files._secure_write(path, payload, append=False)

    def _save_workbook(self, path: str, workbook: Any) -> tuple[Path, str]:
        buffer = BytesIO()
        workbook.save(buffer)
        payload = buffer.getvalue()
        target = self._save_bytes(path, payload)
        return target, hashlib.sha256(payload).hexdigest()

    def _write_path_or_error(self, path: str) -> str | ToolResult:
        try:
            logical = self._files._logical_path(path)
        except ValueError as exc:
            return ToolResult(success=False, error=str(exc))
        decision = self.guard.check_path_operation(str(logical), "write")
        if decision.decision == SecurityDecisionType.BLOCK:
            return ToolResult(success=False, error=decision.reason)
        return path

    def _resolve(self, path: str) -> Path:
        p = Path(path)
        if p.is_absolute():
            resolved = p.resolve()
        else:
            resolved = (self.workspace / p).resolve()
        try:
            resolved.relative_to(self.workspace)
        except ValueError as e:
            raise ValueError(f"Path escapes workspace: {path}") from e
        return resolved

    def _relative_workspace_path(self, path: str) -> Path:
        """Return a workspace-relative path without following final-component symlinks."""
        if not isinstance(path, str) or not path or "\x00" in path:
            raise ValueError("Invalid workspace path")
        if path.startswith("~"):
            raise ValueError("Home-relative paths are not allowed")
        candidate = Path(path)
        if candidate.is_absolute():
            try:
                relative = candidate.relative_to(self.workspace)
            except ValueError as exc:
                raise ValueError(f"Path escapes workspace: {path}") from exc
        else:
            relative = candidate
        parts = tuple(part for part in relative.parts if part not in ("", "."))
        if any(part == ".." for part in parts):
            raise ValueError(f"Path escapes workspace: {path}")
        return Path(*parts) if parts else Path(".")

    def _logical_input_target(self, path: str) -> Path:
        """Return the guarded logical target without reopening Work snapshot bytes."""
        if not _is_work_runtime():
            return self._resolve(path)
        from js_work.scoped_tools import WorkOfficeInput

        if not isinstance(path, WorkOfficeInput):
            raise ValueError("Work Office input snapshot required")
        return self.workspace / self._relative_workspace_path(str(path))

    @contextmanager
    def _materialized_input(self, path: str) -> Iterator[Path]:
        """Yield the ordinary path or a descriptor-bound Work snapshot input."""
        if not _is_work_runtime():
            yield self._resolve(path)
            return

        from js_work.file_scope import WorkOwnerFileScope, current_work_identity
        from js_work.scoped_tools import WorkOfficeInput

        if not isinstance(path, WorkOfficeInput):
            raise ValueError("Work Office input snapshot required")
        owner, session_id = current_work_identity()
        scope = WorkOwnerFileScope(
            self.workspace,
            owner=owner,
            session_id=session_id,
        )
        with scope.materialize_snapshot(path._work_office_snapshot()) as materialized:
            yield materialized

    def _open_csv_fd(self, path: str) -> tuple[int, Path]:
        """Open a CSV via trusted-root fd-relative open; caller must close the fd."""
        if _is_work_runtime():
            logical = self._logical_input_target(path)
            with self._materialized_input(path) as materialized:
                descriptor = os.dup(materialized.fileno())  # type: ignore[attr-defined]
                os.lseek(descriptor, 0, os.SEEK_SET)
                return descriptor, logical
        relative = self._relative_workspace_path(path)
        if relative == Path(".") or not relative.name:
            raise ValueError("A file name is required")
        required_dir_fd = (os.open, os.stat)
        if (
            not hasattr(os, "O_DIRECTORY")
            or not hasattr(os, "O_NOFOLLOW")
            or not hasattr(os, "O_CLOEXEC")
            or any(function not in os.supports_dir_fd for function in required_dir_fd)
        ):
            raise RuntimeError("Secure workspace filesystem primitives are unavailable")
        directory_flags = os.O_RDONLY | os.O_DIRECTORY | os.O_NOFOLLOW | os.O_CLOEXEC
        file_flags = os.O_RDONLY | os.O_NOFOLLOW | os.O_CLOEXEC
        current_fd = os.open(self.workspace, directory_flags)
        try:
            for component in relative.parts[:-1]:
                next_fd = os.open(component, directory_flags, dir_fd=current_fd)
                metadata = os.fstat(next_fd)
                if not stat.S_ISDIR(metadata.st_mode):
                    os.close(next_fd)
                    raise ValueError("Workspace path parent is not a directory")
                os.close(current_fd)
                current_fd = next_fd
            try:
                final_meta = os.stat(relative.name, dir_fd=current_fd, follow_symlinks=False)
            except FileNotFoundError as exc:
                raise FileNotFoundError(path) from exc
            if stat.S_ISLNK(final_meta.st_mode):
                raise ValueError("Symlinks are not allowed for workspace file operations")
            file_fd = os.open(relative.name, file_flags, dir_fd=current_fd)
        finally:
            os.close(current_fd)
        return file_fd, self.workspace / relative

    def get_specs(self) -> list[ToolSpec]:
        return [
            ToolSpec(
                name="csv_read",
                description="Read data from a CSV file. Returns JSON array of rows.",
                parameters=[
                    ToolParam("path", "string", "Path to CSV file (relative to workspace)"),
                    ToolParam(
                        "encoding", "string", "File encoding (default: utf-8)", required=False
                    ),
                    ToolParam(
                        "delimiter", "string", "Column delimiter (default: comma)", required=False
                    ),
                ],
                read_only=True,
            ),
            ToolSpec(
                name="csv_write",
                description="Write data to a CSV file.",
                parameters=[
                    ToolParam("path", "string", "Path to CSV file"),
                    ToolParam("data", "string", "JSON array of rows to write"),
                    ToolParam(
                        "encoding", "string", "File encoding (default: utf-8)", required=False
                    ),
                    ToolParam(
                        "delimiter", "string", "Column delimiter (default: comma)", required=False
                    ),
                ],
                dangerous=True,
            ),
            ToolSpec(
                name="excel_read",
                description="Read data from an Excel file (.xlsx). Returns JSON array of rows.",
                parameters=[
                    ToolParam("path", "string", "Path to Excel file (relative to workspace)"),
                    ToolParam(
                        "sheet", "string", "Sheet name (default: first sheet)", required=False
                    ),
                    ToolParam("start_row", "integer", "1-based start row", required=False),
                    ToolParam("end_row", "integer", "1-based end row (inclusive)", required=False),
                    ToolParam(
                        "start_col", "string", "Start column letter (e.g. 'A')", required=False
                    ),
                    ToolParam("end_col", "string", "End column letter (e.g. 'Z')", required=False),
                ],
                read_only=True,
            ),
            ToolSpec(
                name="excel_write",
                description="Write data to an Excel file. Creates the file if it doesn't exist.",
                parameters=[
                    ToolParam("path", "string", "Path to Excel file"),
                    ToolParam("sheet", "string", "Sheet name (default: 'Sheet1')", required=False),
                    ToolParam(
                        "data", "string", 'JSON array of rows to write (e.g. [["A","B"],[1,2]])'
                    ),
                    ToolParam(
                        "start_cell",
                        "string",
                        "Start cell (e.g. 'A1', default: 'A1')",
                        required=False,
                    ),
                    ToolParam(
                        "append",
                        "boolean",
                        "Append to existing sheet instead of overwriting",
                        required=False,
                    ),
                ],
                dangerous=True,
            ),
            ToolSpec(
                name="excel_merge",
                description=(
                    "Merge data from one Excel file into another at a specific location. "
                    "Useful for combining data: e.g. copy rows from source file and paste into target file at column E."
                ),
                parameters=[
                    ToolParam("source_path", "string", "Source Excel file path"),
                    ToolParam("target_path", "string", "Target Excel file path"),
                    ToolParam(
                        "output_path",
                        "string",
                        "New output path; required by JS Agent Work to preserve the target",
                        required=False,
                    ),
                    ToolParam("source_sheet", "string", "Source sheet name", required=False),
                    ToolParam("target_sheet", "string", "Target sheet name", required=False),
                    ToolParam(
                        "source_range",
                        "string",
                        "Range like 'A1:D10' or leave empty for all data",
                        required=False,
                    ),
                    ToolParam(
                        "target_start_cell",
                        "string",
                        "Target start cell, e.g. 'E1' (default: 'A1')",
                        required=False,
                    ),
                    ToolParam(
                        "include_header",
                        "boolean",
                        "Include header row from source",
                        required=False,
                    ),
                ],
                dangerous=True,
            ),
            ToolSpec(
                name="excel_create",
                description="Create a new blank Excel file with optional headers.",
                parameters=[
                    ToolParam("path", "string", "Output file path"),
                    ToolParam(
                        "sheet_name", "string", "Sheet name (default: 'Sheet1')", required=False
                    ),
                    ToolParam("headers", "string", "JSON array of column headers", required=False),
                ],
                dangerous=True,
            ),
            ToolSpec(
                name="pdf_generate",
                description="Generate a PDF file from tabular data.",
                parameters=[
                    ToolParam("path", "string", "Output PDF file path"),
                    ToolParam("title", "string", "Document title", required=False),
                    ToolParam("data", "string", "JSON array of rows (first row is header)"),
                    ToolParam("page_size", "string", "A4 or LETTER (default: A4)", required=False),
                ],
                dangerous=True,
            ),
        ]

    async def csv_read(
        self,
        path: str,
        encoding: str = "utf-8",
        delimiter: str = ",",
    ) -> ToolResult:
        try:
            csv_encoding = _validate_csv_encoding(
                encoding,
                work_runtime=_is_work_runtime(),
            )
        except (TypeError, ValueError) as exc:
            return ToolResult(success=False, error=str(exc), metadata={"complete": False})

        try:
            relative = self._relative_workspace_path(path)
            target = self.workspace / relative
        except ValueError as e:
            return ToolResult(success=False, error=str(e))
        decision = self.guard.check_path_operation(str(target), "read")
        if decision.decision == SecurityDecisionType.BLOCK:
            return ToolResult(success=False, error=decision.reason)

        suffix = target.suffix.lower()
        if suffix not in {".csv", ".tsv", ".txt"}:
            return ToolResult(
                success=False,
                error="csv_read only accepts .csv/.tsv/.txt delimited text files",
                metadata={"complete": False},
            )

        import csv

        limits = self.limits
        task = asyncio.current_task()
        runtime = current_runtime_context()
        cancel_token = None if runtime is None else runtime.cancel_token

        def _cancelled() -> bool:
            if task is not None and (task.cancelled() or task.cancelling()):
                return True
            return _cancel_requested(cancel_token)

        def _serialize_rows(
            rows: list[list[str]],
            *,
            bytes_read: int,
            pending_high_water: int,
            max_pending_chars: int,
        ) -> ToolResult:
            """Serialize with cancel + output-budget checks so work stops after cancel."""
            budget = limits.tool_output_budget_chars
            if _cancelled():
                raise asyncio.CancelledError
            # Compact lower bound refuses oversized work before indent=2 dumps.
            compact_bound = 2
            for index, row in enumerate(rows):
                if index % 32 == 0 and _cancelled():
                    raise asyncio.CancelledError
                compact_bound += len(json.dumps(row, ensure_ascii=False)) + (1 if index else 0)
                if compact_bound > budget:
                    return ToolResult(
                        success=False,
                        error="CSV output exceeds tool output budget",
                        metadata={
                            "complete": False,
                            "rows_read": len(rows),
                            "output_chars": compact_bound,
                            "bytes_read": bytes_read,
                            "error_class": "csv_output_budget_exceeded",
                        },
                    )
            if _cancelled():
                raise asyncio.CancelledError
            output = json.dumps(rows, ensure_ascii=False, indent=2)
            if len(output) > budget:
                return ToolResult(
                    success=False,
                    error="CSV output exceeds tool output budget",
                    metadata={
                        "complete": False,
                        "rows_read": len(rows),
                        "output_chars": len(output),
                        "bytes_read": bytes_read,
                        "error_class": "csv_output_budget_exceeded",
                    },
                )
            return ToolResult(
                success=True,
                output=output,
                metadata={
                    "rows": len(rows),
                    "columns": len(rows[0]) if rows else 0,
                    "complete": True,
                    "bytes_read": bytes_read,
                    "pending_high_water": pending_high_water,
                    "max_pending_chars": max_pending_chars,
                },
            )

        def _parse() -> ToolResult:
            if _cancelled():
                raise asyncio.CancelledError
            try:
                fd, _logical = self._open_csv_fd(path)
            except FileNotFoundError:
                return ToolResult(success=False, error=f"File not found: {path}")
            except OSError as exc:
                return ToolResult(success=False, error=f"Cannot open CSV: {exc}")
            except ValueError as exc:
                return ToolResult(success=False, error=str(exc), metadata={"complete": False})

            with _CSV_FIELD_SIZE_LOCK:
                previous_limit = csv.field_size_limit()
                try:
                    # Cap a single CSV field so the parser cannot allocate unbounded memory.
                    csv.field_size_limit(max(128, limits.csv_read_max_field_chars))
                    try:
                        before = os.fstat(fd)
                        if not stat.S_ISREG(before.st_mode):
                            return ToolResult(
                                success=False,
                                error="csv_read target must be a regular file",
                                metadata={"complete": False},
                            )
                        file_size = int(before.st_size)
                        if file_size > limits.csv_read_max_bytes:
                            return ToolResult(
                                success=False,
                                error=(
                                    f"CSV file exceeds size limit "
                                    f"({file_size} > {limits.csv_read_max_bytes} bytes)"
                                ),
                                metadata={
                                    "complete": False,
                                    "bytes_read": 0,
                                    "bytes": file_size,
                                },
                            )
                        rows: list[list[str]] = []
                        total_cells = 0
                        fingerprint_before = _csv_file_fingerprint(before)
                        from js.tools import office as office_mod

                        reader_io = office_mod._BinaryIncrementalCSVReader(
                            fd,
                            csv_encoding,
                            max_bytes=limits.csv_read_max_bytes,
                            max_field_chars=limits.csv_read_max_field_chars,
                            max_columns=limits.csv_read_max_columns,
                            expected_fingerprint=fingerprint_before,
                        )
                        fd = -1
                        try:
                            csv_reader = csv.reader(reader_io, delimiter=delimiter)
                            for row_index, row in enumerate(csv_reader):
                                if row_index % 16 == 0 and _cancelled():
                                    raise asyncio.CancelledError
                                bytes_seen = reader_io.bytes_read
                                if bytes_seen > limits.csv_read_max_bytes:
                                    return ToolResult(
                                        success=False,
                                        error=(
                                            "CSV content exceeds byte limit during parse "
                                            f"({limits.csv_read_max_bytes} bytes)"
                                        ),
                                        metadata={
                                            "complete": False,
                                            "rows_read": len(rows),
                                            "bytes_read": bytes_seen,
                                            "error_class": "csv_byte_budget_exceeded",
                                            "truncated": True,
                                        },
                                    )
                                if len(row) > limits.csv_read_max_columns:
                                    return ToolResult(
                                        success=False,
                                        error=(
                                            f"CSV exceeds column limit "
                                            f"({len(row)} > {limits.csv_read_max_columns})"
                                        ),
                                        metadata={
                                            "complete": False,
                                            "rows_read": len(rows),
                                            "bytes_read": bytes_seen,
                                            "error_class": "csv_column_limit_exceeded",
                                        },
                                    )
                                for field in row:
                                    if len(field) > limits.csv_read_max_field_chars:
                                        return ToolResult(
                                            success=False,
                                            error=(
                                                "CSV field exceeds length limit "
                                                f"({limits.csv_read_max_field_chars} chars)"
                                            ),
                                            metadata={
                                                "complete": False,
                                                "rows_read": len(rows),
                                                "bytes_read": bytes_seen,
                                                "error_class": "csv_field_limit_exceeded",
                                            },
                                        )
                                total_cells += len(row)
                                if total_cells > limits.csv_read_max_cells:
                                    return ToolResult(
                                        success=False,
                                        error=(
                                            "CSV exceeds cell limit "
                                            f"({limits.csv_read_max_cells} cells)"
                                        ),
                                        metadata={
                                            "complete": False,
                                            "rows_read": len(rows),
                                            "bytes_read": bytes_seen,
                                            "error_class": "csv_cell_limit_exceeded",
                                        },
                                    )
                                if len(rows) + 1 > limits.csv_read_max_rows:
                                    return ToolResult(
                                        success=False,
                                        error=(
                                            "CSV exceeds row limit "
                                            f"({limits.csv_read_max_rows} rows)"
                                        ),
                                        metadata={
                                            "complete": False,
                                            "rows_read": len(rows),
                                            "bytes_read": bytes_seen,
                                            "error_class": "csv_row_limit_exceeded",
                                        },
                                    )
                                rows.append(row)
                            reader_io._decoder.decode(b"", final=True)
                            after = os.fstat(reader_io.fileno())
                            bytes_read = reader_io.bytes_read
                            fingerprint_after = _csv_file_fingerprint(after)
                            if (
                                fingerprint_before != fingerprint_after
                                or bytes_read != before.st_size
                                or reader_io.changed
                            ):
                                return ToolResult(
                                    success=False,
                                    error="CSV file changed during read",
                                    metadata={
                                        "complete": False,
                                        "rows_read": len(rows),
                                        "bytes_read": bytes_read,
                                        "error_class": "csv_file_changed",
                                        "changed": True,
                                    },
                                )
                        except UnicodeDecodeError as exc:
                            return ToolResult(
                                success=False,
                                error=(
                                    f"Encoding error: {csv_encoding} does not match file content. "
                                    f"Try a different allowlisted encoding. ({exc})"
                                ),
                                metadata={
                                    "complete": False,
                                    "bytes_read": reader_io.bytes_read,
                                    "error_class": "csv_encoding_error",
                                },
                            )
                        except csv.Error as exc:
                            message = str(exc)
                            field_limit_hit = (
                                "field larger than field limit" in message.lower()
                                or "field limit" in message.lower()
                            )
                            return ToolResult(
                                success=False,
                                error=message,
                                metadata={
                                    "complete": False,
                                    "bytes_read": reader_io.bytes_read,
                                    "error_class": (
                                        "csv_field_limit_exceeded"
                                        if field_limit_hit
                                        else "csv_parse_error"
                                    ),
                                },
                            )
                        except ValueError as exc:
                            message = str(exc)
                            if "byte limit" in message:
                                return ToolResult(
                                    success=False,
                                    error=(
                                        "CSV content exceeds byte limit during parse "
                                        f"({limits.csv_read_max_bytes} bytes)"
                                    ),
                                    metadata={
                                        "complete": False,
                                        "rows_read": len(rows),
                                        "bytes_read": reader_io.bytes_read,
                                        "error_class": "csv_byte_budget_exceeded",
                                        "truncated": True,
                                    },
                                )
                            if (
                                "physical line exceeds" in message
                                or "pending buffer exceeds" in message
                            ):
                                return ToolResult(
                                    success=False,
                                    error=message,
                                    metadata={
                                        "complete": False,
                                        "rows_read": len(rows),
                                        "bytes_read": reader_io.bytes_read,
                                        "error_class": "csv_line_limit_exceeded",
                                    },
                                )
                            if "changed during read" in message:
                                return ToolResult(
                                    success=False,
                                    error="CSV file changed during read",
                                    metadata={
                                        "complete": False,
                                        "rows_read": len(rows),
                                        "bytes_read": reader_io.bytes_read,
                                        "error_class": "csv_file_changed",
                                        "changed": True,
                                    },
                                )
                            raise
                        finally:
                            reader_io.close()
                        if _cancelled():
                            raise asyncio.CancelledError
                        return _serialize_rows(
                            rows,
                            bytes_read=reader_io.bytes_read,
                            pending_high_water=reader_io.pending_high_water,
                            max_pending_chars=reader_io.max_pending_chars,
                        )
                    finally:
                        if fd >= 0:
                            os.close(fd)
                finally:
                    csv.field_size_limit(previous_limit)

        try:
            return await asyncio.to_thread(_parse)
        except asyncio.CancelledError:
            raise
        except UnicodeDecodeError as exc:
            return ToolResult(
                success=False,
                error=(
                    f"Encoding error: {csv_encoding} does not match file content "
                    f"({exc}). Try a different allowlisted encoding."
                ),
                metadata={"complete": False},
            )
        except Exception as e:
            msg = str(e)
            if "codec can't decode" in msg.lower() or "decode" in msg.lower():
                return ToolResult(
                    success=False,
                    error=f"Encoding error: {csv_encoding} does not match file content. {msg}",
                    metadata={"complete": False},
                )
            return ToolResult(success=False, error=msg, metadata={"complete": False})

    async def csv_write(
        self,
        path: str,
        data: str = "",
        encoding: str = "utf-8",
        delimiter: str = ",",
    ) -> ToolResult:
        guarded = self._write_path_or_error(path)
        if isinstance(guarded, ToolResult):
            return guarded

        try:
            import csv

            rows_data: list[list[Any]] = json.loads(data) if data else []
            if _is_work_runtime():
                rows_data = _normalize_work_cell_values(rows_data)
                from js_work.routines.office_safety import escape_work_csv_rows

                rows_data = escape_work_csv_rows(rows_data)
                target = self._resolve(path)
                target.parent.mkdir(parents=True, exist_ok=True)

                def _write_csv(staged: Path) -> None:
                    with staged.open("w", encoding=encoding, newline="") as handle:
                        writer = csv.writer(handle, delimiter=delimiter)
                        writer.writerows(rows_data)

                _publish_work_artifact(target, _write_csv, validate_xlsx=False)
            else:
                buf = StringIO()
                writer = csv.writer(buf, delimiter=delimiter)
                writer.writerows(_escape_formula_rows(rows_data))
                target = self._save_bytes(path, buf.getvalue().encode(encoding))

            return ToolResult(
                success=True,
                output=f"Written {len(rows_data)} rows to {path}",
                metadata={"path": str(target), "rows": len(rows_data)},
            )
        except Exception as e:
            return ToolResult(success=False, error=str(e))

    @staticmethod
    def _parse_cell(value: Any) -> Any:
        """Preserve native types (int, float, bool) where possible."""
        if value is None:
            return ""
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return value
        return str(value)

    async def excel_read(
        self,
        path: str,
        sheet: str | None = None,
        start_row: int = 0,
        end_row: int = 0,
        start_col: str = "",
        end_col: str = "",
    ) -> ToolResult:
        try:
            target = self._logical_input_target(path)
        except ValueError as e:
            return ToolResult(success=False, error=str(e))
        decision = self.guard.check_path_operation(str(target), "read")
        if decision.decision == SecurityDecisionType.BLOCK:
            return ToolResult(success=False, error=decision.reason)

        wb = None
        inputs = ExitStack()
        try:
            try:
                from openpyxl import load_workbook
            except ImportError as _e:
                raise ImportError(_OFFICE_EXTRA_MSG) from _e
            try:
                from openpyxl.utils import column_index_from_string
            except ImportError as _e:
                raise ImportError(_OFFICE_EXTRA_MSG) from _e

            if not _is_work_runtime() and not target.exists():
                return ToolResult(success=False, error=f"File not found: {path}")

            workbook_input = inputs.enter_context(self._materialized_input(path))
            if _is_work_runtime():
                _validate_work_xlsx(workbook_input)

            wb = load_workbook(workbook_input, data_only=True, read_only=True)
            ws = wb[sheet] if sheet else wb.active
            if ws is None:
                return ToolResult(success=False, error="Sheet not found")

            min_row = start_row if start_row > 0 else 1
            max_row = end_row if end_row > 0 else ws.max_row
            min_col = column_index_from_string(start_col) if start_col else 1
            max_col = column_index_from_string(end_col) if end_col else ws.max_column

            rows: list[list[Any]] = []
            for row in ws.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
                values_only=True,
            ):
                rows.append([self._parse_cell(cell) for cell in row])

            return ToolResult(
                success=True,
                output=json.dumps(rows, ensure_ascii=False, indent=2),
                metadata={"rows": len(rows), "columns": len(rows[0]) if rows else 0},
            )
        except KeyError as e:
            return ToolResult(success=False, error=f"Sheet not found: {e}")
        except Exception as e:
            return ToolResult(success=False, error=str(e))
        finally:
            if wb is not None:
                wb.close()
            inputs.close()

    @staticmethod
    def _parse_cell_ref(ref: str) -> tuple[str, str]:
        """Split Excel cell reference like 'A1' or 'BC123' into (col_letters, row_num)."""
        col = ""
        row = ""
        for ch in ref:
            if ch.isalpha():
                if row:
                    raise ValueError(f"Invalid cell reference: {ref}")
                col += ch
            elif ch.isdigit():
                row += ch
            else:
                raise ValueError(f"Invalid cell reference: {ref}")
        if not col or not row:
            raise ValueError(f"Invalid cell reference: {ref}")
        return col, row

    async def excel_write(
        self,
        path: str,
        sheet: str | None = None,
        data: str = "",
        start_cell: str = "A1",
        append: bool = False,
    ) -> ToolResult:
        guarded = self._write_path_or_error(path)
        if isinstance(guarded, ToolResult):
            return guarded

        wb = None
        try:
            try:
                from openpyxl import Workbook, load_workbook
            except ImportError as _e:
                raise ImportError(_OFFICE_EXTRA_MSG) from _e
            try:
                from openpyxl.utils import column_index_from_string
            except ImportError as _e:
                raise ImportError(_OFFICE_EXTRA_MSG) from _e

            sheet_name = sheet or "Sheet1"
            rows_data: list[list[Any]] = json.loads(data) if data else []
            if not isinstance(rows_data, list) or any(not isinstance(r, list) for r in rows_data):
                return ToolResult(success=False, error="data must be a JSON array of arrays")

            work_runtime = _is_work_runtime()
            if work_runtime:
                rows_data = _normalize_work_cell_values(rows_data)

            existing: bytes | None = None
            if not work_runtime:
                try:
                    existing, _, _, _ = self._files._secure_read_bytes(
                        path,
                        max_bytes=max(self.limits.file_read_max_chars * 4, 8 * 1024 * 1024),
                    )
                except FileNotFoundError:
                    existing = None

            if existing is not None:
                wb = load_workbook(BytesIO(existing))
            else:
                wb = Workbook()
                # openpyxl default sheet title may be "Sheet" (not "Sheet1").
                if wb.active is not None:
                    wb.active.title = sheet_name

            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(title=sheet_name)

            col_letter, row_num_str = self._parse_cell_ref(start_cell)
            start_col = column_index_from_string(col_letter)
            start_row_num = int(row_num_str)

            if append:
                start_row_num = max(start_row_num, ws.max_row + 1)

            for r_idx, row in enumerate(rows_data, start=start_row_num):
                for c_idx, value in enumerate(row, start=start_col):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    if work_runtime:
                        _write_work_excel_cell(cell, value)
                    else:
                        cell.value = _escape_formula_text(value)

            if work_runtime:
                target = self._resolve(path)
                target.parent.mkdir(parents=True, exist_ok=True)
                _publish_work_artifact(
                    target,
                    lambda staged: wb.save(str(staged)),
                    validate_xlsx=True,
                )
                content_sha256 = hashlib.sha256(target.read_bytes()).hexdigest()
            else:
                target, content_sha256 = self._save_workbook(path, wb)

            return ToolResult(
                success=True,
                output=f"Written {len(rows_data)} rows to {path}",
                metadata={
                    "path": str(target),
                    "rows": len(rows_data),
                    "content_sha256": content_sha256,
                },
            )
        except ValueError as e:
            return ToolResult(success=False, error=str(e))
        except Exception as e:
            return ToolResult(success=False, error=str(e))
        finally:
            if wb is not None:
                wb.close()

    async def excel_merge(
        self,
        source_path: str,
        target_path: str,
        output_path: str = "",
        source_sheet: str | None = None,
        target_sheet: str | None = None,
        source_range: str = "",
        target_start_cell: str = "A1",
        include_header: bool = True,
    ) -> ToolResult:
        try:
            source = self._logical_input_target(source_path)
            target = self._logical_input_target(target_path)
            if output_path:
                write_check = self._write_path_or_error(output_path)
                if isinstance(write_check, ToolResult):
                    return write_check
                output = self._files._logical_path(output_path)
            else:
                output = None
                write_check = self._write_path_or_error(target_path)
                if isinstance(write_check, ToolResult):
                    return write_check
        except ValueError as e:
            return ToolResult(success=False, error=str(e))
        work_runtime = _is_work_runtime()
        if work_runtime and output is None:
            return ToolResult(
                success=False,
                error="output_path is required so JS Agent Work never overwrites the target workbook",
            )
        path_operations = [(source, "read"), (target, "read" if work_runtime else "write")]
        if output is not None:
            path_operations.append((output, "write"))
        for p, op in path_operations:
            decision = self.guard.check_path_operation(str(p), op)
            if decision.decision == SecurityDecisionType.BLOCK:
                return ToolResult(success=False, error=decision.reason)

        src_wb = None
        tgt_wb = None
        inputs = ExitStack()
        try:
            try:
                from openpyxl import load_workbook
            except ImportError as _e:
                raise ImportError(_OFFICE_EXTRA_MSG) from _e
            try:
                from openpyxl.utils import column_index_from_string, range_boundaries
            except ImportError as _e:
                raise ImportError(_OFFICE_EXTRA_MSG) from _e

            if not work_runtime and not source.exists():
                return ToolResult(success=False, error=f"Source not found: {source_path}")
            if not work_runtime and not target.exists():
                return ToolResult(success=False, error=f"Target not found: {target_path}")

            source_input = inputs.enter_context(self._materialized_input(source_path))
            target_input = inputs.enter_context(self._materialized_input(target_path))
            if work_runtime:
                assert output is not None
                _validate_work_xlsx(source_input)
                _validate_work_xlsx(target_input)

            src_wb = load_workbook(source_input, data_only=True, read_only=True)
            src_ws = src_wb[source_sheet] if source_sheet else src_wb.active
            if src_ws is None:
                return ToolResult(success=False, error="Source sheet not found")

            tgt_wb = load_workbook(target_input)
            tgt_ws = tgt_wb[target_sheet] if target_sheet else tgt_wb.active
            if tgt_ws is None:
                return ToolResult(success=False, error="Target sheet not found")

            # Determine source data range
            if source_range:
                min_col, min_row, max_col, max_row = range_boundaries(source_range)
            else:
                min_row, min_col = 1, 1
                max_row = src_ws.max_row
                max_col = src_ws.max_column

            if not include_header and min_row == 1 and max_row is not None and max_row > 1:
                min_row += 1

            col_letter, row_num_str = self._parse_cell_ref(target_start_cell)
            tgt_start_col = column_index_from_string(col_letter)
            tgt_start_row = int(row_num_str)

            rows_copied = 0
            for r_idx, row in enumerate(
                src_ws.iter_rows(
                    min_row=min_row,
                    max_row=max_row,
                    min_col=min_col,
                    max_col=max_col,
                    values_only=True,
                ),
                start=tgt_start_row,
            ):
                for c_idx, value in enumerate(row, start=tgt_start_col):
                    cell = tgt_ws.cell(row=r_idx, column=c_idx)
                    if work_runtime:
                        _write_work_excel_cell(cell, value)
                    else:
                        cell.value = _escape_formula_text(value)
                rows_copied += 1

            if work_runtime:
                assert output is not None
                _publish_work_artifact(
                    output,
                    lambda staged: tgt_wb.save(str(staged)),
                    validate_xlsx=True,
                )
            else:
                write_rel = output_path if output_path else target_path
                written, _digest = self._save_workbook(write_rel, tgt_wb)
                if output is not None:
                    output = written

            return ToolResult(
                success=True,
                output=(
                    f"Merged {rows_copied} rows from {source_path} into "
                    f"{output_path or target_path} at {target_start_cell}"
                ),
                metadata={
                    "rows_copied": rows_copied,
                    **({"output_path": str(output)} if output is not None else {}),
                },
            )
        except ValueError as e:
            return ToolResult(success=False, error=str(e))
        except Exception as e:
            return ToolResult(success=False, error=str(e))
        finally:
            if src_wb is not None:
                src_wb.close()
            if tgt_wb is not None:
                tgt_wb.close()
            inputs.close()

    async def excel_create(
        self,
        path: str,
        sheet_name: str = "Sheet1",
        headers: str = "",
    ) -> ToolResult:
        guarded = self._write_path_or_error(path)
        if isinstance(guarded, ToolResult):
            return guarded

        wb = None
        try:
            try:
                from openpyxl import Workbook
            except ImportError as _e:
                raise ImportError(_OFFICE_EXTRA_MSG) from _e

            wb = Workbook()
            ws = wb.active
            if ws is None:
                ws = wb.create_sheet(title=sheet_name)
            else:
                ws.title = sheet_name

            if headers:
                hdrs: list[Any] = json.loads(headers)
                if not isinstance(hdrs, list):
                    return ToolResult(success=False, error="headers must be a JSON array")
                work_runtime = _is_work_runtime()
                if work_runtime:
                    hdrs = _normalize_work_cell_values(hdrs)
                for c_idx, h in enumerate(hdrs, start=1):
                    cell = ws.cell(row=1, column=c_idx)
                    if work_runtime:
                        _write_work_excel_cell(cell, h)
                    else:
                        cell.value = _escape_formula_text(h)

            if _is_work_runtime():
                target = self._resolve(path)
                target.parent.mkdir(parents=True, exist_ok=True)
                _publish_work_artifact(
                    target,
                    lambda staged: wb.save(str(staged)),
                    validate_xlsx=True,
                )
            else:
                self._save_workbook(path, wb)

            return ToolResult(success=True, output=f"Created Excel file: {path}")
        except Exception as e:
            return ToolResult(success=False, error=str(e))
        finally:
            if wb is not None:
                wb.close()

    async def pdf_generate(
        self,
        path: str,
        title: str = "",
        data: str = "",
        page_size: str = "A4",
    ) -> ToolResult:
        guarded = self._write_path_or_error(path)
        if isinstance(guarded, ToolResult):
            return guarded

        try:
            try:
                from reportlab.lib import colors
                from reportlab.lib.pagesizes import A4, LETTER
                from reportlab.lib.styles import getSampleStyleSheet
                from reportlab.platypus import (
                    Paragraph,
                    SimpleDocTemplate,
                    Spacer,
                    Table,
                    TableStyle,
                )
            except ImportError as _e:
                raise ImportError(_PDF_EXTRA_MSG) from _e

            rows_data: list[list[Any]] = json.loads(data) if data else []
            if not rows_data:
                return ToolResult(success=False, error="No data provided")

            size = A4 if page_size.upper() == "A4" else LETTER

            elements: list[Any] = []
            styles = getSampleStyleSheet()

            if title:
                elements.append(Paragraph(title, styles["Title"]))
                elements.append(Spacer(1, 12))

            # Build table
            table = Table(rows_data)
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f0f0f0")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 10),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                        ("BACKGROUND", (0, 1), (-1, -1), colors.white),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        ("FONTSIZE", (0, 1), (-1, -1), 9),
                    ]
                )
            )
            elements.append(table)
            if _is_work_runtime():
                target = self._resolve(path)
                target.parent.mkdir(parents=True, exist_ok=True)

                def _write_pdf(staged: Path) -> None:
                    SimpleDocTemplate(str(staged), pagesize=size).build(elements)

                _publish_work_artifact(target, _write_pdf, validate_xlsx=False)
            else:
                buffer = BytesIO()
                SimpleDocTemplate(buffer, pagesize=size).build(elements)
                self._save_bytes(path, buffer.getvalue())

            return ToolResult(success=True, output=f"Generated PDF: {path}")
        except Exception as e:
            return ToolResult(success=False, error=str(e))

    def register_all(self, registry: Any) -> None:
        """Register all office tools."""
        for spec in self.get_specs():
            if spec.name == "csv_read":
                registry.register(spec, self.csv_read)
            elif spec.name == "csv_write":
                registry.register(spec, self.csv_write)
            elif spec.name == "excel_read":
                registry.register(spec, self.excel_read)
            elif spec.name == "excel_write":
                registry.register(spec, self.excel_write)
            elif spec.name == "excel_merge":
                registry.register(spec, self.excel_merge)
            elif spec.name == "excel_create":
                registry.register(spec, self.excel_create)
            elif spec.name == "pdf_generate":
                registry.register(spec, self.pdf_generate)
