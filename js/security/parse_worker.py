"""Isolated document-parser worker. Invoked only by bounded_parse."""

from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any

_UNSAFE_PDF_KEYS = {
    "/AA",
    "/EmbeddedFiles",
    "/EF",
    "/JS",
    "/JavaScript",
    "/Launch",
    "/OpenAction",
    "/RichMedia",
    "/SubmitForm",
    "/URI",
    "/XFA",
}
_UNSAFE_PDF_ACTIONS = {
    "/GoToE",
    "/GoToR",
    "/ImportData",
    "/JavaScript",
    "/Launch",
    "/Rendition",
    "/RichMediaExecute",
    "/SubmitForm",
    "/URI",
}
_UNSAFE_PDF_SUBTYPES = {
    "/3D",
    "/FileAttachment",
    "/Movie",
    "/RichMedia",
    "/Screen",
    "/Sound",
}
_MAX_PDF_OBJECTS = 100_000


def _extract_pdf(path: Path, max_pages: int, max_output: int) -> str:
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    texts: list[str] = []
    total = 0
    for index, page in enumerate(reader.pages):
        if index >= max_pages:
            break
        text = page.extract_text() or ""
        if not text:
            continue
        remaining = max_output - total
        if remaining <= 0:
            break
        chunk = text[:remaining]
        texts.append(chunk)
        total += len(chunk.encode("utf-8"))
        if total >= max_output:
            break
    return "\n".join(texts)


def _extract_xlsx(path: Path, max_cells: int, max_output: int) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(str(path), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        if sheet is None:
            return ""
        lines: list[str] = []
        cells = 0
        total = 0
        for row in sheet.iter_rows(values_only=True):
            values = ["" if cell is None else str(cell) for cell in row]
            cells += len(values)
            if cells > max_cells:
                break
            line = "\t".join(values)
            encoded = (line + "\n").encode("utf-8")
            if total + len(encoded) > max_output:
                break
            lines.append(line)
            total += len(encoded)
        return "\n".join(lines)
    finally:
        workbook.close()


def _validate_pdf_active_content(reader: Any) -> None:
    stack: list[Any] = [reader.trailer.get("/Root")]
    seen: set[int] = set()
    visited = 0
    while stack:
        item = stack.pop()
        if item is None:
            continue
        if not isinstance(item, (dict, list, tuple, str, bytes, int, float, bool)):
            getter = getattr(item, "get_object", None)
            if callable(getter):
                item = getter()
        marker = id(item)
        if marker in seen:
            continue
        seen.add(marker)
        visited += 1
        if visited > _MAX_PDF_OBJECTS:
            raise ValueError("PDF active content graph exceeds the safety limit")
        if isinstance(item, dict):
            for raw_key, value in item.items():
                key = str(raw_key)
                if key in _UNSAFE_PDF_KEYS:
                    raise ValueError("PDF active content is not supported")
                if key == "/S" and str(value) in _UNSAFE_PDF_ACTIONS:
                    raise ValueError("PDF active content is not supported")
                if key == "/Subtype" and str(value) in _UNSAFE_PDF_SUBTYPES:
                    raise ValueError("PDF active content is not supported")
                stack.append(value)
        elif isinstance(item, (list, tuple)):
            stack.extend(item)


def _extract_work_pdf(path: Path, max_pages: int, max_output: int) -> dict[str, Any]:
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    if reader.is_encrypted:
        return {"ok": False, "error": "encrypted PDFs are not supported"}
    try:
        _validate_pdf_active_content(reader)
    except ValueError as exc:
        return {"ok": False, "error": str(exc)}
    pages: list[dict[str, Any]] = []
    remaining = max_output
    total_pages = len(reader.pages)
    truncated = total_pages > max_pages
    for index, page in enumerate(reader.pages[:max_pages], start=1):
        text = (page.extract_text() or "").strip()
        if len(text) > remaining:
            text = text[:remaining]
            truncated = True
        pages.append({"page": index, "text": text})
        remaining -= len(text)
        if remaining <= 0:
            if index < total_pages:
                truncated = True
            break
    return {
        "ok": True,
        "page_count": total_pages,
        "pages": pages,
        "truncated": truncated,
    }


def main(argv: list[str]) -> int:
    if len(argv) != 6:
        return 2
    _prog, kind, snapshot, max_pages_s, max_cells_s, max_output_s = argv
    try:
        max_pages = int(max_pages_s)
        max_cells = int(max_cells_s)
        max_output = int(max_output_s)
    except ValueError:
        return 2
    path = Path(snapshot)
    if not path.is_file():
        return 2
    try:
        if kind == "pdf":
            payload = _extract_pdf(path, max_pages, max_output).encode("utf-8")
        elif kind == "xlsx":
            payload = _extract_xlsx(path, max_cells, max_output).encode("utf-8")
        elif kind == "work_pdf":
            payload = json.dumps(
                _extract_work_pdf(path, max_pages, max_output),
                ensure_ascii=False,
            ).encode("utf-8")
        else:
            return 2
    except Exception:
        return 1
    if len(payload) > max_output:
        payload = payload[:max_output]
    sys.stdout.buffer.write(payload)
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
