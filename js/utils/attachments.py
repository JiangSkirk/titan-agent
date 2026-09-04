"""Attachment parsing helpers extracted from JSAgent."""

from __future__ import annotations

from pathlib import Path
from typing import Any, BinaryIO, cast

from js.utils.log import get_logger

logger = get_logger("js.core.attachments")

# Fail-closed extraction limits for hostile or oversized attachments.
MAX_PDF_PAGES = 200
MAX_EXCEL_ROWS = 10_000
MAX_EXCEL_TEXT_BYTES = 5 * 1024 * 1024  # 5 MB of accumulated cell text


class AttachmentLimitError(ValueError):
    """Raised when attachment content exceeds an extraction safety limit."""


def format_size(size_bytes: int) -> str:
    """Format byte size to human-readable string."""
    size = float(size_bytes)
    for unit in ["B", "KB", "MB", "GB"]:
        if size < 1024:
            return f"{size:.1f} {unit}"
        size /= 1024
    return f"{size:.1f} TB"


def _library_source(source: Path | BinaryIO) -> str | BinaryIO:
    return str(source) if isinstance(source, Path) else source


def _rewind(source: Path | BinaryIO) -> None:
    if not isinstance(source, Path):
        source.seek(0)


def extract_pdf_text(path: Path | BinaryIO) -> str:
    """Best-effort PDF-to-text extraction.

    Fail-closed: raises :class:`AttachmentLimitError` when the document exceeds
    ``MAX_PDF_PAGES`` pages instead of silently truncating.
    """
    try:
        from pypdf import PdfReader
    except ImportError as _e:
        raise ImportError("Install js-agent[pdf] to extract PDF text.") from _e
    try:
        _rewind(path)
        reader = PdfReader(_library_source(path))
        page_count = len(reader.pages)
        if page_count > MAX_PDF_PAGES:
            raise AttachmentLimitError(
                f"PDF exceeds page limit ({page_count} > {MAX_PDF_PAGES} pages)"
            )
        texts: list[str] = []
        for page in reader.pages:
            text = page.extract_text()
            if text:
                texts.append(text)
        return "\n".join(texts)
    except AttachmentLimitError:
        raise
    except Exception:
        logger.warning("PDF extraction failed", exc_info=True)
    try:
        import pdfplumber
    except ImportError as _e:
        raise ImportError("Install js-agent[pdf] to extract PDF text.") from _e
    try:
        texts2: list[str] = []
        _rewind(path)
        with pdfplumber.open(cast("Any", _library_source(path))) as pdf:
            page_count = len(pdf.pages)
            if page_count > MAX_PDF_PAGES:
                raise AttachmentLimitError(
                    f"PDF exceeds page limit ({page_count} > {MAX_PDF_PAGES} pages)"
                )
            for pdf_page in pdf.pages:
                text = pdf_page.extract_text()
                if text:
                    texts2.append(text)
        return "\n".join(texts2)
    except AttachmentLimitError:
        raise
    except Exception:
        return ""


def extract_excel_text(path: Path | BinaryIO) -> str:
    """Best-effort Excel-to-text extraction with smart header detection.

    Fail-closed: raises :class:`AttachmentLimitError` when the sheet exceeds
    ``MAX_EXCEL_ROWS`` rows or ``MAX_EXCEL_TEXT_BYTES`` of cell text instead of
    silently truncating.
    """
    try:
        import pandas as pd
    except ImportError as _e:
        raise ImportError("Install js-agent[office] to extract Excel text.") from _e
    try:
        _rewind(path)
        df = pd.read_excel(
            _library_source(path),
            header=None,
            engine="openpyxl",
            nrows=MAX_EXCEL_ROWS + 1,
        )
        if len(df) > MAX_EXCEL_ROWS:
            raise AttachmentLimitError(
                f"Excel exceeds row limit ({len(df)} > {MAX_EXCEL_ROWS} rows)"
            )
        if len(df) == 0:
            return "(空表格)"

        header_rows: list[int] = []
        data_start = 0
        for i in range(min(20, len(df))):
            row = df.iloc[i]
            first_val = row.iloc[0] if len(row) > 0 else None
            is_data_start = False
            if pd.notna(first_val):
                try:
                    if (
                        isinstance(first_val, (int, float))
                        and first_val > 0
                        and first_val == int(first_val)
                    ):
                        is_data_start = True
                except (ValueError, TypeError):
                    logger.warning("Header detection numeric parse failed", exc_info=True)
            if is_data_start and i > 1:
                data_start = i
                break
            str_vals = [
                str(v) for v in row if pd.notna(v) and isinstance(v, str) and str(v).strip()
            ]
            non_null_count = sum(1 for v in row if pd.notna(v))
            total_cols = len(row)
            first_col_str = str(first_val).strip() if pd.notna(first_val) else ""
            if len(str_vals) >= 2:
                is_meta = (
                    (non_null_count == 1 and len(first_col_str) > 30)
                    or (len(first_col_str) > 40)
                    or (non_null_count / total_cols < 0.20)
                )
                if not is_meta:
                    header_rows.append(i)

        if data_start == 0:
            data_start = max(header_rows) + 1 if header_rows else 0

        if data_start > 0:
            header_rows = [h for h in header_rows if h < data_start]
            for candidate in range(max(0, data_start - 3), data_start):
                if candidate not in header_rows:
                    non_null = [v for v in df.iloc[candidate] if pd.notna(v)]
                    if len(non_null) >= 3:
                        header_rows.append(candidate)
            header_rows = sorted(header_rows)

        headers: list[str] = []
        seen_names: dict[str, int] = {}
        for col in range(len(df.columns)):
            parts: list[str] = []
            for hr in header_rows:
                val = df.iloc[hr, col]
                if pd.notna(val):
                    s = str(val).strip()
                    if s and s.lower() not in ("nan", "none"):
                        parts.append(s)
            seen: set[str] = set()
            unique: list[str] = []
            for p in parts:
                if p not in seen:
                    seen.add(p)
                    unique.append(p)
            name = " / ".join(unique) if unique else f"Col_{col}"
            if name in seen_names:
                seen_names[name] += 1
                name = f"{name} ({seen_names[name]})"
            else:
                seen_names[name] = 0
            headers.append(name)

        data = df.iloc[data_start:].copy()
        if len(data) == 0:
            return "(无数据行)"
        data.columns = headers

        first_col = data.iloc[:, 0] if len(data.columns) > 0 else pd.Series()
        non_null = first_col.dropna()
        if len(non_null) > 0:
            numeric_looking = 0
            for v in non_null:
                s = str(v).strip()
                if (
                    s.replace(".", "", 1)
                    .replace(",", "")
                    .replace("%", "")
                    .replace("$", "")
                    .replace("-", "", 1)
                    .isdigit()
                ):
                    numeric_looking += 1
            numeric_ratio = numeric_looking / len(non_null)
            if numeric_ratio > 0.5:
                mask = first_col.apply(
                    lambda x: (
                        pd.isna(x)
                        or str(x)
                        .strip()
                        .replace(".", "", 1)
                        .replace(",", "")
                        .replace("%", "")
                        .replace("$", "")
                        .replace("-", "", 1)
                        .isdigit()
                    )
                )
                data = data[mask]

        data = data.dropna(how="all")

        if len(data) == 0:
            return "(无数据行)"

        empty_indices = [i for i in range(len(data.columns)) if data.iloc[:, i].isna().all()]
        if empty_indices:
            data = data.drop(data.columns[empty_indices], axis=1)

        sparse_indices = [
            i
            for i in range(len(data.columns))
            if str(data.columns[i]).startswith("Col_") and data.iloc[:, i].isna().mean() > 0.95
        ]
        if sparse_indices:
            data = data.drop(data.columns[sparse_indices], axis=1)

        max_rows = 50
        lines: list[str] = []
        lines.append(f"表格: {len(data)} 行 x {len(data.columns)} 列")
        lines.append(f"列名: {', '.join(str(c) for c in data.columns)}")
        lines.append("")
        display_df = data.head(max_rows).fillna("")
        lines.append(display_df.to_string(index=False))
        if len(data) > max_rows:
            lines.append("")
            lines.append(f"... 共 {len(data)} 行，以上显示前 {max_rows} 行")

        lines.append("")
        lines.append("数据汇总:")
        for keyword in ("QTY", "PCS", "G.W", "N.W", "VOLUME", "AMOUNT", "TOTAL"):
            for actual_col in data.columns:
                if keyword in str(actual_col).upper():
                    total = pd.to_numeric(data[actual_col], errors="coerce").sum()
                    if pd.notna(total) and total > 0:
                        lines.append(f"  {actual_col} 合计 = {total:.0f}")

        for keyword in ("STYLE", "COLOR", "SIZE", "MODEL"):
            cols = [c for c in data.columns if keyword in str(c).upper()]
            for c in cols:
                uniq = data[c].dropna().unique()
                if len(uniq) > 0 and len(uniq) <= 20:
                    lines.append(
                        f"  {c}: {', '.join(str(x) for x in uniq[:10])}"
                        + (f" ...等{len(uniq)}种" if len(uniq) > 10 else "")
                    )

        result = "\n".join(lines)
        result_bytes = len(result.encode("utf-8"))
        if result_bytes > MAX_EXCEL_TEXT_BYTES:
            raise AttachmentLimitError(
                f"Excel text exceeds byte limit ({result_bytes} > "
                f"{MAX_EXCEL_TEXT_BYTES} bytes)"
            )
        return result
    except AttachmentLimitError:
        raise
    except Exception:
        logger.warning("Excel extraction failed", exc_info=True)

    try:
        from openpyxl import load_workbook
    except ImportError as _e:
        raise ImportError("Install js-agent[office] to extract Excel text.") from _e
    try:
        _rewind(path)
        wb = load_workbook(_library_source(path), data_only=True, read_only=True)
        ws = wb.active
        if ws is None:
            return ""
        lines2: list[str] = []
        total_bytes = 0
        for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_index > MAX_EXCEL_ROWS:
                raise AttachmentLimitError(
                    f"Excel exceeds row limit ({row_index} > {MAX_EXCEL_ROWS} rows)"
                )
            line = "\t".join(str(c) if c is not None else "" for c in row)
            total_bytes += len(line.encode("utf-8")) + 1
            if total_bytes > MAX_EXCEL_TEXT_BYTES:
                raise AttachmentLimitError(
                    f"Excel text exceeds byte limit ({total_bytes} > "
                    f"{MAX_EXCEL_TEXT_BYTES} bytes)"
                )
            lines2.append(line)
        return "\n".join(lines2)
    except AttachmentLimitError:
        raise
    except Exception:
        return ""
