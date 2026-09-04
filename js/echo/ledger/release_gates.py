from __future__ import annotations

import ast
import base64
import binascii
import hashlib
import hmac
import importlib.metadata
import io
import json
import math
import os
import platform
import re
import shutil
import stat
import subprocess
import sys
import tarfile
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from datetime import UTC, date, datetime
from functools import lru_cache
from pathlib import Path
from statistics import median
from typing import Any

from cryptography.exceptions import InvalidSignature
from cryptography.hazmat.primitives.asymmetric.ed25519 import Ed25519PublicKey

from js.echo.ledger.sandbox_backend import EchoSandboxBackend
from js.echo.ledger.security_matrix import run_security_matrix
from js.echo.ledger.strict_json import (
    StrictJSONError,
    strict_load_object,
    strict_load_object_bytes,
    strict_loads,
)
from js.echo.release_probes import run_echo_release_probes

from .slo_contract import SLO_CONTRACT


def _safe_finite_float(value: object) -> float | None:
    """Coerce a JSON number; never raise on huge ints / non-finite values."""
    if isinstance(value, bool) or not isinstance(value, int | float):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError, OverflowError):
        return None
    if not math.isfinite(number):
        return None
    return number


@dataclass(frozen=True)
class ReleaseReadinessReport:
    internal_ready: bool
    stable_ready: bool
    passed: tuple[str, ...]
    internal_blockers: tuple[str, ...]
    external_blockers: tuple[str, ...]


@dataclass(frozen=True)
class EchoIPBoundaryReport:
    ok: bool
    findings: tuple[str, ...]


_INTERNAL_EVIDENCE = {
    "origin_ledger": Path("ORIGIN_LEDGER.md"),
    "third_party_notices": Path("THIRD_PARTY_NOTICES.md"),
    "codeowners": Path(".github/CODEOWNERS"),
    "adr": Path("docs/adr/0001-echo-ledger-boundary.md"),
    "rfc_template": Path("docs/rfc/echo-ledger-major-change-template.md"),
    "echo_self_developed_boundary": Path("docs/echo/ECHO_SELF_DEVELOPED_BOUNDARY.md"),
    "echo_unified_execution_contract": Path("docs/echo/ECHO_UNIFIED_EXECUTION_CONTRACT.md"),
}

_EXTERNAL_EVIDENCE = {
    "legal_fto_review": Path("docs/security/LEGAL_FTO_REVIEW.md"),
    "clean_room_reviewer": Path("docs/security/CLEAN_ROOM_REVIEW.md"),
    "external_security_audit": Path("docs/security/EXTERNAL_SECURITY_AUDIT.md"),
    "redteam_report": Path("docs/security/REDTEAM_REPORT.md"),
}

_EXTERNAL_APPROVAL_FIELDS = (
    "Status",
    "Reviewer",
    "Date",
    "Evidence-Name",
    "Scope-Commit",
    "SBOM-SHA256",
    "UV-Lock-SHA256",
    "Approval-Reference",
    "Signature",
)
_SIGNED_EXTERNAL_APPROVAL_FIELDS = _EXTERNAL_APPROVAL_FIELDS[:-1]
_SCOPE_COMMIT_PATTERN = re.compile(r"[0-9a-f]{40}", re.IGNORECASE)
_SHA256_PATTERN = re.compile(r"[0-9a-f]{64}", re.IGNORECASE)
_EXTERNAL_APPROVAL_CANONICAL_VERSION = "JS-ECHO-EXTERNAL-APPROVAL-V1"

_STABLE_ARTIFACT_EVIDENCE = {
    "sbom_spdx": Path("docs/security/SBOM.spdx.json"),
    "license_scan": Path("docs/security/LICENSE_SCAN.md"),
    "echo_slo_benchmark": Path("docs/security/ECHO_SLO_BENCHMARK.json"),
}
_LIVE_ACCEPTANCE_EVIDENCE = Path("docs/security/ECHO_LIVE_ACCEPTANCE.json")
_ECHO_BASELINE_COMMIT = "65cc545e3ec893f5bab62d356514643f14456a58"
_ECHO_BASELINE_PROVENANCE_SCHEMA = "echo-old-baseline-provenance-v2"
_ECHO_BASELINE_TOKENIZER_METHOD = "tiktoken_cl100k_base_canonical_json"
_ECHO_BASELINE_TOKENIZER_ENCODING = "cl100k_base"
_ECHO_BASELINE_HISTORY_MESSAGES = 40
_ECHO_BASELINE_HISTORY_WORDS_PER_MESSAGE = 80
_ECHO_BASELINE_HISTORY_MARKER_PREFIX = "benchmark long history message "
_RELEASE_SOURCE_DIGEST_VERSION = b"ECHO-RELEASE-SOURCE-V2\0"
_RELEASE_SOURCE_SURFACE_META_VERSION = b"ECHO-RELEASE-SURFACE-META-V1\0"
_TOKENIZER_TREE_DIGEST_VERSION = b"ECHO-TOKENIZER-TREE-V1\0"
# Directory roots only once — do not also list nested files already covered by rglob.
_RELEASE_SOURCE_DIGEST_SURFACES = (
    Path(".github"),
    Path(".gitignore"),
    Path("Dockerfile"),
    Path("LICENSE"),
    Path("ORIGIN_LEDGER.md"),
    Path("README.md"),
    Path("THIRD_PARTY_NOTICES.md"),
    Path("benchmarks"),
    Path("docs/adr/0001-echo-ledger-boundary.md"),
    Path("docs/echo/ECHO_10_ROUND_AUDIT.md"),
    Path("docs/echo/ECHO_FINAL_REPLACEMENT_REPORT.md"),
    Path("docs/echo/ECHO_SELF_DEVELOPED_BOUNDARY.md"),
    Path("docs/echo/ECHO_UNIFIED_EXECUTION_CONTRACT.md"),
    Path("docs/rfc/echo-ledger-major-change-template.md"),
    Path("docs/security/ECHO_BASELINE_65CC545.json"),
    Path("docs/security/ECHO_E2E_LEDGER_PUBKEY.json"),
    Path("docs/security/ECHO_SOAK_INTEGRITY_PUBKEY.json"),
    Path("docs/security/LICENSE_SCAN.md"),
    Path("docs/security/SBOM.spdx.json"),
    Path("desktop"),
    Path("js"),
    Path("js_work"),
    Path("pyproject.toml"),
    Path("resources"),
    Path("scripts"),
    Path("tests"),
    Path("uv.lock"),
)
_RELEASE_SOURCE_DIGEST_EXCLUDE = frozenset(
    {
        # Evidence artifacts that bind TO this digest - exclude to avoid circular hash.
        Path("docs/security/ECHO_LIVE_ACCEPTANCE.json"),
        Path("docs/security/ECHO_SLO_BENCHMARK.json"),
        Path("docs/security/ECHO_ISOLATED_VENV_E2E.json"),
        # Generated audit reports bind the SLO hash; keep them out of the runtime digest
        # while still validating them via the independent audit-report gate.
        Path("docs/echo/ECHO_10_ROUND_AUDIT.md"),
        Path("docs/echo/ECHO_FINAL_REPLACEMENT_REPORT.md"),
    }
)
_RELEASE_SOURCE_ALLOWED_EMPTY_FILES = frozenset({Path("tests/echo/__init__.py")})
_RELEASE_SOURCE_BINARY_SUFFIXES = frozenset({".icns", ".png", ".ttf", ".woff2"})
_RELEASE_SOURCE_DESKTOP_GENERATED_PARTS = frozenset(
    {
        ".cache",
        ".mypy_cache",
        ".pytest_cache",
        ".ruff_cache",
        "__pycache__",
        "binaries",
        "cache",
        "caches",
        "node_modules",
        "target",
    }
)
_RELEASE_SOURCE_DESKTOP_GENERATED_ROOTS = (Path("desktop/src-tauri/gen"),)
_ISOLATED_VENV_E2E_EVIDENCE = Path("docs/security/ECHO_ISOLATED_VENV_E2E.json")
_ISOLATED_VENV_E2E_SCHEMA_VERSION = "isolated-venv-e2e-v8"
_WORK_LEDGER_CHAIN_TYPES: tuple[str, ...] = (
    "permit",
    "outbox",
    "outbox_claimed",
    "receipt",
    "merge",
)
_ISOLATED_VENV_E2E_SERVER_STEP = "server HTTP+WS+attachment+work E2E"
_ISOLATED_VENV_E2E_REQUIRED_STEPS: tuple[str, ...] = (
    "build: python -m build --no-isolation (wheel+sdist)",
    "wheel: create venv",
    "wheel: pip install artifact offline (echo-tokenizer,office)",
    "wheel: pip check",
    "wheel: import js/js_work from venv site-packages",
    "wheel: tokenizer loads offline from vendored cache",
    "wheel: CLI js --help",
    "wheel: CLI js work --help",
    "wheel: CLI js-work --help",
    "wheel: CLI python -m js_work --help",
    f"wheel: {_ISOLATED_VENV_E2E_SERVER_STEP}",
    "sdist: create venv",
    "sdist: pip install build backends offline",
    "sdist: pip install artifact offline (echo-tokenizer,office)",
    "sdist: pip check",
    "sdist: import js/js_work from venv site-packages",
    "sdist: tokenizer loads offline from vendored cache",
    "sdist: CLI js --help",
    "sdist: CLI js work --help",
    "sdist: CLI js-work --help",
    "sdist: CLI python -m js_work --help",
    f"sdist: {_ISOLATED_VENV_E2E_SERVER_STEP}",
)
_AUDIT_REPORT_EVIDENCE = (
    Path("docs/echo/ECHO_10_ROUND_AUDIT.md"),
    Path("docs/echo/ECHO_FINAL_REPLACEMENT_REPORT.md"),
)
_RELEASE_CANDIDATE_SURFACES = (
    Path(".gitattributes"),
    Path(".gitignore"),
    Path(".gitmodules"),
    Path(".github/workflows/ci.yml"),
    Path(".github/workflows/release-smoke.yml"),
    Path("README.md"),
    Path("js"),
    Path("js_work"),
    Path("pyproject.toml"),
    Path("scripts/generate_release_evidence.py"),
    Path("scripts/release_smoke.py"),
    Path("scripts/verify_installed_artifact.py"),
    Path("uv.lock"),
    *_INTERNAL_EVIDENCE.values(),
    *_STABLE_ARTIFACT_EVIDENCE.values(),
    _LIVE_ACCEPTANCE_EVIDENCE,
    *_AUDIT_REPORT_EVIDENCE,
)

_ECHO_IP_CODE_ROOTS = (Path("js/echo"),)
_ECHO_IP_CODE_ALLOWLIST = (Path("js/echo/ledger/release_gates.py"),)
_scan_root: Path = Path(".")
_ECHO_IP_TEXT_ROOTS = (
    Path("ORIGIN_LEDGER.md"),
    Path("THIRD_PARTY_NOTICES.md"),
    Path("docs/echo"),
    Path("docs/security"),
)
_CLEAN_ROOM_AVOIDANCE_DOC = Path("docs/security/ECHO_2_CLEAN_ROOM.md")
_DISALLOWED_PROJECT_API_TOKENS = (
    "StateGraph",
    "MessagesState",
    "START",
    "END",
    "add_node",
    "add_edge",
    "compile",
    "AgentWorkflow",
    "AgentChat",
    "AssistantAgent",
    "UserProxyAgent",
    "RoundRobinGroupChat",
    "SelectorGroupChat",
    "function_tool",
    "handoff",
    "guardrail",
    "tripwire",
    "RunResult",
    "Workflow",
    "StartEvent",
    "StopEvent",
    "FunctionAgent",
    "CodeAgent",
    "ToolCallingAgent",
    "MultiStepAgent",
    "LocalPythonExecutor",
    "EventStream",
    "ActionEvent",
    "ObservationEvent",
    "Conversation",
    "Runtime",
    "action-observation",
)
_UNVERIFIABLE_SELF_DEVELOPED_CLAIMS = (
    "100% 自研",
    "100%自研",
    "完全自研",
    "全部自研",
    "无侵权",
    "不会侵权",
    "侵权风险为0",
    "侵权风险为 0",
    "zero infringement risk",
    "guaranteed non-infringing",
    "legally original",
    "freedom to operate approved",
)


def verify_release_readiness(
    root: Path,
    *,
    require_audit_reports: bool = True,
    require_live_acceptance: bool = True,
) -> ReleaseReadinessReport:
    passed: list[str] = []
    internal_blockers: list[str] = []
    external_blockers: list[str] = []

    for name, relative_path in _INTERNAL_EVIDENCE.items():
        if _has_content(root / relative_path):
            passed.append(name)
        else:
            internal_blockers.append(f"{name}_missing")

    matrix = run_security_matrix()
    if matrix.ok and matrix.total == 25:
        passed.append("security_matrix_25")
    else:
        internal_blockers.append("security_matrix_25_failed")

    probe = EchoSandboxBackend(workspace=root).probe()
    if probe.real_process_backend:
        passed.append("real_sandbox_backend")
    else:
        internal_blockers.append("real_sandbox_backend_missing")

    ip_boundary = verify_echo_ip_boundary(root)
    if ip_boundary.ok:
        passed.append("echo_ip_boundary")
    else:
        internal_blockers.append("echo_ip_boundary_failed")

    echo_probes = run_echo_release_probes()
    passed.extend(echo_probes.passed)
    internal_blockers.extend(echo_probes.failed)

    if require_live_acceptance:
        live_acceptance_path = root / _LIVE_ACCEPTANCE_EVIDENCE
        if _valid_echo_live_acceptance(root, live_acceptance_path):
            passed.append("echo_live_acceptance_60m")
        elif live_acceptance_path.is_file():
            internal_blockers.append("echo_live_acceptance_invalid")
        else:
            internal_blockers.append("echo_live_acceptance_missing")

    release_candidate_head = _release_candidate_head(root)
    for name, relative_path in _EXTERNAL_EVIDENCE.items():
        path = root / relative_path
        if not _has_content(path):
            external_blockers.append(f"{name}_missing")
        elif _is_external_approval(root, name, path, release_candidate_head):
            passed.append(name)
        else:
            external_blockers.append(f"{name}_pending")

    for name, relative_path in _STABLE_ARTIFACT_EVIDENCE.items():
        path = root / relative_path
        if not _has_content(path):
            external_blockers.append(f"{name}_missing")
        elif _has_unresolved_artifact_marker(path):
            external_blockers.append(f"{name}_pending")
        elif not _stable_artifact_is_valid(name, path):
            external_blockers.append(f"{name}_invalid")
        else:
            passed.append(name)

    if require_audit_reports:
        if _audit_reports_match_benchmark(root):
            passed.append("echo_audit_reports_bound")
        else:
            internal_blockers.append("echo_audit_reports_stale")

    # SLO must bind the current runtime source digest (not merely security matrix).
    slo_path = root / _STABLE_ARTIFACT_EVIDENCE["echo_slo_benchmark"]
    if not _valid_echo_slo_benchmark(slo_path):
        if "echo_slo_benchmark" in passed:
            # demote: stable-artifact loop may have treated unresolved markers only
            pass
        if "echo_slo_benchmark_digest_unbound" not in internal_blockers:
            # Prefer a specific internal blocker when SLO exists but digest diverges.
            if slo_path.is_file():
                internal_blockers.append("echo_slo_benchmark_digest_unbound")
            else:
                internal_blockers.append("echo_slo_benchmark_missing")

    e2e_path = root / _ISOLATED_VENV_E2E_EVIDENCE
    if _valid_isolated_venv_e2e(root, e2e_path):
        passed.append("isolated_venv_e2e")
    elif e2e_path.is_file():
        internal_blockers.append("isolated_venv_e2e_invalid")
    else:
        internal_blockers.append("isolated_venv_e2e_missing")

    # Deduplicate while preserving order
    passed = list(dict.fromkeys(passed))
    internal_blockers = list(dict.fromkeys(internal_blockers))

    internal_ready = not internal_blockers
    stable_ready = internal_ready and not external_blockers
    return ReleaseReadinessReport(
        internal_ready=internal_ready,
        stable_ready=stable_ready,
        passed=tuple(passed),
        internal_blockers=tuple(internal_blockers),
        external_blockers=tuple(external_blockers),
    )


CI_DEFERRED_INTERNAL_BLOCKERS = frozenset(
    {
        "echo_slo_benchmark_digest_unbound",
        "isolated_venv_e2e_invalid",
        "isolated_venv_e2e_missing",
    }
)


def filter_ci_deferred_internal_blockers(
    blockers: Sequence[str],
    *,
    github_actions: bool | None = None,
) -> tuple[str, ...]:
    """Drop digest-bound evidence blockers that shared CI cannot rebind.

    ``verify_release_readiness`` stays honest. GitHub Actions smoke still
    requires security matrix, real sandbox, and IP boundary; SLO JSON and
    isolated-venv E2E are generated on a developer machine and rebound there.
    """

    ordered = tuple(blockers)
    if github_actions is None:
        github_actions = os.environ.get("GITHUB_ACTIONS") == "true"
    if not github_actions:
        return ordered
    return tuple(item for item in ordered if item not in CI_DEFERRED_INTERNAL_BLOCKERS)


def _iter_release_source_files(root: Path) -> list[Path]:
    """Walk the release-source digest surfaces with the shared file filters."""
    resolved_root = root.resolve()
    seen: set[str] = set()
    files: list[Path] = []
    for relative in _RELEASE_SOURCE_DIGEST_SURFACES:
        candidate = resolved_root / relative
        if candidate.is_file() and not candidate.is_symlink():
            key = candidate.relative_to(resolved_root).as_posix()
            if key not in seen and _release_source_member_included(Path(key)):
                seen.add(key)
                files.append(candidate)
        elif candidate.is_dir():
            for path in candidate.rglob("*"):
                if not path.is_file() or path.is_symlink():
                    continue
                key = path.relative_to(resolved_root).as_posix()
                if key in seen or not _release_source_member_included(Path(key)):
                    continue
                seen.add(key)
                files.append(path)
    return sorted(files, key=lambda item: item.relative_to(resolved_root).as_posix())


class ReleaseSourceIntegrityError(ValueError):
    """A deterministic blocker found before release-gate work starts."""


def validate_release_source_integrity(root: Path) -> None:
    """Fail closed on structurally invalid files in the release digest surfaces."""
    resolved_root = root.resolve()
    findings: list[str] = []

    def relative_path(path: Path) -> Path:
        return path.relative_to(resolved_root)

    def is_ignored(relative: Path) -> bool:
        return not _release_source_member_included(relative)

    def validate_file(path: Path) -> None:
        relative = relative_path(path)
        try:
            payload = path.read_bytes()
        except OSError:
            findings.append(f"{relative.as_posix()}: unreadable")
            return
        if not payload:
            if relative not in _RELEASE_SOURCE_ALLOWED_EMPTY_FILES:
                findings.append(f"{relative.as_posix()}: empty")
            return
        if relative.parts[0] == "resources" or relative.suffix in _RELEASE_SOURCE_BINARY_SUFFIXES:
            return
        try:
            source = payload.decode("utf-8")
        except UnicodeDecodeError:
            findings.append(f"{relative.as_posix()}: invalid UTF-8")
            return
        if relative.suffix == ".py":
            try:
                ast.parse(source, filename=relative.as_posix())
            except (SyntaxError, ValueError, TypeError, MemoryError, RecursionError):
                findings.append(f"{relative.as_posix()}: invalid Python syntax")

    def walk_directory(directory: Path) -> int:
        regular_files = 0
        try:
            with os.scandir(directory) as iterator:
                entries = sorted(iterator, key=lambda entry: entry.name)
        except OSError:
            findings.append(f"{relative_path(directory).as_posix()}: unreadable directory")
            return 0
        for entry in entries:
            path = Path(entry.path)
            relative = relative_path(path)
            if is_ignored(relative):
                continue
            try:
                mode = entry.stat(follow_symlinks=False).st_mode
            except OSError:
                findings.append(f"{relative.as_posix()}: unreadable")
                continue
            if stat.S_ISLNK(mode):
                findings.append(f"{relative.as_posix()}: symlink")
            elif stat.S_ISDIR(mode):
                regular_files += walk_directory(path)
            elif stat.S_ISREG(mode):
                regular_files += 1
                validate_file(path)
            else:
                findings.append(f"{relative.as_posix()}: special file")
        return regular_files

    for relative in _RELEASE_SOURCE_DIGEST_SURFACES:
        if relative in _RELEASE_SOURCE_DIGEST_EXCLUDE:
            continue
        candidate = resolved_root / relative
        try:
            mode = candidate.lstat().st_mode
        except FileNotFoundError:
            findings.append(f"{relative.as_posix()}: missing")
            continue
        except OSError:
            findings.append(f"{relative.as_posix()}: unreadable")
            continue
        if stat.S_ISLNK(mode):
            findings.append(f"{relative.as_posix()}: symlink")
        elif stat.S_ISREG(mode):
            validate_file(candidate)
        elif stat.S_ISDIR(mode):
            if walk_directory(candidate) == 0:
                findings.append(f"{relative.as_posix()}: empty directory")
        else:
            findings.append(f"{relative.as_posix()}: special file")

    if findings:
        details = "; ".join(sorted(dict.fromkeys(findings)))
        raise ReleaseSourceIntegrityError(f"release source integrity preflight failed: {details}")


def release_source_digest(root: Path) -> str:
    """Hash release-relevant source and tests without depending on git state."""
    resolved_root = root.resolve()
    digest = hashlib.sha256(_RELEASE_SOURCE_DIGEST_VERSION)
    for path in _iter_release_source_files(resolved_root):
        relative_bytes = path.relative_to(resolved_root).as_posix().encode("utf-8")
        payload = path.read_bytes()
        digest.update(len(relative_bytes).to_bytes(8, "big"))
        digest.update(relative_bytes)
        digest.update(len(payload).to_bytes(8, "big"))
        digest.update(payload)
    return digest.hexdigest()


def release_source_surface_metadata_fingerprint(root: Path) -> str:
    """Hash release-surface path metadata (not contents) for soak drift detection."""
    resolved_root = root.resolve()
    digest = hashlib.sha256(_RELEASE_SOURCE_SURFACE_META_VERSION)
    for path in _iter_release_source_files(resolved_root):
        relative_bytes = path.relative_to(resolved_root).as_posix().encode("utf-8")
        stat = path.stat()
        digest.update(len(relative_bytes).to_bytes(8, "big"))
        digest.update(relative_bytes)
        digest.update(int(stat.st_size).to_bytes(8, "big", signed=False))
        digest.update(int(stat.st_mtime_ns).to_bytes(8, "big", signed=True))
        digest.update(int(stat.st_ctime_ns).to_bytes(8, "big", signed=True))
        digest.update(int(stat.st_ino).to_bytes(8, "big", signed=False))
        digest.update(int(stat.st_mode).to_bytes(4, "big", signed=False))
    return digest.hexdigest()


def tokenizer_resource_digest(root: Path) -> str:
    """Versioned content hash of the vendored tokenizer resource tree."""
    resolved_root = root.resolve()
    tokenizer_root = resolved_root / "resources" / "tokenizer"
    digest = hashlib.sha256(_TOKENIZER_TREE_DIGEST_VERSION)
    if not tokenizer_root.is_dir():
        return digest.hexdigest()
    files = sorted(
        (
            path
            for path in tokenizer_root.rglob("*")
            if path.is_file() and not path.is_symlink() and path.name != ".DS_Store"
        ),
        key=lambda item: item.relative_to(resolved_root).as_posix(),
    )
    for path in files:
        relative_bytes = path.relative_to(resolved_root).as_posix().encode("utf-8")
        payload = path.read_bytes()
        digest.update(len(relative_bytes).to_bytes(8, "big"))
        digest.update(relative_bytes)
        digest.update(len(payload).to_bytes(8, "big"))
        digest.update(payload)
    return digest.hexdigest()


def _audit_reports_match_benchmark(root: Path) -> bool:
    benchmark_path = root / _STABLE_ARTIFACT_EVIDENCE["echo_slo_benchmark"]
    if not benchmark_path.is_file():
        return False
    digest = hashlib.sha256(benchmark_path.read_bytes()).hexdigest()
    marker = f"Benchmark SHA-256: `{digest}`"
    return all(
        (root / relative_path).is_file()
        and marker in (root / relative_path).read_text(encoding="utf-8", errors="replace")
        for relative_path in _AUDIT_REPORT_EVIDENCE
    )


def verify_echo_ip_boundary(root: Path) -> EchoIPBoundaryReport:
    """Scan Echo release surfaces for obvious clean-room boundary violations.

    This is an engineering gate, not a legal opinion. It prevents accidental
    import/API-shape drift toward well-known agent frameworks and blocks
    unverifiable public claims such as "100% self-developed" or "no
    infringement".
    """
    global _scan_root
    _scan_root = root
    findings: list[str] = []
    for path in _iter_echo_ip_code_files(root):
        _scan_for_project_api_tokens(path, findings)
        _scan_for_unverifiable_claims(path, findings)
    for path in _iter_echo_ip_text_files(root):
        _scan_for_unverifiable_claims(path, findings)
    return EchoIPBoundaryReport(ok=not findings, findings=tuple(findings))


def _iter_echo_ip_code_files(root: Path) -> tuple[Path, ...]:
    files: list[Path] = []
    for relative_root in _ECHO_IP_CODE_ROOTS:
        directory = root / relative_root
        if directory.is_dir():
            files.extend(
                path
                for path in directory.rglob("*.py")
                if "__pycache__" not in path.parts
                and path.relative_to(root) not in _ECHO_IP_CODE_ALLOWLIST
            )
    return tuple(sorted(files))


def _iter_echo_ip_text_files(root: Path) -> tuple[Path, ...]:
    files: list[Path] = []
    for relative_root in _ECHO_IP_TEXT_ROOTS:
        path = root / relative_root
        if path.is_file():
            files.append(path)
        elif path.is_dir():
            files.extend(
                child
                for child in path.rglob("*.md")
                if child.relative_to(root) != _CLEAN_ROOM_AVOIDANCE_DOC
            )
    return tuple(sorted(files))


# Narrow per-file token exemptions for false positives caused by generic
# English words that collide with disallowed framework API names.
# Each entry is (relative_path, token, reason).  The file is still scanned
# for ALL other disallowed tokens; only the named token is suppressed.
#
# handoff_vault.py defines an internal MAC domain constant
# "js-agent:handoff-vault:v1" — "handoff" here is a generic English noun
# meaning "transfer of responsibility", not the AutoGen/AgentScope
# "handoff()" registration API that the IP gate is designed to block.
_TOKEN_EXEMPTIONS: dict[tuple[str, str], str] = {
    ("js/echo/handoff_vault.py", "handoff"): (
        "internal vault MAC domain; not a framework handoff() registration API"
    ),
}


def _scan_for_project_api_tokens(path: Path, findings: list[str]) -> None:
    text = path.read_text(encoding="utf-8")
    try:
        rel = str(path.relative_to(_scan_root))
    except ValueError:
        rel = str(path)
    for line_no, line in enumerate(text.splitlines(), start=1):
        for token in _DISALLOWED_PROJECT_API_TOKENS:
            if _project_api_token_present(token, line):
                if (rel, token) in _TOKEN_EXEMPTIONS:
                    continue
                findings.append(f"{path}:{line_no}: disallowed project API token {token}")


def _project_api_token_present(token: str, line: str) -> bool:
    if token == "action-observation":
        return "action-observation" in line
    if token == "compile":
        return re.search(r"(?<!\.)\bcompile\b", line) is not None
    if token in {"Workflow", "Conversation", "Runtime"}:
        return (
            re.search(
                rf"\b(class\s+{token}\b|(?<![A-Za-z0-9_]){token}\s*\()",
                line,
            )
            is not None
        )
    return re.search(rf"\b{re.escape(token)}\b", line) is not None


def _scan_for_unverifiable_claims(path: Path, findings: list[str]) -> None:
    text = path.read_text(encoding="utf-8")
    lower_text = text.lower()
    for phrase in _UNVERIFIABLE_SELF_DEVELOPED_CLAIMS:
        if phrase.lower() not in lower_text:
            continue
        for line_no, line in enumerate(text.splitlines(), start=1):
            if phrase.lower() in line.lower():
                findings.append(f"{path}:{line_no}: unverifiable self-developed claim")


def _has_content(path: Path) -> bool:
    return path.is_file() and bool(path.read_text(encoding="utf-8").strip())


def _is_external_approval(
    root: Path,
    evidence_name: str,
    path: Path,
    release_candidate_head: str | None,
) -> bool:
    text = path.read_text(encoding="utf-8")
    fields = _front_matter_fields(text)
    if _has_duplicate_approval_fields(text):
        return False
    if any(not fields.get(field, "").strip() for field in _EXTERNAL_APPROVAL_FIELDS):
        return False
    reviewer = fields["Reviewer"]
    reviewed_on = fields["Date"]
    if fields["Status"] != "APPROVED" or "PENDING" in text:
        return False
    try:
        date.fromisoformat(reviewed_on)
    except ValueError:
        return False
    if fields["Evidence-Name"] != evidence_name:
        return False
    if _SCOPE_COMMIT_PATTERN.fullmatch(fields["Scope-Commit"]) is None:
        return False
    if release_candidate_head is None or not hmac.compare_digest(
        fields["Scope-Commit"], release_candidate_head
    ):
        return False
    if not _approval_artifacts_match(root, fields):
        return False
    trusted_key = _trusted_reviewer_key(reviewer)
    signature = _detached_signature(fields["Signature"])
    if trusted_key is None or signature is None:
        return False
    try:
        trusted_key.verify(signature, _canonical_external_approval_payload(fields))
    except (InvalidSignature, ValueError):
        return False
    return True


def _release_candidate_head(root: Path) -> str | None:
    repository_root = _git_output(root, "rev-parse", "--show-toplevel")
    head = _git_output(root, "rev-parse", "--verify", "HEAD^{commit}")
    if repository_root is None or head is None:
        return None
    if Path(repository_root).resolve() != root.resolve():
        return None
    if _SCOPE_COMMIT_PATTERN.fullmatch(head) is None:
        return None
    status = _git_output(
        root,
        "status",
        "--porcelain=v1",
        "--untracked-files=all",
        "--ignore-submodules=none",
        "--",
        *(path.as_posix() for path in _RELEASE_CANDIDATE_SURFACES),
    )
    if status is None or status:
        return None
    return head


def _git_output(root: Path, *args: str) -> str | None:
    environment = os.environ.copy()
    for name in (
        "GIT_ALTERNATE_OBJECT_DIRECTORIES",
        "GIT_DIR",
        "GIT_INDEX_FILE",
        "GIT_OBJECT_DIRECTORY",
        "GIT_WORK_TREE",
    ):
        environment.pop(name, None)
    environment["GIT_LITERAL_PATHSPECS"] = "1"
    environment["GIT_OPTIONAL_LOCKS"] = "0"
    try:
        result = subprocess.run(
            ["git", *args],
            cwd=root,
            check=False,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            env=environment,
            timeout=5.0,
        )
    except (OSError, subprocess.TimeoutExpired):
        return None
    if result.returncode != 0:
        return None
    return result.stdout.strip()


def _git_bytes(root: Path, *args: str, timeout: float = 15.0) -> bytes | None:
    """Read immutable Git object bytes without locks, checkout, or worktree mutation."""
    environment = os.environ.copy()
    for name in (
        "GIT_ALTERNATE_OBJECT_DIRECTORIES",
        "GIT_DIR",
        "GIT_INDEX_FILE",
        "GIT_OBJECT_DIRECTORY",
        "GIT_WORK_TREE",
    ):
        environment.pop(name, None)
    environment["GIT_LITERAL_PATHSPECS"] = "1"
    environment["GIT_OPTIONAL_LOCKS"] = "0"
    try:
        result = subprocess.run(
            ["git", *args],
            cwd=root,
            check=False,
            capture_output=True,
            env=environment,
            timeout=timeout,
        )
    except (OSError, subprocess.TimeoutExpired):
        return None
    if result.returncode != 0:
        return None
    return result.stdout


def _approval_artifacts_match(root: Path, fields: Mapping[str, str]) -> bool:
    sbom_sha256 = fields["SBOM-SHA256"]
    lock_sha256 = fields["UV-Lock-SHA256"]
    if (
        _SHA256_PATTERN.fullmatch(sbom_sha256) is None
        or _SHA256_PATTERN.fullmatch(lock_sha256) is None
    ):
        return False
    try:
        current_sbom_sha256 = hashlib.sha256(
            (root / _STABLE_ARTIFACT_EVIDENCE["sbom_spdx"]).read_bytes()
        ).hexdigest()
        current_lock_sha256 = hashlib.sha256((root / "uv.lock").read_bytes()).hexdigest()
    except OSError:
        return False
    return hmac.compare_digest(sbom_sha256.lower(), current_sbom_sha256) and hmac.compare_digest(
        lock_sha256.lower(), current_lock_sha256
    )


def _trusted_reviewer_key(reviewer: str) -> Ed25519PublicKey | None:
    raw_keys = os.environ.get("JS_ECHO_TRUSTED_REVIEW_KEYS")
    if not raw_keys:
        return None
    try:
        configured_keys = strict_loads(raw_keys)
    except (StrictJSONError, ValueError):
        return None
    if not isinstance(configured_keys, dict):
        return None
    encoded_key = configured_keys.get(reviewer)
    if not isinstance(encoded_key, str):
        return None
    try:
        raw_key = base64.b64decode(encoded_key, validate=True)
        return Ed25519PublicKey.from_public_bytes(raw_key)
    except (ValueError, binascii.Error):
        return None


def _detached_signature(encoded_signature: str) -> bytes | None:
    try:
        signature = base64.b64decode(encoded_signature, validate=True)
    except (ValueError, binascii.Error):
        return None
    return signature if len(signature) == 64 else None


def _canonical_external_approval_payload(fields: Mapping[str, str]) -> bytes:
    """Return the versioned, fixed-order payload for detached Ed25519 approval signatures."""
    try:
        lines = [
            _EXTERNAL_APPROVAL_CANONICAL_VERSION,
            *(f"{field}: {fields[field]}" for field in _SIGNED_EXTERNAL_APPROVAL_FIELDS),
            "",
        ]
    except KeyError as exc:
        raise ValueError(f"external approval is missing {exc.args[0]}") from exc
    return "\n".join(lines).encode("utf-8")


def _front_matter_fields(text: str) -> dict[str, str]:
    fields: dict[str, str] = {}
    for raw_line in text.splitlines():
        if ":" not in raw_line:
            continue
        key, value = raw_line.split(":", 1)
        key = key.strip()
        if key in _EXTERNAL_APPROVAL_FIELDS:
            fields[key] = value.strip()
    return fields


def _has_duplicate_approval_fields(text: str) -> bool:
    seen: set[str] = set()
    for raw_line in text.splitlines():
        if ":" not in raw_line:
            continue
        key, _ = raw_line.split(":", 1)
        key = key.strip()
        if key not in _EXTERNAL_APPROVAL_FIELDS:
            continue
        if key in seen:
            return True
        seen.add(key)
    return False


def _has_unresolved_artifact_marker(path: Path) -> bool:
    text = path.read_text(encoding="utf-8")
    explicit_markers = {"PENDING", "TODO", "TBD", "PLACEHOLDER", "NOT GENERATED"}
    for line in text.splitlines():
        upper = line.strip().upper()
        if upper in explicit_markers:
            return True
        if upper.startswith(
            (
                "STATUS: PENDING",
                "STATUS: TODO",
                "STATUS: TBD",
                "STATUS: PLACEHOLDER",
                "STATUS: NOT GENERATED",
            )
        ):
            return True
    return any(
        line.strip().startswith(("UNRESOLVED:", "BLOCKER:", "- UNRESOLVED", "- BLOCKER"))
        for line in text.upper().splitlines()
    )


def _stable_artifact_is_valid(name: str, path: Path) -> bool:
    if name == "sbom_spdx":
        return _valid_spdx_sbom(path)
    if name == "license_scan":
        return _valid_license_scan(path)
    if name == "echo_slo_benchmark":
        return _valid_echo_slo_benchmark(path)
    return True


def _live_json_matches(actual: object, expected: object) -> bool:
    try:
        actual_json = json.dumps(
            actual,
            sort_keys=True,
            separators=(",", ":"),
            allow_nan=False,
        )
        expected_json = json.dumps(
            expected,
            sort_keys=True,
            separators=(",", ":"),
            allow_nan=False,
        )
    except (TypeError, ValueError, OverflowError):
        return False
    return hmac.compare_digest(actual_json, expected_json)


def _live_canonical_sha256(value: object) -> str | None:
    try:
        payload = json.dumps(
            value,
            sort_keys=True,
            separators=(",", ":"),
            allow_nan=False,
        ).encode()
    except (TypeError, ValueError, OverflowError):
        return None
    return hashlib.sha256(payload).hexdigest()


def _live_minute_medians(
    points: list[tuple[float, int]],
    *,
    start: float,
) -> list[tuple[float, int]]:
    buckets: dict[int, list[int]] = {}
    for elapsed, value in points:
        if elapsed < start:
            continue
        minute = int(elapsed // 60)
        buckets.setdefault(minute, []).append(value)
    return [
        (minute * 60.0 + 30.0, int(median(values))) for minute, values in sorted(buckets.items())
    ]


def _live_theil_sen_growth(points: list[tuple[float, int]]) -> float:
    slopes: list[float] = []
    for index, (left_elapsed, left_value) in enumerate(points):
        for right_elapsed, right_value in points[index + 1 :]:
            elapsed = right_elapsed - left_elapsed
            if elapsed > 0:
                slopes.append((right_value - left_value) / elapsed)
    if not slopes:
        return 0.0
    return round(float(median(slopes)) * 60 / (1024 * 1024), 3)


def _live_plateau_growth(
    points: list[tuple[float, int]],
    *,
    duration_seconds: float,
    warmup_seconds: float,
) -> float:
    reference_start = max(warmup_seconds, duration_seconds * 0.5)
    reference_end = min(duration_seconds - 60.0, reference_start + 300.0)
    reference_values = [
        value for elapsed, value in points if reference_start <= elapsed <= reference_end
    ]
    final_values = [
        value for elapsed, value in points if elapsed >= max(0.0, duration_seconds - 60.0)
    ]
    if not reference_values or not final_values:
        return math.inf
    return (float(median(final_values)) - float(median(reference_values))) / (1024 * 1024)


def _live_required_sample_count(duration_seconds: float) -> int:
    return max(
        2,
        math.ceil(
            max(0.0, duration_seconds)
            / _LIVE_RESOURCE_SAMPLE_INTERVAL_SECONDS
            * _LIVE_RESOURCE_MIN_SAMPLE_RATIO
        ),
    )


def _live_resource_report(
    samples: list[dict[str, Any]],
    *,
    process_name: str,
    duration_seconds: float,
    max_rss_bytes: int,
    max_growth_mib_per_minute: float,
) -> dict[str, object]:
    points = [
        (float(sample["elapsed_seconds"]), int(sample["rss_bytes"][process_name]))
        for sample in samples
    ]
    required_count = _live_required_sample_count(duration_seconds)
    coverage = points[-1][0] - points[0][0]
    max_gap = max(
        (right[0] - left[0] for left, right in zip(points, points[1:], strict=False)),
        default=math.inf,
    )
    minimum_coverage = max(0.0, duration_seconds - _LIVE_RESOURCE_MAX_SAMPLE_GAP_SECONDS)
    sample_integrity_ok = bool(
        len(points) >= required_count
        and coverage >= minimum_coverage
        and max_gap <= _LIVE_RESOURCE_MAX_SAMPLE_GAP_SECONDS
    )
    warmup_seconds = 900.0 if duration_seconds >= 3600.0 else duration_seconds * 0.5
    minute_medians = _live_minute_medians(points, start=0.0)
    stable_points = _live_minute_medians(points, start=warmup_seconds)
    growth = _live_theil_sen_growth(stable_points)
    tail_growth = _live_theil_sen_growth(stable_points[-3:])
    plateau_growth = _live_plateau_growth(
        points,
        duration_seconds=duration_seconds,
        warmup_seconds=warmup_seconds,
    )
    peak = max(value for _elapsed, value in points)
    stability_enforced = duration_seconds >= _LIVE_RESOURCE_STABILITY_MIN_SECONDS
    peak_ok = peak <= max_rss_bytes
    growth_ok = not stability_enforced or bool(
        growth <= max_growth_mib_per_minute
        and plateau_growth <= _LIVE_RESOURCE_MAX_PLATEAU_GROWTH_MIB
    )
    return {
        "sample_count": len(points),
        "required_sample_count": required_count,
        "sample_coverage_seconds": round(coverage, 3),
        "max_sample_gap_seconds": round(max_gap, 3),
        "sample_integrity_ok": sample_integrity_ok,
        "start_rss_mib": round(points[0][1] / (1024 * 1024), 3),
        "final_rss_mib": round(points[-1][1] / (1024 * 1024), 3),
        "peak_rss_mib": round(peak / (1024 * 1024), 3),
        "growth_mib_per_minute": growth,
        "tail_growth_mib_per_minute": tail_growth,
        "plateau_growth_mib": (round(plateau_growth, 3) if math.isfinite(plateau_growth) else None),
        "stability_window_start_seconds": round(warmup_seconds, 3),
        "minute_medians": [
            {
                "elapsed_seconds": round(elapsed, 3),
                "rss_mib": round(value / (1024 * 1024), 3),
            }
            for elapsed, value in minute_medians
        ],
        "peak_within_limit": peak_ok,
        "growth_within_limit": growth_ok,
        "ok": peak_ok and growth_ok and sample_integrity_ok,
    }


def _live_storage_report(
    samples: list[dict[str, Any]],
    *,
    product_name: str,
    duration_seconds: float,
    max_state_bytes: int,
    max_session_partitions: int,
) -> dict[str, object]:
    points = [
        (float(sample["elapsed_seconds"]), dict(sample["storage"][product_name]))
        for sample in samples
    ]
    elapsed_points = [elapsed for elapsed, _evidence in points]
    required_count = _live_required_sample_count(duration_seconds)
    coverage = elapsed_points[-1] - elapsed_points[0]
    max_gap = max(
        (right - left for left, right in zip(elapsed_points, elapsed_points[1:], strict=False)),
        default=math.inf,
    )
    minimum_coverage = max(0.0, duration_seconds - _LIVE_RESOURCE_MAX_SAMPLE_GAP_SECONDS)
    sample_integrity_ok = bool(
        len(points) >= required_count
        and coverage >= minimum_coverage
        and max_gap <= _LIVE_RESOURCE_MAX_SAMPLE_GAP_SECONDS
    )
    total_points = [(elapsed, int(evidence["total_bytes"])) for elapsed, evidence in points]
    partition_points = [
        (elapsed, int(evidence["partition_storage_bytes"])) for elapsed, evidence in points
    ]
    warmup_seconds = 900.0 if duration_seconds >= 3600.0 else duration_seconds * 0.5
    total_growth = _live_theil_sen_growth(_live_minute_medians(total_points, start=warmup_seconds))
    partition_growth = _live_theil_sen_growth(
        _live_minute_medians(partition_points, start=warmup_seconds)
    )
    total_plateau = _live_plateau_growth(
        total_points,
        duration_seconds=duration_seconds,
        warmup_seconds=warmup_seconds,
    )
    partition_plateau = _live_plateau_growth(
        partition_points,
        duration_seconds=duration_seconds,
        warmup_seconds=warmup_seconds,
    )
    peak_total = max(value for _elapsed, value in total_points)
    max_active_partitions = max(
        int(evidence["max_active_session_partitions_per_owner"]) for _elapsed, evidence in points
    )
    retention_errors = sorted(
        {
            str(error)
            for _elapsed, evidence in points
            for error in evidence["retention_checkpoint_errors"]
        }
    )
    retirement_observations = [
        {str(owner) for owner in evidence["incomplete_retirements"]}
        for _elapsed, evidence in points
    ]
    transient_retirements = sorted(set().union(*retirement_observations))
    marker_observations = [
        {
            str(marker)
            for marker in (
                evidence["incomplete_retirement_markers"] or evidence["incomplete_retirements"]
            )
        }
        for _elapsed, evidence in points
    ]
    stale_markers: set[str] = set(marker_observations[-1])
    for previous, current in zip(marker_observations, marker_observations[1:], strict=False):
        stale_markers.update(previous & current)
    stale_retirements = sorted({marker.split("|", 1)[0] for marker in stale_markers})
    component_names = sorted(
        {
            str(component)
            for _elapsed, evidence in points
            for component in evidence["component_bytes"]
        }
    )
    component_growth: dict[str, dict[str, object]] = {}
    for component in component_names:
        component_points = [
            (elapsed, int(evidence["component_bytes"].get(component, 0)))
            for elapsed, evidence in points
        ]
        component_plateau = _live_plateau_growth(
            component_points,
            duration_seconds=duration_seconds,
            warmup_seconds=warmup_seconds,
        )
        component_growth[component] = {
            "start_mib": round(component_points[0][1] / (1024 * 1024), 3),
            "final_mib": round(component_points[-1][1] / (1024 * 1024), 3),
            "growth_mib_per_minute": _live_theil_sen_growth(
                _live_minute_medians(component_points, start=warmup_seconds)
            ),
            "plateau_growth_mib": (
                round(component_plateau, 3) if math.isfinite(component_plateau) else None
            ),
        }
    bounds_ok = bool(
        peak_total <= max_state_bytes
        and max_active_partitions <= max_session_partitions
        and not retention_errors
        and not stale_retirements
    )
    stability_enforced = duration_seconds >= _LIVE_RESOURCE_STABILITY_MIN_SECONDS
    growth_ok = not stability_enforced or bool(
        total_growth <= _LIVE_MAX_STATE_GROWTH_MIB_PER_MINUTE
        and partition_growth <= _LIVE_MAX_PARTITION_GROWTH_MIB_PER_MINUTE
        and total_plateau <= _LIVE_STORAGE_MAX_PLATEAU_GROWTH_MIB
        and partition_plateau <= _LIVE_PARTITION_MAX_PLATEAU_GROWTH_MIB
    )
    return {
        "sample_count": len(points),
        "required_sample_count": required_count,
        "sample_coverage_seconds": round(coverage, 3),
        "max_sample_gap_seconds": round(max_gap, 3),
        "sample_integrity_ok": sample_integrity_ok,
        "start_total_mib": round(total_points[0][1] / (1024 * 1024), 3),
        "final_total_mib": round(total_points[-1][1] / (1024 * 1024), 3),
        "peak_total_mib": round(peak_total / (1024 * 1024), 3),
        "total_growth_mib_per_minute": total_growth,
        "partition_growth_mib_per_minute": partition_growth,
        "total_plateau_growth_mib": (
            round(total_plateau, 3) if math.isfinite(total_plateau) else None
        ),
        "partition_plateau_growth_mib": (
            round(partition_plateau, 3) if math.isfinite(partition_plateau) else None
        ),
        "max_active_session_partitions_per_owner": max_active_partitions,
        "retention_checkpoint_errors": retention_errors,
        "transient_retirement_observations": transient_retirements,
        "stale_incomplete_retirement_markers": sorted(stale_markers),
        "stale_incomplete_retirements": stale_retirements,
        "incomplete_retirements": stale_retirements,
        "component_growth": component_growth,
        "bounds_within_limit": bounds_ok,
        "growth_within_limit": growth_ok,
        "ok": sample_integrity_ok and bounds_ok and growth_ok,
    }


def _valid_live_sample(sample: object, *, previous_elapsed: float | None) -> float | None:
    if not isinstance(sample, dict):
        return None
    elapsed = _safe_finite_float(sample.get("elapsed_seconds"))
    if (
        elapsed is None
        or elapsed < 0
        or (previous_elapsed is not None and elapsed <= previous_elapsed)
    ):
        return None
    rss_bytes = sample.get("rss_bytes")
    process_counts = sample.get("process_counts")
    storage = sample.get("storage")
    required_names = {"js_agent", "js_work"}
    if (
        not isinstance(rss_bytes, dict)
        or set(rss_bytes) != required_names
        or not isinstance(process_counts, dict)
        or set(process_counts) != required_names
        or not isinstance(storage, dict)
        or set(storage) != required_names
    ):
        return None
    for name in required_names:
        rss = rss_bytes.get(name)
        process_count = process_counts.get(name)
        if (
            isinstance(rss, bool)
            or not isinstance(rss, int)
            or rss < 0
            or isinstance(process_count, bool)
            or not isinstance(process_count, int)
            or process_count <= 0
        ):
            return None
        evidence = storage.get(name)
        if not isinstance(evidence, dict):
            return None
        integer_fields = (
            "total_bytes",
            "file_count",
            "partition_storage_bytes",
            "partition_file_count",
            "max_active_session_partitions_per_owner",
            "retired_session_partition_count",
            "retention_checkpoint_bytes",
        )
        if any(
            isinstance(evidence.get(field), bool)
            or not isinstance(evidence.get(field), int)
            or int(evidence[field]) < 0
            for field in integer_fields
        ):
            return None
        for field in (
            "retention_checkpoint_errors",
            "incomplete_retirements",
            "incomplete_retirement_markers",
        ):
            values = evidence.get(field)
            if not isinstance(values, list) or not all(isinstance(value, str) for value in values):
                return None
        components = evidence.get("component_bytes")
        if (
            not isinstance(components, dict)
            or set(components) != set(_LIVE_STORAGE_COMPONENT_NAMES)
            or any(
                isinstance(value, bool) or not isinstance(value, int) or value < 0
                for value in components.values()
            )
            or sum(components.values()) != evidence["total_bytes"]
            or evidence["partition_storage_bytes"] != components["echo_partitions"]
            or evidence["partition_storage_bytes"] > evidence["total_bytes"]
            or evidence["partition_file_count"] > evidence["file_count"]
            or evidence["retention_checkpoint_bytes"] > evidence["total_bytes"]
        ):
            return None
    return elapsed


def _valid_live_resource_evidence(data: dict[str, Any], *, duration_seconds: float) -> bool:
    max_rss_bytes = data.get("max_rss_bytes")
    max_state_bytes = data.get("max_state_bytes")
    max_session_partitions = data.get("max_session_partitions_per_owner")
    max_growth = _safe_finite_float(data.get("max_rss_growth_mib_per_minute"))
    if (
        isinstance(max_rss_bytes, bool)
        or max_rss_bytes != _LIVE_DEFAULT_MAX_RSS_BYTES
        or isinstance(max_state_bytes, bool)
        or max_state_bytes != _LIVE_DEFAULT_MAX_STATE_BYTES
        or isinstance(max_session_partitions, bool)
        or max_session_partitions != _LIVE_DEFAULT_MAX_SESSION_PARTITIONS_PER_OWNER
        or max_growth != _LIVE_DEFAULT_MAX_RSS_GROWTH_MIB_PER_MINUTE
    ):
        return False
    resources = data.get("resources")
    storage = data.get("storage_stability")
    if not isinstance(resources, dict) or not isinstance(storage, dict):
        return False
    resource_samples = resources.get("samples")
    storage_samples = storage.get("samples")
    if (
        not isinstance(resource_samples, list)
        or not resource_samples
        or not isinstance(storage_samples, list)
        or not _live_json_matches(resource_samples, storage_samples)
    ):
        return False
    previous_elapsed: float | None = None
    for sample in resource_samples:
        previous_elapsed = _valid_live_sample(sample, previous_elapsed=previous_elapsed)
        if previous_elapsed is None:
            return False
    if (
        resources.get("samples_truncated") is not False
        or resources.get("recorded_sample_count") != len(resource_samples)
        or resources.get("omitted_sample_count") != 0
        or storage.get("samples_truncated") is not False
        or storage.get("recorded_sample_count") != len(storage_samples)
        or storage.get("omitted_sample_count") != 0
    ):
        return False
    resource_thresholds = {
        "stability_enforced": True,
        "max_rss_bytes": max_rss_bytes,
        "max_growth_mib_per_minute": max_growth,
        "max_plateau_growth_mib": _LIVE_RESOURCE_MAX_PLATEAU_GROWTH_MIB,
        "sample_interval_seconds": _LIVE_RESOURCE_SAMPLE_INTERVAL_SECONDS,
        "max_sample_gap_seconds": _LIVE_RESOURCE_MAX_SAMPLE_GAP_SECONDS,
        "min_sample_ratio": _LIVE_RESOURCE_MIN_SAMPLE_RATIO,
        "stability_min_seconds": _LIVE_RESOURCE_STABILITY_MIN_SECONDS,
    }
    if any(
        not _live_json_matches(resources.get(key), value)
        for key, value in resource_thresholds.items()
    ):
        return False
    processes = resources.get("processes")
    if not isinstance(processes, dict) or set(processes) != {"js_agent", "js_work"}:
        return False
    process_ok = True
    for process_name in ("js_agent", "js_work"):
        evidence = processes.get(process_name)
        if not isinstance(evidence, dict):
            return False
        expected = _live_resource_report(
            resource_samples,
            process_name=process_name,
            duration_seconds=duration_seconds,
            max_rss_bytes=max_rss_bytes,
            max_growth_mib_per_minute=max_growth,
        )
        if any(not _live_json_matches(evidence.get(key), value) for key, value in expected.items()):
            return False
        process_ok = process_ok and expected["ok"] is True
    if resources.get("ok") is not process_ok:
        return False
    storage_thresholds = {
        "stability_enforced": True,
        "max_state_bytes": max_state_bytes,
        "max_total_growth_mib_per_minute": _LIVE_MAX_STATE_GROWTH_MIB_PER_MINUTE,
        "max_partition_growth_mib_per_minute": _LIVE_MAX_PARTITION_GROWTH_MIB_PER_MINUTE,
        "max_total_plateau_growth_mib": _LIVE_STORAGE_MAX_PLATEAU_GROWTH_MIB,
        "max_partition_plateau_growth_mib": _LIVE_PARTITION_MAX_PLATEAU_GROWTH_MIB,
        "max_active_session_partitions_per_owner": max_session_partitions,
        "sample_interval_seconds": _LIVE_RESOURCE_SAMPLE_INTERVAL_SECONDS,
        "max_sample_gap_seconds": _LIVE_RESOURCE_MAX_SAMPLE_GAP_SECONDS,
        "min_sample_ratio": _LIVE_RESOURCE_MIN_SAMPLE_RATIO,
        "stability_min_seconds": _LIVE_RESOURCE_STABILITY_MIN_SECONDS,
    }
    if any(
        not _live_json_matches(storage.get(key), value) for key, value in storage_thresholds.items()
    ):
        return False
    products = storage.get("products")
    if not isinstance(products, dict) or set(products) != {"js_agent", "js_work"}:
        return False
    storage_ok = True
    for product_name in ("js_agent", "js_work"):
        evidence = products.get(product_name)
        if not isinstance(evidence, dict):
            return False
        expected = _live_storage_report(
            resource_samples,
            product_name=product_name,
            duration_seconds=duration_seconds,
            max_state_bytes=max_state_bytes,
            max_session_partitions=max_session_partitions,
        )
        if any(not _live_json_matches(evidence.get(key), value) for key, value in expected.items()):
            return False
        storage_ok = storage_ok and expected["ok"] is True
    return resources.get("ok") is True and storage.get("ok") is storage_ok is True


def _valid_echo_live_acceptance(root: Path, path: Path) -> bool:
    """Validate a source-bound, uninterrupted one-hour local acceptance artifact."""
    try:
        data = strict_load_object(path)
    except (OSError, json.JSONDecodeError, StrictJSONError, ValueError):
        return False
    if data.get("schema_version") != "echo-live-acceptance-v4":
        return False
    source_digest = data.get("source_digest")
    current_digest = release_source_digest(root)
    if not isinstance(source_digest, str) or not hmac.compare_digest(
        source_digest,
        current_digest,
    ):
        return False
    acceptance_pid = data.get("acceptance_pid")
    if (
        isinstance(acceptance_pid, bool)
        or not isinstance(acceptance_pid, int)
        or acceptance_pid <= 0
    ):
        return False
    duration = data.get("duration_seconds")
    concurrency = data.get("concurrency")
    duration_number = _safe_finite_float(duration)
    if (
        data.get("ok") is not True
        or data.get("network") != "local-only"
        or duration_number is None
        or duration_number < 3600.0
        or isinstance(concurrency, bool)
        or not isinstance(concurrency, int)
        or concurrency < 2
    ):
        return False
    source_integrity = data.get("source_integrity")
    if not isinstance(source_integrity, dict):
        return False
    check_count = source_integrity.get("check_count")
    expected_integrity_digest = source_integrity.get("expected_digest")
    final_integrity_digest = source_integrity.get("final_digest")
    expected_meta = source_integrity.get("expected_metadata_fingerprint")
    final_meta = source_integrity.get("final_metadata_fingerprint")
    if (
        source_integrity.get("ok") is not True
        or source_integrity.get("drifted") is not False
        or isinstance(check_count, bool)
        or not isinstance(check_count, int)
        or check_count <= 0
        or not isinstance(expected_integrity_digest, str)
        or not hmac.compare_digest(expected_integrity_digest, source_digest)
        or not isinstance(final_integrity_digest, str)
        or not hmac.compare_digest(final_integrity_digest, current_digest)
        or not isinstance(expected_meta, str)
        or not isinstance(final_meta, str)
        or _SHA256_PATTERN.fullmatch(expected_meta) is None
        or _SHA256_PATTERN.fullmatch(final_meta) is None
        or not hmac.compare_digest(expected_meta, final_meta)
    ):
        return False
    integrity_checks = source_integrity.get("checks")
    chain_root = source_integrity.get("chain_root")
    if not isinstance(integrity_checks, list) or len(integrity_checks) != check_count:
        return False
    running_root = bytes(32)
    prior_monotonic: float | None = None
    prior_wall: datetime | None = None
    setup_bias: float | None = None
    receipt_start = _parse_utc_timestamp(data.get("started_utc"))
    receipt_end = _parse_utc_timestamp(data.get("finished_utc"))
    if receipt_start is None or receipt_end is None or receipt_end < receipt_start:
        return False
    wall_span_seconds = (receipt_end - receipt_start).total_seconds()
    if not math.isfinite(wall_span_seconds):
        return False
    if wall_span_seconds < duration_number - _SOAK_COVERAGE_TOLERANCE_SECONDS:
        return False
    first_mono: float | None = None
    first_wall: datetime | None = None
    for index, check in enumerate(integrity_checks, start=1):
        if not isinstance(check, dict) or check.get("index") != index:
            return False
        digest = check.get("source_digest")
        metadata_digest = check.get("metadata_fingerprint")
        monotonic_s = check.get("monotonic_s")
        wall_utc = _parse_utc_timestamp(check.get("wall_utc"))
        mono = _safe_finite_float(monotonic_s)
        if (
            not isinstance(digest, str)
            or not hmac.compare_digest(digest, source_digest)
            or not isinstance(metadata_digest, str)
            or not hmac.compare_digest(metadata_digest, expected_meta)
            or mono is None
            or mono < 0
            or wall_utc is None
            or wall_utc < receipt_start
            or wall_utc > receipt_end
        ):
            return False
        wall_from_start = (wall_utc - receipt_start).total_seconds()
        if not math.isfinite(wall_from_start):
            return False
        bias = wall_from_start - mono
        if index == 1:
            if mono > _SOAK_FIRST_MONOTONIC_MAX_SECONDS:
                return False
            # Setup may precede the first integrity sample; bias must be small and stable.
            if (
                bias < -_SOAK_WALL_MONO_ALIGN_TOLERANCE_SECONDS
                or bias > _SOAK_SETUP_BIAS_MAX_SECONDS
            ):
                return False
            setup_bias = bias
            first_mono = mono
            first_wall = wall_utc
        elif setup_bias is None or abs(bias - setup_bias) > _SOAK_WALL_MONO_ALIGN_TOLERANCE_SECONDS:
            return False
        if prior_monotonic is not None:
            interval = mono - prior_monotonic
            wall_interval = (wall_utc - prior_wall).total_seconds() if prior_wall else interval
            if prior_wall is not None and wall_utc < prior_wall:
                return False
            if not math.isfinite(interval) or not math.isfinite(wall_interval):
                return False
            if not (2.5 <= interval <= 15.0 and abs(wall_interval - interval) <= 2.0):
                return False
        prior_monotonic = mono
        prior_wall = wall_utc
        canonical = json.dumps(
            {
                "index": index,
                "metadata_fingerprint": metadata_digest,
                "monotonic_s": monotonic_s,
                "source_digest": digest,
                "wall_utc": check.get("wall_utc"),
            },
            sort_keys=True,
            separators=(",", ":"),
        ).encode()
        running_root = hashlib.sha256(running_root + canonical).digest()
    source_check_chain_root = source_integrity.get("source_check_chain_root")
    resources_for_chain = data.get("resources")
    storage_for_chain = data.get("storage_stability")
    resource_samples_for_chain = (
        resources_for_chain.get("samples") if isinstance(resources_for_chain, dict) else None
    )
    storage_samples_for_chain = (
        storage_for_chain.get("samples") if isinstance(storage_for_chain, dict) else None
    )
    resource_sample_root = _live_canonical_sha256(resource_samples_for_chain)
    storage_sample_root = _live_canonical_sha256(storage_samples_for_chain)
    resource_sample_count = (
        len(resource_samples_for_chain) if isinstance(resource_samples_for_chain, list) else None
    )
    storage_sample_count = (
        len(storage_samples_for_chain) if isinstance(storage_samples_for_chain, list) else None
    )
    if (
        not isinstance(source_check_chain_root, str)
        or not hmac.compare_digest(source_check_chain_root, running_root.hex())
        or source_integrity.get("chain_binding") != _LIVE_RESOURCE_CHAIN_BINDING
        or resource_sample_root is None
        or storage_sample_root is None
        or source_integrity.get("resource_sample_count") != resource_sample_count
        or source_integrity.get("storage_sample_count") != storage_sample_count
        or source_integrity.get("resource_sample_root") != resource_sample_root
        or source_integrity.get("storage_sample_root") != storage_sample_root
    ):
        return False
    sample_binding = {
        "binding_version": _LIVE_RESOURCE_CHAIN_BINDING,
        "resource_sample_count": resource_sample_count,
        "resource_sample_root": resource_sample_root,
        "storage_sample_count": storage_sample_count,
        "storage_sample_root": storage_sample_root,
    }
    sample_binding_json = json.dumps(
        sample_binding,
        sort_keys=True,
        separators=(",", ":"),
    ).encode()
    running_root = hashlib.sha256(running_root + sample_binding_json).digest()
    if not isinstance(chain_root, str) or not hmac.compare_digest(chain_root, running_root.hex()):
        return False
    if (
        prior_monotonic is None
        or prior_wall is None
        or setup_bias is None
        or first_mono is None
        or first_wall is None
    ):
        return False
    # Single total coverage budget: first/last startup+cleanup must not stack into 2x slack.
    check_span = prior_monotonic - first_mono
    wall_check_span = (prior_wall - first_wall).total_seconds()
    if not math.isfinite(check_span) or not math.isfinite(wall_check_span):
        return False
    minimum_coverage = duration_number - _SOAK_COVERAGE_TOLERANCE_SECONDS
    if check_span < minimum_coverage or wall_check_span < minimum_coverage:
        return False
    if prior_monotonic < minimum_coverage:
        return False
    # Final sample must be near finished_utc (allow brief cleanup after last check).
    cleanup_seconds = (receipt_end - prior_wall).total_seconds()
    if not math.isfinite(cleanup_seconds):
        return False
    if cleanup_seconds < -_SOAK_WALL_MONO_ALIGN_TOLERANCE_SECONDS:
        return False
    if cleanup_seconds > _SOAK_FINAL_CLEANUP_MAX_SECONDS:
        return False
    if wall_span_seconds < minimum_coverage:
        return False
    frozen_key_path = root / "docs" / "security" / "ECHO_SOAK_INTEGRITY_PUBKEY.json"
    try:
        frozen_key = strict_load_object(frozen_key_path)
        public_raw = base64.b64decode(str(frozen_key["public_key_b64"]), validate=True)
        signature = base64.b64decode(
            str(source_integrity["chain_root_signature_b64"]), validate=True
        )
    except (
        OSError,
        KeyError,
        TypeError,
        ValueError,
        json.JSONDecodeError,
        StrictJSONError,
        binascii.Error,
    ):
        return False
    fingerprint = hashlib.sha256(public_raw).hexdigest()
    if (
        frozen_key.get("algorithm") != "Ed25519"
        or frozen_key.get("fingerprint_sha256") != fingerprint
        or source_integrity.get("pubkey_fingerprint") != fingerprint
    ):
        return False
    try:
        Ed25519PublicKey.from_public_bytes(public_raw).verify(signature, chain_root.encode("ascii"))
    except (InvalidSignature, ValueError):
        return False
    if duration_number >= 3600.0 and check_count < math.floor(duration_number / 5.0) - 40:
        # Allow ~200s of probe overhead: each integrity sample hashes the release
        # surface, so the effective period is slightly longer than 5.0s under load.
        return False

    products = data.get("products")
    required_product_checks = (
        "status_primary_healthy",
        "chat",
        "stream",
        "tool_continue",
        "attachment",
        "cross_owner_rejected",
        "secret_rejected",
        "cancel",
        "provider_error_single_terminal",
        "storage_within_limit",
    )
    if not isinstance(products, dict) or set(products) != {"js_agent", "js_work"}:
        return False
    for product_name in ("js_agent", "js_work"):
        checks = products.get(product_name)
        if not isinstance(checks, dict) or any(
            checks.get(check) is not True for check in required_product_checks
        ):
            return False

    provider = data.get("provider")
    if not isinstance(provider, dict) or provider.get("classification_complete") is not True:
        return False
    soak = data.get("soak")
    if not isinstance(soak, dict):
        return False
    sample_count = soak.get("sample_count")
    terminal = soak.get("terminal")
    if (
        not isinstance(sample_count, int)
        or isinstance(sample_count, bool)
        or sample_count <= 0
        or soak.get("success") != sample_count
        or soak.get("failures") != 0
        or soak.get("crosstalk") != 0
        or soak.get("http_5xx") != 0
        or not isinstance(terminal, dict)
        or terminal.get("done") != sample_count
        or terminal.get("error") != 0
    ):
        return False
    for field in ("requested_seconds", "active_elapsed_seconds"):
        value = soak.get(field)
        number = _safe_finite_float(value)
        if number is None or number < 3600.0:
            return False
    # active_elapsed must agree with integrity check span (no inflated active with short span).
    first_check = integrity_checks[0]
    last_check = integrity_checks[-1]
    if not isinstance(first_check, dict) or not isinstance(last_check, dict):
        return False
    last_mono = _safe_finite_float(last_check.get("monotonic_s"))
    first_mono_val = _safe_finite_float(first_check.get("monotonic_s"))
    if last_mono is None or first_mono_val is None:
        return False
    check_span = last_mono - first_mono_val
    active_elapsed = _safe_finite_float(soak.get("active_elapsed_seconds"))
    if active_elapsed is None:
        return False
    if abs(active_elapsed - check_span) > _SOAK_COVERAGE_TOLERANCE_SECONDS:
        return False

    if not _valid_live_resource_evidence(data, duration_seconds=duration_number):
        return False
    resources = data["resources"]
    processes = resources["processes"]

    cleanup = data.get("cleanup")
    children = cleanup.get("children") if isinstance(cleanup, dict) else None
    if (
        not isinstance(cleanup, dict)
        or cleanup.get("all_processes_stopped") is not True
        or cleanup.get("graceful") is not True
        or cleanup.get("errors") != []
        or not isinstance(children, dict)
        or set(children) != {"fake-provider", "js_agent", "js_work"}
    ):
        return False
    if not all(
        isinstance(child, dict)
        and child.get("stopped") is True
        and child.get("forced_kill") is False
        and child.get("stop_error") is None
        for child in children.values()
    ):
        return False
    process_tree = data.get("process_tree")
    wrapper = process_tree.get("wrapper") if isinstance(process_tree, dict) else None
    acceptance = process_tree.get("acceptance") if isinstance(process_tree, dict) else None
    process_children = process_tree.get("children") if isinstance(process_tree, dict) else None

    def _valid_process_snapshot(snapshot: object) -> bool:
        if not isinstance(snapshot, dict):
            return False
        pid = snapshot.get("pid")
        ppid = snapshot.get("ppid")
        started = _safe_finite_float(snapshot.get("create_time"))
        return bool(
            isinstance(pid, int)
            and not isinstance(pid, bool)
            and pid > 0
            and isinstance(ppid, int)
            and not isinstance(ppid, bool)
            and ppid >= 0
            and started is not None
            and started > 0
        )

    if (
        not isinstance(wrapper, dict)
        or not _valid_process_snapshot(wrapper)
        or not isinstance(acceptance, dict)
        or not _valid_process_snapshot(acceptance)
        or not isinstance(process_children, dict)
        or set(process_children) != {"fake-provider", "js_agent", "js_work"}
        or acceptance.get("pid") != acceptance_pid
        or acceptance.get("ppid") != wrapper.get("pid")
    ):
        return False
    if not all(
        _valid_process_snapshot(snapshot) and snapshot.get("ppid") == acceptance_pid
        for snapshot in process_children.values()
    ):
        return False
    for process_name in ("js_agent", "js_work"):
        resource_process = processes.get(process_name)
        child_process = process_children.get(process_name)
        if (
            not isinstance(resource_process, dict)
            or not isinstance(child_process, dict)
            or resource_process.get("pid") != child_process.get("pid")
        ):
            return False
    start_epoch = receipt_start.timestamp()
    end_epoch = receipt_end.timestamp()
    return all(
        start_epoch - 5.0 <= float(snapshot["create_time"]) <= end_epoch
        for snapshot in (acceptance, *process_children.values())
    )


def _valid_spdx_sbom(path: Path) -> bool:
    try:
        data = strict_load_object(path)
    except (OSError, json.JSONDecodeError, StrictJSONError, ValueError):
        return False
    packages = data.get("packages")
    if data.get("spdxVersion") != "SPDX-2.3" or not isinstance(packages, list):
        return False
    return any(
        isinstance(package, dict)
        and package.get("SPDXID") == "SPDXRef-Package-js-agent"
        and package.get("licenseDeclared") == "MIT"
        for package in packages
    )


def _valid_license_scan(path: Path) -> bool:
    text = path.read_text(encoding="utf-8")
    return (
        "Status: COMPLETE_LOCAL_SCAN_EXTERNAL_REVIEW_REQUIRED" in text
        and "Packages scanned:" in text
        and "| Package | Version | License metadata |" in text
    )


def _parse_utc_timestamp(value: object) -> datetime | None:
    if not isinstance(value, str) or not value.strip():
        return None
    text = value.strip()
    if text.endswith("Z"):
        text = text[:-1] + "+00:00"
    try:
        parsed = datetime.fromisoformat(text)
    except ValueError:
        return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=UTC)
    return parsed.astimezone(UTC)


def _path_is_within(path: Path, root: Path) -> bool:
    try:
        path.resolve().relative_to(root.resolve())
    except ValueError:
        return False
    return True


def _normalize_sha256_ref(value: str) -> str:
    """Return a bare 64-hex digest from ``sha256:<hex>`` or bare ``<hex>`` forms."""
    text = value.strip().lower()
    if text.startswith("sha256:"):
        text = text.removeprefix("sha256:")
    return text


def _expand_isolated_e2e_argv(argv: Sequence[object], *, root: Path) -> list[str] | None:
    expanded: list[str] = []
    for item in argv:
        if not isinstance(item, str) or not item:
            return None
        extra = ""
        path_part = item
        if item.startswith("{repo_root}") and "[" in item:
            path_part, extra_tail = item.split("[", 1)
            extra = "[" + extra_tail
        expanded_item = _expand_repo_root_token(path_part, root=root)
        if expanded_item is None:
            return None
        expanded.append(expanded_item + extra)
    return expanded


def _valid_isolated_venv_e2e_step(
    step: object,
    *,
    source_digest: str,
    root: Path,
    evidence_root: Path,
) -> bool:
    if not isinstance(step, dict):
        return False
    if step.get("ok") is not True:
        return False
    if step.get("exit_code") != 0:
        return False
    argv = step.get("argv")
    if (
        not isinstance(argv, list)
        or not argv
        or not all(isinstance(item, str) and item for item in argv)
    ):
        return False
    cwd = step.get("cwd")
    if not isinstance(cwd, str) or not cwd.strip():
        return False
    expanded_cwd = _expand_repo_root_token(cwd, root=root)
    if expanded_cwd is None:
        return False
    expanded_argv = _expand_isolated_e2e_argv(argv, root=root)
    if expanded_argv is None:
        return False
    cwd_path = Path(expanded_cwd)
    if (
        not cwd_path.is_absolute()
        or str(cwd_path.resolve()) != expanded_cwd
        or not cwd_path.is_dir()
    ):
        return False
    if not any(
        _path_is_within(cwd_path, approved)
        for approved in (root.resolve(), evidence_root.resolve())
    ):
        return False
    stdout = step.get("stdout_tail")
    stderr = step.get("stderr_tail")
    if not isinstance(stdout, str) or not isinstance(stderr, str):
        return False
    capture_values: dict[str, Path] = {}
    expanded_captures: dict[str, str] = {}
    for field in ("stdout_path", "stderr_path", "step_receipt_path"):
        raw_path = step.get(field)
        expanded_path = _expand_repo_root_token(raw_path, root=root)
        if expanded_path is None:
            return False
        capture_path = Path(expanded_path).resolve()
        try:
            capture_path.relative_to(evidence_root.resolve())
        except ValueError:
            return False
        if not capture_path.is_file():
            return False
        capture_values[field] = capture_path
        expanded_captures[field] = expanded_path
    stdout_payload = capture_values["stdout_path"].read_bytes()
    stderr_payload = capture_values["stderr_path"].read_bytes()
    if (
        step.get("stdout_sha256") != hashlib.sha256(stdout_payload).hexdigest()
        or step.get("stderr_sha256") != hashlib.sha256(stderr_payload).hexdigest()
        or stdout != stdout_payload.decode("utf-8", errors="replace")[-4000:]
        or stderr != stderr_payload.decode("utf-8", errors="replace")[-4000:]
    ):
        return False
    try:
        step_receipt = strict_load_object(capture_values["step_receipt_path"])
    except (OSError, json.JSONDecodeError, StrictJSONError, ValueError):
        return False
    receipt_expected: dict[str, object] = {
        "argv": expanded_argv,
        "cwd": expanded_cwd,
        "exit_code": step.get("exit_code"),
        "stdout_sha256": step.get("stdout_sha256"),
        "stderr_sha256": step.get("stderr_sha256"),
        "stdout_path": expanded_captures["stdout_path"],
        "stderr_path": expanded_captures["stderr_path"],
        "step_receipt_path": expanded_captures["step_receipt_path"],
    }
    if any(step_receipt.get(field) != receipt_expected[field] for field in receipt_expected):
        return False
    argv = expanded_argv
    if any(Path(argument).name == "true" for argument in argv):
        return False
    name = step.get("step")
    if not isinstance(name, str):
        return False

    def _python_command(command: Sequence[object]) -> bool:
        if not command or not isinstance(command[0], str) or "python" not in Path(command[0]).name:
            return False
        executable = Path(command[0])
        if not executable.is_absolute():
            executable = root / executable
        if not executable.is_file():
            return False
        return any(
            os.path.commonpath((str(executable.absolute()), str(approved.absolute())))
            == str(approved.absolute())
            for approved in (root.resolve(), evidence_root.resolve())
        )

    if name == _ISOLATED_VENV_E2E_REQUIRED_STEPS[0]:
        if (
            not _python_command(argv)
            or argv[1:3] != ["-m", "build"]
            or argv[-1] != "--no-isolation"
            or len(argv) != 6
            or argv[3] != "--outdir"
            or cwd_path != root.resolve()
        ):
            return False
    else:
        kind = "wheel" if name.startswith("wheel:") else "sdist"
        if not name.startswith(f"{kind}:") or cwd_path.name != f"install-{kind}":
            return False
        venv_dir = cwd_path / "venv"
        venv_python = str(venv_dir / "bin" / "python")
        suffix = name.removeprefix(f"{kind}: ")
        if suffix == "create venv":
            if not _python_command(argv) or argv[1:] != ["-m", "venv", str(venv_dir)]:
                return False
        elif suffix == "pip install build backends offline":
            if len(argv) != 13:
                return False
            if (
                argv
                != [
                    venv_python,
                    "-m",
                    "pip",
                    "install",
                    "--no-index",
                    "--find-links",
                    argv[6],
                    "--no-input",
                    "hatchling",
                    "pathspec",
                    "packaging",
                    "trove-classifiers",
                    "pluggy",
                ]
                or not Path(str(argv[6])).is_dir()
                or not _path_is_within(Path(str(argv[6])), evidence_root)
            ):
                return False
        elif suffix == "pip install artifact offline (echo-tokenizer,office)":
            if len(argv) not in {11, 12}:
                return False
            expected_prefix = [
                venv_python,
                "-m",
                "pip",
                "install",
                "--no-index",
                "--find-links",
            ]
            expected_middle = [argv[6], "--no-input", "--report", argv[9]]
            if (
                argv[:6] != expected_prefix
                or argv[6:10] != expected_middle
                or not Path(str(argv[6])).is_dir()
                or not _path_is_within(Path(str(argv[6])), evidence_root)
                or not Path(str(argv[9])).is_file()
            ):
                return False
            artifact_arg = str(argv[-1])
            expected_suffix = ".whl[echo-tokenizer,office]"
            if kind == "sdist":
                expected_suffix = ".tar.gz[echo-tokenizer,office]"
                if argv[10:-1] != ["--no-build-isolation"]:
                    return False
            elif argv[10:-1]:
                return False
            if not artifact_arg.endswith(expected_suffix):
                return False
            artifact_path = Path(artifact_arg.split("[", 1)[0]).resolve()
            if not artifact_path.is_file():
                return False
            report = step.get("pip_report")
            if not isinstance(report, dict):
                return False
            report_path_raw = _expand_repo_root_token(report.get("path"), root=root)
            if report_path_raw is None:
                return False
            report_path = Path(report_path_raw).resolve()
            if (
                not report_path.is_file()
                or not _path_is_within(report_path, evidence_root.resolve())
                or report.get("sha256") != hashlib.sha256(report_path.read_bytes()).hexdigest()
                or not isinstance(report.get("packages"), list)
                or not report["packages"]
            ):
                return False
        elif suffix == "pip check":
            if argv != [venv_python, "-m", "pip", "check"]:
                return False
        elif suffix == "import js/js_work from venv site-packages":
            if (
                len(argv) != 3
                or argv[:2] != [venv_python, "-c"]
                or "import js, js_work" not in str(argv[2])
                or "site-packages" not in str(argv[2])
            ):
                return False
            import_evidence = step.get("import_evidence")
            modules = import_evidence.get("modules") if isinstance(import_evidence, dict) else None
            if (
                not isinstance(import_evidence, dict)
                or import_evidence.get("errors") != []
                or not isinstance(modules, dict)
                or set(modules) != {"js", "js_work"}
            ):
                return False
            for module_evidence in modules.values():
                if not isinstance(module_evidence, dict):
                    return False
                module_file_raw = _expand_repo_root_token(module_evidence.get("file"), root=root)
                site_packages_raw = _expand_repo_root_token(
                    module_evidence.get("site_packages"),
                    root=root,
                )
                if module_file_raw is None or site_packages_raw is None:
                    return False
                module_file = Path(module_file_raw).resolve()
                site_packages = Path(site_packages_raw).resolve()
                claimed_sha = module_evidence.get("file_sha256")
                if (
                    not module_file.is_file()
                    or not site_packages.is_dir()
                    or not _path_is_within(module_file, site_packages)
                    or not _path_is_within(site_packages, cwd_path)
                    or not isinstance(claimed_sha, str)
                    or not hmac.compare_digest(
                        claimed_sha, hashlib.sha256(module_file.read_bytes()).hexdigest()
                    )
                ):
                    return False
        elif suffix == "tokenizer loads offline from vendored cache":
            if (
                len(argv) != 3
                or argv[:2] != [venv_python, "-c"]
                or "tiktoken_counter_factory" not in str(argv[2])
            ):
                return False
        elif suffix == "CLI js --help":
            if argv != [str(venv_dir / "bin" / "js"), "--help"]:
                return False
        elif suffix == "CLI js work --help":
            if argv != [str(venv_dir / "bin" / "js"), "work", "--help"]:
                return False
        elif suffix == "CLI js-work --help":
            if argv != [str(venv_dir / "bin" / "js-work"), "--help"]:
                return False
        elif suffix == "CLI python -m js_work --help":
            if argv != [venv_python, "-m", "js_work", "--help"]:
                return False
        elif suffix == _ISOLATED_VENV_E2E_SERVER_STEP:
            if argv != [venv_python, str(cwd_path / "server_e2e.py")]:
                return False
        else:
            return False
    started = _parse_utc_timestamp(step.get("started_utc"))
    finished = _parse_utc_timestamp(step.get("finished_utc"))
    if started is None or finished is None or finished < started:
        return False
    bound_digest = step.get("source_digest")
    return isinstance(bound_digest, str) and hmac.compare_digest(bound_digest, source_digest)


def _owner_path_slug(owner_key_hash: str) -> str:
    digest = hashlib.sha256(owner_key_hash.encode("utf-8")).hexdigest()
    return f"o_{digest[:24]}"


def _session_path_slug(session_id: str) -> str:
    digest = hashlib.sha256(session_id.encode("utf-8")).hexdigest()
    return f"s_{digest[:24]}"


def _work_ledger_chain_record_dict(record: object) -> dict[str, object]:
    from js.echo.ledger.journal import CommitRecord

    if not isinstance(record, CommitRecord):
        raise TypeError("expected CommitRecord")
    return {
        "seq": record.seq,
        "record_type": record.record_type,
        "tenant_id": record.tenant_id,
        "run_id": record.run_id,
        "payload": record.payload,
        "prev_hash": record.prev_hash,
        "record_hash": record.record_hash,
    }


def _work_ledger_chain_record_from_dict(row: object) -> object:
    from js.echo.ledger.journal import CommitRecord

    if not isinstance(row, dict):
        raise TypeError("expected chain record dict")
    return CommitRecord(
        seq=int(row["seq"]),
        record_type=str(row["record_type"]),
        tenant_id=str(row["tenant_id"]),
        run_id=str(row["run_id"]),
        payload=dict(row["payload"]),
        prev_hash=str(row["prev_hash"]),
        record_hash=str(row["record_hash"]),
        mac=b"",
    )


def _verify_work_ledger_chain_records(records: tuple[object, ...]) -> bool:
    """Verify permit→…→merge lifecycle records.

    Records must be self-consistent (hash payload) and strictly increasing in seq.
    Contiguous ``prev_hash`` linkage is *not* required: real journals may insert
    approval or other records between lifecycle steps while remaining MAC-valid.
    """
    from js.echo.ledger._hashing import stable_hash
    from js.echo.ledger.journal import CommitRecord

    if len(records) != len(_WORK_LEDGER_CHAIN_TYPES):
        return False
    effect_id: str | None = None
    run_id: str | None = None
    tenant_id: str | None = None
    previous_seq: int | None = None
    for expected_type, record in zip(_WORK_LEDGER_CHAIN_TYPES, records, strict=True):
        if not isinstance(record, CommitRecord):
            return False
        if record.record_type != expected_type:
            return False
        recomputed_hash = stable_hash(record.hash_payload())
        if record.record_hash != recomputed_hash:
            return False
        if previous_seq is not None and record.seq <= previous_seq:
            return False
        previous_seq = record.seq
        payload = record.payload if isinstance(record.payload, dict) else {}
        row_effect = str(payload.get("effect_id") or "")
        if not row_effect:
            return False
        if effect_id is None:
            effect_id = row_effect
            run_id = record.run_id
            tenant_id = record.tenant_id
        elif row_effect != effect_id or record.run_id != run_id or record.tenant_id != tenant_id:
            return False
    return True


def _effect_lifecycle_chain(
    records: tuple[object, ...],
    effect_id: str,
) -> tuple[object, ...] | None:
    from js.echo.ledger.journal import CommitRecord

    by_type: dict[str, CommitRecord] = {}
    for record in records:
        if not isinstance(record, CommitRecord):
            continue
        payload = record.payload if isinstance(record.payload, dict) else {}
        if str(payload.get("effect_id") or "") != effect_id:
            continue
        if record.record_type in _WORK_LEDGER_CHAIN_TYPES:
            by_type[record.record_type] = record
    if set(by_type) != set(_WORK_LEDGER_CHAIN_TYPES):
        return None
    chain = tuple(by_type[record_type] for record_type in _WORK_LEDGER_CHAIN_TYPES)
    if not all(chain[index].seq < chain[index + 1].seq for index in range(len(chain) - 1)):
        return None
    return chain


def _find_work_effect_chain(
    records: tuple[object, ...],
    *,
    run_id: str,
    tool_name: str,
) -> tuple[tuple[Any, ...], str, str] | None:
    from js.echo.ledger.journal import CommitRecord

    for record in reversed(records):
        if not isinstance(record, CommitRecord):
            continue
        if record.record_type != "merge" or record.run_id != run_id:
            continue
        payload = record.payload if isinstance(record.payload, dict) else {}
        effect_id = str(payload.get("effect_id") or "")
        if not effect_id or str(payload.get("status") or "") != "ok":
            continue
        chain = _effect_lifecycle_chain(records, effect_id)
        if chain is None:
            continue
        permit = chain[0]
        if not isinstance(permit, CommitRecord):
            continue
        permit_payload = permit.payload if isinstance(permit.payload, dict) else {}
        tool_effect = permit_payload.get("tool_effect")
        if not isinstance(tool_effect, dict):
            continue
        if tool_effect.get("tool_name") != tool_name:
            continue
        lease_id = str(tool_effect.get("lease_id") or "")
        if not lease_id:
            continue
        return chain, effect_id, lease_id
    return None


def _read_work_journal_records(journal_path: Path) -> list[object]:
    from js.echo.ledger.journal import _read_file_records

    return list(_read_file_records(journal_path))


def build_work_ledger_receipt_binding(
    *,
    journal_path: Path,
    mac_key: bytes,
    state_dir: Path,
    owner: str,
    session: str,
    product_id: str,
    run_id: str,
    tool_name: str,
) -> dict[str, object] | None:
    """Extract durable permit→claim→receipt→merge binding from a verified journal."""
    from js.echo.ledger.journal import verify_file

    report = verify_file(journal_path, mac_key=mac_key)
    if not report.ok:
        return None
    records = tuple(_read_work_journal_records(journal_path))
    located = _find_work_effect_chain(records, run_id=run_id, tool_name=tool_name)
    if located is None:
        return None
    chain, effect_id, lease_id = located
    if not _verify_work_ledger_chain_records(chain):
        return None
    from js.echo.ledger.journal import CommitRecord

    merge_record = chain[-1]
    receipt_record = chain[-2]
    if not isinstance(merge_record, CommitRecord) or not isinstance(receipt_record, CommitRecord):
        return None
    receipt_payload = receipt_record.payload if isinstance(receipt_record.payload, dict) else {}
    terminal_status = str(receipt_payload.get("status") or "")
    if terminal_status != "ok":
        return None
    try:
        journal_relative_path = str(journal_path.resolve().relative_to(state_dir.resolve()))
    except ValueError:
        return None
    return {
        "journal_relative_path": journal_relative_path,
        "journal_sha256": hashlib.sha256(journal_path.read_bytes()).hexdigest(),
        "ledger_sequence": merge_record.seq,
        "record_hash": merge_record.record_hash,
        "effect_id": effect_id,
        "run_id": run_id,
        "lease_id": lease_id,
        "tool_name": tool_name,
        "terminal_status": terminal_status,
        "owner": owner,
        "session": session,
        "product_id": product_id,
        "lease_consumed": True,
        "ledger_chain": [_work_ledger_chain_record_dict(record) for record in chain],
    }


def _verify_work_ledger_receipt_binding(
    receipt: Mapping[str, object],
    *,
    evidence_root: Path | None = None,
    repo_root: Path | None = None,
) -> bool:
    owner = receipt.get("owner")
    session = receipt.get("session")
    run_id = receipt.get("run_id")
    tool_name = receipt.get("tool_name")
    lease_id = receipt.get("lease_id")
    effect_id = receipt.get("effect_id")
    terminal_status = receipt.get("terminal_status")
    journal_relative_path = receipt.get("journal_relative_path")
    journal_sha256 = receipt.get("journal_sha256")
    ledger_sequence = receipt.get("ledger_sequence")
    record_hash = receipt.get("record_hash")
    ledger_chain = receipt.get("ledger_chain")
    if not all(
        isinstance(value, str) and value.strip()
        for value in (
            owner,
            session,
            run_id,
            tool_name,
            lease_id,
            effect_id,
            terminal_status,
            journal_relative_path,
            record_hash,
        )
    ):
        return False
    if not isinstance(journal_sha256, str) or len(journal_sha256) != 64:
        return False
    if not isinstance(ledger_sequence, int) or isinstance(ledger_sequence, bool):
        return False
    if not isinstance(ledger_chain, list) or len(ledger_chain) != len(_WORK_LEDGER_CHAIN_TYPES):
        return False
    try:
        chain = tuple(_work_ledger_chain_record_from_dict(row) for row in ledger_chain)
    except (KeyError, TypeError, ValueError):
        return False
    if not _verify_work_ledger_chain_records(chain):
        return False
    from js.echo.ledger.journal import CommitRecord

    merge_record = chain[-1]
    receipt_record = chain[-2]
    permit_record = chain[0]
    if (
        not isinstance(merge_record, CommitRecord)
        or not isinstance(receipt_record, CommitRecord)
        or not isinstance(permit_record, CommitRecord)
    ):
        return False
    if merge_record.seq != ledger_sequence or merge_record.record_hash != record_hash:
        return False
    permit_payload = permit_record.payload if isinstance(permit_record.payload, dict) else {}
    tool_effect = permit_payload.get("tool_effect")
    if not isinstance(tool_effect, dict):
        return False
    if (
        tool_effect.get("tool_name") != tool_name
        or str(tool_effect.get("lease_id") or "") != lease_id
        or str(permit_payload.get("effect_id") or "") != effect_id
    ):
        return False
    for record in chain:
        if not isinstance(record, CommitRecord):
            return False
        if record.run_id != run_id or record.tenant_id != owner:
            return False
    receipt_payload = receipt_record.payload if isinstance(receipt_record.payload, dict) else {}
    merge_payload = merge_record.payload if isinstance(merge_record.payload, dict) else {}
    if (
        str(receipt_payload.get("effect_id") or "") != effect_id
        or str(merge_payload.get("effect_id") or "") != effect_id
        or str(receipt_payload.get("status") or "") != terminal_status
        or str(merge_payload.get("status") or "") != terminal_status
        or terminal_status != "ok"
    ):
        return False
    if receipt.get("product_id") != "js-work":
        return False
    if receipt.get("lease_consumed") is not True:
        return False
    from js.echo.ledger.service import _scope_partition_slugs

    product_slug, owner_slug, session_slug = _scope_partition_slugs(
        tenant_id=str(owner),
        product_id=str(receipt["product_id"]),
        session_id=str(session),
    )
    expected_journal_relative_path = (
        f"echo/ledger/partitions/{product_slug}/{owner_slug}/{session_slug}/chat.jsonl"
    )
    if journal_relative_path != expected_journal_relative_path:
        return False
    if evidence_root is None:
        return True

    journal_evidence_path = receipt.get("journal_evidence_path")
    if not isinstance(journal_evidence_path, str) or not journal_evidence_path.strip():
        return False
    evidence_root = evidence_root.resolve()
    journal_path = (evidence_root / journal_evidence_path).resolve()
    try:
        journal_path.relative_to(evidence_root)
    except ValueError:
        return False
    try:
        journal_payload = journal_path.read_bytes()
    except OSError:
        return False
    if not hmac.compare_digest(hashlib.sha256(journal_payload).hexdigest(), journal_sha256):
        return False

    from js.echo.ledger._hashing import stable_hash
    from js.echo.ledger.journal import GENESIS_HASH, CommitRecord

    trusted_records = tuple(_read_work_journal_records(journal_path))
    previous = GENESIS_HASH
    for index, record in enumerate(trusted_records):
        if (
            not isinstance(record, CommitRecord)
            or record.seq != index
            or record.prev_hash != previous
            or record.record_hash != stable_hash(record.hash_payload())
        ):
            return False
        previous = record.record_hash
    frozen_path: Path | None = None
    if repo_root is not None:
        candidate = repo_root.resolve() / "docs" / "security" / "ECHO_E2E_LEDGER_PUBKEY.json"
        if candidate.is_file():
            frozen_path = candidate
    if frozen_path is None:
        public_key_path = evidence_root.parent
        while (
            public_key_path != public_key_path.parent
            and not (
                public_key_path / "docs" / "security" / "ECHO_E2E_LEDGER_PUBKEY.json"
            ).is_file()
        ):
            public_key_path = public_key_path.parent
        frozen_path = public_key_path / "docs" / "security" / "ECHO_E2E_LEDGER_PUBKEY.json"
    try:
        frozen = strict_load_object(frozen_path)
        raw_public = base64.b64decode(str(frozen["public_key_b64"]), validate=True)
        signature = base64.b64decode(str(receipt["ledger_signature_b64"]), validate=True)
    except (
        OSError,
        KeyError,
        TypeError,
        ValueError,
        json.JSONDecodeError,
        StrictJSONError,
        binascii.Error,
    ):
        return False
    fingerprint = hashlib.sha256(raw_public).hexdigest()
    if (
        frozen.get("algorithm") != "Ed25519"
        or frozen.get("fingerprint_sha256") != fingerprint
        or receipt.get("pubkey_fingerprint") != fingerprint
    ):
        return False
    signature_payload = {
        field: receipt.get(field)
        for field in (
            "journal_sha256",
            "arguments_sha256",
            "output_sha256",
            "product_id",
            "owner",
            "session",
            "run_id",
            "effect_id",
        )
    }
    if any(not isinstance(value, str) or not value for value in signature_payload.values()):
        return False
    canonical_signature_payload = json.dumps(
        signature_payload, sort_keys=True, separators=(",", ":")
    ).encode()
    try:
        Ed25519PublicKey.from_public_bytes(raw_public).verify(
            signature, canonical_signature_payload
        )
    except (InvalidSignature, ValueError):
        return False
    located = _find_work_effect_chain(
        trusted_records,
        run_id=str(run_id),
        tool_name=str(tool_name),
    )
    if located is None:
        return False
    trusted_chain, trusted_effect_id, trusted_lease_id = located
    if (
        trusted_effect_id != effect_id
        or trusted_lease_id != lease_id
        or [_work_ledger_chain_record_dict(item) for item in trusted_chain] != ledger_chain
    ):
        return False
    trusted_permit, trusted_outbox, trusted_claim, trusted_receipt, trusted_merge = trusted_chain
    if not all(
        isinstance(item, CommitRecord)
        for item in (trusted_permit, trusted_outbox, trusted_claim, trusted_receipt, trusted_merge)
    ):
        return False
    permit_payload = trusted_permit.payload
    tool_effect = permit_payload.get("tool_effect") if isinstance(permit_payload, dict) else None
    outbox_payload = trusted_outbox.payload
    claim_payload = trusted_claim.payload
    trusted_receipt_payload = trusted_receipt.payload
    trusted_merge_payload = trusted_merge.payload
    if not (
        isinstance(tool_effect, dict)
        and isinstance(outbox_payload, dict)
        and isinstance(claim_payload, dict)
        and isinstance(trusted_receipt_payload, dict)
        and isinstance(trusted_merge_payload, dict)
    ):
        return False
    args_hash = str(tool_effect.get("args_hash") or "")
    outbox_id = str(outbox_payload.get("outbox_id") or "")
    seal_id = str(permit_payload.get("seal_id") or "")
    permit_seal = permit_payload.get("seal")
    outbox_seal = outbox_payload.get("seal")
    if (
        tool_effect.get("product_id") != receipt.get("product_id")
        or tool_effect.get("session_id") != session
        or tool_effect.get("tool_name") != tool_name
        or str(tool_effect.get("lease_id") or "") != lease_id
        or not args_hash
        or str(outbox_payload.get("sealed_input_ref") or "") != args_hash
        or not seal_id
        or not isinstance(permit_seal, dict)
        or permit_seal != outbox_seal
        or permit_seal.get("seal_id") != seal_id
        or permit_seal.get("effect_id") != effect_id
        or permit_seal.get("tenant_id") != owner
        or not outbox_id
        or str(claim_payload.get("outbox_id") or "") != outbox_id
        or str(trusted_receipt_payload.get("outbox_id") or "") != outbox_id
        or not str(
            trusted_receipt_payload.get("output_hash")
            or trusted_receipt_payload.get("output_ref")
            or ""
        )
        or _normalize_sha256_ref(str(trusted_receipt_payload.get("output_hash") or ""))
        != _normalize_sha256_ref(str(receipt.get("output_sha256") or ""))
        or len(_normalize_sha256_ref(str(receipt.get("output_sha256") or ""))) != 64
        or args_hash != receipt.get("arguments_sha256")
        or str(trusted_receipt_payload.get("status") or "") != terminal_status
        or str(trusted_merge_payload.get("status") or "") != terminal_status
        or trusted_merge.seq != ledger_sequence
        or trusted_merge.record_hash != record_hash
    ):
        return False
    return receipt.get("status") == terminal_status and receipt.get("terminal") is (
        terminal_status == "ok"
    )


def _read_xlsx_first_row_cells(path: Path) -> list[list[str]] | None:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        cells = [[str(cell.value) for cell in row] for row in sheet.iter_rows(min_row=1, max_row=1)]
        workbook.close()
    except OSError:
        return None
    return cells


def _valid_isolated_venv_e2e_work_receipt(
    receipt: object,
    *,
    evidence_root: Path | None = None,
    work_output: Mapping[str, object] | None = None,
    repo_root: Path | None = None,
) -> bool:
    if not isinstance(receipt, dict):
        return False
    if receipt.get("product_id") != "js-work":
        return False
    owner = receipt.get("owner")
    session = receipt.get("session")
    run_id = receipt.get("run_id")
    tool_name = receipt.get("tool_name")
    lease_id = receipt.get("lease_id")
    output_path = receipt.get("output_path")
    output_sha256 = receipt.get("output_sha256")
    output_cells = receipt.get("output_cells")
    if not all(
        isinstance(value, str) and value.strip()
        for value in (owner, session, run_id, tool_name, lease_id)
    ):
        return False
    if tool_name != "excel_write":
        return False
    if receipt.get("lease_consumed") is not True:
        return False
    if receipt.get("journal_records_added") is not True:
        return False
    if receipt.get("status") != "ok":
        return False
    if receipt.get("terminal") is not True:
        return False
    if receipt.get("output_exists") is not True:
        return False
    if not isinstance(output_sha256, str) or len(output_sha256) != 64:
        return False
    if not isinstance(output_path, str) or not output_path.strip():
        return False
    if not isinstance(output_cells, list) or len(output_cells) != 1:
        return False
    row = output_cells[0]
    if not isinstance(row, list) or row != ["iso", "e2e", "leased"]:
        return False
    owner_root = f"owners/{_owner_path_slug(str(owner))}/{_session_path_slug(str(session))}"
    if owner_root not in output_path.replace("\\", "/"):
        return False
    if not _verify_work_ledger_receipt_binding(
        receipt, evidence_root=evidence_root, repo_root=repo_root
    ):
        return False
    if work_output is None or evidence_root is None:
        return True
    if not isinstance(work_output, dict):
        return False
    relative_path = work_output.get("path")
    claimed_sha = work_output.get("sha256")
    claimed_cells = work_output.get("cells")
    if not isinstance(relative_path, str) or not relative_path.strip():
        return False
    if not isinstance(claimed_sha, str) or len(claimed_sha) != 64:
        return False
    if claimed_cells != [["iso", "e2e", "leased"]]:
        return False
    artifact_path = (evidence_root / relative_path).resolve()
    if not artifact_path.is_file():
        return False
    actual_sha = hashlib.sha256(artifact_path.read_bytes()).hexdigest()
    if not hmac.compare_digest(actual_sha, claimed_sha):
        return False
    if not hmac.compare_digest(actual_sha, output_sha256):
        return False
    reread_cells = _read_xlsx_first_row_cells(artifact_path)
    return reread_cells == [["iso", "e2e", "leased"]]


def _valid_isolated_venv_e2e_work_output(
    work_output: object,
    *,
    evidence_root: Path,
) -> bool:
    if not isinstance(work_output, dict):
        return False
    relative_path = work_output.get("path")
    claimed_sha = work_output.get("sha256")
    claimed_bytes = work_output.get("bytes")
    claimed_cells = work_output.get("cells")
    if not isinstance(relative_path, str) or not relative_path.strip():
        return False
    if not isinstance(claimed_sha, str) or len(claimed_sha) != 64:
        return False
    if isinstance(claimed_bytes, bool) or not isinstance(claimed_bytes, int) or claimed_bytes < 0:
        return False
    if claimed_cells != [["iso", "e2e", "leased"]]:
        return False
    artifact_path = (evidence_root / relative_path).resolve()
    try:
        artifact_path.relative_to(evidence_root.resolve())
    except ValueError:
        return False
    if not artifact_path.is_file():
        return False
    actual_sha = hashlib.sha256(artifact_path.read_bytes()).hexdigest()
    if not hmac.compare_digest(actual_sha, claimed_sha):
        return False
    if artifact_path.stat().st_size != claimed_bytes:
        return False
    reread_cells = _read_xlsx_first_row_cells(artifact_path)
    return reread_cells == [["iso", "e2e", "leased"]]


def _valid_isolated_venv_e2e_provider_detail(
    detail: object,
    *,
    evidence_root: Path | None = None,
    work_output: Mapping[str, object] | None = None,
    repo_root: Path | None = None,
) -> bool:
    if not isinstance(detail, dict):
        return False
    if detail.get("chat_status") != 200:
        return False
    provider_calls = detail.get("provider_calls")
    if (
        not isinstance(provider_calls, int)
        or isinstance(provider_calls, bool)
        or provider_calls <= 0
    ):
        return False
    if detail.get("attachment_consumed") is not True:
        return False
    if detail.get("attachment_marker_in_messages") is not True:
        return False
    if detail.get("all_provider_hosts_loopback") is not True:
        return False
    unexpected = detail.get("unexpected_provider_calls")
    if unexpected != []:
        return False
    scenarios = detail.get("provider_scenarios")
    if not isinstance(scenarios, dict):
        return False
    for scenario in ("chat", "attachment", "tool"):
        count = scenarios.get(scenario)
        if not isinstance(count, int) or isinstance(count, bool) or count <= 0:
            return False
    if detail.get("ws_terminal_ok") is not True:
        return False
    if detail.get("ws_ok") is False or detail.get("ws_terminal") == "error":
        return False
    for frame_key in (
        "ws_saw_token",
        "ws_saw_thinking",
        "ws_saw_tool_call",
        "ws_saw_done",
    ):
        if detail.get(frame_key) is not True:
            return False
    return _valid_isolated_venv_e2e_work_receipt(
        detail.get("work_receipt"),
        evidence_root=evidence_root,
        work_output=work_output,
        repo_root=repo_root,
    )


def _resolve_isolated_e2e_evidence_root(root: Path, data: Mapping[str, object]) -> Path | None:
    evidence_root = data.get("evidence_root")
    expanded = _expand_repo_root_token(evidence_root, root=root)
    if expanded is None:
        return None
    # Accept repo-relative, {repo_root}/rel, and absolute (external evidence) paths.
    candidate = Path(expanded)
    if not candidate.is_absolute():
        candidate = (root / expanded).resolve()
        resolved_root = root.resolve()
        try:
            candidate.relative_to(resolved_root)
        except ValueError:
            return None
    return candidate if candidate.is_dir() else None


def _valid_isolated_venv_e2e_artifact_entry(
    entry: object,
    *,
    evidence_root: Path,
    kind: str,
) -> bool:
    if not isinstance(entry, dict):
        return False
    relative_path = entry.get("path")
    claimed_sha = entry.get("sha256")
    claimed_bytes = entry.get("bytes")
    if not isinstance(relative_path, str) or not relative_path.strip():
        return False
    if Path(relative_path).is_absolute():
        return False
    if not isinstance(claimed_sha, str) or len(claimed_sha) != 64:
        return False
    if isinstance(claimed_bytes, bool) or not isinstance(claimed_bytes, int) or claimed_bytes < 0:
        return False
    artifact_path = (evidence_root / relative_path).resolve()
    try:
        artifact_path.relative_to(evidence_root.resolve())
    except ValueError:
        return False
    if not artifact_path.is_file():
        return False
    payload = artifact_path.read_bytes()
    actual_sha = hashlib.sha256(payload).hexdigest()
    if not hmac.compare_digest(actual_sha, claimed_sha):
        return False
    if artifact_path.stat().st_size != claimed_bytes:
        return False
    if kind == "wheel" and artifact_path.suffix != ".whl":
        return False
    return not (kind == "sdist" and not artifact_path.name.endswith(".tar.gz"))


def _valid_isolated_venv_e2e(root: Path, path: Path) -> bool:
    """Validate offline isolated wheel/sdist E2E evidence bound to source digest."""
    try:
        data = strict_load_object(path)
    except (OSError, json.JSONDecodeError, StrictJSONError, ValueError):
        return False
    if data.get("ok") is not True:
        return False
    if data.get("schema_version") != _ISOLATED_VENV_E2E_SCHEMA_VERSION:
        return False
    if data.get("offline") is not True:
        return False
    source_digest = data.get("source_digest")
    if not isinstance(source_digest, str) or not hmac.compare_digest(
        source_digest,
        release_source_digest(root),
    ):
        return False
    evidence_root = _resolve_isolated_e2e_evidence_root(root, data)
    if evidence_root is None:
        return False
    artifacts = data.get("artifacts")
    if not isinstance(artifacts, dict):
        return False
    for kind in ("wheel", "sdist"):
        entry = artifacts.get(kind)
        if not _valid_isolated_venv_e2e_artifact_entry(
            entry,
            evidence_root=evidence_root,
            kind=kind,
        ):
            return False
    work_output = data.get("work_output")
    work_outputs = data.get("work_outputs")
    if (
        not isinstance(work_output, dict)
        or not isinstance(work_outputs, dict)
        or set(work_outputs) != {"wheel", "sdist"}
    ):
        return False
    for kind in ("wheel", "sdist"):
        kind_output = work_outputs.get(kind)
        if not _valid_isolated_venv_e2e_work_output(
            kind_output,
            evidence_root=evidence_root,
        ):
            return False
        if (
            not isinstance(kind_output, dict)
            or kind_output.get("path") != f"e2e/work/{kind}/iso-e2e.xlsx"
        ):
            return False
    if work_output != work_outputs["sdist"]:
        return False
    manifest = data.get("manifest")
    if not isinstance(manifest, list):
        return False
    manifest_paths: list[str] = []
    for item in manifest:
        if not _valid_isolated_venv_e2e_artifact_entry(
            item,
            evidence_root=evidence_root,
            kind="internal",
        ):
            return False
        manifest_paths.append(str(item["path"]))
    expected_manifest_paths = {str(artifacts[kind]["path"]) for kind in ("wheel", "sdist")}
    for kind in ("wheel", "sdist"):
        expected_manifest_paths.update(
            {
                f"e2e/work/{kind}/iso-e2e.xlsx",
                f"e2e/work/{kind}/ledger.journal",
            }
        )
    if len(manifest_paths) != len(set(manifest_paths)) or set(manifest_paths) != (
        expected_manifest_paths
    ):
        return False

    results = data.get("results")
    if not isinstance(results, list):
        return False
    if len(results) != len(_ISOLATED_VENV_E2E_REQUIRED_STEPS):
        return False

    expected_digest = release_source_digest(root)
    step_names: list[str] = []
    for step in results:
        if not isinstance(step, dict):
            return False
        name = step.get("step")
        if not isinstance(name, str) or not name.strip():
            return False
        step_names.append(name)
    if step_names != list(_ISOLATED_VENV_E2E_REQUIRED_STEPS):
        return False
    for required, step in zip(_ISOLATED_VENV_E2E_REQUIRED_STEPS, results, strict=True):
        if step.get("step") != required:
            return False
        if not _valid_isolated_venv_e2e_step(
            step,
            source_digest=expected_digest,
            root=root,
            evidence_root=evidence_root,
        ):
            return False

    pip_field = data.get("pip_check")
    if not isinstance(pip_field, dict):
        return False
    for kind in ("wheel", "sdist"):
        pip_step = next(item for item in results if item.get("step") == f"{kind}: pip check")
        field_entry = pip_field.get(kind)
        step_ok = pip_step.get("ok") is True and pip_step.get("exit_code") == 0
        if isinstance(field_entry, dict):
            field_ok = field_entry.get("ok") is True and field_entry.get("exit_code") == 0
        else:
            field_ok = False
        if not step_ok or not field_ok:
            return False

    for kind in ("wheel", "sdist"):
        server_step = next(
            item
            for item in results
            if item.get("step") == f"{kind}: {_ISOLATED_VENV_E2E_SERVER_STEP}"
        )
        detail = server_step.get("detail") if isinstance(server_step.get("detail"), dict) else {}
        if not _valid_isolated_venv_e2e_provider_detail(
            detail,
            evidence_root=evidence_root,
            work_output=work_outputs[kind],
            repo_root=root,
        ):
            return False

    for step in results:
        detail = step.get("detail") if isinstance(step.get("detail"), dict) else {}
        chat = detail.get("chat_status", step.get("chat_status"))
        if chat is not None and chat != 200:
            return False
        if detail.get("ws_ok") is False or step.get("ws_ok") is False:
            return False
        if detail.get("ws_terminal_ok") is False or step.get("ws_terminal_ok") is False:
            return False
        if detail.get("ws_terminal") == "error" or step.get("ws_terminal") == "error":
            return False

    # Round 8.10: non-secret key provenance must prove random external-temp lifecycle.
    provenance_path = evidence_root / "e2e" / "E2E_KEY_PROVENANCE.json"
    try:
        provenance = strict_load_object(provenance_path)
        from js.echo.ledger.e2e_signing import assert_provenance_destroyed

        assert_provenance_destroyed(provenance)
        frozen_key = strict_load_object(root / "docs" / "security" / "ECHO_E2E_LEDGER_PUBKEY.json")
    except (OSError, json.JSONDecodeError, StrictJSONError, ValueError, RuntimeError):
        return False
    return bool(provenance.get("public_fingerprint") == frozen_key.get("fingerprint_sha256"))


def _resolve_repo_root(path: Path) -> Path | None:
    """Locate the repository root from an evidence or docs artifact path.

    ``docs/security/<file>`` keeps the historical ``parents[2]`` binding so
    synthetic readiness fixtures under a temporary root continue to work.
    Evidence paths such as ``…/slo/slo_run_N.json`` walk upward for
    ``pyproject.toml`` + ``js/``.
    """
    resolved = path.resolve()
    if (
        len(resolved.parents) >= 2
        and resolved.parents[0].name == "security"
        and resolved.parents[1].name == "docs"
    ):
        return resolved.parents[2]
    for candidate in (resolved, *resolved.parents):
        if (candidate / "pyproject.toml").is_file() and (candidate / "js").is_dir():
            return candidate
    return None


def _slo_latency_summary(values: list[float]) -> dict[str, float | int]:
    ordered = sorted(values)
    p95_index = min(len(ordered) - 1, max(0, round((len(ordered) - 1) * 0.95)))
    p95_ms = round(ordered[p95_index], 3)
    max_ms = max(round(max(values), 3), p95_ms)
    return {
        "n": len(values),
        "min_ms": round(min(values), 3),
        "mean_ms": round(sum(values) / len(values), 3),
        "p50_ms": round(float(median(values)), 3),
        "p95_ms": p95_ms,
        "max_ms": max_ms,
    }


def _valid_ws_stream_timing_receipt(
    timing: object,
    *,
    top_level: bool,
    aggregate_first_p95: float | None = None,
    aggregate_terminal_p95: float | None = None,
    group_first_p95s: list[float] | None = None,
    group_terminal_p95s: list[float] | None = None,
) -> tuple[float, float] | None:
    if not isinstance(timing, dict) or set(timing) != {
        "evidence_schema_version",
        "clock",
        "configured_cadence_ms",
        "first_text_token_semantics",
        "timing_receipts",
        "status_latency",
        "thinking_latency",
        "first_text_token_latency",
        "usage_latency",
        "terminal_latency",
        "provider_stream_event_calls",
        "provider_stream_completed",
        "provider_stream_cancelled",
        "failures",
        "resilience",
    }:
        return None
    if (
        timing.get("evidence_schema_version") != "echo-ws-stream-timing-v1"
        or timing.get("clock") != "time.perf_counter_ns"
        or timing.get("configured_cadence_ms") != {"first_text_token": 5.0, "inter_text_token": 1.0}
        or timing.get("first_text_token_semantics")
        != "first non-empty websocket token frame after send"
        or timing.get("failures") != []
    ):
        return None
    receipts = timing.get("timing_receipts")
    if not isinstance(receipts, list) or len(receipts) != SLO_CONTRACT.benchmark_measured:
        return None
    offset_names = ("status", "thinking", "first_text_token", "usage", "terminal")
    offset_values: dict[str, list[float]] = {name: [] for name in offset_names}
    expected_frames = ["status", "thinking", "token", "token", "token", "usage", "done"]
    for receipt in receipts:
        if not isinstance(receipt, dict) or set(receipt) != {
            "send_monotonic_ns",
            "clock",
            "frame_offsets_ms",
            "frame_types",
            "terminal_count",
            "terminal_type",
        }:
            return None
        send_ns = receipt.get("send_monotonic_ns")
        offsets = receipt.get("frame_offsets_ms")
        if (
            isinstance(send_ns, bool)
            or not isinstance(send_ns, int)
            or send_ns <= 0
            or receipt.get("clock") != "time.perf_counter_ns"
            or not isinstance(offsets, dict)
            or set(offsets) != set(offset_names)
            or receipt.get("frame_types") != expected_frames
            or receipt.get("terminal_count") != 1
            or receipt.get("terminal_type") != "done"
        ):
            return None
        ordered_offsets: list[float] = []
        for name in offset_names:
            value = _safe_finite_float(offsets.get(name))
            if value is None or value < 0:
                return None
            ordered_offsets.append(value)
            offset_values[name].append(value)
        if any(
            left > right for left, right in zip(ordered_offsets, ordered_offsets[1:], strict=False)
        ):
            return None
    for name in offset_names:
        field = f"{name}_latency"
        summary = timing.get(field)
        expected_summary = _slo_latency_summary(offset_values[name])
        if top_level and name in {"first_text_token", "terminal"}:
            if not isinstance(summary, dict):
                return None
            expected_keys = set(expected_summary) | {"group_p95_ms"}
            aggregate_p95 = (
                aggregate_first_p95 if name == "first_text_token" else aggregate_terminal_p95
            )
            group_p95s = group_first_p95s if name == "first_text_token" else group_terminal_p95s
            if (
                set(summary) != expected_keys
                or aggregate_p95 is None
                or group_p95s is None
                or not _live_json_matches(summary.get("p95_ms"), aggregate_p95)
                or not _live_json_matches(summary.get("group_p95_ms"), group_p95s)
                or any(
                    not _live_json_matches(summary.get(key), value)
                    for key, value in expected_summary.items()
                    if key != "p95_ms"
                )
            ):
                return None
        elif not _live_json_matches(summary, expected_summary):
            return None
    sample_count = len(receipts)
    if (
        timing.get("provider_stream_event_calls") != sample_count
        or timing.get("provider_stream_completed") != sample_count
        or timing.get("provider_stream_cancelled") != 0
    ):
        return None
    resilience = timing.get("resilience")
    if not isinstance(resilience, dict) or set(resilience) != {
        "single_terminal_all_ok",
        "slow_consumer",
        "disconnect",
    }:
        return None
    slow_consumer = resilience.get("slow_consumer")
    disconnect = resilience.get("disconnect")
    if (
        resilience.get("single_terminal_all_ok") is not True
        or slow_consumer
        != {
            "ok": True,
            "consumer_pause_ms": 10.0,
            "bounded_max_frames": 7,
            "received_frame_count": 7,
            "terminal_count": 1,
            "terminal_type": "done",
        }
        or not isinstance(disconnect, dict)
        or set(disconnect)
        != {
            "ok",
            "status_received",
            "provider_started",
            "provider_cancelled",
            "terminal_frames_after_disconnect",
            "bounded_wait_ms",
            "max_wait_ms",
        }
    ):
        return None
    bounded_wait = _safe_finite_float(disconnect.get("bounded_wait_ms"))
    max_wait = _safe_finite_float(disconnect.get("max_wait_ms"))
    if (
        disconnect.get("ok") is not True
        or disconnect.get("status_received") is not True
        or disconnect.get("provider_started") is not True
        or disconnect.get("provider_cancelled") is not True
        or disconnect.get("terminal_frames_after_disconnect") != 0
        or bounded_wait is None
        or bounded_wait < 0
        or max_wait != 1_000.0
        or bounded_wait > max_wait
    ):
        return None
    first_p95 = _safe_finite_float(timing["first_text_token_latency"].get("p95_ms"))
    terminal_p95 = _safe_finite_float(timing["terminal_latency"].get("p95_ms"))
    if (
        first_p95 is None
        or first_p95 > SLO_CONTRACT.ws_first_token_p95_ms
        or terminal_p95 is None
        or terminal_p95 > SLO_CONTRACT.ws_terminal_p95_ms
    ):
        return None
    return first_p95, terminal_p95


def _slo_canonical_sha256(value: object) -> str | None:
    try:
        encoded = json.dumps(
            value,
            ensure_ascii=False,
            separators=(",", ":"),
            sort_keys=True,
            allow_nan=False,
        ).encode("utf-8")
    except (TypeError, ValueError, OverflowError):
        return None
    return hashlib.sha256(encoded).hexdigest()


@lru_cache(maxsize=1)
def _expected_compaction_semantic_fields() -> dict[str, object]:
    effect_id = "benchmark-effect-1"
    expected: list[dict[str, object]] = [
        {
            "record_type": "outbox",
            "tenant_id": "bench",
            "run_id": "compact-effect",
            "payload": {
                "outbox_id": "benchmark-outbox-1",
                "effect_id": effect_id,
                "seal": {
                    "action_kind": "tool.file_write",
                    "replay_class": "non_idempotent",
                },
            },
        },
        {
            "record_type": "merge",
            "tenant_id": "bench",
            "run_id": "compact-effect",
            "payload": {"effect_id": effect_id},
        },
    ]
    for index in range(1_000):
        decision: dict[str, object] = {
            "record_type": "decision",
            "tenant_id": "bench",
            "run_id": f"compact-{index}",
            "payload": {"decision_id": f"c{index}", "idx": index},
        }
        expected.append(decision)
    sample_indices = [0, 1, 2, 501, 901, 1_001]
    sampled = [expected[index] for index in sample_indices]
    active_tail = expected[-100:]
    logical_digest = _slo_canonical_sha256(expected)
    sampled_digest = _slo_canonical_sha256(sampled)
    active_digest = _slo_canonical_sha256(active_tail)
    tombstone_digest = _slo_canonical_sha256([effect_id])
    if None in {logical_digest, sampled_digest, active_digest, tombstone_digest}:
        raise RuntimeError("deterministic compaction digest construction failed")
    return {
        "schema_version": "echo-compaction-semantic-receipt-v1",
        "semantic_verification_outside_timed_interval": True,
        "expected_logical_record_count": 1_002,
        "logical_record_count": 1_002,
        "expected_active_record_count": 101,
        "active_record_count": 101,
        "active_record_types": ["snapshot_anchor"] + ["decision"] * 100,
        "expected_retained_record_count": 100,
        "retained_record_count": 100,
        "archive_chain_verified": True,
        "archive_chain_errors": [],
        "archive_generation_count": 1,
        "archive_generation": 1,
        "archive_cumulative_record_count": 1_002,
        "archive_cumulative_tombstone_count": 1,
        "tombstones": [effect_id],
        "tombstone_sha256": tombstone_digest,
        "archived_effect_lookup_ok": True,
        "expected_logical_payload_sha256": logical_digest,
        "logical_payload_sha256": logical_digest,
        "logical_payload_equivalent": True,
        "sample_indices": sample_indices,
        "expected_sampled_payload_sha256": sampled_digest,
        "sampled_payload_sha256": sampled_digest,
        "sampled_payload_equivalent": True,
        "expected_active_payload_sha256": active_digest,
        "active_payload_sha256": active_digest,
        "active_payload_equivalent": True,
        "post_compaction_bad_tail_recovery_ok": True,
        "corrupt_tail_quarantine_count": 1,
        "ok": True,
    }


def _valid_compaction_semantic_receipt(recovery: object) -> tuple[str, float] | None:
    if not isinstance(recovery, dict) or set(recovery) != {
        "journal_replay_10k_record_count",
        "journal_replay_10k_records_s",
        "bad_tail_recovery_ok",
        "compaction_record_count",
        "compaction_latency_ms",
        "compaction_ok",
        "compaction_semantic_receipt_sha256",
        "compaction_semantics",
    }:
        return None
    latency = _safe_finite_float(recovery.get("compaction_latency_ms"))
    if (
        recovery.get("compaction_record_count") != 101
        or recovery.get("compaction_ok") is not True
        or latency is None
        or latency < 0
        or latency > SLO_CONTRACT.compaction_ms
    ):
        return None
    receipt = recovery.get("compaction_semantics")
    if not isinstance(receipt, dict):
        return None
    expected = _expected_compaction_semantic_fields()
    exact_keys = set(expected) | {
        "archive_ref_sha256",
        "active_journal_sha256",
        "archive_sha256",
        "receipt_sha256",
    }
    if set(receipt) != exact_keys or any(
        not _live_json_matches(receipt.get(key), value) for key, value in expected.items()
    ):
        return None
    for field in ("archive_ref_sha256", "active_journal_sha256", "archive_sha256"):
        value = receipt.get(field)
        if not isinstance(value, str) or _SHA256_PATTERN.fullmatch(value) is None:
            return None
    claimed_digest = receipt.get("receipt_sha256")
    bound_receipt = {key: value for key, value in receipt.items() if key != "receipt_sha256"}
    recomputed_digest = _slo_canonical_sha256(bound_receipt)
    if (
        not isinstance(claimed_digest, str)
        or recomputed_digest is None
        or not hmac.compare_digest(claimed_digest, recomputed_digest)
        or recovery.get("compaction_semantic_receipt_sha256") != claimed_digest
    ):
        return None
    return claimed_digest, latency


def _baseline_workload_corpus_digest() -> str:
    history = [
        {
            "role": "user" if index % 2 == 0 else "assistant",
            "content": f"{_ECHO_BASELINE_HISTORY_MARKER_PREFIX}{index} "
            + ("context " * _ECHO_BASELINE_HISTORY_WORDS_PER_MESSAGE),
        }
        for index in range(_ECHO_BASELINE_HISTORY_MESSAGES)
    ]
    canonical = json.dumps(
        {
            "history": history,
            "long_session_id": "old-long-{group}-{index}",
            "long_request": "benchmark api full {index}",
            "short_session_id": "old-short-{group}-{index}",
            "short_request": "benchmark api short {index}",
        },
        ensure_ascii=False,
        separators=(",", ":"),
        sort_keys=True,
    ).encode()
    return hashlib.sha256(canonical).hexdigest()


def _release_source_member_included(relative: Path) -> bool:
    if relative in _RELEASE_SOURCE_DIGEST_EXCLUDE:
        return False
    if (
        "__pycache__" in relative.parts
        or relative.suffix in {".pyc", ".pyo"}
        or relative.name == ".DS_Store"
    ):
        return False
    if relative.parts and relative.parts[0] == "desktop":
        if relative.name == ".embedded_source_digest":
            return False
        if any(part in _RELEASE_SOURCE_DESKTOP_GENERATED_PARTS for part in relative.parts[1:]):
            return False
        if any(
            relative == generated or generated in relative.parents
            for generated in _RELEASE_SOURCE_DESKTOP_GENERATED_ROOTS
        ):
            return False
    return any(
        relative == surface or surface in relative.parents
        for surface in _RELEASE_SOURCE_DIGEST_SURFACES
    )


def _canonical_git_archive_member(name: str) -> Path | None:
    """Return one unambiguous repository-relative POSIX archive path."""
    if (
        not name
        or name.startswith("/")
        or "\\" in name
        or "\x00" in name
        or re.match(r"[A-Za-z]:", name) is not None
    ):
        return None
    parts = name.split("/")
    if any(part in {"", ".", ".."} for part in parts):
        return None
    relative = Path(*parts)
    if relative.is_absolute() or relative.as_posix() != name:
        return None
    return relative


def _git_release_source_digest(root: Path, commit: str) -> str | None:
    """Recompute the V2 source digest from immutable Git blobs, never a checkout."""
    archive = _git_bytes(root, "archive", "--format=tar", commit, timeout=30.0)
    if archive is None:
        return None
    digest = hashlib.sha256(_RELEASE_SOURCE_DIGEST_VERSION)
    try:
        with tarfile.open(fileobj=io.BytesIO(archive), mode="r:") as stream:
            seen: set[str] = set()
            included_files: list[tuple[tarfile.TarInfo, Path]] = []
            for member in stream.getmembers():
                relative = _canonical_git_archive_member(member.name)
                if relative is None or member.name in seen:
                    return None
                seen.add(member.name)
                if not _release_source_member_included(relative):
                    continue
                if member.isdir():
                    continue
                if not member.isfile():
                    return None
                included_files.append((member, relative))
            members = sorted(
                included_files,
                key=lambda item: item[1].as_posix(),
            )
            for member, relative in members:
                extracted = stream.extractfile(member)
                if extracted is None:
                    return None
                payload = extracted.read()
                relative_bytes = relative.as_posix().encode("utf-8")
                digest.update(len(relative_bytes).to_bytes(8, "big"))
                digest.update(relative_bytes)
                digest.update(len(payload).to_bytes(8, "big"))
                digest.update(payload)
    except (OSError, tarfile.TarError, UnicodeError, ValueError):
        return None
    return digest.hexdigest()


@lru_cache(maxsize=1)
def _current_baseline_runtime_identity() -> tuple[str, str, str, str, str] | None:
    try:
        tiktoken_version = importlib.metadata.version("tiktoken")
        executable_sha256 = hashlib.sha256(Path(sys.executable).resolve().read_bytes()).hexdigest()
    except (OSError, importlib.metadata.PackageNotFoundError):
        return None
    platform_identity = platform.platform(aliased=True, terse=False)
    return (
        tiktoken_version,
        platform.python_implementation(),
        platform.python_version(),
        executable_sha256,
        platform_identity,
    )


def _valid_old_baseline_evidence(
    evidence: Mapping[str, Any],
    *,
    root: Path,
    baseline_script_sha256: str,
) -> bool:
    """Verify the external v2 receipt against Git objects and current harness inputs."""
    expected_top_keys = {
        "api_full_agent",
        "commit",
        "environment",
        "failures",
        "import_root",
        "iterations",
        "limitations",
        "methodology",
        "paid_provider_calls",
        "prompt_tokens",
        "provenance",
        "run_summaries",
        "runs",
        "schema_version",
        "script_sha256",
        "short_prompt_tokens",
        "source",
        "source_digest",
        "tree",
        "warmup",
    }
    if set(evidence) != expected_top_keys:
        return False
    provenance = evidence.get("provenance")
    if not isinstance(provenance, dict) or set(provenance) != {
        "baseline_script_sha256",
        "commit",
        "harness_root",
        "harness_sha256",
        "import_root",
        "import_root_sha256",
        "interpreter",
        "measured_root",
        "platform",
        "schema_version",
        "source_digest",
        "source_digest_algorithm",
        "tokenizer",
        "tree",
        "uv_lock_sha256",
        "workload",
    }:
        return False
    workload = provenance.get("workload")
    tokenizer = provenance.get("tokenizer")
    interpreter = provenance.get("interpreter")
    platform_receipt = provenance.get("platform")
    environment = evidence.get("environment")
    if (
        not isinstance(workload, dict)
        or set(workload) != {"corpus_sha256", "history_message_count", "history_words_per_message"}
        or not isinstance(tokenizer, dict)
        or set(tokenizer)
        != {
            "encoding",
            "method",
            "resource_digest_algorithm",
            "resource_root",
            "resource_tree_sha256",
            "tiktoken_version",
        }
        or not isinstance(interpreter, dict)
        or set(interpreter) != {"executable_sha256", "implementation", "version"}
        or not isinstance(platform_receipt, dict)
        or set(platform_receipt) != {"identity", "identity_sha256"}
        or not isinstance(environment, dict)
        or set(environment) != {"platform", "python"}
    ):
        return False

    repository_root = _git_output(root, "rev-parse", "--show-toplevel")
    commit = _git_output(root, "rev-parse", "--verify", f"{_ECHO_BASELINE_COMMIT}^{{commit}}")
    tree = _git_output(root, "rev-parse", "--verify", f"{_ECHO_BASELINE_COMMIT}^{{tree}}")
    if (
        repository_root is None
        or Path(repository_root).resolve() != root.resolve()
        or commit != _ECHO_BASELINE_COMMIT
        or tree is None
        or _SCOPE_COMMIT_PATTERN.fullmatch(tree) is None
    ):
        return False
    measured_source_digest = _git_release_source_digest(root, commit)
    old_uv_lock = _git_bytes(
        root,
        "show",
        "--no-ext-diff",
        "--no-textconv",
        f"{commit}:uv.lock",
    )
    old_import_root = _git_bytes(
        root,
        "show",
        "--no-ext-diff",
        "--no-textconv",
        f"{commit}:js/__init__.py",
    )
    runtime_identity = _current_baseline_runtime_identity()
    if (
        measured_source_digest is None
        or old_uv_lock is None
        or old_import_root is None
        or runtime_identity is None
    ):
        return False
    (
        tiktoken_version,
        python_implementation,
        python_version,
        executable_sha256,
        platform_identity,
    ) = runtime_identity
    measured_root_value = provenance.get("measured_root")
    import_root_value = provenance.get("import_root")
    if not isinstance(measured_root_value, str) or not isinstance(import_root_value, str):
        return False
    measured_root = Path(measured_root_value)
    import_root = Path(import_root_value)
    if not measured_root.is_absolute() or not import_root.is_absolute():
        return False
    try:
        import_relative = import_root.relative_to(measured_root)
    except ValueError:
        return False
    if import_relative != Path("js/__init__.py"):
        return False

    expected_tokenizer_sha256 = tokenizer_resource_digest(root)
    expected_platform_sha256 = hashlib.sha256(platform_identity.encode("utf-8")).hexdigest()
    if (
        evidence.get("schema_version") != "echo-old-baseline-v2"
        or evidence.get("source") != "independent_clean_commit_export"
        or evidence.get("commit") != commit
        or evidence.get("tree") != tree
        or evidence.get("source_digest") != measured_source_digest
        or evidence.get("iterations") != SLO_CONTRACT.benchmark_measured
        or evidence.get("warmup") != SLO_CONTRACT.benchmark_warmup
        or evidence.get("runs") != SLO_CONTRACT.benchmark_groups
        or evidence.get("paid_provider_calls") != 0
        or evidence.get("failures") != []
        or evidence.get("script_sha256") != baseline_script_sha256
        or evidence.get("import_root") != import_root_value
        or provenance.get("schema_version") != _ECHO_BASELINE_PROVENANCE_SCHEMA
        or provenance.get("commit") != commit
        or provenance.get("tree") != tree
        or provenance.get("source_digest_algorithm")
        != _RELEASE_SOURCE_DIGEST_VERSION.decode("ascii").rstrip("\0")
        or provenance.get("source_digest") != measured_source_digest
        or provenance.get("uv_lock_sha256") != hashlib.sha256(old_uv_lock).hexdigest()
        or _expand_repo_root_token(provenance.get("harness_root"), root=root) != str(root.resolve())
        or provenance.get("baseline_script_sha256") != baseline_script_sha256
        or provenance.get("harness_sha256") != baseline_script_sha256
        or provenance.get("import_root_sha256") != hashlib.sha256(old_import_root).hexdigest()
        or workload.get("history_message_count") != _ECHO_BASELINE_HISTORY_MESSAGES
        or workload.get("history_words_per_message") != _ECHO_BASELINE_HISTORY_WORDS_PER_MESSAGE
        or workload.get("corpus_sha256") != _baseline_workload_corpus_digest()
        or tokenizer.get("method") != _ECHO_BASELINE_TOKENIZER_METHOD
        or tokenizer.get("encoding") != _ECHO_BASELINE_TOKENIZER_ENCODING
        or tokenizer.get("tiktoken_version") != tiktoken_version
        or tokenizer.get("resource_digest_algorithm")
        != _TOKENIZER_TREE_DIGEST_VERSION.decode("ascii").rstrip("\0")
        or _expand_repo_root_token(tokenizer.get("resource_root"), root=root)
        != str((root / "resources" / "tokenizer").resolve())
        or tokenizer.get("resource_tree_sha256") != expected_tokenizer_sha256
        or interpreter.get("implementation") != python_implementation
        or interpreter.get("version") != python_version
        or interpreter.get("executable_sha256") != executable_sha256
        or platform_receipt.get("identity") != platform_identity
        or platform_receipt.get("identity_sha256") != expected_platform_sha256
        or environment.get("python") != python_version
        or environment.get("platform") != platform.platform()
    ):
        return False

    expected_markers = [
        f"{_ECHO_BASELINE_HISTORY_MARKER_PREFIX}{index}"
        for index in range(_ECHO_BASELINE_HISTORY_MESSAGES)
    ]
    expected_marker_counts = dict.fromkeys(expected_markers, 1)
    expected_marker_sha256 = hashlib.sha256(
        json.dumps(expected_markers, ensure_ascii=False, separators=(",", ":")).encode()
    ).hexdigest()
    empty_marker_sha256 = hashlib.sha256(b"[]").hexdigest()
    run_summaries = evidence.get("run_summaries")
    if not isinstance(run_summaries, list) or len(run_summaries) != SLO_CONTRACT.benchmark_groups:
        return False
    group_means: list[float] = []
    group_p50s: list[float] = []
    group_p95s: list[float] = []
    group_maxes: list[float] = []
    long_p50s: list[float] = []
    long_p95s: list[float] = []
    short_p50s: list[float] = []
    short_p95s: list[float] = []
    for group_index, summary in enumerate(run_summaries, start=1):
        if not isinstance(summary, dict) or set(summary) != {
            "api_full_agent",
            "failures",
            "group",
            "long_provider_payload_evidence",
            "prompt_tokens",
            "short_prompt_tokens",
            "short_provider_payload_evidence",
        }:
            return False
        latency = summary.get("api_full_agent")
        long_tokens = summary.get("prompt_tokens")
        short_tokens = summary.get("short_prompt_tokens")
        long_receipts = summary.get("long_provider_payload_evidence")
        short_receipts = summary.get("short_provider_payload_evidence")
        if (
            summary.get("group") != group_index
            or summary.get("failures") != []
            or not isinstance(latency, dict)
            or set(latency) != {"max_ms", "mean_ms", "p50_ms", "p95_ms"}
            or not isinstance(long_tokens, dict)
            or set(long_tokens) != {"method", "p50", "p95", "source"}
            or not isinstance(short_tokens, dict)
            or set(short_tokens) != {"method", "p50", "p95", "source"}
            or not isinstance(long_receipts, list)
            or len(long_receipts) != SLO_CONTRACT.benchmark_measured
            or not isinstance(short_receipts, list)
            or len(short_receipts) != SLO_CONTRACT.benchmark_measured
        ):
            return False
        latency_numbers = [_safe_finite_float(latency.get(key)) for key in latency]
        long_numbers = [_safe_finite_float(long_tokens.get(key)) for key in ("p50", "p95")]
        short_numbers = [_safe_finite_float(short_tokens.get(key)) for key in ("p50", "p95")]
        if (
            any(number is None or number <= 0 for number in latency_numbers)
            or any(number is None or number <= 0 for number in long_numbers)
            or any(number is None or number <= 0 for number in short_numbers)
            or long_tokens.get("source") != "tokenizer"
            or long_tokens.get("method") != _ECHO_BASELINE_TOKENIZER_METHOD
            or short_tokens.get("source") != "tokenizer"
            or short_tokens.get("method") != _ECHO_BASELINE_TOKENIZER_METHOD
        ):
            return False
        mean_ms = _safe_finite_float(latency.get("mean_ms"))
        p50_ms = _safe_finite_float(latency.get("p50_ms"))
        p95_ms = _safe_finite_float(latency.get("p95_ms"))
        max_ms = _safe_finite_float(latency.get("max_ms"))
        long_p50 = _safe_finite_float(long_tokens.get("p50"))
        long_p95 = _safe_finite_float(long_tokens.get("p95"))
        short_p50 = _safe_finite_float(short_tokens.get("p50"))
        short_p95 = _safe_finite_float(short_tokens.get("p95"))
        if (
            mean_ms is None
            or p50_ms is None
            or p95_ms is None
            or max_ms is None
            or long_p50 is None
            or long_p95 is None
            or short_p50 is None
            or short_p95 is None
            or p50_ms > p95_ms
            or p95_ms > max_ms
            or mean_ms > max_ms
            or long_p50 > long_p95
            or short_p50 > short_p95
            or long_p50 <= short_p50
            or long_p95 <= short_p95
        ):
            return False
        for receipt, is_long in [
            *((item, True) for item in long_receipts),
            *((item, False) for item in short_receipts),
        ]:
            if not isinstance(receipt, dict) or set(receipt) != {
                "history_marker_count",
                "history_marker_counts",
                "history_marker_sha256",
                "message_count",
                "message_identity_sha256",
                "provider_payload_sha256",
            }:
                return False
            if (
                _SHA256_PATTERN.fullmatch(str(receipt.get("message_identity_sha256", ""))) is None
                or _SHA256_PATTERN.fullmatch(str(receipt.get("provider_payload_sha256", "")))
                is None
            ):
                return False
            if is_long:
                if (
                    receipt.get("message_count") != 42
                    or receipt.get("history_marker_count") != _ECHO_BASELINE_HISTORY_MESSAGES
                    or receipt.get("history_marker_counts") != expected_marker_counts
                    or receipt.get("history_marker_sha256") != expected_marker_sha256
                ):
                    return False
            elif (
                receipt.get("message_count") != 2
                or receipt.get("history_marker_count") != 0
                or receipt.get("history_marker_counts") != {}
                or receipt.get("history_marker_sha256") != empty_marker_sha256
            ):
                return False
        group_means.append(mean_ms)
        group_p50s.append(p50_ms)
        group_p95s.append(p95_ms)
        group_maxes.append(max_ms)
        long_p50s.append(long_p50)
        long_p95s.append(long_p95)
        short_p50s.append(short_p50)
        short_p95s.append(short_p95)

    top_latency = evidence.get("api_full_agent")
    top_long = evidence.get("prompt_tokens")
    top_short = evidence.get("short_prompt_tokens")
    if (
        not isinstance(top_latency, dict)
        or set(top_latency) != {"group_p95_ms", "max_ms", "mean_ms", "p50_ms", "p95_ms"}
        or not isinstance(top_long, dict)
        or set(top_long) != {"method", "p50", "p95", "source"}
        or not isinstance(top_short, dict)
        or set(top_short) != {"method", "p50", "p95", "source"}
    ):
        return False
    return bool(
        top_long.get("source") == "tokenizer"
        and top_long.get("method") == _ECHO_BASELINE_TOKENIZER_METHOD
        and top_short.get("source") == "tokenizer"
        and top_short.get("method") == _ECHO_BASELINE_TOKENIZER_METHOD
        and _live_json_matches(top_latency.get("group_p95_ms"), group_p95s)
        and _live_json_matches(top_latency.get("mean_ms"), round(float(median(group_means)), 3))
        and _live_json_matches(top_latency.get("p50_ms"), round(float(median(group_p50s)), 3))
        and _live_json_matches(top_latency.get("p95_ms"), round(float(median(group_p95s)), 3))
        and _live_json_matches(top_latency.get("max_ms"), round(max(group_maxes), 3))
        and _live_json_matches(top_long.get("p50"), round(float(median(long_p50s)), 3))
        and _live_json_matches(top_long.get("p95"), round(float(median(long_p95s)), 3))
        and _live_json_matches(top_short.get("p50"), round(float(median(short_p50s)), 3))
        and _live_json_matches(top_short.get("p95"), round(float(median(short_p95s)), 3))
    )


def _valid_echo_slo_benchmark(path: Path, *, root: Path | None = None) -> bool:
    try:
        data = strict_load_object(path)
    except (OSError, json.JSONDecodeError, StrictJSONError, ValueError):
        return False

    def _summaries_are_ordered(value: object) -> bool:
        if isinstance(value, list):
            return all(_summaries_are_ordered(item) for item in value)
        if not isinstance(value, dict):
            return True
        for p95_key, max_key in (("p95_ms", "max_ms"), ("p95", "max")):
            if p95_key in value and max_key in value:
                # Echo-level ws_stream_timing summaries carry an aggregate median
                # in p95_ms alongside per-run statistics in other fields.
                # The aggregate median can exceed a single group's max, so the
                # p95<=max invariant only applies when group_p95_ms is absent.
                if "group_p95_ms" in value:
                    continue
                p95 = value.get(p95_key)
                maximum = value.get(max_key)
                p95_n = _safe_finite_float(p95)
                max_n = _safe_finite_float(maximum)
                if p95_n is None or max_n is None or p95_n > max_n:
                    return False
        return all(_summaries_are_ordered(item) for item in value.values())

    if not _summaries_are_ordered(data):
        return False
    matrix = data.get("security_matrix")
    if not isinstance(matrix, dict):
        return False
    if not (matrix.get("ok") is True and matrix.get("passed") == 25 and matrix.get("total") == 25):
        return False
    metadata = data.get("metadata")
    if not isinstance(metadata, dict) or metadata.get("slo_contract") != SLO_CONTRACT.as_dict():
        return False
    resolved_root = root.resolve() if root is not None else _resolve_repo_root(path)
    if resolved_root is None:
        return False
    benchmark_script = resolved_root / "scripts" / "echo_architecture_benchmark.py"
    try:
        benchmark_script_sha256 = hashlib.sha256(benchmark_script.read_bytes()).hexdigest()
    except OSError:
        return False
    if metadata.get("benchmark_script_sha256") != benchmark_script_sha256:
        return False
    expected_digest = release_source_digest(resolved_root)
    # Prefer top-level source_digest; fall back to metadata for transitional fixtures.
    source_digest = data.get("source_digest")
    if not isinstance(source_digest, str):
        source_digest = metadata.get("source_digest")
    if not isinstance(source_digest, str) or not hmac.compare_digest(
        source_digest, expected_digest
    ):
        return False
    expected_tokenizer = tokenizer_resource_digest(resolved_root)
    if (
        metadata.get("tokenizer_tree_digest_version")
        != _TOKENIZER_TREE_DIGEST_VERSION.decode("ascii").rstrip("\0")
        or metadata.get("tokenizer_resource_sha256") != expected_tokenizer
    ):
        return False
    if (
        metadata.get("runs") != SLO_CONTRACT.benchmark_groups
        or metadata.get("iterations") != SLO_CONTRACT.benchmark_measured
        or metadata.get("warmup") != SLO_CONTRACT.benchmark_warmup
    ):
        return False
    aggregate = data.get("aggregate")
    run_summaries = data.get("run_summaries")
    if (
        not isinstance(aggregate, dict)
        or aggregate.get("group_count") != SLO_CONTRACT.benchmark_groups
        or not isinstance(run_summaries, list)
        or len(run_summaries) != SLO_CONTRACT.benchmark_groups
    ):
        return False
    modes = data.get("modes")
    token_comparison = data.get("token_comparison")
    recovery = data.get("recovery_probes")
    if not (
        isinstance(modes, dict)
        and isinstance(modes.get("echo"), dict)
        and isinstance(token_comparison, dict)
        and isinstance(recovery, dict)
    ):
        return False
    latency_limits = {
        scenario: thresholds["p95_ms"]
        for scenario, thresholds in SLO_CONTRACT.benchmark_latency_thresholds().items()
    }

    def _number(value: object) -> float | None:
        result = _safe_finite_float(value)
        if result is None or result < 0:
            return None
        return result

    echo_mode = modes["echo"]
    aggregate_latency = aggregate.get("latency_p95_median_ms")
    if not isinstance(aggregate_latency, dict):
        return False
    for scenario, limit in latency_limits.items():
        latency = echo_mode.get(scenario, {}).get("latency")
        if not isinstance(latency, dict):
            return False
        samples = latency.get("n")
        p95_ms = _number(aggregate_latency.get(scenario))
        if not isinstance(samples, int) or isinstance(samples, bool) or samples <= 0:
            return False
        if p95_ms is None or p95_ms > limit:
            return False

    group_first_p95s: list[float] = []
    group_terminal_p95s: list[float] = []
    for summary in run_summaries:
        if not isinstance(summary, dict):
            return False
        timing_result = _valid_ws_stream_timing_receipt(
            summary.get("ws_stream_timing"),
            top_level=False,
        )
        if timing_result is None:
            return False
        first_p95, terminal_p95 = timing_result
        group_first_p95s.append(first_p95)
        group_terminal_p95s.append(terminal_p95)
    aggregate_first_p95 = _number(aggregate.get("ws_first_token_p95_median_ms"))
    aggregate_terminal_p95 = _number(aggregate.get("ws_terminal_p95_median_ms"))
    if (
        aggregate_first_p95 is None
        or aggregate_first_p95 > SLO_CONTRACT.ws_first_token_p95_ms
        or aggregate_terminal_p95 is None
        or aggregate_terminal_p95 > SLO_CONTRACT.ws_terminal_p95_ms
        or not _live_json_matches(
            aggregate.get("ws_first_token_p95_runs_ms"),
            group_first_p95s,
        )
        or not _live_json_matches(
            aggregate.get("ws_terminal_p95_runs_ms"),
            group_terminal_p95s,
        )
        or not _live_json_matches(
            aggregate_first_p95,
            round(float(median(group_first_p95s)), 3),
        )
        or not _live_json_matches(
            aggregate_terminal_p95,
            round(float(median(group_terminal_p95s)), 3),
        )
        or _valid_ws_stream_timing_receipt(
            echo_mode.get("ws_stream_timing"),
            top_level=True,
            aggregate_first_p95=aggregate_first_p95,
            aggregate_terminal_p95=aggregate_terminal_p95,
            group_first_p95s=group_first_p95s,
            group_terminal_p95s=group_terminal_p95s,
        )
        is None
    ):
        return False

    token_source = token_comparison.get("token_source")
    if token_source not in {"provider_actual", "tokenizer"}:
        return False
    prompt_p95 = _number(token_comparison.get("api_full_agent_prompt_p95_echo"))
    prompt_limit = _number(token_comparison.get("api_full_agent_prompt_p95_limit"))
    if prompt_p95 is None or prompt_limit is None or prompt_p95 > prompt_limit:
        return False
    if token_comparison.get("api_full_agent_prompt_within_limit") is not True:
        return False

    journal_append_p95 = _number(aggregate.get("journal_append_p95_max_ms"))
    if journal_append_p95 is None or journal_append_p95 > SLO_CONTRACT.journal_append_p95_ms:
        return False
    replay_seconds = _number(aggregate.get("replay_10k_max_seconds"))
    compaction_ms = _number(aggregate.get("compaction_max_ms"))
    replay_count = aggregate.get("replay_10k_record_count_min")
    if not (
        isinstance(replay_count, int)
        and not isinstance(replay_count, bool)
        and replay_count >= 10_000
        and replay_seconds is not None
        and replay_seconds <= SLO_CONTRACT.replay_10k_seconds
        and aggregate.get("bad_tail_all_ok") is True
        and aggregate.get("compaction_all_ok") is True
        and compaction_ms is not None
        and compaction_ms <= SLO_CONTRACT.compaction_ms
    ):
        return False
    top_compaction = _valid_compaction_semantic_receipt(recovery)
    if top_compaction is None:
        return False
    group_compaction_digests: list[str] = []
    group_compaction_latencies: list[float] = []
    for summary in run_summaries:
        if not isinstance(summary, dict):
            return False
        compaction_result = _valid_compaction_semantic_receipt(summary.get("recovery"))
        if compaction_result is None:
            return False
        semantic_digest, compaction_latency = compaction_result
        group_compaction_digests.append(semantic_digest)
        group_compaction_latencies.append(compaction_latency)
    if (
        aggregate.get("compaction_semantics_all_ok") is not True
        or not _live_json_matches(
            aggregate.get("compaction_semantic_receipt_sha256s"),
            group_compaction_digests,
        )
        or not _live_json_matches(compaction_ms, round(max(group_compaction_latencies), 3))
        or not _live_json_matches(top_compaction[1], compaction_ms)
    ):
        return False

    def _valid_concurrency(probe: object) -> bool:
        if not isinstance(probe, dict):
            return False
        peak_rss = _number(probe.get("peak_rss_mb"))
        runtime_peak = probe.get("runtime_peak_inflight")
        expected_total = SLO_CONTRACT.concurrency_workers * SLO_CONTRACT.concurrency_rounds
        receipts = probe.get("request_receipts")
        provider_events = probe.get("provider_call_events")
        if (
            probe.get("evidence_schema_version") != "echo-concurrency-evidence-v1"
            or not isinstance(receipts, list)
            or len(receipts) != expected_total
            or not isinstance(provider_events, list)
            or len(provider_events) != expected_total * 2
            or probe.get("failures") != []
        ):
            return False
        receipt_payload = {
            "request_receipts": receipts,
            "provider_call_events": provider_events,
        }
        receipt_sha256 = _live_canonical_sha256(receipt_payload)
        if receipt_sha256 is None or probe.get("receipt_sha256") != receipt_sha256:
            return False
        expected_requests = {
            (round_index, worker_index): (
                f"benchmark-concurrency:{round_index}:{worker_index}",
                f"echo-concurrency-{round_index}-{worker_index}",
                (
                    f"benchmark-concurrency:{round_index}:{worker_index}"
                    f"|benchmark-isolation-secret:{round_index}:{worker_index}"
                ),
            )
            for round_index in range(SLO_CONTRACT.concurrency_rounds)
            for worker_index in range(SLO_CONTRACT.concurrency_workers)
        }
        observed_requests: set[tuple[int, int]] = set()
        recomputed_completed = 0
        recomputed_5xx = 0
        recomputed_crosstalk = 0
        for receipt in receipts:
            if not isinstance(receipt, dict):
                return False
            round_index = receipt.get("round")
            worker_index = receipt.get("worker")
            status_code = receipt.get("status_code")
            if (
                isinstance(round_index, bool)
                or not isinstance(round_index, int)
                or isinstance(worker_index, bool)
                or not isinstance(worker_index, int)
                or isinstance(status_code, bool)
                or not isinstance(status_code, int)
            ):
                return False
            request_key = (round_index, worker_index)
            expected = expected_requests.get(request_key)
            if expected is None or request_key in observed_requests:
                return False
            observed_requests.add(request_key)
            marker, session_id, response_text = expected
            if (
                receipt.get("session_id") != session_id
                or receipt.get("expected_response") != response_text
            ):
                return False
            if status_code >= 500:
                recomputed_5xx += 1
            crosstalk = bool(
                receipt.get("observed_session_id") != session_id
                or receipt.get("observed_response") != response_text
            )
            if crosstalk:
                recomputed_crosstalk += 1
            if status_code == 200 and not crosstalk:
                recomputed_completed += 1
            if marker not in response_text:
                return False
        if observed_requests != set(expected_requests):
            return False
        active_requests: set[str] = set()
        ended_requests: set[str] = set()
        recomputed_peak = 0
        expected_markers = {value[0] for value in expected_requests.values()}
        for sequence, event in enumerate(provider_events, start=1):
            if not isinstance(event, dict) or event.get("sequence") != sequence:
                return False
            phase = event.get("phase")
            request_id = event.get("request_id")
            if not isinstance(request_id, str) or request_id not in expected_markers:
                return False
            if phase == "start":
                if request_id in active_requests or request_id in ended_requests:
                    return False
                active_requests.add(request_id)
                recomputed_peak = max(recomputed_peak, len(active_requests))
            elif phase == "end":
                if request_id not in active_requests:
                    return False
                active_requests.remove(request_id)
                ended_requests.add(request_id)
            else:
                return False
        if active_requests or ended_requests != expected_markers:
            return False
        return bool(
            probe.get("submitted_concurrency") == SLO_CONTRACT.concurrency_workers
            and probe.get("rounds") == SLO_CONTRACT.concurrency_rounds
            and probe.get("total_requests") == expected_total
            and probe.get("completed_ok") == recomputed_completed == expected_total
            and probe.get("http_5xx_count") == recomputed_5xx == 0
            and probe.get("crosstalk_count") == recomputed_crosstalk == 0
            and probe.get("isolation_checks") == len(observed_requests) == expected_total
            and probe.get("overlap_layer") == "real_gated_provider_calls"
            and probe.get("execution_model") == "single_process_async_asgi"
            and isinstance(runtime_peak, int)
            and not isinstance(runtime_peak, bool)
            and runtime_peak == recomputed_peak
            and runtime_peak >= SLO_CONTRACT.concurrency_workers
            and peak_rss is not None
            and peak_rss <= SLO_CONTRACT.max_rss_mb
        )

    top_level_concurrency = data.get("concurrency_probe")
    if not _valid_concurrency(top_level_concurrency):
        return False
    for index, summary in enumerate(run_summaries, start=1):
        if (
            not isinstance(summary, dict)
            or summary.get("group") != index
            or not _valid_concurrency(summary.get("concurrency"))
        ):
            return False
    first_group = run_summaries[0]
    if not isinstance(first_group, dict) or not _live_json_matches(
        top_level_concurrency,
        first_group.get("concurrency"),
    ):
        return False

    baseline = data.get("baseline_comparison")
    if not isinstance(baseline, dict):
        return False
    latency_comparison = baseline.get("api_full_agent")
    token_comparison_baseline = baseline.get("prompt_tokens")
    short_token_comparison = baseline.get("short_prompt_tokens")
    aggregate_long_tokens = aggregate.get("long_prompt_tokens")
    aggregate_short_tokens = aggregate.get("short_prompt_tokens")
    if (
        not isinstance(latency_comparison, dict)
        or not isinstance(token_comparison_baseline, dict)
        or not isinstance(short_token_comparison, dict)
        or not isinstance(aggregate_long_tokens, dict)
        or not isinstance(aggregate_short_tokens, dict)
    ):
        return False
    old_p95 = _number(latency_comparison.get("old_p95_ms"))
    echo_p95 = _number(latency_comparison.get("echo_p95_ms"))
    claimed_delta = latency_comparison.get("p95_delta_pct")
    old_token_p50 = _number(token_comparison_baseline.get("old_p50"))
    echo_token_p50 = _number(token_comparison_baseline.get("echo_p50"))
    claimed_p50_reduction = token_comparison_baseline.get("p50_reduction_pct")
    old_token_p95 = _number(token_comparison_baseline.get("old_p95"))
    echo_token_p95 = _number(token_comparison_baseline.get("echo_p95"))
    claimed_p95_reduction = token_comparison_baseline.get("reduction_pct")
    old_short_p50 = _number(short_token_comparison.get("old_p50"))
    echo_short_p50 = _number(short_token_comparison.get("echo_p50"))
    claimed_short_p50_increase = short_token_comparison.get("p50_increase_pct")
    old_short_p95 = _number(short_token_comparison.get("old_p95"))
    echo_short_p95 = _number(short_token_comparison.get("echo_p95"))
    claimed_short_p95_increase = short_token_comparison.get("p95_increase_pct")
    baseline_artifact_name = baseline.get("baseline_artifact")
    if baseline_artifact_name != "ECHO_BASELINE_65CC545.json":
        return False
    baseline_artifact = resolved_root / "docs" / "security" / baseline_artifact_name
    try:
        baseline_artifact_bytes = baseline_artifact.read_bytes()
        baseline_evidence = strict_load_object_bytes(baseline_artifact_bytes)
    except (OSError, UnicodeDecodeError, json.JSONDecodeError, StrictJSONError, ValueError):
        return False
    baseline_evidence_latency = baseline_evidence.get("api_full_agent")
    baseline_evidence_tokens = baseline_evidence.get("prompt_tokens")
    baseline_evidence_short_tokens = baseline_evidence.get("short_prompt_tokens")
    if (
        not isinstance(baseline_evidence_latency, dict)
        or not isinstance(baseline_evidence_tokens, dict)
        or not isinstance(baseline_evidence_short_tokens, dict)
    ):
        return False
    evidence_sha256 = hashlib.sha256(baseline_artifact_bytes).hexdigest()
    baseline_script = resolved_root / "benchmarks" / "old_architecture_baseline.py"
    try:
        baseline_script_sha256 = hashlib.sha256(baseline_script.read_bytes()).hexdigest()
    except OSError:
        return False
    if not _valid_old_baseline_evidence(
        baseline_evidence,
        root=resolved_root,
        baseline_script_sha256=baseline_script_sha256,
    ):
        return False
    if (
        baseline.get("valid") is not True
        or baseline.get("source") != "independent_clean_commit_export"
        or baseline.get("baseline_commit") != _ECHO_BASELINE_COMMIT
        or baseline.get("iterations") != SLO_CONTRACT.benchmark_measured
        or baseline.get("warmup") != SLO_CONTRACT.benchmark_warmup
        or baseline.get("runs") != SLO_CONTRACT.benchmark_groups
        or baseline.get("paid_provider_calls") != 0
        or baseline.get("baseline_artifact_sha256") != evidence_sha256
        or baseline.get("baseline_script_sha256") != baseline_script_sha256
        or baseline_evidence.get("schema_version") != "echo-old-baseline-v2"
        or baseline_evidence.get("source") != "independent_clean_commit_export"
        or baseline_evidence.get("commit") != baseline.get("baseline_commit")
        or baseline_evidence.get("iterations") != baseline.get("iterations")
        or baseline_evidence.get("warmup") != baseline.get("warmup")
        or baseline_evidence.get("runs") != baseline.get("runs")
        or baseline_evidence.get("paid_provider_calls") != 0
        or baseline_evidence.get("script_sha256") != baseline_script_sha256
        or baseline_evidence.get("failures") != []
        or _number(baseline_evidence_latency.get("p95_ms")) != old_p95
        or _number(baseline_evidence_tokens.get("p50")) != old_token_p50
        or _number(baseline_evidence_tokens.get("p95")) != old_token_p95
        or _number(baseline_evidence_short_tokens.get("p50")) != old_short_p50
        or _number(baseline_evidence_short_tokens.get("p95")) != old_short_p95
        or old_p95 is None
        or old_p95 <= 0
        or echo_p95 is None
        or echo_p95 > old_p95
        or echo_p95 != _number(aggregate_latency.get("api_full_agent"))
        or isinstance(claimed_delta, bool)
        or not isinstance(claimed_delta, int | float)
        or old_token_p50 is None
        or old_token_p50 <= 0
        or echo_token_p50 is None
        or old_token_p95 is None
        or old_token_p95 <= 0
        or echo_token_p95 is None
        or echo_token_p50 != _number(aggregate_long_tokens.get("p50_median"))
        or echo_token_p95 != _number(aggregate_long_tokens.get("p95_median"))
        or isinstance(claimed_p50_reduction, bool)
        or not isinstance(claimed_p50_reduction, int | float)
        or isinstance(claimed_p95_reduction, bool)
        or not isinstance(claimed_p95_reduction, int | float)
        or old_short_p50 is None
        or old_short_p50 <= 0
        or echo_short_p50 is None
        or old_short_p95 is None
        or old_short_p95 <= 0
        or echo_short_p95 is None
        or echo_short_p50 != _number(aggregate_short_tokens.get("p50_median"))
        or echo_short_p95 != _number(aggregate_short_tokens.get("p95_median"))
        or isinstance(claimed_short_p50_increase, bool)
        or not isinstance(claimed_short_p50_increase, int | float)
        or isinstance(claimed_short_p95_increase, bool)
        or not isinstance(claimed_short_p95_increase, int | float)
        or token_comparison_baseline.get("source") != "tokenizer"
        or token_comparison_baseline.get("method") != baseline_evidence_tokens.get("method")
        or short_token_comparison.get("source") != "tokenizer"
        or short_token_comparison.get("method") != baseline_evidence_short_tokens.get("method")
    ):
        return False
    calculated_delta = (echo_p95 - old_p95) / old_p95 * 100.0
    p50_reduction = (old_token_p50 - echo_token_p50) / old_token_p50 * 100.0
    p95_reduction = (old_token_p95 - echo_token_p95) / old_token_p95 * 100.0
    short_p50_increase = (echo_short_p50 - old_short_p50) / old_short_p50 * 100.0
    short_p95_increase = (echo_short_p95 - old_short_p95) / old_short_p95 * 100.0
    return bool(
        abs(float(claimed_delta) - calculated_delta) <= 0.01
        and abs(float(claimed_p50_reduction) - p50_reduction) <= 0.01
        and abs(float(claimed_p95_reduction) - p95_reduction) <= 0.01
        and abs(float(claimed_short_p50_increase) - short_p50_increase) <= 0.01
        and abs(float(claimed_short_p95_increase) - short_p95_increase) <= 0.01
        and p50_reduction >= SLO_CONTRACT.long_context_min_reduction_pct
        and p95_reduction >= SLO_CONTRACT.long_context_min_reduction_pct
        and short_p50_increase <= SLO_CONTRACT.short_context_max_increase_pct
        and short_p95_increase <= SLO_CONTRACT.short_context_max_increase_pct
    )


LOCAL_GATE_RECEIPT_SCHEMA_VERSION = "js-agent-local-gate-receipt-v4"
LOCAL_GATE_SPEC_VERSION = "js-agent-local-gate-spec-v3"
TOOLCHAIN_LOCK_SCHEMA_VERSION = "js-agent-toolchain-lock-v1"
READINESS_RESULT_SENTINEL = "JS_AGENT_READINESS_V1="
READINESS_RESULT_SCHEMA_VERSION = "js-agent-readiness-result-v1"
RELEASE_RESULT_SENTINEL = "JS_AGENT_RELEASE_RESULT_V1="
RELEASE_RESULT_SCHEMA_VERSION = "js-agent-release-result-v1"
RELEASE_RESULT_FIELDS = frozenset({"schema_version", "ok", "gate"})
_DESKTOP_RELEASE_BINDING_FIELDS = frozenset(
    {"desktop_manifest_sha256", "app_tree_sha256", "app_sha256"}
)
_TAURI_RELEASE_BINDING_FIELDS = _DESKTOP_RELEASE_BINDING_FIELDS | frozenset(
    {"result_sha256", "harness_sha256"}
)
READINESS_RESULT_FIELDS = frozenset(
    {"schema_version", "source_digest", "internal_ready", "internal_blockers"}
)
# Soak timing: allow probe startup/cleanup but reject compressed/late-start forgeries.
_LIVE_DEFAULT_MAX_STATE_BYTES = 512 * 1024 * 1024
_LIVE_DEFAULT_MAX_RSS_BYTES = 512 * 1024 * 1024
_LIVE_DEFAULT_MAX_RSS_GROWTH_MIB_PER_MINUTE = 0.5
_LIVE_DEFAULT_MAX_SESSION_PARTITIONS_PER_OWNER = 64
_LIVE_RESOURCE_SAMPLE_INTERVAL_SECONDS = 5.0
_LIVE_RESOURCE_STABILITY_MIN_SECONDS = 600.0
_LIVE_RESOURCE_MAX_SAMPLE_GAP_SECONDS = 15.0
_LIVE_RESOURCE_MIN_SAMPLE_RATIO = 0.9
_LIVE_RESOURCE_MAX_PLATEAU_GROWTH_MIB = 16.0
_LIVE_MAX_STATE_GROWTH_MIB_PER_MINUTE = 0.5
_LIVE_MAX_PARTITION_GROWTH_MIB_PER_MINUTE = 0.05
_LIVE_STORAGE_MAX_PLATEAU_GROWTH_MIB = 32.0
_LIVE_PARTITION_MAX_PLATEAU_GROWTH_MIB = 4.0
_LIVE_RESOURCE_CHAIN_BINDING = "echo-live-resource-samples-v1"
_LIVE_STORAGE_COMPONENT_NAMES = (
    "audit",
    "compression_feedback",
    "learning",
    "lifecycle",
    "memory_enhanced",
    "metacognition",
    "prompt_optimization",
    "quality",
    "review_capsules",
    "token_stats",
    "echo_partitions",
    "events",
    "other",
)
_SOAK_FIRST_MONOTONIC_MAX_SECONDS = 15.0
_SOAK_WALL_MONO_ALIGN_TOLERANCE_SECONDS = 2.0
# Single total coverage tolerance — do not stack first/last windows into 2x slack.
_SOAK_COVERAGE_TOLERANCE_SECONDS = 15.0
_SOAK_WALL_DURATION_TOLERANCE_SECONDS = _SOAK_COVERAGE_TOLERANCE_SECONDS
_SOAK_GATE_RECEIPT_MIN_SECONDS = 3600.0 - _SOAK_COVERAGE_TOLERANCE_SECONDS
_SOAK_FINAL_CLEANUP_MAX_SECONDS = 30.0
_SOAK_SETUP_BIAS_MAX_SECONDS = 30.0
_SUPERVISED_SOAK_SCHEMA_VERSION = "js-agent-supervised-soak-v1"
_SUPERVISED_OVERLAY_SCHEMA_VERSION = "js-agent-tauri-overlay-v1"
_SUPERVISED_SOAK_COMBINED_RELATIVE = Path("soak/supervised_soak.combined.json")
_SUPERVISED_SOAK_CORE_RAW_RELATIVE = Path("soak/echo_core_soak.raw.json")
_SUPERVISED_SOAK_OVERLAY_RAW_RELATIVE = Path("soak/tauri_overlay.raw.json")
_SUPERVISED_COUNTER_KEYS = frozenset(
    {
        "mode_switches",
        "app_restarts",
        "sidecar_recoveries",
        "ws_cancel_cycles",
        "r4_ops",
        "r6_ops",
    }
)
_SUPERVISED_COMBINED_FIELDS = frozenset(
    {
        "schema_version",
        "ok",
        "started_utc",
        "finished_utc",
        "duration_seconds",
        "elapsed_seconds",
        "source_digest",
        "metadata_fingerprint",
        "core",
        "overlay",
        "combined_sha256",
    }
)
_SUPERVISED_CORE_FIELDS = frozenset({"exit_code", "raw_sha256", "ok"})
_SUPERVISED_OVERLAY_SUMMARY_FIELDS = frozenset(
    {
        "exit_code",
        "raw_sha256",
        "ok",
        "targets",
        "counters",
        "targets_met",
        "cycles",
        "heartbeat_count",
        "max_heartbeat_gap_s",
        "max_heartbeat_gap_limit_s",
        "chain_root",
        "desktop_manifest_sha256",
        "app_tree_sha256",
        "app_sha256",
    }
)
_SUPERVISED_OVERLAY_FIELDS = frozenset(
    {
        "schema_version",
        "ok",
        "started_utc",
        "finished_utc",
        "duration_seconds",
        "elapsed_seconds",
        "source_digest",
        "metadata_fingerprint",
        "acceptance_pid",
        "targets",
        "counters",
        "targets_met",
        "cycles",
        "heartbeats",
        "heartbeat_count",
        "max_heartbeat_gap_s",
        "max_heartbeat_gap_limit_s",
        "chain_root",
        "errors",
        "app_path",
        "harness_exec",
        "desktop_manifest_path",
        "desktop_manifest_sha256",
        "app_tree_sha256",
        "app_sha256",
    }
)
_SUPERVISED_HEARTBEAT_FIELDS = frozenset(
    {
        "index",
        "monotonic_s",
        "wall_utc",
        "note",
        "counters",
        "source_digest",
        "prev_chain",
        "chain",
    }
)
_LOCAL_GATE_DURATION_TOLERANCE_SECONDS = 1.0
_EVIDENCE_DIR_TOKEN = "{evidence_dir}"
_REPO_ROOT_TOKEN = "{repo_root}"
_SOURCE_DIGEST_TOKEN = "{source_digest}"


def _expand_repo_root_token(value: object, *, root: Path) -> str | None:
    """Expand ``{repo_root}`` / ``{repo_root}/rel``; leave other strings unchanged."""

    if not isinstance(value, str) or not value:
        return None
    resolved = root.resolve()
    if value == _REPO_ROOT_TOKEN:
        return str(resolved)
    prefix = f"{_REPO_ROOT_TOKEN}/"
    if value.startswith(prefix):
        suffix = value.removeprefix(prefix)
        if not suffix or suffix.startswith("/") or ".." in Path(suffix).parts:
            return None
        # Do not follow the final symlink: venv python receipts record
        # `{repo_root}/.venv/bin/python`, not the Homebrew Cellar target.
        return str(resolved / suffix)
    return value


_LOCAL_GATE_PARSER_KINDS = frozenset(
    {
        "git_diff",
        "ruff",
        "mypy",
        "pytest",
        "slo_json",
        "soak_json",
        "e2e_json",
        "python_exit",
        "release_markers",
        "readiness_json",
        "generic",
    }
)
_TOOLCHAIN_VENV_BINARIES = ("python", "ruff", "mypy")


@dataclass(frozen=True)
class LocalGateOutputParseRules:
    parser: str = "generic"
    require_exit_code_zero: bool = True
    stderr_must_be_empty: bool = True

    def __post_init__(self) -> None:
        if self.parser not in _LOCAL_GATE_PARSER_KINDS:
            raise ValueError(f"unsupported local gate parser {self.parser!r}")


@dataclass(frozen=True)
class LocalGateSpec:
    gate_name: str
    argv: tuple[str, ...]
    coverage_scope: tuple[str, ...]
    output_parse: LocalGateOutputParseRules = LocalGateOutputParseRules()


def _normalize_gate_path_token(path_text: str, *, root: Path, evidence_dir: Path) -> str:
    candidate = Path(path_text)
    resolved = candidate.resolve() if candidate.is_absolute() else (root / candidate).resolve()
    evidence_resolved = evidence_dir.resolve()
    root_resolved = root.resolve()
    try:
        relative = resolved.relative_to(evidence_resolved)
        return f"{_EVIDENCE_DIR_TOKEN}/{relative.as_posix()}"
    except ValueError:
        pass
    try:
        relative = resolved.relative_to(root_resolved)
        return f"{_REPO_ROOT_TOKEN}/{relative.as_posix()}"
    except ValueError:
        return resolved.as_posix()


def normalize_gate_argv(
    argv: Sequence[str],
    *,
    root: Path,
    evidence_dir: Path,
    source_digest: str | None = None,
) -> tuple[str, ...]:
    normalized: list[str] = []
    for item in argv:
        if source_digest is not None and item == source_digest:
            normalized.append(_SOURCE_DIGEST_TOKEN)
            continue
        if (
            ("/" in item or item.startswith("."))
            and not item.startswith("-")
            and item not in {".venv/bin/python", ".venv/bin/ruff", ".venv/bin/mypy"}
        ):
            normalized.append(
                _normalize_gate_path_token(item, root=root, evidence_dir=evidence_dir)
            )
        else:
            normalized.append(item)
    return tuple(normalized)


def _materialize_gate_argv_template(
    template: tuple[str, ...],
    *,
    root: Path,
    evidence_dir: Path,
    source_digest: str,
) -> tuple[str, ...]:
    materialized: list[str] = []
    for item in template:
        if item == _EVIDENCE_DIR_TOKEN:
            materialized.append(str(evidence_dir.resolve()))
            continue
        if item.startswith(f"{_EVIDENCE_DIR_TOKEN}/"):
            suffix = item.removeprefix(f"{_EVIDENCE_DIR_TOKEN}/")
            materialized.append(str((evidence_dir / suffix).resolve()))
            continue
        if item == _REPO_ROOT_TOKEN:
            materialized.append(str(root.resolve()))
            continue
        if item.startswith(f"{_REPO_ROOT_TOKEN}/"):
            suffix = item.removeprefix(f"{_REPO_ROOT_TOKEN}/")
            materialized.append(str((root / suffix).resolve()))
            continue
        if item == _SOURCE_DIGEST_TOKEN:
            materialized.append(source_digest)
            continue
        materialized.append(item)
    return tuple(materialized)


def canonical_gate_capture_paths(gate_name: str, evidence_dir: Path) -> tuple[Path, Path]:
    gates_dir = evidence_dir.resolve() / "gates"
    return (
        gates_dir / f"{gate_name}.stdout.txt",
        gates_dir / f"{gate_name}.stderr.txt",
    )


def expand_path_tokens(
    raw: str,
    *,
    root: Path,
    evidence_dir: Path,
    home: Path | None = None,
) -> Path:
    """Expand <REPO_ROOT>/<EVIDENCE_ROOT>/<HOME> tokens used in sanitized receipts."""
    text = raw
    text = text.replace("<EVIDENCE_ROOT>", str(evidence_dir.resolve()))
    text = text.replace("<REPO_ROOT>", str(root.resolve()))
    text = text.replace("<HOME>", str((home or Path.home()).resolve()))
    return Path(text)


def _executable_version(exe: Path) -> str | None:
    try:
        result = subprocess.run(
            [str(exe), "--version"],
            capture_output=True,
            text=True,
            check=False,
            timeout=30,
        )
    except (OSError, subprocess.SubprocessError):
        return None
    output = (result.stdout or result.stderr or "").strip()
    if not output:
        return None
    return output.splitlines()[0]


def _toolchain_entry_for_path(exe: Path) -> dict[str, str] | None:
    try:
        resolved = exe.resolve()
        if not resolved.is_file():
            return None
        payload = resolved.read_bytes()
    except OSError:
        return None
    entry: dict[str, str] = {
        "realpath": str(resolved),
        "sha256": hashlib.sha256(payload).hexdigest(),
    }
    version = _executable_version(resolved)
    if version is not None:
        entry["version"] = version
    return entry


def build_gate_toolchain_manifest(root: Path) -> dict[str, dict[str, str]]:
    """Hash ``.venv/bin/{python,ruff,mypy}`` and record ``git`` when present."""
    manifest: dict[str, dict[str, str]] = {}
    venv_bin = root / ".venv" / "bin"
    for name in _TOOLCHAIN_VENV_BINARIES:
        entry = _toolchain_entry_for_path(venv_bin / name)
        if entry is not None:
            manifest[f".venv/bin/{name}"] = entry
    git_path = shutil.which("git")
    if git_path is not None:
        entry = _toolchain_entry_for_path(Path(git_path))
        if entry is not None:
            manifest["git"] = entry
    return manifest


def _tool_provenance(name: str, executable: Path) -> str:
    if name == "python":
        return f"interpreter:{Path(sys.executable).resolve()}"
    if name in {"ruff", "mypy"}:
        try:
            distribution = importlib.metadata.distribution(name)
            return f"package:{Path(str(distribution.locate_file(''))).resolve()}"
        except importlib.metadata.PackageNotFoundError:
            return "venv-bin"
    return f"path:{executable.parent.resolve()}"


def build_frozen_toolchain_lock(root: Path) -> dict[str, object]:
    """Snapshot every executable authorized to produce final local-gate evidence."""
    resolved_root = root.resolve()
    executable_paths = {
        "python": resolved_root / ".venv" / "bin" / "python",
        "ruff": resolved_root / ".venv" / "bin" / "ruff",
        "mypy": resolved_root / ".venv" / "bin" / "mypy",
    }
    git_path = shutil.which("git")
    if git_path is not None:
        executable_paths["git"] = Path(git_path)
    tools: dict[str, dict[str, str]] = {}
    for name, executable in executable_paths.items():
        entry = _toolchain_entry_for_path(executable)
        if entry is None or not entry.get("version"):
            continue
        entry["provenance"] = _tool_provenance(name, executable)
        tools[name] = entry
    return {
        "schema_version": TOOLCHAIN_LOCK_SCHEMA_VERSION,
        "tools": tools,
    }


def _canonical_json_sha256(payload: Mapping[str, object]) -> str:
    body = json.dumps(payload, sort_keys=True, separators=(",", ":")).encode()
    return hashlib.sha256(body).hexdigest()


def _valid_frozen_toolchain_lock(payload: object) -> bool:
    if not isinstance(payload, dict):
        return False
    if payload.get("schema_version") != TOOLCHAIN_LOCK_SCHEMA_VERSION:
        return False
    tools = payload.get("tools")
    if not isinstance(tools, dict) or set(tools) != {"python", "ruff", "mypy", "git"}:
        return False
    for name, entry in tools.items():
        if not isinstance(name, str) or not isinstance(entry, dict):
            return False
        if any(
            not isinstance(entry.get(field), str) or not str(entry[field]).strip()
            for field in ("realpath", "sha256", "version", "provenance")
        ):
            return False
        if _SHA256_PATTERN.fullmatch(str(entry["sha256"])) is None:
            return False
        path = Path(str(entry["realpath"]))
        if not path.is_file() or hashlib.sha256(path.read_bytes()).hexdigest() != entry["sha256"]:
            return False
        version = str(entry["version"]).lower()
        if name == "python" and "python" not in version:
            return False
        if name == "ruff" and "ruff" not in version:
            return False
        if name == "mypy" and "mypy" not in version:
            return False
        if name == "git" and "git version" not in version:
            return False
    return True


def write_toolchain_lock(evidence_dir: Path, root: Path) -> dict[str, object]:
    """Create the immutable toolchain lock used by every receipt in an evidence pack."""
    payload = build_frozen_toolchain_lock(root)
    if not _valid_frozen_toolchain_lock(payload):
        raise ValueError("complete trustworthy python/ruff/mypy/git toolchain is required")
    target = evidence_dir.resolve() / "TOOLCHAIN.lock.json"
    target.parent.mkdir(parents=True, exist_ok=True)
    body = json.dumps(payload, indent=2, sort_keys=True) + "\n"
    temporary = target.with_name(f".{target.name}.{os.getpid()}.tmp")
    temporary.write_text(body, encoding="utf-8")
    os.replace(temporary, target)
    return payload


def build_receipt_toolchain_for_argv(
    root: Path,
    argv: Sequence[str],
) -> dict[str, dict[str, str]]:
    if not argv:
        return {}
    manifest = build_gate_toolchain_manifest(root)
    argv0 = argv[0]
    if argv0 in manifest:
        return {argv0: manifest[argv0]}
    if "/" in argv0 or argv0.startswith("."):
        candidate = Path(argv0)
        resolved = candidate if candidate.is_absolute() else (root / candidate)
        entry = _toolchain_entry_for_path(resolved)
        if entry is not None:
            return {argv0: entry}
    return {}


def _toolchain_entries_match(
    receipt_entry: Mapping[str, object],
    live_entry: Mapping[str, str],
    *,
    root: Path,
    evidence_dir: Path | None = None,
) -> bool:
    evidence = evidence_dir or root
    for field in ("realpath", "sha256", "version"):
        if field not in live_entry:
            continue
        receipt_value = receipt_entry.get(field)
        if not isinstance(receipt_value, str):
            return False
        if field == "realpath":
            expanded = str(
                expand_path_tokens(receipt_value, root=root, evidence_dir=evidence).resolve()
            )
            if expanded != live_entry[field]:
                return False
        elif receipt_value != live_entry[field]:
            return False
    return True


def _valid_receipt_toolchain(
    receipt: Mapping[str, object],
    *,
    root: Path,
    argv: Sequence[str],
    evidence_dir: Path | None = None,
) -> bool:
    toolchain = receipt.get("toolchain")
    if not isinstance(toolchain, dict):
        return False
    if not argv:
        return False
    argv0 = argv[0]
    if argv0 == "git":
        live = build_gate_toolchain_manifest(root).get("git")
        receipt_entry = toolchain.get("git")
        if not isinstance(receipt_entry, dict) or live is None:
            return False
        return _toolchain_entries_match(
            receipt_entry, live, root=root, evidence_dir=evidence_dir
        ) and set(toolchain) == {"git"}
    live_toolchain = build_receipt_toolchain_for_argv(root, argv)
    if not live_toolchain:
        return False
    if set(toolchain.keys()) != set(live_toolchain.keys()):
        return False
    for key, live_entry in live_toolchain.items():
        receipt_entry = toolchain.get(key)
        if not isinstance(receipt_entry, dict):
            return False
        if not _toolchain_entries_match(
            receipt_entry, live_entry, root=root, evidence_dir=evidence_dir
        ):
            return False
    return True


def _valid_toolchain_semantics(
    parser: str,
    toolchain: Mapping[str, object],
    argv: Sequence[str],
) -> bool:
    if not argv:
        return False
    argv0 = argv[0]
    entry = toolchain.get(argv0)
    if argv0 == "git":
        entry = toolchain.get("git")
    if not isinstance(entry, dict):
        return False
    version = entry.get("version")
    if not isinstance(version, str) or not version.strip():
        return False
    lowered = version.lower()
    if parser == "ruff":
        return "ruff" in lowered
    if parser == "mypy":
        return "mypy" in lowered
    if parser in {
        "pytest",
        "python_exit",
        "slo_json",
        "soak_json",
        "e2e_json",
        "readiness_json",
    }:
        return "python" in lowered
    if parser == "git_diff":
        return "git version" in lowered
    return True


def format_readiness_result_line(
    report: ReleaseReadinessReport,
    *,
    source_digest: str,
) -> str:
    """Emit the unique machine-readable readiness result line for gate stdout."""
    if _SHA256_PATTERN.fullmatch(source_digest) is None or source_digest != source_digest.lower():
        raise ValueError("source_digest must be a lowercase 64-hex SHA-256")
    payload = {
        "schema_version": READINESS_RESULT_SCHEMA_VERSION,
        "source_digest": source_digest,
        "internal_ready": bool(report.internal_ready),
        "internal_blockers": [str(item) for item in report.internal_blockers],
    }
    return READINESS_RESULT_SENTINEL + _canonical_json_text(payload)


def _canonical_json_text(payload: Mapping[str, object]) -> str:
    from js.echo.ledger.strict_json import canonical_json_text

    return canonical_json_text(payload)


_RELEASE_MARKER_GATES: frozenset[str] = frozenset(
    {"release_smoke", "echo_full_audit", "desktop_build", "tauri_webview_lifecycle"}
)


def format_release_result_line(
    *,
    gate: str,
    ok: bool = True,
    bindings: Mapping[str, str] | None = None,
) -> str:
    """Emit the unique trailing release-marker sentinel for smoke/audit gates."""
    if gate not in _RELEASE_MARKER_GATES:
        raise ValueError(f"gate {gate!r} is not a registered release_markers gate")
    if gate != str(gate) or gate.strip() != gate:
        raise ValueError("release marker gate must be an exact registered name")
    payload = {
        "schema_version": RELEASE_RESULT_SCHEMA_VERSION,
        "ok": bool(ok),
        "gate": gate,
    }
    if bindings is not None:
        expected = (
            _TAURI_RELEASE_BINDING_FIELDS
            if gate == "tauri_webview_lifecycle"
            else _DESKTOP_RELEASE_BINDING_FIELDS
            if gate == "desktop_build"
            else frozenset()
        )
        if not ok or set(bindings) != expected:
            raise ValueError("release marker artifact bindings are not exact for gate")
        if any(
            not isinstance(value, str)
            or _SHA256_PATTERN.fullmatch(value) is None
            or value != value.lower()
            for value in bindings.values()
        ):
            raise ValueError("release marker artifact binding must be lowercase SHA-256")
        payload["bindings"] = dict(bindings)
    return RELEASE_RESULT_SENTINEL + _canonical_json_text(payload)


def _parse_release_result_payload(
    stdout_text: str,
    *,
    expected_gate: str | None,
) -> dict[str, object] | None:
    if expected_gate is None:
        return None
    if expected_gate not in _RELEASE_MARKER_GATES or expected_gate.strip() != expected_gate:
        return None
    if stdout_text.count(RELEASE_RESULT_SENTINEL) != 1:
        return None
    lines = stdout_text.splitlines()
    non_empty = [line for line in lines if line.strip()]
    if not non_empty:
        return None
    last = non_empty[-1].strip()
    if not last.startswith(RELEASE_RESULT_SENTINEL):
        return None
    encoded = last[len(RELEASE_RESULT_SENTINEL) :]
    try:
        payload = strict_loads(encoded)
    except (StrictJSONError, ValueError, TypeError):
        return None
    if not isinstance(payload, dict):
        return None
    if encoded != _canonical_json_text(payload):
        return None
    if payload.get("schema_version") != RELEASE_RESULT_SCHEMA_VERSION:
        return None
    if type(payload.get("ok")) is not bool:
        return None
    gate = payload.get("gate")
    if not isinstance(gate, str) or gate != expected_gate:
        return None
    expected_binding_fields = (
        _TAURI_RELEASE_BINDING_FIELDS
        if gate == "tauri_webview_lifecycle"
        else _DESKTOP_RELEASE_BINDING_FIELDS
        if gate == "desktop_build"
        else None
    )
    expected_fields = RELEASE_RESULT_FIELDS | (
        {"bindings"} if payload.get("ok") is True and expected_binding_fields is not None else set()
    )
    if set(payload) != expected_fields:
        return None
    parsed: dict[str, object] = {
        "schema_version": payload["schema_version"],
        "ok": payload["ok"],
        "gate": gate,
    }
    if expected_binding_fields is not None and payload.get("ok") is True:
        bindings = payload.get("bindings")
        if not isinstance(bindings, dict) or set(bindings) != expected_binding_fields:
            return None
        if any(
            not isinstance(value, str)
            or _SHA256_PATTERN.fullmatch(value) is None
            or value != value.lower()
            for value in bindings.values()
        ):
            return None
        parsed["bindings"] = dict(bindings)
    return parsed


def _loads_reject_duplicate_keys(text: str) -> object:
    """Backward-compatible alias — all evidence JSON uses strict_loads."""
    return strict_loads(text)


def _parse_readiness_result_payload(stdout_text: str) -> dict[str, object] | None:
    """Parse the unique trailing readiness sentinel; reject trailing/multi/junk JSON."""
    # Raw occurrence count across the entire stdout — not just line-start matches.
    if stdout_text.count(READINESS_RESULT_SENTINEL) != 1:
        return None
    lines = stdout_text.splitlines()
    non_empty = [line for line in lines if line.strip()]
    if not non_empty:
        return None
    last_non_empty_index = max(i for i, line in enumerate(lines) if line.strip())
    raw = lines[last_non_empty_index].strip()
    if not raw.startswith(READINESS_RESULT_SENTINEL):
        return None
    encoded = raw[len(READINESS_RESULT_SENTINEL) :]
    try:
        payload = _loads_reject_duplicate_keys(encoded)
    except (json.JSONDecodeError, ValueError, TypeError):
        return None
    if not isinstance(payload, dict) or set(payload) != READINESS_RESULT_FIELDS:
        return None
    # Exact canonical serialization — reject alternate spacing/key order/unicode escapes.
    if encoded != _canonical_json_text(payload):
        return None
    schema = payload.get("schema_version")
    digest = payload.get("source_digest")
    ready = payload.get("internal_ready")
    blockers = payload.get("internal_blockers")
    if schema != READINESS_RESULT_SCHEMA_VERSION:
        return None
    if not isinstance(digest, str) or _SHA256_PATTERN.fullmatch(digest) is None:
        return None
    if digest != digest.lower():
        return None
    if type(ready) is not bool:
        return None
    if not isinstance(blockers, list) or not all(isinstance(item, str) for item in blockers):
        return None
    if ready is True and blockers:
        return None
    if ready is False and not blockers:
        return None
    return {
        "schema_version": schema,
        "source_digest": digest,
        "internal_ready": ready,
        "internal_blockers": list(blockers),
    }


def _is_lower_sha256(value: object) -> bool:
    return bool(
        isinstance(value, str)
        and value == value.lower()
        and _SHA256_PATTERN.fullmatch(value) is not None
    )


def _is_single_link_regular_file(path: Path) -> bool:
    try:
        file_stat = path.lstat()
    except OSError:
        return False
    return bool(stat.S_ISREG(file_stat.st_mode) and file_stat.st_nlink == 1)


def _supervised_targets(duration_seconds: float) -> dict[str, int]:
    scale = max(duration_seconds / 3600.0, 0.0)
    return {
        "mode_switches": max(1, int(30 * scale)) if duration_seconds >= 60 else 1,
        "app_restarts": max(1, int(6 * scale)) if duration_seconds >= 120 else 1,
        "sidecar_recoveries": max(1, int(3 * scale)) if duration_seconds >= 180 else 0,
        "ws_cancel_cycles": max(1, int(30 * scale)) if duration_seconds >= 60 else 1,
        "r4_ops": max(1, int(12 * scale)) if duration_seconds >= 120 else 0,
        "r6_ops": max(1, int(12 * scale)) if duration_seconds >= 180 else 0,
    }


def _valid_supervised_counters(value: object) -> bool:
    return bool(
        isinstance(value, dict)
        and set(value) == _SUPERVISED_COUNTER_KEYS
        and all(
            isinstance(item, int) and not isinstance(item, bool) and item >= 0
            for item in value.values()
        )
    )


def _valid_supervised_overlay_report(
    data: object,
    *,
    root: Path,
    evidence_dir: Path,
    expected_source_digest: str,
    expected_metadata_fingerprint: str,
) -> bool:
    if not isinstance(data, dict) or set(data) != _SUPERVISED_OVERLAY_FIELDS:
        return False
    duration = _safe_finite_float(data.get("duration_seconds"))
    elapsed = _safe_finite_float(data.get("elapsed_seconds"))
    started = _parse_utc_timestamp(data.get("started_utc"))
    finished = _parse_utc_timestamp(data.get("finished_utc"))
    acceptance_pid = data.get("acceptance_pid")
    if (
        data.get("schema_version") != _SUPERVISED_OVERLAY_SCHEMA_VERSION
        or data.get("ok") is not True
        or data.get("errors") != []
        or duration != 3600.0
        or elapsed is None
        or elapsed < duration - _SOAK_COVERAGE_TOLERANCE_SECONDS
        or started is None
        or finished is None
        or finished < started
        or (finished - started).total_seconds() < duration - _SOAK_COVERAGE_TOLERANCE_SECONDS
        or not isinstance(acceptance_pid, int)
        or isinstance(acceptance_pid, bool)
        or acceptance_pid <= 0
        or data.get("source_digest") != expected_source_digest
        or data.get("metadata_fingerprint") != expected_metadata_fingerprint
        or not _is_lower_sha256(data.get("source_digest"))
        or not _is_lower_sha256(data.get("metadata_fingerprint"))
    ):
        return False

    expected_targets = _supervised_targets(duration)
    targets = data.get("targets")
    counters = data.get("counters")
    if (
        not isinstance(targets, dict)
        or not isinstance(counters, dict)
        or not _valid_supervised_counters(targets)
        or targets != expected_targets
        or not _valid_supervised_counters(counters)
        or data.get("targets_met") is not True
        or any(counters[key] < expected_targets[key] for key in _SUPERVISED_COUNTER_KEYS)
    ):
        return False
    cycles = data.get("cycles")
    if not isinstance(cycles, int) or isinstance(cycles, bool) or cycles <= 0:
        return False

    resolved_evidence = evidence_dir.resolve()
    app_path = resolved_evidence / "desktop-build/artifacts/JS Agent.app"
    manifest_path = resolved_evidence / "desktop-build/manifest.json"
    harness_path = (
        resolved_evidence
        / "harness/JS Agent UI Test Harness.app/Contents/MacOS/js-agent-ui-test-harness"
    )
    if (
        data.get("app_path") != str(app_path)
        or data.get("desktop_manifest_path") != str(manifest_path)
        or data.get("harness_exec") != str(harness_path)
    ):
        return False
    try:
        from scripts.run_tauri_webview_gate import _manifest_bindings

        bindings = _manifest_bindings(
            app_path=app_path,
            manifest_path=manifest_path,
            repo_root=root.resolve(),
        )
    except (OSError, RuntimeError, TypeError, ValueError):
        return False
    if bindings is None or any(
        not _is_lower_sha256(data.get(field))
        or not hmac.compare_digest(str(data.get(field)), bindings[field])
        for field in ("desktop_manifest_sha256", "app_tree_sha256", "app_sha256")
    ):
        return False

    heartbeats = data.get("heartbeats")
    heartbeat_count = data.get("heartbeat_count")
    gap_limit = _safe_finite_float(data.get("max_heartbeat_gap_limit_s"))
    declared_max_gap = _safe_finite_float(data.get("max_heartbeat_gap_s"))
    if (
        not isinstance(heartbeats, list)
        or not isinstance(heartbeat_count, int)
        or isinstance(heartbeat_count, bool)
        or heartbeat_count != len(heartbeats)
        or heartbeat_count < 2
        or gap_limit != 15.0
        or declared_max_gap is None
        or declared_max_gap < 0
        or declared_max_gap > gap_limit + 0.5
    ):
        return False

    running_chain = bytes(32)
    prior_monotonic: float | None = None
    prior_wall: datetime | None = None
    setup_bias: float | None = None
    prior_counters = dict.fromkeys(_SUPERVISED_COUNTER_KEYS, 0)
    computed_max_gap = 0.0
    for index, heartbeat in enumerate(heartbeats, start=1):
        if not isinstance(heartbeat, dict) or set(heartbeat) != _SUPERVISED_HEARTBEAT_FIELDS:
            return False
        monotonic_s = _safe_finite_float(heartbeat.get("monotonic_s"))
        wall_utc = _parse_utc_timestamp(heartbeat.get("wall_utc"))
        note = heartbeat.get("note")
        heartbeat_counters = heartbeat.get("counters")
        if (
            heartbeat.get("index") != index
            or monotonic_s is None
            or monotonic_s < 0
            or wall_utc is None
            or wall_utc < started
            or wall_utc > finished
            or not isinstance(note, str)
            or not note
            or len(note) > 128
            or heartbeat.get("source_digest") != expected_source_digest
            or not isinstance(heartbeat_counters, dict)
            or not _valid_supervised_counters(heartbeat_counters)
            or heartbeat.get("prev_chain") != running_chain.hex()
        ):
            return False
        if index == 1:
            if note != "overlay_start" or monotonic_s > _SOAK_SETUP_BIAS_MAX_SECONDS:
                return False
            if any(heartbeat_counters[key] != 0 for key in _SUPERVISED_COUNTER_KEYS):
                return False
            setup_bias = (wall_utc - started).total_seconds() - monotonic_s
            if (
                setup_bias < -_SOAK_WALL_MONO_ALIGN_TOLERANCE_SECONDS
                or setup_bias > _SOAK_SETUP_BIAS_MAX_SECONDS
            ):
                return False
        elif (
            setup_bias is None
            or abs((wall_utc - started).total_seconds() - monotonic_s - setup_bias)
            > _SOAK_WALL_MONO_ALIGN_TOLERANCE_SECONDS
        ):
            return False
        if any(heartbeat_counters[key] < prior_counters[key] for key in _SUPERVISED_COUNTER_KEYS):
            return False
        if prior_monotonic is not None:
            gap = monotonic_s - prior_monotonic
            if gap < 0 or gap > gap_limit + 0.5:
                return False
            computed_max_gap = max(computed_max_gap, gap)
        if prior_wall is not None and wall_utc < prior_wall:
            return False
        if prior_wall is not None and prior_monotonic is not None:
            wall_gap = (wall_utc - prior_wall).total_seconds()
            if abs(wall_gap - (monotonic_s - prior_monotonic)) > (
                _SOAK_WALL_MONO_ALIGN_TOLERANCE_SECONDS
            ):
                return False
        canonical = {
            field: heartbeat[field] for field in _SUPERVISED_HEARTBEAT_FIELDS if field != "chain"
        }
        running_chain = hashlib.sha256(
            json.dumps(canonical, sort_keys=True, separators=(",", ":")).encode("utf-8")
        ).digest()
        if heartbeat.get("chain") != running_chain.hex():
            return False
        prior_monotonic = monotonic_s
        prior_wall = wall_utc
        prior_counters = dict(heartbeat_counters)

    cleanup_seconds = (
        (finished - prior_wall).total_seconds() if prior_wall is not None else math.inf
    )
    return not (
        prior_monotonic is None
        or prior_monotonic < duration - _SOAK_COVERAGE_TOLERANCE_SECONDS
        or cleanup_seconds < -_SOAK_WALL_MONO_ALIGN_TOLERANCE_SECONDS
        or cleanup_seconds > _SOAK_FINAL_CLEANUP_MAX_SECONDS
        or prior_counters != counters
        or data.get("chain_root") != running_chain.hex()
        or not _is_lower_sha256(data.get("chain_root"))
        or abs(round(computed_max_gap, 3) - declared_max_gap) > 0.001
    )


def _valid_supervised_soak_artifact(
    *,
    root: Path,
    evidence_dir: Path,
    path: Path,
    expected_source_digest: str,
) -> bool:
    resolved_evidence = evidence_dir.resolve()
    expected_path = resolved_evidence / _SUPERVISED_SOAK_COMBINED_RELATIVE
    core_raw_path = resolved_evidence / _SUPERVISED_SOAK_CORE_RAW_RELATIVE
    overlay_raw_path = resolved_evidence / _SUPERVISED_SOAK_OVERLAY_RAW_RELATIVE
    if path.resolve() != expected_path or not all(
        _is_single_link_regular_file(candidate)
        for candidate in (expected_path, core_raw_path, overlay_raw_path)
    ):
        return False
    try:
        combined = strict_load_object(expected_path)
        overlay_raw = strict_load_object(overlay_raw_path)
    except (OSError, ValueError, StrictJSONError):
        return False
    if set(combined) != _SUPERVISED_COMBINED_FIELDS:
        return False
    claimed_combined_sha = combined.get("combined_sha256")
    unsigned_combined = {key: value for key, value in combined.items() if key != "combined_sha256"}
    actual_combined_sha = hashlib.sha256(
        json.dumps(
            unsigned_combined,
            sort_keys=True,
            separators=(",", ":"),
            ensure_ascii=False,
        ).encode("utf-8")
    ).hexdigest()
    duration = _safe_finite_float(combined.get("duration_seconds"))
    elapsed = _safe_finite_float(combined.get("elapsed_seconds"))
    started = _parse_utc_timestamp(combined.get("started_utc"))
    finished = _parse_utc_timestamp(combined.get("finished_utc"))
    current_metadata = release_source_surface_metadata_fingerprint(root.resolve())
    if (
        combined.get("schema_version") != _SUPERVISED_SOAK_SCHEMA_VERSION
        or combined.get("ok") is not True
        or not _is_lower_sha256(claimed_combined_sha)
        or not hmac.compare_digest(str(claimed_combined_sha), actual_combined_sha)
        or duration != 3600.0
        or elapsed is None
        or elapsed < duration - _SOAK_COVERAGE_TOLERANCE_SECONDS
        or started is None
        or finished is None
        or finished < started
        or (finished - started).total_seconds() < duration - _SOAK_COVERAGE_TOLERANCE_SECONDS
        or combined.get("source_digest") != expected_source_digest
        or release_source_digest(root.resolve()) != expected_source_digest
        or combined.get("metadata_fingerprint") != current_metadata
        or not _is_lower_sha256(current_metadata)
    ):
        return False

    core = combined.get("core")
    overlay = combined.get("overlay")
    if (
        not isinstance(core, dict)
        or set(core) != _SUPERVISED_CORE_FIELDS
        or core.get("exit_code") != 0
        or core.get("ok") is not True
        or not _is_lower_sha256(core.get("raw_sha256"))
        or core.get("raw_sha256") != _sha256_file(core_raw_path)
        or not isinstance(overlay, dict)
        or set(overlay) != _SUPERVISED_OVERLAY_SUMMARY_FIELDS
        or overlay.get("exit_code") != 0
        or overlay.get("ok") is not True
        or not _is_lower_sha256(overlay.get("raw_sha256"))
        or overlay.get("raw_sha256") != _sha256_file(overlay_raw_path)
    ):
        return False
    if not _valid_echo_live_acceptance(root.resolve(), core_raw_path):
        return False
    if not _valid_supervised_overlay_report(
        overlay_raw,
        root=root.resolve(),
        evidence_dir=resolved_evidence,
        expected_source_digest=expected_source_digest,
        expected_metadata_fingerprint=current_metadata,
    ):
        return False
    overlay_summary_fields = _SUPERVISED_OVERLAY_SUMMARY_FIELDS - {
        "exit_code",
        "raw_sha256",
    }
    if any(overlay.get(field) != overlay_raw.get(field) for field in overlay_summary_fields):
        return False
    overlay_started = _parse_utc_timestamp(overlay_raw.get("started_utc"))
    overlay_finished = _parse_utc_timestamp(overlay_raw.get("finished_utc"))
    return bool(
        overlay_started is not None
        and overlay_finished is not None
        and started <= overlay_started
        and finished >= overlay_finished
    )


def _artifact_path_from_argv(
    argv: Sequence[str],
    spec: LocalGateSpec,
    *,
    root: Path,
    evidence_dir: Path,
    source_digest: str,
) -> Path | None:
    parser = spec.output_parse.parser
    argv_list = list(argv)
    if spec.gate_name == "soak_3600" and parser == "soak_json":
        return evidence_dir.resolve() / _SUPERVISED_SOAK_COMBINED_RELATIVE
    if parser in {"slo_json", "soak_json"}:
        flag = "--output"
    elif parser == "e2e_json":
        flag = "--json"
    elif spec.gate_name == "echo_full_audit":
        # Bind the final audit markdown (default or explicit --output).
        flag = "--output"
        try:
            index = argv_list.index(flag)
            raw_path = argv_list[index + 1]
        except (ValueError, IndexError):
            from js.echo.ledger.final_evidence import default_echo_full_audit_artifact

            return default_echo_full_audit_artifact(root)
        if raw_path == _SOURCE_DIGEST_TOKEN:
            return None
        materialized = _materialize_gate_argv_template(
            (raw_path,),
            root=root,
            evidence_dir=evidence_dir,
            source_digest=source_digest,
        )[0]
        path = Path(materialized)
        return path if path.is_absolute() else (root / path).resolve()
    else:
        return None
    try:
        index = argv_list.index(flag)
        raw_path = argv_list[index + 1]
    except (ValueError, IndexError):
        return None
    if raw_path == _SOURCE_DIGEST_TOKEN:
        return None
    materialized = _materialize_gate_argv_template(
        (raw_path,),
        root=root,
        evidence_dir=evidence_dir,
        source_digest=source_digest,
    )[0]
    return Path(materialized)


def parse_gate_stdout(
    parser: str,
    stdout_text: str,
    *,
    exit_code: int,
    require_exit_code_zero: bool,
    expected_gate: str | None = None,
) -> dict[str, object]:
    result: dict[str, object] = {"parser": parser, "ok": False}
    if require_exit_code_zero and exit_code != 0:
        result["reason"] = "nonzero_exit"
        return result
    if parser == "git_diff":
        result["ok"] = stdout_text.strip() == ""
        return result
    if parser == "ruff":
        stripped = re.sub(r"\x1b\[[0-9;]*m", "", stdout_text).strip()
        result["ok"] = stripped == "" or stripped == "All checks passed!"
        return result
    if parser == "mypy":
        stripped = stdout_text.strip()
        silent_failure = bool(
            re.search(r"\b(?:traceback|internal error|error:|failed|aborted)\b", stripped, re.I)
        )
        success_text = (
            not stripped or "Success: no issues found" in stripped or "Found 0 errors" in stripped
        )
        result.update(
            {
                "success_text": success_text,
                "silent_failure_pattern": silent_failure,
                "ok": exit_code == 0 and success_text and not silent_failure,
            }
        )
        return result
    if parser == "pytest":
        parsed_text = "\n".join(
            line for line in stdout_text.splitlines() if not line.lstrip().startswith("#")
        )
        counts = {
            name: sum(int(value) for value in re.findall(rf"(\d+)\s+{name}\b", parsed_text))
            for name in ("passed", "failed", "error", "skipped")
        }
        result["counts"] = counts
        result["ok"] = (
            exit_code == 0
            and counts["passed"] > 0
            and counts["failed"] == 0
            and counts["error"] == 0
            and "FAILED" not in parsed_text
            and "ERROR" not in parsed_text
        )
        return result
    if parser == "release_markers":
        payload = _parse_release_result_payload(stdout_text, expected_gate=expected_gate)
        markers = len(re.findall(r"(?m)^[ \t]*\[OK\]\s+\S", stdout_text))
        json_ok = (
            isinstance(payload, dict)
            and payload.get("ok") is True
            and expected_gate is not None
            and payload.get("gate") == expected_gate
        )
        result.update(
            {
                "ok_markers": markers,
                "json_ok": json_ok,
                "payload": payload,
                "expected_gate": expected_gate,
                # Sentinel + exact gate identity are mandatory.
                "ok": json_ok,
            }
        )
        return result
    if parser == "readiness_json":
        payload = _parse_readiness_result_payload(stdout_text)
        valid = (
            isinstance(payload, dict)
            and payload.get("internal_ready") is True
            and payload.get("internal_blockers") == []
        )
        result.update({"payload": payload, "ok": valid})
        return result
    result["ok"] = parser in {
        "python_exit",
        "generic",
        "slo_json",
        "soak_json",
        "e2e_json",
    }
    return result


def _parse_gate_stdout(
    parser: str,
    stdout_text: str,
    *,
    exit_code: int,
    require_exit_code_zero: bool,
    expected_gate: str | None = None,
) -> bool:
    return (
        parse_gate_stdout(
            parser,
            stdout_text,
            exit_code=exit_code,
            require_exit_code_zero=require_exit_code_zero,
            expected_gate=expected_gate,
        ).get("ok")
        is True
    )


def _valid_gate_artifact(
    spec: LocalGateSpec,
    *,
    root: Path,
    evidence_dir: Path,
    argv: Sequence[str],
    source_digest: str,
    artifact_sha256: str | None,
) -> bool:
    parser = spec.output_parse.parser
    if parser not in {"slo_json", "soak_json", "e2e_json"}:
        return True
    if not isinstance(artifact_sha256, str) or len(artifact_sha256) != 64:
        return False
    artifact_path = _artifact_path_from_argv(
        argv,
        spec,
        root=root,
        evidence_dir=evidence_dir,
        source_digest=source_digest,
    )
    if artifact_path is None or not artifact_path.is_file():
        return False
    digest = _sha256_file(artifact_path)
    if digest != artifact_sha256:
        return False
    if parser == "slo_json":
        return _valid_echo_slo_benchmark(artifact_path, root=root)
    if parser == "soak_json":
        if spec.gate_name == "soak_3600":
            return _valid_supervised_soak_artifact(
                root=root,
                evidence_dir=evidence_dir,
                path=artifact_path,
                expected_source_digest=source_digest,
            )
        return _valid_echo_live_acceptance(root, artifact_path)
    if parser == "e2e_json":
        return _valid_isolated_venv_e2e(root, artifact_path)
    return False


def build_local_gate_specs(*, evidence_dir: Path) -> dict[str, LocalGateSpec]:
    evidence = evidence_dir.as_posix()
    slo_output = tuple(f"{_EVIDENCE_DIR_TOKEN}/slo/slo_run_{index}.json" for index in range(1, 6))
    specs: dict[str, LocalGateSpec] = {
        "git_diff_check": LocalGateSpec(
            gate_name="git_diff_check",
            argv=("git", "diff", "--check"),
            coverage_scope=(".github", "js", "js_work", "tests", "scripts"),
            output_parse=LocalGateOutputParseRules(parser="git_diff"),
        ),
        "ruff": LocalGateSpec(
            gate_name="ruff",
            argv=(".venv/bin/ruff", "check", "js/", "js_work/", "tests/"),
            coverage_scope=("js/", "js_work/", "tests/"),
            output_parse=LocalGateOutputParseRules(parser="ruff"),
        ),
        "mypy": LocalGateSpec(
            gate_name="mypy",
            argv=(".venv/bin/mypy", "js/", "js_work/", "--no-error-summary"),
            coverage_scope=("js/", "js_work/"),
            output_parse=LocalGateOutputParseRules(parser="mypy"),
        ),
        "pytest_targeted_round84": LocalGateSpec(
            gate_name="pytest_targeted_round84",
            argv=(
                ".venv/bin/python",
                "-m",
                "pytest",
                "tests/test_csv_pending_budget_round84.py",
                "tests/test_csv_pending_budget_round83.py",
                "tests/test_isolated_product_e2e_round84.py",
                "tests/test_isolated_product_e2e_round83.py",
                "tests/test_local_gate_receipt_round84.py",
                "tests/test_signer_nonregular_round83.py",
                "tests/work/test_excel_merge_formula_round83.py",
                "tests/work/test_csv_formula_literal_round83.py",
                "-q",
                "--tb=line",
            ),
            coverage_scope=("tests/test_local_gate_receipt_round84.py",),
            output_parse=LocalGateOutputParseRules(parser="pytest"),
        ),
        "pytest_targeted_round85": LocalGateSpec(
            gate_name="pytest_targeted_round85",
            argv=(
                ".venv/bin/python",
                "-m",
                "pytest",
                "tests/test_local_gate_receipt_round84.py",
                "tests/test_local_gate_receipt_round85.py",
                "tests/test_isolated_product_e2e_round84.py",
                "tests/test_isolated_product_e2e_round85.py",
                "-q",
                "--tb=line",
            ),
            coverage_scope=(
                "tests/test_local_gate_receipt_round85.py",
                "tests/test_isolated_product_e2e_round85.py",
            ),
            output_parse=LocalGateOutputParseRules(parser="pytest"),
        ),
        "pytest_targeted_round86": LocalGateSpec(
            gate_name="pytest_targeted_round86",
            argv=(
                ".venv/bin/python",
                "-m",
                "pytest",
                "tests/test_csv_unified_budget_round86.py",
                "tests/test_local_gate_receipt_round86.py",
                "tests/test_isolated_product_e2e_round86.py",
                "tests/test_work_ledger_archive_round86.py",
                "tests/test_soak_source_integrity_round86.py",
                "-q",
                "--tb=line",
            ),
            coverage_scope=(
                "tests/test_local_gate_receipt_round86.py",
                "tests/test_isolated_product_e2e_round86.py",
                "tests/test_work_ledger_archive_round86.py",
                "tests/test_soak_source_integrity_round86.py",
            ),
            output_parse=LocalGateOutputParseRules(parser="pytest"),
        ),
        "pytest_targeted_round87": LocalGateSpec(
            gate_name="pytest_targeted_round87",
            argv=(
                ".venv/bin/python",
                "-m",
                "pytest",
                "tests/test_csv_unified_budget_round87.py",
                "tests/test_local_gate_receipt_round87.py",
                "tests/test_isolated_product_e2e_round87.py",
                "tests/test_work_ledger_archive_round87.py",
                "tests/test_soak_source_integrity_round87.py",
                "-q",
                "--tb=line",
            ),
            coverage_scope=(
                "tests/test_csv_unified_budget_round87.py",
                "tests/test_local_gate_receipt_round87.py",
                "tests/test_isolated_product_e2e_round87.py",
                "tests/test_work_ledger_archive_round87.py",
                "tests/test_soak_source_integrity_round87.py",
            ),
            output_parse=LocalGateOutputParseRules(parser="pytest"),
        ),
        "pytest_targeted_round88": LocalGateSpec(
            gate_name="pytest_targeted_round88",
            argv=(
                ".venv/bin/python",
                "-m",
                "pytest",
                "tests/test_local_gate_receipt_round88.py",
                "tests/test_soak_source_integrity_round88.py",
                "tests/test_final_validator_readonly_round88.py",
                "-q",
                "--tb=line",
            ),
            coverage_scope=(
                "tests/test_local_gate_receipt_round88.py",
                "tests/test_soak_source_integrity_round88.py",
                "tests/test_final_validator_readonly_round88.py",
            ),
            output_parse=LocalGateOutputParseRules(parser="pytest"),
        ),
        "pytest_targeted_round89": LocalGateSpec(
            gate_name="pytest_targeted_round89",
            argv=(
                ".venv/bin/python",
                "-m",
                "pytest",
                "tests/test_local_gate_receipt_round89.py",
                "tests/test_soak_source_integrity_round89.py",
                "tests/test_e2e_ephemeral_signing_round89.py",
                "tests/test_evidence_export_round89.py",
                "-q",
                "--tb=line",
            ),
            coverage_scope=(
                "tests/test_local_gate_receipt_round89.py",
                "tests/test_soak_source_integrity_round89.py",
                "tests/test_e2e_ephemeral_signing_round89.py",
                "tests/test_evidence_export_round89.py",
            ),
            output_parse=LocalGateOutputParseRules(parser="pytest"),
        ),
        "pytest_targeted_round810": LocalGateSpec(
            gate_name="pytest_targeted_round810",
            argv=(
                ".venv/bin/python",
                "-m",
                "pytest",
                "tests/test_local_gate_receipt_round810.py",
                "tests/test_strict_json_round810.py",
                "tests/test_soak_nonfinite_round810.py",
                "tests/test_e2e_key_lifecycle_round810.py",
                "tests/test_e2e_ephemeral_signing_round89.py",
                "tests/test_evidence_export_round810.py",
                "tests/test_evidence_export_round89.py",
                "-q",
                "--tb=line",
            ),
            coverage_scope=(
                "tests/test_local_gate_receipt_round810.py",
                "tests/test_strict_json_round810.py",
                "tests/test_soak_nonfinite_round810.py",
                "tests/test_e2e_key_lifecycle_round810.py",
                "tests/test_e2e_ephemeral_signing_round89.py",
                "tests/test_evidence_export_round810.py",
                "tests/test_evidence_export_round89.py",
            ),
            output_parse=LocalGateOutputParseRules(parser="pytest"),
        ),
        "pytest_targeted_round811": LocalGateSpec(
            gate_name="pytest_targeted_round811",
            argv=(
                ".venv/bin/python",
                "-m",
                "pytest",
                "tests/test_e2e_key_destroy_round811.py",
                "tests/test_release_marker_binding_round811.py",
                "tests/test_strict_json_integer_bounds_round811.py",
                "tests/test_manifest_closure_round811.py",
                "tests/test_export_receipt_closure_round811.py",
                "tests/test_archive_scan_binding_round811.py",
                "tests/test_e2e_key_lifecycle_round810.py",
                "tests/test_e2e_ephemeral_signing_round89.py",
                "tests/test_evidence_export_round810.py",
                "tests/test_strict_json_round810.py",
                "tests/test_local_gate_receipt_round810.py",
                "-q",
                "--tb=line",
            ),
            coverage_scope=(
                "tests/test_e2e_key_destroy_round811.py",
                "tests/test_release_marker_binding_round811.py",
                "tests/test_strict_json_integer_bounds_round811.py",
                "tests/test_manifest_closure_round811.py",
                "tests/test_export_receipt_closure_round811.py",
                "tests/test_archive_scan_binding_round811.py",
                "tests/test_e2e_key_lifecycle_round810.py",
                "tests/test_e2e_ephemeral_signing_round89.py",
                "tests/test_evidence_export_round810.py",
                "tests/test_strict_json_round810.py",
                "tests/test_local_gate_receipt_round810.py",
            ),
            output_parse=LocalGateOutputParseRules(parser="pytest"),
        ),
        "pytest_targeted_round812": LocalGateSpec(
            gate_name="pytest_targeted_round812",
            argv=(
                ".venv/bin/python",
                "-m",
                "pytest",
                "tests/test_e2e_key_destroy_round811.py",
                "tests/test_e2e_key_destroy_race_round812.py",
                "tests/test_e2e_key_prepare_rollback_round812.py",
                "tests/test_archive_verifier_rescan_round812.py",
                "tests/test_archive_artifact_set_round812.py",
                "tests/test_release_marker_binding_round811.py",
                "tests/test_strict_json_integer_bounds_round811.py",
                "tests/test_manifest_closure_round811.py",
                "tests/test_export_receipt_closure_round811.py",
                "tests/test_archive_scan_binding_round811.py",
                "tests/test_e2e_key_lifecycle_round810.py",
                "tests/test_e2e_ephemeral_signing_round89.py",
                "tests/test_evidence_export_round810.py",
                "tests/test_strict_json_round810.py",
                "tests/test_local_gate_receipt_round810.py",
                "-q",
                "--tb=line",
            ),
            coverage_scope=(
                "tests/test_e2e_key_destroy_round811.py",
                "tests/test_e2e_key_destroy_race_round812.py",
                "tests/test_e2e_key_prepare_rollback_round812.py",
                "tests/test_archive_verifier_rescan_round812.py",
                "tests/test_archive_artifact_set_round812.py",
                "tests/test_release_marker_binding_round811.py",
                "tests/test_strict_json_integer_bounds_round811.py",
                "tests/test_manifest_closure_round811.py",
                "tests/test_export_receipt_closure_round811.py",
                "tests/test_archive_scan_binding_round811.py",
                "tests/test_e2e_key_lifecycle_round810.py",
                "tests/test_e2e_ephemeral_signing_round89.py",
                "tests/test_evidence_export_round810.py",
                "tests/test_strict_json_round810.py",
                "tests/test_local_gate_receipt_round810.py",
            ),
            output_parse=LocalGateOutputParseRules(parser="pytest"),
        ),
        "pytest_targeted_round813": LocalGateSpec(
            gate_name="pytest_targeted_round813",
            argv=(
                ".venv/bin/python",
                "-m",
                "pytest",
                "tests/test_e2e_key_destroy_round811.py",
                "tests/test_e2e_key_destroy_race_round812.py",
                "tests/test_e2e_key_destroy_close_round813.py",
                "tests/test_e2e_key_prepare_rollback_round812.py",
                "tests/test_e2e_key_prepare_rollback_round813.py",
                "tests/test_archive_verifier_rescan_round812.py",
                "tests/test_archive_artifact_set_round812.py",
                "tests/test_archive_full_tree_round813.py",
                "tests/test_release_marker_binding_round811.py",
                "tests/test_strict_json_integer_bounds_round811.py",
                "tests/test_manifest_closure_round811.py",
                "tests/test_export_receipt_closure_round811.py",
                "tests/test_archive_scan_binding_round811.py",
                "tests/test_e2e_key_lifecycle_round810.py",
                "tests/test_e2e_ephemeral_signing_round89.py",
                "tests/test_evidence_export_round810.py",
                "tests/test_strict_json_round810.py",
                "tests/test_local_gate_receipt_round810.py",
                "-q",
                "--tb=line",
            ),
            coverage_scope=(
                "tests/test_e2e_key_destroy_round811.py",
                "tests/test_e2e_key_destroy_race_round812.py",
                "tests/test_e2e_key_destroy_close_round813.py",
                "tests/test_e2e_key_prepare_rollback_round812.py",
                "tests/test_e2e_key_prepare_rollback_round813.py",
                "tests/test_archive_verifier_rescan_round812.py",
                "tests/test_archive_artifact_set_round812.py",
                "tests/test_archive_full_tree_round813.py",
                "tests/test_release_marker_binding_round811.py",
                "tests/test_strict_json_integer_bounds_round811.py",
                "tests/test_manifest_closure_round811.py",
                "tests/test_export_receipt_closure_round811.py",
                "tests/test_archive_scan_binding_round811.py",
                "tests/test_e2e_key_lifecycle_round810.py",
                "tests/test_e2e_ephemeral_signing_round89.py",
                "tests/test_evidence_export_round810.py",
                "tests/test_strict_json_round810.py",
                "tests/test_local_gate_receipt_round810.py",
            ),
            output_parse=LocalGateOutputParseRules(parser="pytest"),
        ),
        "pytest_targeted_round815": LocalGateSpec(
            gate_name="pytest_targeted_round815",
            argv=(
                ".venv/bin/python",
                "-m",
                "pytest",
                "tests/test_baseline_provenance_round815.py",
                "tests/test_release_validator_resource_round815.py",
                "tests/test_source_integrity_preflight_round815.py",
                "tests/work/test_work_output_staging_round815.py",
                "tests/work/test_work_report_rollback_round815.py",
                "tests/work/test_work_shared_office_snapshot_round815.py",
                "-q",
                "--tb=line",
            ),
            coverage_scope=(
                "tests/test_baseline_provenance_round815.py",
                "tests/test_release_validator_resource_round815.py",
                "tests/test_source_integrity_preflight_round815.py",
                "tests/work/test_work_output_staging_round815.py",
                "tests/work/test_work_report_rollback_round815.py",
                "tests/work/test_work_shared_office_snapshot_round815.py",
            ),
            output_parse=LocalGateOutputParseRules(parser="pytest"),
        ),
        "pytest_work": LocalGateSpec(
            gate_name="pytest_work",
            argv=(".venv/bin/python", "-m", "pytest", "tests/work/", "-q", "--tb=line"),
            coverage_scope=("tests/work/",),
            output_parse=LocalGateOutputParseRules(parser="pytest"),
        ),
        "pytest_full_not_playwright": LocalGateSpec(
            gate_name="pytest_full_not_playwright",
            argv=(
                ".venv/bin/python",
                "-m",
                "pytest",
                "tests/",
                "-m",
                "not playwright",
                "-q",
                "--tb=line",
            ),
            coverage_scope=("tests/",),
            output_parse=LocalGateOutputParseRules(parser="pytest"),
        ),
        "pytest_playwright": LocalGateSpec(
            gate_name="pytest_playwright",
            argv=(
                ".venv/bin/python",
                "-m",
                "pytest",
                "-m",
                "playwright",
                "tests/e2e",
                "-q",
                "--tb=line",
            ),
            coverage_scope=("tests/e2e",),
            output_parse=LocalGateOutputParseRules(parser="pytest"),
        ),
        "release_smoke": LocalGateSpec(
            gate_name="release_smoke",
            argv=(".venv/bin/python", "scripts/release_smoke.py", "--all"),
            coverage_scope=("scripts/release_smoke.py",),
            output_parse=LocalGateOutputParseRules(parser="release_markers"),
        ),
        "echo_full_audit": LocalGateSpec(
            gate_name="echo_full_audit",
            argv=(".venv/bin/python", "scripts/echo_full_audit.py"),
            coverage_scope=("scripts/echo_full_audit.py", "js/echo/"),
            output_parse=LocalGateOutputParseRules(parser="release_markers"),
        ),
        "soak_3600": LocalGateSpec(
            gate_name="soak_3600",
            # Parent-supervised Echo core soak + Tauri lifecycle overlay.
            # Core still writes docs/security/ECHO_LIVE_ACCEPTANCE.json for soak_json.
            argv=(
                ".venv/bin/python",
                "-u",
                "scripts/run_supervised_soak.py",
                "--duration-seconds",
                "3600",
                "--concurrency",
                "2",
                "--output",
                f"{_REPO_ROOT_TOKEN}/docs/security/ECHO_LIVE_ACCEPTANCE.json",
                "--evidence-dir",
                _EVIDENCE_DIR_TOKEN,
                "--app-path",
                f"{_EVIDENCE_DIR_TOKEN}/desktop-build/artifacts/JS Agent.app",
                "--harness-path",
                f"{_EVIDENCE_DIR_TOKEN}/harness/JS Agent UI Test Harness.app",
            ),
            coverage_scope=(
                "scripts/run_supervised_soak.py",
                "scripts/echo_live_acceptance.py",
                "desktop/tests/harness/",
            ),
            output_parse=LocalGateOutputParseRules(parser="soak_json", stderr_must_be_empty=True),
        ),
        "isolated_venv_e2e": LocalGateSpec(
            gate_name="isolated_venv_e2e",
            argv=(
                ".venv/bin/python",
                "scripts/isolated_venv_e2e.py",
                "--source-digest",
                _SOURCE_DIGEST_TOKEN,
                "--wheelhouse",
                f"{_EVIDENCE_DIR_TOKEN}/wheelhouse",
                "--evidence-dir",
                _EVIDENCE_DIR_TOKEN,
                "--json",
                f"{_EVIDENCE_DIR_TOKEN}/e2e/ECHO_ISOLATED_VENV_E2E.json",
                "--log",
                f"{_EVIDENCE_DIR_TOKEN}/e2e/isolated_venv_e2e.out",
            ),
            coverage_scope=("scripts/isolated_venv_e2e.py",),
            output_parse=LocalGateOutputParseRules(parser="e2e_json"),
        ),
        "strict_readiness": LocalGateSpec(
            gate_name="strict_readiness",
            argv=(
                ".venv/bin/python",
                "-c",
                "from pathlib import Path; "
                "from js.echo.ledger.release_gates import ("
                "format_readiness_result_line, release_source_digest, verify_release_readiness"
                "); "
                "import sys; "
                "root = Path('.').resolve(); "
                "digest = release_source_digest(root); "
                "report = verify_release_readiness(root); "
                "print(format_readiness_result_line(report, source_digest=digest)); "
                "sys.exit(0 if report.internal_ready else 1)",
            ),
            coverage_scope=("js/echo/ledger/release_gates.py",),
            output_parse=LocalGateOutputParseRules(parser="readiness_json"),
        ),
        "desktop_build": LocalGateSpec(
            gate_name="desktop_build",
            argv=(
                ".venv/bin/python",
                "scripts/run_desktop_build_gate.py",
                "--evidence-dir",
                _EVIDENCE_DIR_TOKEN,
            ),
            coverage_scope=("desktop/", "scripts/run_desktop_build_gate.py"),
            output_parse=LocalGateOutputParseRules(parser="release_markers"),
        ),
        "tauri_webview_lifecycle": LocalGateSpec(
            gate_name="tauri_webview_lifecycle",
            argv=(
                ".venv/bin/python",
                "scripts/run_tauri_webview_gate.py",
                "--evidence-dir",
                _EVIDENCE_DIR_TOKEN,
                "--app-path",
                f"{_EVIDENCE_DIR_TOKEN}/desktop-build/artifacts/JS Agent.app",
            ),
            coverage_scope=(
                "desktop/src-tauri/",
                "scripts/run_tauri_webview_gate.py",
            ),
            output_parse=LocalGateOutputParseRules(parser="release_markers"),
        ),
    }
    for index, output_path in enumerate(slo_output, start=1):
        specs[f"slo_run_{index}"] = LocalGateSpec(
            gate_name=f"slo_run_{index}",
            argv=(
                ".venv/bin/python",
                "scripts/echo_architecture_benchmark.py",
                "--iterations",
                "50",
                "--warmup",
                "10",
                "--enforce-slo",
                "--baseline",
                f"{_REPO_ROOT_TOKEN}/docs/security/ECHO_BASELINE_65CC545.json",
                "--output",
                output_path,
            ),
            coverage_scope=("scripts/echo_architecture_benchmark.py", "js/echo/"),
            output_parse=LocalGateOutputParseRules(parser="slo_json"),
        )
    _ = evidence
    return specs


def get_local_gate_spec(gate_name: str, *, evidence_dir: Path) -> LocalGateSpec | None:
    return build_local_gate_specs(evidence_dir=evidence_dir).get(gate_name)


def expected_gate_argv(
    spec: LocalGateSpec,
    *,
    root: Path,
    evidence_dir: Path,
    source_digest: str,
) -> tuple[str, ...]:
    return _materialize_gate_argv_template(
        spec.argv,
        root=root,
        evidence_dir=evidence_dir,
        source_digest=source_digest,
    )


def argv_matches_gate_spec(
    argv: Sequence[str],
    spec: LocalGateSpec,
    *,
    root: Path,
    evidence_dir: Path,
    source_digest: str,
) -> bool:
    expected = expected_gate_argv(
        spec,
        root=root,
        evidence_dir=evidence_dir,
        source_digest=source_digest,
    )
    normalized_actual = normalize_gate_argv(
        argv,
        root=root,
        evidence_dir=evidence_dir,
        source_digest=source_digest,
    )
    normalized_expected = normalize_gate_argv(
        expected,
        root=root,
        evidence_dir=evidence_dir,
        source_digest=source_digest,
    )
    return normalized_actual == normalized_expected


def _receipt_duration_consistent(receipt: Mapping[str, object]) -> bool:
    start = _parse_utc_timestamp(receipt.get("start_utc"))
    end = _parse_utc_timestamp(receipt.get("end_utc"))
    duration = receipt.get("duration_seconds")
    if start is None or end is None:
        return False
    if end < start:
        return False
    number = _safe_finite_float(duration)
    if number is None or number < 0:
        return False
    delta = (end - start).total_seconds()
    return abs(number - delta) <= _LOCAL_GATE_DURATION_TOLERANCE_SECONDS


def _capture_paths_under_evidence(
    receipt: Mapping[str, object],
    *,
    evidence_dir: Path,
    root: Path | None = None,
) -> bool:
    evidence_resolved = evidence_dir.resolve()
    resolved_root = (root or evidence_dir).resolve()
    for field in ("stdout_path", "stderr_path"):
        raw = receipt.get(field)
        if not isinstance(raw, str) or not raw.strip():
            return False
        path = expand_path_tokens(raw, root=resolved_root, evidence_dir=evidence_dir).resolve()
        try:
            path.relative_to(evidence_resolved)
        except ValueError:
            return False
    return True


# Hard local gates that must have passing final receipts. ``verify_release_readiness``
# is tracked separately and never substitutes for ruff/mypy/pytest receipts.
REQUIRED_FINAL_LOCAL_GATES: tuple[str, ...] = (
    "git_diff_check",
    "ruff",
    "mypy",
    "pytest_targeted_round815",
    "pytest_work",
    "pytest_full_not_playwright",
    "pytest_playwright",
    # Latency/packaging gates bind digest-scoped artifacts before release_smoke's
    # echo_ledger check (which requires internal_ready).
    "slo_run_1",
    "slo_run_2",
    "slo_run_3",
    "slo_run_4",
    "slo_run_5",
    "soak_3600",
    "isolated_venv_e2e",
    "release_smoke",
    # Desktop build and real Tauri WebView lifecycle must pass for product readiness.
    "desktop_build",
    "tauri_webview_lifecycle",
    # Audit after soak/e2e so Internal release ready reflects live acceptance.
    "echo_full_audit",
    "strict_readiness",
)

_INDEPENDENT_GATE_RECEIPTS: frozenset[str] = frozenset(
    {
        "ruff",
        "mypy",
        "pytest_full_not_playwright",
        "pytest_work",
        "pytest_playwright",
        "desktop_build",
        "tauri_webview_lifecycle",
    }
)


@dataclass(frozen=True)
class FinalLocalGateEvidenceReport:
    all_local_gates_passed: bool
    passed_gates: tuple[str, ...]
    blockers: tuple[str, ...]
    product_internal_ready: bool = False


def snapshot_final_gate_inputs(root: Path, evidence_dir: Path) -> dict[str, object]:
    """Archive validator inputs so a timestamped evidence bundle is self-describing."""
    resolved_root = root.resolve()
    target = evidence_dir.resolve() / "validator_inputs"
    target.mkdir(parents=True, exist_ok=True)
    entries: dict[str, dict[str, object]] = {}
    for label, relative in (
        ("baseline_script", Path("benchmarks/old_architecture_baseline.py")),
        ("benchmark_script", Path("scripts/echo_architecture_benchmark.py")),
    ):
        source = resolved_root / relative
        if not source.is_file():
            entries[label] = {"source": relative.as_posix(), "present": False}
            continue
        destination = target / relative.name
        shutil.copy2(source, destination)
        entries[label] = {
            "source": relative.as_posix(),
            "path": destination.relative_to(evidence_dir.resolve()).as_posix(),
            "sha256": hashlib.sha256(destination.read_bytes()).hexdigest(),
            "present": True,
        }
    tokenizer_payload = {
        "schema_version": "js-agent-tokenizer-digest-evidence-v1",
        "digest_version": _TOKENIZER_TREE_DIGEST_VERSION.decode("ascii").rstrip("\0"),
        "tokenizer_resource_sha256": tokenizer_resource_digest(resolved_root),
    }
    tokenizer_path = target / "tokenizer.digest.json"
    tokenizer_path.write_text(
        json.dumps(tokenizer_payload, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    entries["tokenizer_digest"] = {
        "path": tokenizer_path.relative_to(evidence_dir.resolve()).as_posix(),
        "sha256": hashlib.sha256(tokenizer_path.read_bytes()).hexdigest(),
        "present": True,
    }
    return {
        "schema_version": "js-agent-final-gate-inputs-v1",
        "artifacts": entries,
    }


def write_final_validator_receipt(
    evidence_dir: Path,
    *,
    summary_payload: Mapping[str, object],
    validator_payload: Mapping[str, object],
) -> None:
    """Write the gate summary and its final validator receipt together."""
    resolved_evidence = evidence_dir.resolve()
    resolved_evidence.mkdir(parents=True, exist_ok=True)
    (resolved_evidence / "gate_run_summary.json").write_text(
        json.dumps(dict(summary_payload), indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    (resolved_evidence / "final_validator.receipt.json").write_text(
        json.dumps(dict(validator_payload), indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def _sha256_file(path: Path) -> str | None:
    try:
        return hashlib.sha256(path.read_bytes()).hexdigest()
    except OSError:
        return None


def _valid_desktop_release_bindings(
    parse_result: object,
    *,
    gate_name: str,
    root: Path,
    evidence_dir: Path,
    expected_source_digest: str,
) -> bool:
    if gate_name not in {"desktop_build", "tauri_webview_lifecycle"}:
        return True
    if not isinstance(parse_result, dict):
        return False
    payload = parse_result.get("payload")
    if not isinstance(payload, dict):
        return False
    bindings = payload.get("bindings")
    expected_fields = (
        _TAURI_RELEASE_BINDING_FIELDS
        if gate_name == "tauri_webview_lifecycle"
        else _DESKTOP_RELEASE_BINDING_FIELDS
    )
    if not isinstance(bindings, dict) or set(bindings) != expected_fields:
        return False

    manifest_path = evidence_dir.resolve() / "desktop-build/manifest.json"
    app_path = evidence_dir.resolve() / "desktop-build/artifacts/JS Agent.app"
    executable = app_path / "Contents/MacOS/js-agent-desktop"
    try:
        from desktop.build_driver import _sha256_tree, verify_manifest

        if verify_manifest(manifest_path, repo_root=root.resolve()):
            return False
        manifest = strict_load_object(manifest_path)
        artifacts = manifest.get("artifacts")
        if not isinstance(artifacts, dict):
            return False
        rust_main = artifacts.get("rust_main")
        app_tree = artifacts.get("app_tree")
        if not isinstance(rust_main, dict) or not isinstance(app_tree, dict):
            return False
        if (
            manifest.get("source_digest") != expected_source_digest
            or rust_main.get("path") != "artifacts/JS Agent.app/Contents/MacOS/js-agent-desktop"
            or app_tree.get("path") != "artifacts/JS Agent.app"
        ):
            return False
        actual_app_sha = _sha256_file(executable)
        actual_tree_sha = _sha256_tree(app_path)
        actual_manifest_sha = _sha256_file(manifest_path)
        import plistlib

        info = plistlib.loads((app_path / "Contents/Info.plist").read_bytes())
    except (OSError, RuntimeError, TypeError, ValueError, StrictJSONError):
        return False
    actual = {
        "desktop_manifest_sha256": actual_manifest_sha,
        "app_tree_sha256": actual_tree_sha,
        "app_sha256": actual_app_sha,
    }
    if (
        info.get("CFBundleIdentifier") != "com.titan.js-agent"
        or info.get("CFBundleExecutable") != "js-agent-desktop"
        or rust_main.get("sha256") != actual_app_sha
        or app_tree.get("sha256") != actual_tree_sha
        or any(
            not isinstance(value, str)
            or not hmac.compare_digest(value, str(bindings.get(field, "")))
            for field, value in actual.items()
        )
    ):
        return False
    if gate_name == "desktop_build":
        return True

    result_path = evidence_dir.resolve() / "tauri-webview/result.json"
    harness_path = (
        evidence_dir.resolve()
        / "harness/JS Agent UI Test Harness.app/Contents/MacOS/js-agent-ui-test-harness"
    )
    try:
        result = strict_load_object(result_path)
        from scripts.run_tauri_webview_gate import (
            _RESULT_FIELDS,
            EXPECTED_BUNDLE_IDENTIFIER,
            RESULT_SCHEMA_VERSION,
            _trusted_harness_hash,
            _valid_scenarios,
        )

        result_sha = _sha256_file(result_path)
        harness_sha = _trusted_harness_hash(
            harness_bundle=harness_path.parent.parent.parent,
            harness_exec=harness_path,
            repo_root=root.resolve(),
        )
    except (OSError, ValueError, StrictJSONError):
        return False
    if (
        set(result) != _RESULT_FIELDS
        or result.get("schema_version") != RESULT_SCHEMA_VERSION
        or result.get("ok") is not True
        or result.get("status") != "passed"
        or result.get("accessibility_authorized") is not True
        or result.get("bundle_identifier") != EXPECTED_BUNDLE_IDENTIFIER
        or result.get("app_sha256") != actual_app_sha
        or result.get("app_tree_sha256") != actual_tree_sha
        or result.get("desktop_manifest_sha256") != actual_manifest_sha
        or result.get("harness_sha256") != harness_sha
        or not _valid_scenarios(result.get("scenarios"))
        or result_sha != bindings.get("result_sha256")
        or harness_sha != bindings.get("harness_sha256")
    ):
        return False
    nonce = result.get("nonce")
    started = _parse_utc_timestamp(result.get("started_utc"))
    finished = _parse_utc_timestamp(result.get("finished_utc"))
    return bool(
        isinstance(nonce, str)
        and re.fullmatch(r"[0-9a-f]{64}", nonce)
        and started is not None
        and finished is not None
        and started <= finished
    )


_RECEIPT_REQUIRED_KEYS: frozenset[str] = frozenset(
    {
        "schema_version",
        "gate_spec_version",
        "gate_name",
        "argv",
        "normalized_argv",
        "coverage_scope",
        "output_parse",
        "toolchain",
        "toolchain_lock_sha256",
        "toolchain_before",
        "toolchain_after",
        "parse_result",
        "cwd",
        "evidence_dir",
        "start_utc",
        "end_utc",
        "duration_seconds",
        "source_digest_before",
        "source_digest_after",
        "exit_code",
        "stdout_path",
        "stderr_path",
        "stdout_sha256",
        "stderr_sha256",
        "passed",
    }
)


def _valid_local_gate_receipt(
    receipt: object,
    *,
    root: Path,
    expected_source_digest: str,
    evidence_dir: Path,
    gate_spec: LocalGateSpec | None = None,
) -> bool:
    """Return True when a receipt is eligible for final evidence success."""
    if not isinstance(receipt, dict):
        return False
    if receipt.get("schema_version") != LOCAL_GATE_RECEIPT_SCHEMA_VERSION:
        return False
    if receipt.get("gate_spec_version") != LOCAL_GATE_SPEC_VERSION:
        return False
    gate_name = receipt.get("gate_name")
    if not isinstance(gate_name, str) or not gate_name.strip():
        return False
    spec = gate_spec or get_local_gate_spec(gate_name, evidence_dir=evidence_dir)
    if spec is None or spec.gate_name != gate_name:
        return False
    # Closed-set key validation: reject receipts with unknown fields.
    required_keys = _RECEIPT_REQUIRED_KEYS.copy()
    artifact_sha = receipt.get("artifact_sha256")
    if artifact_sha is not None:
        required_keys = required_keys | {"artifact_sha256"}
    if receipt.get("source_drift") is True:
        required_keys = required_keys | {"source_drift"}
    if receipt.get("toolchain_drift") is True:
        required_keys = required_keys | {"toolchain_drift"}
    if set(receipt.keys()) != required_keys:
        return False
    coverage_scope = receipt.get("coverage_scope")
    if coverage_scope != list(spec.coverage_scope):
        return False
    output_parse = receipt.get("output_parse")
    if not isinstance(output_parse, dict):
        return False
    if output_parse.get("require_exit_code_zero") is not spec.output_parse.require_exit_code_zero:
        return False
    if output_parse.get("stderr_must_be_empty") is not spec.output_parse.stderr_must_be_empty:
        return False
    parser = output_parse.get("parser")
    if parser != spec.output_parse.parser:
        return False
    normalized_argv = receipt.get("normalized_argv")
    argv = receipt.get("argv")
    if (
        not isinstance(argv, list)
        or not argv
        or not all(isinstance(item, str) and item for item in argv)
    ):
        return False
    if not isinstance(normalized_argv, list) or not all(
        isinstance(item, str) and item for item in normalized_argv
    ):
        return False
    expected_normalized = normalize_gate_argv(
        argv,
        root=root,
        evidence_dir=evidence_dir,
        source_digest=expected_source_digest,
    )
    if tuple(normalized_argv) != expected_normalized:
        return False
    if not argv_matches_gate_spec(
        argv,
        spec,
        root=root,
        evidence_dir=evidence_dir,
        source_digest=expected_source_digest,
    ):
        return False
    cwd = receipt.get("cwd")
    if not isinstance(cwd, str) or not cwd.strip():
        return False
    cwd_path = expand_path_tokens(cwd, root=root, evidence_dir=evidence_dir)
    if cwd_path.resolve() != root.resolve():
        return False
    if not _receipt_duration_consistent(receipt):
        return False
    digest_before = receipt.get("source_digest_before")
    digest_after = receipt.get("source_digest_after")
    if not isinstance(digest_before, str) or not isinstance(digest_after, str):
        return False
    if len(digest_before) != 64 or len(digest_after) != 64:
        return False
    if not hmac.compare_digest(digest_before, expected_source_digest):
        return False
    if not hmac.compare_digest(digest_after, expected_source_digest):
        return False
    if receipt.get("source_drift") is True:
        return False
    exit_code = receipt.get("exit_code")
    if not isinstance(exit_code, int) or isinstance(exit_code, bool):
        return False
    stdout_path = receipt.get("stdout_path")
    stderr_path = receipt.get("stderr_path")
    stdout_sha256 = receipt.get("stdout_sha256")
    stderr_sha256 = receipt.get("stderr_sha256")
    if not isinstance(stdout_path, str) or not stdout_path.strip():
        return False
    if not isinstance(stderr_path, str) or not stderr_path.strip():
        return False
    if not isinstance(stdout_sha256, str) or not isinstance(stderr_sha256, str):
        return False
    if len(stdout_sha256) != 64 or len(stderr_sha256) != 64:
        return False
    expected_stdout, expected_stderr = canonical_gate_capture_paths(gate_name, evidence_dir)
    stdout_file = expand_path_tokens(stdout_path, root=root, evidence_dir=evidence_dir)
    stderr_file = expand_path_tokens(stderr_path, root=root, evidence_dir=evidence_dir)
    if (
        stdout_file.resolve() != expected_stdout.resolve()
        or stderr_file.resolve() != expected_stderr.resolve()
    ):
        return False
    if not _capture_paths_under_evidence(receipt, evidence_dir=evidence_dir, root=root):
        return False
    if not stdout_file.is_file() or not stderr_file.is_file():
        return False
    if _sha256_file(stdout_file) != stdout_sha256:
        return False
    if _sha256_file(stderr_file) != stderr_sha256:
        return False
    if spec.output_parse.stderr_must_be_empty and stderr_file.stat().st_size != 0:
        return False
    if not _valid_receipt_toolchain(receipt, root=root, argv=argv, evidence_dir=evidence_dir):
        return False
    lock_path = evidence_dir.resolve() / "TOOLCHAIN.lock.json"
    try:
        lock_bytes = lock_path.read_bytes()
        frozen_lock = strict_load_object_bytes(lock_bytes)
    except (OSError, json.JSONDecodeError, StrictJSONError, ValueError):
        return False
    if not _valid_frozen_toolchain_lock(frozen_lock):
        return False
    lock_sha = receipt.get("toolchain_lock_sha256")
    before = receipt.get("toolchain_before")
    after = receipt.get("toolchain_after")
    # Receipts may store path-tokenized toolchain locks for privacy.
    from js.echo.ledger.evidence_export import redact_text as _redact_text

    def _redact_lock(value: object) -> object:
        if isinstance(value, str):
            return _redact_text(value, repo_root=root, evidence_root=evidence_dir)
        if isinstance(value, dict):
            return {str(key): _redact_lock(item) for key, item in value.items()}
        if isinstance(value, list):
            return [_redact_lock(item) for item in value]
        return value

    redacted_lock = _redact_lock(frozen_lock)
    if (
        not isinstance(lock_sha, str)
        or not hmac.compare_digest(lock_sha, hashlib.sha256(lock_bytes).hexdigest())
        or before != redacted_lock
        or after != redacted_lock
        or build_frozen_toolchain_lock(root) != frozen_lock
    ):
        return False
    toolchain_obj = receipt.get("toolchain")
    if not isinstance(toolchain_obj, dict) or not _valid_toolchain_semantics(
        spec.output_parse.parser,
        toolchain_obj,
        argv,
    ):
        return False
    artifact_sha256 = receipt.get("artifact_sha256")
    if spec.output_parse.parser in {"slo_json", "soak_json", "e2e_json"}:
        if not _valid_gate_artifact(
            spec,
            root=root,
            evidence_dir=evidence_dir,
            argv=argv,
            source_digest=expected_source_digest,
            artifact_sha256=artifact_sha256 if isinstance(artifact_sha256, str) else None,
        ):
            return False
    elif gate_name == "echo_full_audit":
        # Marker gates normally forbid artifact_sha256; audit is the exception and
        # must bind the final audit markdown SHA so receipts cannot be marker-only.
        if not isinstance(artifact_sha256, str) or len(artifact_sha256) != 64:
            return False
        artifact_path = _artifact_path_from_argv(
            argv,
            spec,
            root=root,
            evidence_dir=evidence_dir,
            source_digest=expected_source_digest,
        )
        if artifact_path is None or not artifact_path.is_file():
            return False
        actual_artifact_sha = _sha256_file(artifact_path)
        if actual_artifact_sha is None or not hmac.compare_digest(
            artifact_sha256, actual_artifact_sha
        ):
            return False
    elif artifact_sha256 is not None:
        return False
    try:
        stdout_text = stdout_file.read_text(encoding="utf-8")
    except OSError:
        return False
    expected_parse_result = parse_gate_stdout(
        spec.output_parse.parser,
        stdout_text,
        exit_code=exit_code,
        require_exit_code_zero=spec.output_parse.require_exit_code_zero,
        expected_gate=gate_name,
    )
    if receipt.get("parse_result") != expected_parse_result:
        return False
    if expected_parse_result.get("ok") is not True:
        return False
    if not _valid_desktop_release_bindings(
        expected_parse_result,
        gate_name=gate_name,
        root=root,
        evidence_dir=evidence_dir,
        expected_source_digest=expected_source_digest,
    ):
        return False
    if spec.output_parse.parser == "readiness_json":
        payload = expected_parse_result.get("payload")
        if not isinstance(payload, dict):
            return False
        payload_digest = payload.get("source_digest")
        if not isinstance(payload_digest, str):
            return False
        if not hmac.compare_digest(payload_digest, expected_source_digest):
            return False
        if not hmac.compare_digest(payload_digest, digest_before):
            return False
        if not hmac.compare_digest(payload_digest, digest_after):
            return False
    if gate_name == "soak_3600":
        receipt_duration = _safe_finite_float(receipt.get("duration_seconds"))
        if receipt_duration is None or receipt_duration < _SOAK_GATE_RECEIPT_MIN_SECONDS:
            return False
    if receipt.get("passed") is not True:
        return False
    return not (spec.output_parse.require_exit_code_zero and exit_code != 0)


def _iter_final_gate_receipt_paths(final_dir: Path) -> tuple[Path, ...]:
    if not final_dir.is_dir():
        return ()
    return tuple(
        sorted(
            path
            for path in final_dir.iterdir()
            if path.is_file() and path.name.endswith(".receipt.json")
        )
    )


def validate_final_local_gate_evidence(
    root: Path,
    *,
    final_dir: Path,
    evidence_dir: Path | None = None,
    expected_source_digest: str | None = None,
) -> FinalLocalGateEvidenceReport:
    """Read-only validation of final local gate receipts.

    Does not create directories, copy files, or write summary/receipt artifacts.
    Callers that need derived summaries must use ``generate_final_local_gate_summary``.
    """
    expected_digest = expected_source_digest or release_source_digest(root)
    resolved_root = root.resolve()
    resolved_evidence = (evidence_dir or final_dir.parent).resolve()
    passed: list[str] = []
    blockers: list[str] = []

    receipt_paths = _iter_final_gate_receipt_paths(final_dir)
    if not receipt_paths:
        blockers.append("final_gate_receipts_missing")

    receipts_by_gate: dict[str, dict[str, object]] = {}
    for path in receipt_paths:
        try:
            payload = strict_load_object(path)
        except (OSError, json.JSONDecodeError, StrictJSONError, ValueError):
            blockers.append(f"{path.name}:invalid_json")
            continue
        gate_name = payload.get("gate_name")
        if not isinstance(gate_name, str) or not gate_name.strip():
            blockers.append(f"{path.name}:missing_gate_name")
            continue
        expected_filename = f"{gate_name}.receipt.json"
        if path.name != expected_filename:
            blockers.append(f"{gate_name}:filename_mismatch")
            continue
        if gate_name in receipts_by_gate:
            blockers.append(f"{gate_name}:duplicate_receipt")
            continue
        receipts_by_gate[gate_name] = payload
        spec = get_local_gate_spec(gate_name, evidence_dir=resolved_evidence)
        if _valid_local_gate_receipt(
            payload,
            root=resolved_root,
            expected_source_digest=expected_digest,
            evidence_dir=resolved_evidence,
            gate_spec=spec,
        ):
            passed.append(gate_name)
        else:
            if payload.get("source_drift") is True:
                blockers.append(f"{gate_name}:source_drift")
            elif payload.get("passed") is not True:
                blockers.append(f"{gate_name}:not_passed")
            else:
                blockers.append(f"{gate_name}:invalid_receipt")

    for gate_name in REQUIRED_FINAL_LOCAL_GATES:
        if gate_name not in receipts_by_gate:
            blockers.append(f"{gate_name}:receipt_missing")

    # Envelope time-ordering: no receipt may finish after the envelope is generated.
    envelope_path = resolved_evidence / "MANIFEST.envelope.json"
    if envelope_path.is_file():
        try:
            envelope = strict_load_object(envelope_path)
            envelope_time = _parse_utc_timestamp(envelope.get("generated_utc"))
        except (OSError, json.JSONDecodeError, StrictJSONError, ValueError):
            envelope_time = None
        if envelope_time is not None:
            for gate_name, payload in receipts_by_gate.items():
                receipt_end = _parse_utc_timestamp(payload.get("end_utc"))
                if receipt_end is not None and receipt_end > envelope_time:
                    blockers.append(f"{gate_name}:receipt_after_envelope")

    readiness = verify_release_readiness(
        root,
        require_audit_reports=False,
        require_live_acceptance=False,
    )
    if readiness.internal_ready and not any(
        gate_name in passed for gate_name in _INDEPENDENT_GATE_RECEIPTS
    ):
        blockers.append("verify_release_readiness_not_substitute_for_local_gates")
    if "strict_readiness" in passed and not readiness.internal_ready:
        blockers.append("strict_readiness:readiness_not_internal_ready")
    if "isolated_venv_e2e" in passed:
        e2e_path = resolved_evidence / "e2e" / "ECHO_ISOLATED_VENV_E2E.json"
        if not _valid_isolated_venv_e2e(resolved_root, e2e_path):
            blockers.append("isolated_venv_e2e:evidence_invalid")

    passed = sorted(set(passed))
    blockers = list(dict.fromkeys(blockers))
    required_passed = all(gate in passed for gate in REQUIRED_FINAL_LOCAL_GATES)
    all_local_gates_passed = required_passed and not blockers
    # product_internal_ready requires Echo internal_ready AND desktop build AND
    # real Tauri WebView lifecycle. This prevents echo_internal_ready from
    # producing a false-green for the desktop product.
    desktop_ready = "desktop_build" in passed and "tauri_webview_lifecycle" in passed
    product_internal_ready = all_local_gates_passed and readiness.internal_ready and desktop_ready
    return FinalLocalGateEvidenceReport(
        all_local_gates_passed=all_local_gates_passed,
        passed_gates=tuple(passed),
        blockers=tuple(blockers),
        product_internal_ready=product_internal_ready,
    )


def generate_final_local_gate_summary(
    root: Path,
    *,
    final_dir: Path,
    evidence_dir: Path,
    expected_source_digest: str | None = None,
    report: FinalLocalGateEvidenceReport | None = None,
) -> tuple[FinalLocalGateEvidenceReport, dict[str, object], dict[str, object]]:
    """Build and write derived gate summary + final validator receipt under evidence_dir.

    Explicit write path — never used by the read-only validator.
    """
    expected_digest = expected_source_digest or release_source_digest(root)
    resolved_evidence = evidence_dir.resolve()
    resolved_root = root.resolve()
    final_report = report or validate_final_local_gate_evidence(
        resolved_root,
        final_dir=final_dir,
        evidence_dir=resolved_evidence,
        expected_source_digest=expected_digest,
    )
    generated_utc = datetime.now(tz=UTC).strftime("%Y-%m-%dT%H:%M:%SZ")
    input_snapshot = snapshot_final_gate_inputs(resolved_root, resolved_evidence)
    summary_payload: dict[str, object] = {
        "schema_version": "js-agent-gate-run-summary-v2",
        "generated_utc": generated_utc,
        "source_digest": expected_digest,
        "required_gates": list(REQUIRED_FINAL_LOCAL_GATES),
        "passed_gates": list(final_report.passed_gates),
        "blockers": list(final_report.blockers),
        "all_local_gates_passed": final_report.all_local_gates_passed,
        "validator_inputs": input_snapshot,
    }
    validator_payload: dict[str, object] = {
        "schema_version": "js-agent-final-validator-receipt-v1",
        "generated_utc": generated_utc,
        "validator": "validate_final_local_gate_evidence",
        "writer": "generate_final_local_gate_summary",
        "source_digest": expected_digest,
        "gate_run_summary_sha256": _canonical_json_sha256(summary_payload),
        "ok": final_report.all_local_gates_passed,
        "blockers": list(final_report.blockers),
        "validator_inputs_sha256": _canonical_json_sha256(input_snapshot),
    }
    write_final_validator_receipt(
        resolved_evidence,
        summary_payload=summary_payload,
        validator_payload=validator_payload,
    )
    return final_report, summary_payload, validator_payload
