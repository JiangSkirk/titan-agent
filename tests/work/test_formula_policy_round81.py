"""Round 8.1 D: Work formula policy — literals vs typed Formula."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from js_work.routines.office_safety import Formula, reject_formula_like_text
from js_work.routines.precise_edit import PreciseExcelEditEngine


def _three_fixtures(tmp_path: Path) -> list[str]:
    """Anonymous generated fixtures with three different workbook shapes."""
    names: list[str] = []

    simple = tmp_path / "fixture_simple.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "S"
    ws["A1"] = "name"
    ws["B1"] = "qty"
    ws["A2"] = "widget"
    ws["B2"] = 3
    wb.save(simple)
    wb.close()
    names.append(simple.name)

    multi = tmp_path / "fixture_multi_sheet.xlsx"
    wb = Workbook()
    ws1 = wb.active
    assert ws1 is not None
    ws1.title = "Orders"
    ws1["A1"] = "id"
    ws2 = wb.create_sheet("Meta")
    ws2["A1"] = "note"
    wb.save(multi)
    wb.close()
    names.append(multi.name)

    wide = tmp_path / "fixture_wide.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Grid"
    for col in range(1, 6):
        for row in range(1, 4):
            ws.cell(row=row, column=col, value=f"r{row}c{col}")
    wb.save(wide)
    wb.close()
    names.append(wide.name)
    return names


def test_plain_formula_like_strings_are_not_rejected() -> None:
    reject_formula_like_text("=1+1", label="value")
    reject_formula_like_text("+123", label="value")
    reject_formula_like_text("-9", label="value")
    reject_formula_like_text("@mention", label="value")


def test_typed_formula_rejects_dangerous_functions() -> None:
    with pytest.raises(ValueError):
        Formula('=WEBSERVICE("http://example.test")')
    with pytest.raises(ValueError):
        Formula('=HYPERLINK("http://example.test","x")')
    with pytest.raises(ValueError):
        Formula("=[Other.xlsx]Sheet1!A1")
    with pytest.raises(ValueError):
        Formula("=UNKNOWNFUNC(1)")


@pytest.mark.parametrize("index", [0, 1, 2])
def test_set_cell_literal_strings_not_executable_formulas(tmp_path: Path, index: int) -> None:
    names = _three_fixtures(tmp_path)
    source_name = names[index]
    sheet = load_workbook(tmp_path / source_name).sheetnames[0]
    engine = PreciseExcelEditEngine(tmp_path)
    engine.apply(
        source_path=source_name,
        output_path=f"out_{index}.xlsx",
        operations=[
            {
                "op": "set_cell",
                "sheet": sheet,
                "cell": "A1",
                "value": "=1+1",
            }
        ],
    )
    wb = load_workbook(tmp_path / f"out_{index}.xlsx", data_only=False)
    try:
        cell = wb[sheet]["A1"]
        assert cell.value == "=1+1"
        assert cell.data_type == "s"
    finally:
        wb.close()


def test_typed_formula_writes_restricted_formula(tmp_path: Path) -> None:
    _three_fixtures(tmp_path)
    engine = PreciseExcelEditEngine(tmp_path)
    engine.apply(
        source_path="fixture_simple.xlsx",
        output_path="formula_ok.xlsx",
        operations=[
            {"op": "set_cell", "sheet": "S", "cell": "B2", "value": 2},
            {"op": "set_cell", "sheet": "S", "cell": "B3", "value": 5},
            {
                "op": "set_cell",
                "sheet": "S",
                "cell": "B4",
                "value": Formula("=SUM(B2:B3)"),
            },
        ],
    )
    wb = load_workbook(tmp_path / "formula_ok.xlsx", data_only=False)
    try:
        cell = wb["S"]["B4"]
        assert cell.value == "=SUM(B2:B3)"
        assert cell.data_type == "f"
    finally:
        wb.close()
