"""Round 8.15 shared Office inputs stay bound to Work-approved bytes."""

from __future__ import annotations

import json
import os
from collections.abc import Callable
from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook, load_workbook

from js.echo.attachment_gate import owner_slug, session_slug
from js.echo.turn_context import RuntimeContext, reset_runtime_context, set_runtime_context
from js.tools.registry import ToolResult
from js_work import file_scope as file_scope_module
from js_work.agent_factory import create_work_agent
from js_work.config import load_work_settings
from js_work.tools import WorkToolProfile


def _save_xlsx(path: Path, rows: list[list[Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()


async def _execute_with_real_single_use_lease(
    agent: Any,
    context: RuntimeContext,
    *,
    tool_name: str,
    arguments: dict[str, Any],
) -> tuple[ToolResult, ToolResult]:
    lease_error, signed = agent._authorize_echo_tool_lease(
        tool_name=tool_name,
        arguments=arguments,
        session_id=context.session_id,
        run_id=context.run_id,
        owner_key_hash=context.owner_key_hash,
    )
    assert lease_error is None
    assert signed is not None
    assert signed.product_id == "js-work"
    assert signed.session_id == context.session_id
    assert signed.owner_key_hash == context.owner_key_hash

    result = await agent.registry.execute(
        context.run_id,
        tool_name,
        arguments,
        echo_mode="on",
        execution_context=signed,
    )
    replay = await agent.registry.execute(
        context.run_id,
        tool_name,
        arguments,
        echo_mode="on",
        execution_context=signed,
    )
    return result, replay


def _install_post_scope_swap(
    agent: Any,
    tool_name: str,
    swap: Callable[[], None],
) -> None:
    swapped = False

    def _swap_after_work_scope(arguments: dict[str, Any]) -> dict[str, Any]:
        nonlocal swapped
        if not swapped:
            swap()
            swapped = True
        return arguments

    assert agent.registry.register_argument_policy(tool_name, _swap_after_work_scope)


@pytest.mark.asyncio
@pytest.mark.parametrize("tool_name", ["csv_read", "excel_read"])
async def test_shared_office_read_consumes_approved_bytes_after_post_policy_replace(
    tmp_path: Path,
    tool_name: str,
) -> None:
    settings = load_work_settings(home=tmp_path)
    agent = create_work_agent(settings=settings, profile=WorkToolProfile.OFFICE)
    context = agent.echo_runtime.build_context(
        channel="test",
        owner_key_hash="owner-a",
        session_id="session-a",
        run_id=f"run-{tool_name}",
    )
    owner_root = settings.workspace / "owners" / owner_slug("owner-a") / session_slug(
        "session-a"
    )
    suffix = ".csv" if tool_name == "csv_read" else ".xlsx"
    source = owner_root / f"approved{suffix}"
    replacement = owner_root / f"replacement{suffix}"
    owner_root.mkdir(parents=True, exist_ok=True)
    if tool_name == "csv_read":
        source.write_text("name,qty\nauthorized,1\n", encoding="utf-8")
        replacement.write_text("name,qty\nreplacement,99\n", encoding="utf-8")
    else:
        _save_xlsx(source, [["name", "qty"], ["authorized", 1]])
        _save_xlsx(replacement, [["name", "qty"], ["replacement", 99]])

    _install_post_scope_swap(agent, tool_name, lambda: os.replace(replacement, source))
    token = set_runtime_context(context)
    try:
        result, replay = await _execute_with_real_single_use_lease(
            agent,
            context,
            tool_name=tool_name,
            arguments={"path": source.name},
        )
    finally:
        reset_runtime_context(token)
        await agent.close()

    assert result.success is True, result.error
    rows = json.loads(result.output)
    expected = ["authorized", "1"] if tool_name == "csv_read" else ["authorized", 1]
    assert rows[1] == expected
    assert replay.success is False
    assert "lease denied" in replay.error.lower()


@pytest.mark.asyncio
async def test_excel_merge_consumes_approved_upload_source_and_private_template(
    tmp_path: Path,
) -> None:
    settings = load_work_settings(home=tmp_path)
    agent = create_work_agent(settings=settings, profile=WorkToolProfile.OFFICE)
    context = agent.echo_runtime.build_context(
        channel="test",
        owner_key_hash="owner-a",
        session_id="session-a",
        run_id="run-merge",
    )
    owner_root = settings.workspace / "owners" / owner_slug("owner-a") / session_slug(
        "session-a"
    )
    upload_root = settings.workspace / "uploads" / owner_slug("owner-a") / session_slug(
        "session-a"
    )
    source = upload_root / "source.xlsx"
    replacement_source = upload_root / "replacement-source.xlsx"
    template = owner_root / "template.xlsx"
    replacement_template = owner_root / "replacement-template.xlsx"
    output = owner_root / "reports" / "merged.xlsx"
    _save_xlsx(source, [["name", "qty"], ["authorized", 1]])
    _save_xlsx(replacement_source, [["name", "qty"], ["replacement", 99]])
    _save_xlsx(template, [["AUTHORIZED-TEMPLATE"]])
    _save_xlsx(replacement_template, [["REPLACEMENT-TEMPLATE"]])

    def _swap_both_inputs() -> None:
        os.replace(replacement_source, source)
        os.replace(replacement_template, template)

    _install_post_scope_swap(agent, "excel_merge", _swap_both_inputs)
    arguments = {
        "source_path": source.relative_to(settings.workspace).as_posix(),
        "target_path": template.name,
        "output_path": "reports/merged.xlsx",
        "target_start_cell": "A2",
    }
    token = set_runtime_context(context)
    try:
        result, replay = await _execute_with_real_single_use_lease(
            agent,
            context,
            tool_name="excel_merge",
            arguments=arguments,
        )
    finally:
        reset_runtime_context(token)
        await agent.close()

    assert result.success is True, result.error
    workbook = load_workbook(output, data_only=True)
    try:
        worksheet = workbook.active
        assert worksheet is not None
        assert worksheet["A1"].value == "AUTHORIZED-TEMPLATE"
        assert worksheet["A3"].value == "authorized"
        assert worksheet["B3"].value == 1
    finally:
        workbook.close()
    assert replay.success is False
    assert "lease denied" in replay.error.lower()


@pytest.mark.asyncio
async def test_shared_office_inputs_reject_cross_owner_and_absolute_paths(
    tmp_path: Path,
) -> None:
    settings = load_work_settings(home=tmp_path)
    agent = create_work_agent(settings=settings, profile=WorkToolProfile.OFFICE)
    owner_a_root = settings.workspace / "owners" / owner_slug("owner-a") / session_slug(
        "session-a"
    )
    owner_b_file = (
        settings.workspace
        / "owners"
        / owner_slug("owner-b")
        / session_slug("session-a")
        / "secret.xlsx"
    )
    _save_xlsx(owner_a_root / "approved.xlsx", [["authorized", 1]])
    _save_xlsx(owner_b_file, [["secret", 99]])

    for index, (path, expected_error) in enumerate(
        (
            (
                owner_b_file.relative_to(settings.workspace).as_posix(),
                "fs_roots denied",
            ),
            (str(owner_a_root / "approved.xlsx"), "absolute work paths"),
        )
    ):
        context = agent.echo_runtime.build_context(
            channel="test",
            owner_key_hash="owner-a",
            session_id="session-a",
            run_id=f"run-denied-{index}",
        )
        token = set_runtime_context(context)
        try:
            result, replay = await _execute_with_real_single_use_lease(
                agent,
                context,
                tool_name="excel_read",
                arguments={"path": path},
            )
        finally:
            reset_runtime_context(token)

        assert result.success is False
        assert result.output == ""
        assert expected_error in result.error.lower()
        assert replay.success is False
        assert replay.output == ""

    await agent.close()


@pytest.mark.asyncio
async def test_csv_snapshot_capture_obeys_the_existing_csv_byte_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    settings = load_work_settings(home=tmp_path)
    settings.tools = settings.tools.model_copy(update={"csv_read_max_bytes": 32})
    agent = create_work_agent(settings=settings, profile=WorkToolProfile.OFFICE)
    context = agent.echo_runtime.build_context(
        channel="test",
        owner_key_hash="owner-a",
        session_id="session-a",
        run_id="run-csv-budget",
    )
    owner_root = settings.workspace / "owners" / owner_slug("owner-a") / session_slug(
        "session-a"
    )
    source = owner_root / "oversize.csv"
    owner_root.mkdir(parents=True, exist_ok=True)
    source.write_bytes(b"x" * 128)
    source_identity = (source.stat().st_dev, source.stat().st_ino)
    bytes_read = 0
    real_read = file_scope_module.os.read

    def _track_source_reads(descriptor: int, size: int) -> bytes:
        nonlocal bytes_read
        chunk = real_read(descriptor, size)
        metadata = os.fstat(descriptor)
        if (metadata.st_dev, metadata.st_ino) == source_identity:
            bytes_read += len(chunk)
        return chunk

    monkeypatch.setattr(file_scope_module.os, "read", _track_source_reads)
    token = set_runtime_context(context)
    try:
        result, _replay = await _execute_with_real_single_use_lease(
            agent,
            context,
            tool_name="csv_read",
            arguments={"path": source.name},
        )
    finally:
        reset_runtime_context(token)
        await agent.close()

    assert result.success is False
    assert "size limit" in result.error.lower()
    assert bytes_read <= settings.tools.csv_read_max_bytes + 1


@pytest.mark.asyncio
async def test_later_argument_policy_cannot_rebind_approved_snapshot(
    tmp_path: Path,
) -> None:
    from js_work.file_scope import WorkOwnerFileScope

    settings = load_work_settings(home=tmp_path)
    agent = create_work_agent(settings=settings, profile=WorkToolProfile.OFFICE)
    context = agent.echo_runtime.build_context(
        channel="test",
        owner_key_hash="owner-a",
        session_id="session-a",
        run_id="run-snapshot-rebind",
    )
    owner_root = settings.workspace / "owners" / owner_slug("owner-a") / session_slug(
        "session-a"
    )
    owner_root.mkdir(parents=True, exist_ok=True)
    (owner_root / "approved.csv").write_text(
        "name,qty\nauthorized,1\n",
        encoding="utf-8",
    )
    (owner_root / "replacement.csv").write_text(
        "name,qty\nreplacement,99\n",
        encoding="utf-8",
    )
    mutation_blocked = False

    def _attempt_rebind(arguments: dict[str, Any]) -> dict[str, Any]:
        nonlocal mutation_blocked
        scope = WorkOwnerFileScope(
            settings.workspace,
            owner="owner-a",
            session_id="session-a",
        )
        replacement = scope.read_routine_input("replacement.csv")
        try:
            arguments["path"].snapshot = replacement
        except AttributeError:
            mutation_blocked = True
        return arguments

    assert agent.registry.register_argument_policy("csv_read", _attempt_rebind)
    token = set_runtime_context(context)
    try:
        result, replay = await _execute_with_real_single_use_lease(
            agent,
            context,
            tool_name="csv_read",
            arguments={"path": "approved.csv"},
        )
    finally:
        reset_runtime_context(token)
        await agent.close()

    assert mutation_blocked is True
    assert result.success is True, result.error
    assert json.loads(result.output)[1] == ["authorized", "1"]
    assert replay.success is False
    assert "lease denied" in replay.error.lower()


@pytest.mark.asyncio
async def test_registry_rejects_different_snapshot_handle_for_the_same_path(
    tmp_path: Path,
) -> None:
    from js_work.file_scope import WorkOwnerFileScope

    settings = load_work_settings(home=tmp_path)
    agent = create_work_agent(settings=settings, profile=WorkToolProfile.OFFICE)
    context = agent.echo_runtime.build_context(
        channel="test",
        owner_key_hash="owner-a",
        session_id="session-a",
        run_id="run-snapshot-replacement",
    )
    owner_root = settings.workspace / "owners" / owner_slug("owner-a") / session_slug(
        "session-a"
    )
    owner_root.mkdir(parents=True, exist_ok=True)
    (owner_root / "approved.csv").write_text(
        "name,qty\nauthorized,1\n",
        encoding="utf-8",
    )
    (owner_root / "replacement.csv").write_text(
        "name,qty\nreplacement,99\n",
        encoding="utf-8",
    )

    def _replace_handle(arguments: dict[str, Any]) -> dict[str, Any]:
        scope = WorkOwnerFileScope(
            settings.workspace,
            owner="owner-a",
            session_id="session-a",
        )
        current = arguments["path"]
        replacement = scope.read_routine_input("replacement.csv")
        arguments["path"] = type(current)(str(current), replacement)
        return arguments

    assert agent.registry.register_argument_policy("csv_read", _replace_handle)
    token = set_runtime_context(context)
    try:
        result, replay = await _execute_with_real_single_use_lease(
            agent,
            context,
            tool_name="csv_read",
            arguments={"path": "approved.csv"},
        )
    finally:
        reset_runtime_context(token)
        await agent.close()

    assert result.success is False
    assert result.output == ""
    assert "snapshot binding" in result.error.lower()
    assert replay.success is False
    assert "lease denied" in replay.error.lower()
