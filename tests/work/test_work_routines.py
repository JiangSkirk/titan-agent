from __future__ import annotations

import asyncio
import json
import shutil
import threading
import zipfile
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from typing import Any

import pytest
from click.testing import CliRunner
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from js_work.agent_factory import create_work_agent
from js_work.cli import main as work_main
from js_work.config import load_work_settings
from js_work.tools import WorkToolProfile


def _save_source_workbook(path: Path, *, conflicting_unit: bool = False) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "表格1"
    ws.append(["序号", "品名", "规格", "数量", "单位", "备注"])
    ws.append([1, "全棉面料", "40S", 10, "米", "白色"])
    ws.append([2, "涤纶面料", "75D", 8.5, "米", "蓝色"])
    ws.append([3, "全棉面料", "40S", 3, "码" if conflicting_unit else "米", "补充"])
    wb.save(path)


def _save_mixed_factory_workbook(path: Path, *, only_auxiliary: bool = False) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "订单明细"
    ws.append(["序号", "品名", "规格", "数量", "单位", "类别", "备注"])
    if only_auxiliary:
        ws.append([1, "纽扣", "18L", 200, "粒", "辅料", "不应进入面料统计"])
        ws.append([2, "拉链", "20CM", 50, "条", "辅料", "不应进入面料统计"])
    else:
        ws.append([1, "全棉面料", "40S", 10, "米", "面料", "白色"])
        ws.append([2, "纽扣", "18L", 200, "粒", "辅料", "不应进入面料统计"])
        ws.append([3, "涤纶面料", "75D", 8.5, "米", "面料", "蓝色"])
        ws.append([4, "全棉面料", "40S", 3, "米", "面料", "补充"])
    wb.save(path)


def _write_mixed_factory_csv(path: Path) -> None:
    path.write_text(
        "\n".join(
            [
                "序号,品名,规格,数量,单位,类别,备注",
                "1,全棉面料,40S,10,米,面料,白色",
                "2,纽扣,18L,200,粒,辅料,不应进入面料统计",
                "3,全棉面料,40S,3,米,面料,补充",
            ]
        )
        + "\n",
        encoding="utf-8",
    )


def _save_template_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "面料统计"
    ws.merge_cells("A1:E1")
    ws["A1"] = "面料统计模板"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].fill = PatternFill("solid", fgColor="D9EAD3")
    ws.column_dimensions["B"].width = 22
    ws.row_dimensions[2].height = 24
    headers = ["面料名称", "规格", "数量", "单位", "备注"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FFF2CC")
    for col in range(1, 6):
        cell = ws.cell(row=3, column=col, value="")
        cell.fill = PatternFill("solid", fgColor="FCE5CD")
    ws["F2"] = "数量合计"
    ws["F3"] = "=SUM(C3:C100)"
    wb.save(path)


def _save_template_with_unsupported_formula(path: Path) -> None:
    _save_template_workbook(path)
    wb = load_workbook(path, data_only=False)
    ws = wb["面料统计"]
    ws["G2"] = "外部查找"
    ws["G3"] = "=VLOOKUP(A3,其它表!A:B,2,FALSE)"
    wb.save(path)


def _add_external_workbook_relationship(path: Path) -> None:
    patched = path.with_name(f"{path.stem}.patched.xlsx")
    relationship_path = "xl/_rels/workbook.xml.rels"
    external = (
        b'<Relationship Id="rIdSynthetic" '
        b'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/externalLink" '
        b'Target="https://example.invalid/private.xlsx" TargetMode="External"/>'
    )
    with zipfile.ZipFile(path) as source, zipfile.ZipFile(patched, "w") as output:
        for info in source.infolist():
            payload = source.read(info.filename)
            if info.filename == relationship_path:
                payload = payload.replace(b"</Relationships>", external + b"</Relationships>")
            output.writestr(info, payload)
    patched.replace(path)


def _save_large_factory_workbook(path: Path, *, rows: int = 101) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "订单明细"
    ws.append(["序号", "品名", "规格", "数量", "单位", "类别", "备注"])
    for idx in range(1, rows + 1):
        ws.append([idx, f"面料{idx}", f"{idx}S", 1, "米", "面料", ""])
    wb.save(path)


def test_routine_store_creates_disabled_draft_and_approves(tmp_path: Path) -> None:
    from js_work.routines import RoutineStatus, WorkRoutineStore

    store = WorkRoutineStore(tmp_path / "state")
    routine = store.create_draft(
        name="面料统计",
        trigger_phrases=["面料统计", "按模板生成面料表"],
        routine_type="spreadsheet_template",
        field_mapping={"面料名称": "品名", "规格": "规格", "数量": "数量", "单位": "单位"},
        validation_rules={"required_fields": ["面料名称", "规格", "数量", "单位"]},
    )

    assert routine.status == RoutineStatus.DRAFT
    assert routine.enabled is False
    assert store.list_routines()[0].routine_id == routine.routine_id

    approved = store.approve(routine.routine_id)

    assert approved.status == RoutineStatus.ENABLED
    assert approved.enabled is True
    assert store.get(routine.routine_id).enabled is True
    assert not (tmp_path / "state" / "skills.db").exists()
    assert not (tmp_path / "state" / "skill_promotions.db").exists()


def test_template_analysis_and_render_preserves_structure_and_style(tmp_path: Path) -> None:
    from js_work.routines.spreadsheet import SpreadsheetTemplateEngine

    template = tmp_path / "template.xlsx"
    output = tmp_path / "output.xlsx"
    _save_template_workbook(template)
    engine = SpreadsheetTemplateEngine(tmp_path)

    analysis = engine.analyze_template("template.xlsx")

    assert analysis.sheet_name == "面料统计"
    assert analysis.header_row == 2
    assert analysis.headers[:5] == ["面料名称", "规格", "数量", "单位", "备注"]
    assert "A1:E1" in analysis.merged_cells
    assert analysis.column_widths["B"] == 22
    assert analysis.formulas["F3"] == "=SUM(C3:C100)"

    engine.render_from_template(
        template_path="template.xlsx",
        output_path="output.xlsx",
        rows=[
            {"面料名称": "全棉面料", "规格": "40S", "数量": 13, "单位": "米", "备注": "白色"},
            {"面料名称": "涤纶面料", "规格": "75D", "数量": 8.5, "单位": "米", "备注": "蓝色"},
        ],
    )

    wb = load_workbook(output, data_only=False)
    ws = wb["面料统计"]
    assert "A1:E1" in {str(rng) for rng in ws.merged_cells.ranges}
    assert ws.column_dimensions["B"].width == 22
    assert ws["A2"].fill.fgColor.rgb == "00FFF2CC"
    assert ws["A3"].fill.fgColor.rgb == "00FCE5CD"
    assert ws["A4"].fill.fgColor.rgb == "00FCE5CD"
    assert ws["A3"].value == "全棉面料"
    assert ws["C4"].value == 8.5
    assert ws["F3"].value == "=SUM(C3:C100)"


def test_spreadsheet_render_and_validation_never_overwrite_existing_artifacts(
    tmp_path: Path,
) -> None:
    from js_work.routines.spreadsheet import SpreadsheetTemplateEngine

    template = tmp_path / "template.xlsx"
    output = tmp_path / "output.xlsx"
    _save_template_workbook(template)
    output.write_bytes(b"keep-output")
    engine = SpreadsheetTemplateEngine(tmp_path)

    with pytest.raises(ValueError, match="already exists"):
        engine.render_from_template(
            template_path="template.xlsx",
            output_path="output.xlsx",
            rows=[],
        )
    assert output.read_bytes() == b"keep-output"

    output.unlink()
    engine.render_from_template(
        template_path="template.xlsx",
        output_path="output.xlsx",
        rows=[],
    )
    report_path = output.with_suffix(".validation.json")
    report_path.write_text("keep-report", encoding="utf-8")
    with pytest.raises(ValueError, match="already exists"):
        engine.validate_output(
            source_path="template.xlsx",
            template_path="template.xlsx",
            output_path="output.xlsx",
            rows=[],
        )
    assert report_path.read_text(encoding="utf-8") == "keep-report"


def test_spreadsheet_engine_rejects_external_ooxml_and_generated_formula_injection(
    tmp_path: Path,
) -> None:
    from js_work.routines.spreadsheet import SpreadsheetTemplateEngine

    unsafe_template = tmp_path / "unsafe-template.xlsx"
    _save_template_workbook(unsafe_template)
    _add_external_workbook_relationship(unsafe_template)
    engine = SpreadsheetTemplateEngine(tmp_path)

    with pytest.raises(ValueError, match="external OOXML relationship"):
        engine.analyze_template("unsafe-template.xlsx")

    safe_template = tmp_path / "safe-template.xlsx"
    _save_template_workbook(safe_template)
    # Formula-like strings are written as literals, never auto-executed formulas.
    engine.render_from_template(
        template_path="safe-template.xlsx",
        output_path="formula-injection.xlsx",
        rows=[{"面料名称": "=2+2", "数量": 1}],
    )
    from openpyxl import load_workbook

    wb = load_workbook(tmp_path / "formula-injection.xlsx", data_only=False)
    try:
        values = [cell.value for row in wb.active.iter_rows() for cell in row]
        assert "=2+2" in values
        for row in wb.active.iter_rows():
            for cell in row:
                if cell.value == "=2+2":
                    assert cell.data_type == "s"
    finally:
        wb.close()


def test_shared_xlsx_validator_detects_input_replacement_during_validation(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    from js_work.routines.office_safety import validate_safe_xlsx
    from js_work.routines.precise_edit import PreciseExcelEditEngine

    source = tmp_path / "source.xlsx"
    replacement = tmp_path / "replacement.xlsx"
    _save_template_workbook(source)
    _save_source_workbook(replacement)
    original_validate = PreciseExcelEditEngine._validate_archive

    def replace_after_archive_check(path: Path) -> None:
        original_validate(path)
        replacement.replace(path)

    monkeypatch.setattr(
        PreciseExcelEditEngine,
        "_validate_archive",
        staticmethod(replace_after_archive_check),
    )

    with pytest.raises(ValueError, match="changed while"):
        validate_safe_xlsx(source)


def test_validate_safe_xlsx_accepts_fd_bound_snapshot(tmp_path: Path) -> None:
    import os

    from js_work.file_scope import MaterializedSnapshotPath
    from js_work.routines.office_safety import validate_safe_xlsx

    source = tmp_path / "source.xlsx"
    _save_template_workbook(source)
    snapshot = MaterializedSnapshotPath(source)
    snapshot._snapshot_fd = os.open(source, os.O_RDONLY | getattr(os, "O_CLOEXEC", 0))
    try:
        validate_safe_xlsx(snapshot)
    finally:
        os.close(snapshot._snapshot_fd)
        snapshot._snapshot_fd = -1


def test_spreadsheet_routine_runner_generates_validation_report(tmp_path: Path) -> None:
    from js_work.routines import WorkRoutineStore
    from js_work.routines.spreadsheet import WorkSpreadsheetRoutineRunner

    source = tmp_path / "source.xlsx"
    template = tmp_path / "template.xlsx"
    output = tmp_path / "reports" / "fabric_stats.xlsx"
    _save_source_workbook(source)
    _save_template_workbook(template)

    store = WorkRoutineStore(tmp_path / "state")
    routine = store.create_draft(
        name="面料统计",
        trigger_phrases=["面料统计"],
        routine_type="spreadsheet_template",
        field_mapping={
            "面料名称": "品名",
            "规格": "规格",
            "数量": "数量",
            "单位": "单位",
            "备注": "备注",
        },
        validation_rules={"required_fields": ["面料名称", "规格", "数量", "单位"]},
    )
    store.approve(routine.routine_id)
    reviewer_seen: dict[str, Any] = {}

    class Reviewer:
        def review(self, summary: dict[str, Any]) -> dict[str, Any]:
            reviewer_seen.update(summary)
            return {"status": "passed", "issues": []}

    result = WorkSpreadsheetRoutineRunner(tmp_path, tmp_path / "state").run(
        routine=store.get(routine.routine_id),
        source_path="source.xlsx",
        template_path="template.xlsx",
        output_path="reports/fabric_stats.xlsx",
        reviewer=Reviewer(),
    )

    assert result.status == "passed"
    assert output.exists()
    report_path = output.with_suffix(".validation.json")
    report = json.loads(report_path.read_text(encoding="utf-8"))
    assert report["status"] == "passed"
    assert report["row_count"] == 2
    assert report["totals"]["数量"] == 21.5
    assert report["source_hash"].startswith("sha256:")
    assert report["template_hash"].startswith("sha256:")
    assert report["output_hash"].startswith("sha256:")
    assert reviewer_seen["row_count"] == 2
    assert "source_rows" not in reviewer_seen

    from js.echo.ledger.journal import FileEchoLedger, verify_file
    from js.echo.ledger.service import EchoSafetyService

    service = EchoSafetyService(state_dir=tmp_path / "state")
    journal_path = service.journal_path_for("js-work-local")
    journal_key = service.journal_key_for("js-work-local")
    verify_report = verify_file(journal_path, mac_key=journal_key)
    records = FileEchoLedger(journal_path, mac_key=journal_key).records
    assert verify_report.ok
    assert any(record.record_type == "merge" for record in records)


def test_spreadsheet_routine_recalculates_formula_cached_values(tmp_path: Path) -> None:
    from js_work.routines import WorkRoutineStore
    from js_work.routines.spreadsheet import WorkSpreadsheetRoutineRunner

    source = tmp_path / "source.xlsx"
    template = tmp_path / "template.xlsx"
    output = tmp_path / "reports" / "fabric_stats.xlsx"
    _save_source_workbook(source)
    _save_template_workbook(template)

    store = WorkRoutineStore(tmp_path / "state")
    routine = store.create_draft(
        name="面料统计",
        trigger_phrases=["面料统计"],
        routine_type="spreadsheet_template",
        field_mapping={
            "面料名称": "品名",
            "规格": "规格",
            "数量": "数量",
            "单位": "单位",
            "备注": "备注",
        },
        validation_rules={"required_fields": ["面料名称", "规格", "数量", "单位"]},
    )
    store.approve(routine.routine_id)

    result = WorkSpreadsheetRoutineRunner(tmp_path, tmp_path / "state").run(
        routine=store.get(routine.routine_id),
        source_path="source.xlsx",
        template_path="template.xlsx",
        output_path="reports/fabric_stats.xlsx",
    )

    wb = load_workbook(output, data_only=True)
    ws = wb["面料统计"]
    assert result.status == "passed"
    assert ws["F3"].value == 21.5


def test_formula_cache_evaluates_supported_formulas_without_libreoffice(tmp_path: Path) -> None:
    from js_work.routines.formula_cache import refresh_formula_caches

    output = tmp_path / "formula-cache.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Calculations"
    ws["A1"] = 10
    ws["A2"] = 5
    ws["B1"] = "=SUM(A1:A2)"
    ws["B2"] = "=SUM(A1:A2)+SUM(A2:A2)"
    ws["B3"] = "=(A1+A2)*2-4/2"
    ws["B4"] = "=VLOOKUP(A1,Other!A:B,2,FALSE)"
    wb.save(output)

    result = refresh_formula_caches(output, soffice=None)

    formula_wb = load_workbook(output, data_only=False)
    values_wb = load_workbook(output, data_only=True)
    try:
        formula_ws = formula_wb["Calculations"]
        values_ws = values_wb["Calculations"]
        assert formula_ws["B1"].value == "=SUM(A1:A2)"
        assert formula_ws["B2"].value == "=SUM(A1:A2)+SUM(A2:A2)"
        assert formula_ws["B3"].value == "=(A1+A2)*2-4/2"
        assert formula_ws["B4"].value == "=VLOOKUP(A1,Other!A:B,2,FALSE)"
        assert values_ws["B1"].value == 15
        assert values_ws["B2"].value == 20
        assert values_ws["B3"].value == 28
        assert values_ws["B4"].value is None
    finally:
        formula_wb.close()
        values_wb.close()
    assert result.cached_count == 3
    assert result.unsupported_formulas == {"Calculations": {"B4": "=VLOOKUP(A1,Other!A:B,2,FALSE)"}}
    assert result.libreoffice_used is False


def test_formula_cache_rejects_oversized_sum_before_visiting_cells() -> None:
    from js_work.routines.formula_cache import evaluate_formula

    class WorksheetThatMustNotBeRead:
        def __getitem__(self, _coordinate: str) -> Any:
            raise AssertionError("oversized formula attempted to read a cell")

        def cell(self, *, row: int, column: int) -> Any:
            raise AssertionError(f"oversized formula attempted to visit cell {row},{column}")

    assert (
        evaluate_formula(
            WorksheetThatMustNotBeRead(),
            "=SUM(A1:XFD1048576)",
        )
        is None
    )


def test_formula_cache_libreoffice_fallback_is_isolated_offline_and_timed(
    tmp_path: Path,
    monkeypatch,
) -> None:
    from js_work.routines import formula_cache

    output = tmp_path / "formula-fallback.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = 1
    ws["B1"] = "=VLOOKUP(A1,Other!A:B,2,FALSE)"
    wb.save(output)
    observed: dict[str, Any] = {}

    def fake_run(command: list[str], **kwargs: Any) -> Any:
        observed["command"] = command
        observed.update(kwargs)
        out_dir = Path(command[command.index("--outdir") + 1])
        source = Path(command[-1])
        shutil.copy2(source, out_dir / source.name)
        return type("Completed", (), {"returncode": 0, "stdout": "", "stderr": ""})()

    monkeypatch.setattr(formula_cache.subprocess, "run", fake_run)
    monkeypatch.setattr(
        formula_cache,
        "_libreoffice_network_sandbox_prefix",
        lambda: ("/trusted/sandbox", "--deny-network"),
    )

    result = formula_cache.refresh_formula_caches(
        output,
        soffice=Path("/fake/soffice"),
        timeout=7,
    )

    assert result.libreoffice_used is True
    assert observed["timeout"] == 7
    assert observed["check"] is False
    assert observed["command"][:2] == ["/trusted/sandbox", "--deny-network"]
    assert "--headless" in observed["command"]
    assert any(part.startswith("-env:UserInstallation=file:") for part in observed["command"])
    assert observed["env"]["NO_PROXY"] == "*"
    assert observed["env"]["no_proxy"] == "*"
    for key in ("HTTP_PROXY", "HTTPS_PROXY", "ALL_PROXY", "http_proxy", "https_proxy", "all_proxy"):
        assert key not in observed["env"]


def test_libreoffice_conversion_fails_closed_without_verified_network_sandbox(
    tmp_path: Path,
    monkeypatch,
) -> None:
    from js_work.routines import formula_cache

    source = tmp_path / "legacy.xls"
    source.write_bytes(b"legacy workbook placeholder")
    called = False

    def forbidden_run(*_args: Any, **_kwargs: Any) -> Any:
        nonlocal called
        called = True
        raise AssertionError("LibreOffice started without a network sandbox")

    monkeypatch.setattr(formula_cache.subprocess, "run", forbidden_run)
    monkeypatch.setattr(
        formula_cache,
        "_libreoffice_network_sandbox_prefix",
        lambda: None,
        raising=False,
    )

    result = formula_cache.run_libreoffice_conversion(
        source,
        tmp_path / "converted",
        soffice=Path("/fake/soffice"),
    )

    assert result.output_path is None
    assert "network sandbox" in result.detail.lower()
    assert called is False


def test_libreoffice_conversion_rejects_stale_same_name_output(
    tmp_path: Path,
    monkeypatch,
) -> None:
    from js_work.routines import formula_cache

    source = tmp_path / "legacy.xls"
    source.write_bytes(b"legacy workbook placeholder")
    output_dir = tmp_path / "converted"
    output_dir.mkdir()
    stale_output = output_dir / "legacy.xlsx"
    stale_output.write_bytes(b"stale")

    def fake_run(_command: list[str], **_kwargs: Any) -> Any:
        return type("Completed", (), {"returncode": 0, "stdout": "", "stderr": ""})()

    monkeypatch.setattr(formula_cache.subprocess, "run", fake_run)
    monkeypatch.setattr(
        formula_cache,
        "_libreoffice_network_sandbox_prefix",
        lambda: ("/trusted/sandbox", "--deny-network"),
        raising=False,
    )

    result = formula_cache.run_libreoffice_conversion(
        source,
        output_dir,
        soffice=Path("/fake/soffice"),
    )

    assert result.output_path is None
    assert not stale_output.exists()


def test_libreoffice_conversion_sanitizes_process_failures(
    tmp_path: Path,
    monkeypatch,
) -> None:
    from js_work.routines import formula_cache

    private_detail = "/Users/private/Documents/customer.xls secret-token"
    source = tmp_path / "legacy.xls"
    source.write_bytes(b"legacy workbook placeholder")
    monkeypatch.setattr(
        formula_cache,
        "_libreoffice_network_sandbox_prefix",
        lambda: ("/trusted/sandbox", "--deny-network"),
        raising=False,
    )

    monkeypatch.setattr(
        formula_cache.subprocess,
        "run",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(OSError(private_detail)),
    )
    os_error = formula_cache.run_libreoffice_conversion(
        source,
        tmp_path / "converted-os-error",
        soffice=Path("/fake/soffice"),
    )

    monkeypatch.setattr(
        formula_cache.subprocess,
        "run",
        lambda *_args, **_kwargs: type(
            "Completed",
            (),
            {"returncode": 1, "stdout": "", "stderr": private_detail},
        )(),
    )
    process_error = formula_cache.run_libreoffice_conversion(
        source,
        tmp_path / "converted-process-error",
        soffice=Path("/fake/soffice"),
    )

    assert os_error.detail == "LibreOffice conversion failed safely"
    assert process_error.detail == "LibreOffice conversion failed safely"
    assert private_detail not in str(os_error)
    assert private_detail not in str(process_error)


def test_spreadsheet_routine_blocks_conflicting_units(tmp_path: Path) -> None:
    from js_work.routines import WorkRoutineStore
    from js_work.routines.spreadsheet import WorkSpreadsheetRoutineRunner

    source = tmp_path / "source.xlsx"
    template = tmp_path / "template.xlsx"
    _save_source_workbook(source, conflicting_unit=True)
    _save_template_workbook(template)

    store = WorkRoutineStore(tmp_path / "state", session_id="session-a")
    routine = store.create_draft(
        name="面料统计",
        trigger_phrases=["面料统计"],
        routine_type="spreadsheet_template",
        field_mapping={"面料名称": "品名", "规格": "规格", "数量": "数量", "单位": "单位"},
        validation_rules={"required_fields": ["面料名称", "规格", "数量", "单位"]},
    )
    store.approve(routine.routine_id)

    result = WorkSpreadsheetRoutineRunner(tmp_path, tmp_path / "state").run(
        routine=store.get(routine.routine_id),
        source_path="source.xlsx",
        template_path="template.xlsx",
        output_path="reports/conflict.xlsx",
    )

    assert result.status == "needs_review"
    assert any(issue["code"] == "unit_conflict" for issue in result.issues)


def test_spreadsheet_routine_filters_auxiliary_rows_and_aggregates_duplicates(
    tmp_path: Path,
) -> None:
    from js_work.routines import WorkRoutineStore
    from js_work.routines.spreadsheet import WorkSpreadsheetRoutineRunner

    source = tmp_path / "source.xlsx"
    template = tmp_path / "template.xlsx"
    output = tmp_path / "reports" / "fabric_stats.xlsx"
    _save_mixed_factory_workbook(source)
    _save_template_workbook(template)

    store = WorkRoutineStore(tmp_path / "state")
    routine = store.create_draft(
        name="面料统计",
        trigger_phrases=["面料统计"],
        routine_type="spreadsheet_template",
        field_mapping={
            "面料名称": "品名",
            "规格": "规格",
            "数量": "数量",
            "单位": "单位",
            "备注": "备注",
        },
        row_filters=[{"field": "类别", "equals": "面料"}],
        aggregation_rules={
            "group_by": ["面料名称", "规格", "单位"],
            "sum_fields": ["数量"],
            "merge_text_fields": ["备注"],
        },
        validation_rules={"required_fields": ["面料名称", "规格", "数量", "单位"]},
    )
    store.approve(routine.routine_id)

    result = WorkSpreadsheetRoutineRunner(tmp_path, tmp_path / "state").run(
        routine=store.get(routine.routine_id),
        source_path="source.xlsx",
        template_path="template.xlsx",
        output_path="reports/fabric_stats.xlsx",
    )

    assert result.status == "passed"
    report = json.loads(output.with_suffix(".validation.json").read_text(encoding="utf-8"))
    assert report["source_row_count"] == 4
    assert report["selected_source_row_count"] == 3
    assert report["excluded_source_row_count"] == 1
    assert report["row_count"] == 2
    assert report["totals"]["数量"] == 21.5
    assert report["formula_checks"]["F3"]["status"] == "passed"
    assert report["excluded_rows"][0]["reason"] == "row_filter_not_matched"

    wb = load_workbook(output, data_only=False)
    ws = wb["面料统计"]
    assert ws["A3"].value == "全棉面料"
    assert ws["C3"].value == 13
    assert ws["E3"].value == "白色；补充"
    assert ws["A4"].value == "涤纶面料"
    assert ws["C4"].value == 8.5
    assert ws["F3"].value == "=SUM(C3:C100)"


def test_spreadsheet_routine_needs_review_when_filters_select_no_rows(
    tmp_path: Path,
) -> None:
    from js_work.routines import WorkRoutineStore
    from js_work.routines.spreadsheet import WorkSpreadsheetRoutineRunner

    source = tmp_path / "source.xlsx"
    template = tmp_path / "template.xlsx"
    _save_mixed_factory_workbook(source, only_auxiliary=True)
    _save_template_workbook(template)

    store = WorkRoutineStore(tmp_path / "state")
    routine = store.create_draft(
        name="面料统计",
        trigger_phrases=["面料统计"],
        routine_type="spreadsheet_template",
        field_mapping={"面料名称": "品名", "规格": "规格", "数量": "数量", "单位": "单位"},
        row_filters=[{"field": "类别", "equals": "面料"}],
        validation_rules={"required_fields": ["面料名称", "规格", "数量", "单位"]},
    )
    store.approve(routine.routine_id)

    result = WorkSpreadsheetRoutineRunner(tmp_path, tmp_path / "state").run(
        routine=store.get(routine.routine_id),
        source_path="source.xlsx",
        template_path="template.xlsx",
        output_path="reports/empty.xlsx",
    )

    assert result.status == "needs_review"
    assert any(issue["code"] == "no_rows_selected" for issue in result.issues)


def test_spreadsheet_routine_accepts_csv_sources(tmp_path: Path) -> None:
    from js_work.routines import WorkRoutineStore
    from js_work.routines.spreadsheet import WorkSpreadsheetRoutineRunner

    source = tmp_path / "source.csv"
    template = tmp_path / "template.xlsx"
    output = tmp_path / "reports" / "from_csv.xlsx"
    _write_mixed_factory_csv(source)
    _save_template_workbook(template)

    store = WorkRoutineStore(tmp_path / "state")
    routine = store.create_draft(
        name="面料统计",
        trigger_phrases=["面料统计"],
        routine_type="spreadsheet_template",
        field_mapping={
            "面料名称": "品名",
            "规格": "规格",
            "数量": "数量",
            "单位": "单位",
            "备注": "备注",
        },
        row_filters=[{"field": "类别", "equals": "面料"}],
        aggregation_rules={
            "group_by": ["面料名称", "规格", "单位"],
            "sum_fields": ["数量"],
            "merge_text_fields": ["备注"],
        },
        validation_rules={"required_fields": ["面料名称", "规格", "数量", "单位"]},
    )
    store.approve(routine.routine_id)

    result = WorkSpreadsheetRoutineRunner(tmp_path, tmp_path / "state").run(
        routine=store.get(routine.routine_id),
        source_path="source.csv",
        template_path="template.xlsx",
        output_path="reports/from_csv.xlsx",
    )

    report = json.loads(output.with_suffix(".validation.json").read_text(encoding="utf-8"))
    assert result.status == "passed"
    assert report["row_count"] == 1
    assert report["totals"]["数量"] == 13


def test_spreadsheet_routine_blocks_unsupported_formula(tmp_path: Path) -> None:
    from js_work.routines import WorkRoutineStore
    from js_work.routines.spreadsheet import WorkSpreadsheetRoutineRunner

    source = tmp_path / "source.xlsx"
    template = tmp_path / "template.xlsx"
    _save_mixed_factory_workbook(source)
    _save_template_with_unsupported_formula(template)

    store = WorkRoutineStore(tmp_path / "state")
    routine = store.create_draft(
        name="面料统计",
        trigger_phrases=["面料统计"],
        routine_type="spreadsheet_template",
        field_mapping={"面料名称": "品名", "规格": "规格", "数量": "数量", "单位": "单位"},
        row_filters=[{"field": "类别", "equals": "面料"}],
        validation_rules={"required_fields": ["面料名称", "规格", "数量", "单位"]},
    )
    store.approve(routine.routine_id)

    with pytest.raises(ValueError, match="unsupported or unknown formula function: VLOOKUP"):
        WorkSpreadsheetRoutineRunner(tmp_path, tmp_path / "state").run(
            routine=store.get(routine.routine_id),
            source_path="source.xlsx",
            template_path="template.xlsx",
            output_path="reports/unsupported_formula.xlsx",
        )


def test_spreadsheet_routine_blocks_formula_range_that_misses_generated_rows(
    tmp_path: Path,
) -> None:
    from js_work.routines import WorkRoutineStore
    from js_work.routines.spreadsheet import WorkSpreadsheetRoutineRunner

    source = tmp_path / "source.xlsx"
    template = tmp_path / "template.xlsx"
    output = tmp_path / "reports" / "range_overflow.xlsx"
    _save_large_factory_workbook(source, rows=101)
    _save_template_workbook(template)

    store = WorkRoutineStore(tmp_path / "state")
    routine = store.create_draft(
        name="面料统计",
        trigger_phrases=["面料统计"],
        routine_type="spreadsheet_template",
        field_mapping={"面料名称": "品名", "规格": "规格", "数量": "数量", "单位": "单位"},
        row_filters=[{"field": "类别", "equals": "面料"}],
        validation_rules={"required_fields": ["面料名称", "规格", "数量", "单位"]},
    )
    store.approve(routine.routine_id)

    result = WorkSpreadsheetRoutineRunner(tmp_path, tmp_path / "state").run(
        routine=store.get(routine.routine_id),
        source_path="source.xlsx",
        template_path="template.xlsx",
        output_path="reports/range_overflow.xlsx",
    )

    report = json.loads(output.with_suffix(".validation.json").read_text(encoding="utf-8"))
    assert result.status == "needs_review"
    assert report["row_count"] == 101
    assert report["formula_checks"]["F3"]["status"] == "formula_range_excludes_data"
    assert any(issue["code"] == "formula_range_excludes_data" for issue in result.issues)


def test_work_agent_office_profile_exposes_routine_tools_only_in_office(tmp_path: Path) -> None:
    office_agent = create_work_agent(
        settings=load_work_settings(home=tmp_path / "office"),
        profile=WorkToolProfile.OFFICE,
    )
    office_names = {tool.name for tool in office_agent.registry.list_tools()}
    assert {
        "excel_template_analyze",
        "excel_extract_table",
        "excel_render_from_template",
        "excel_validate_output",
        "work_routine_preview",
        "work_routine_run",
    } <= office_names
    assert not {"shell", "python", "fleet_collaborate"} & office_names

    safe_agent = create_work_agent(
        settings=load_work_settings(home=tmp_path / "safe"),
        profile=WorkToolProfile.SAFE,
    )
    safe_names = {tool.name for tool in safe_agent.registry.list_tools()}
    assert (
        not {
            "excel_render_from_template",
            "excel_validate_output",
            "work_routine_preview",
            "work_routine_run",
            "file_write",
        }
        & safe_names
    )


def test_work_routine_tool_reports_needs_review_as_successful_execution(tmp_path: Path) -> None:
    from js.echo.attachment_gate import session_slug
    from js.echo.turn_context import RuntimeContext, reset_runtime_context, set_runtime_context
    from js_work.routines import WorkRoutineStore
    from js_work.routines.tools import WorkRoutineTools

    local_root = tmp_path / "local" / session_slug("session-a")
    local_root.mkdir(parents=True)
    source = local_root / "source.xlsx"
    template = local_root / "template.xlsx"
    _save_source_workbook(source, conflicting_unit=True)
    _save_template_workbook(template)

    store = WorkRoutineStore(tmp_path / "state", session_id="session-a")
    routine = store.create_draft(
        name="面料统计",
        trigger_phrases=["面料统计"],
        routine_type="spreadsheet_template",
        field_mapping={"面料名称": "品名", "规格": "规格", "数量": "数量", "单位": "单位"},
        validation_rules={"required_fields": ["面料名称", "规格", "数量", "单位"]},
    )
    store.approve(routine.routine_id)

    context = RuntimeContext(
        product_id="js-work",
        channel="test",
        owner_key_hash="js-work-local",
        session_id="session-a",
        run_id="run-a",
        role="local-user",
        profile="office",
        capabilities=("work_routine_run",),
        workspace=tmp_path,
        state_dir=tmp_path / "state",
    )
    token = set_runtime_context(context)
    try:
        result = asyncio.run(
            WorkRoutineTools(
                workspace=tmp_path,
                state_dir=tmp_path / "state",
            ).work_routine_run(
                routine_id=routine.routine_id,
                source_path="source.xlsx",
                template_path="template.xlsx",
                output_path="reports/needs_review.xlsx",
            )
        )
    finally:
        reset_runtime_context(token)

    assert result.success is True
    assert result.metadata["status"] == "needs_review"
    assert result.metadata["report"] is not None
    assert any(issue["code"] == "unit_conflict" for issue in result.metadata["report"]["issues"])
    payload = json.loads(result.output)
    assert payload["report"]["status"] == "needs_review"


def test_routine_miner_generates_disabled_draft_after_three_repetitions(tmp_path: Path) -> None:
    from js_work.routines import RoutineStatus, WorkRoutineMiner, WorkRoutineStore

    store = WorkRoutineStore(tmp_path / "state")
    miner = WorkRoutineMiner(store)
    draft = None
    for _ in range(3):
        draft = miner.observe(
            task_text="按面料统计模板生成表格",
            routine_type="spreadsheet_template",
            trigger_phrase="面料统计",
            field_mapping={"面料名称": "品名", "数量": "数量"},
            validation_rules={"required_fields": ["面料名称", "数量"]},
        )

    assert draft is not None
    assert draft.status == RoutineStatus.DRAFT
    assert draft.enabled is False
    assert "面料统计" in draft.trigger_phrases


def test_routine_miner_observation_update_is_atomic_under_concurrency(
    tmp_path: Path,
) -> None:
    from js_work.routines import WorkRoutineMiner, WorkRoutineStore

    store = WorkRoutineStore(tmp_path / "state", owner_key_hash="owner-a")
    miner = WorkRoutineMiner(store, threshold=2)
    original_load = miner._load
    barrier = threading.Barrier(8)

    def synchronized_load() -> dict[str, Any]:
        snapshot = original_load()
        barrier.wait(timeout=2.0)
        return snapshot

    miner._load = synchronized_load  # type: ignore[method-assign]

    def observe() -> object:
        return miner.observe(
            task_text="synthetic spreadsheet task",
            routine_type="spreadsheet_template",
            trigger_phrase="synthetic material summary",
            field_mapping={"material": "name", "quantity": "quantity"},
            validation_rules={"required_fields": ["material", "quantity"]},
        )

    with ThreadPoolExecutor(max_workers=8) as pool:
        results = list(pool.map(lambda _index: observe(), range(8)))

    routine_ids = {result.routine_id for result in results if hasattr(result, "routine_id")}
    assert len(routine_ids) == 1
    routines = store.list_routines()
    assert [routine.routine_id for routine in routines] == list(routine_ids)


def test_js_work_cli_routine_commands(tmp_path: Path) -> None:
    runner = CliRunner()

    draft = runner.invoke(
        work_main,
        [
            "--home",
            str(tmp_path),
            "routine",
            "draft",
            "--name",
            "面料统计",
            "--trigger",
            "面料统计",
            "--mapping",
            json.dumps({"面料名称": "品名", "数量": "数量"}, ensure_ascii=False),
        ],
    )
    assert draft.exit_code == 0
    routine_id = json.loads(draft.output)["routine_id"]

    approve = runner.invoke(work_main, ["--home", str(tmp_path), "routine", "approve", routine_id])
    assert approve.exit_code == 0
    assert json.loads(approve.output)["status"] == "enabled"

    listed = runner.invoke(work_main, ["--home", str(tmp_path), "routine", "list"])
    assert listed.exit_code == 0
    assert json.loads(listed.output)["routines"][0]["routine_id"] == routine_id


def test_js_work_cli_routine_dry_run_reports_without_writing_output(tmp_path: Path) -> None:
    from js.echo.attachment_gate import session_slug

    runner = CliRunner()
    workspace = tmp_path / ".js-work" / "workspace"
    local_root = workspace / "local" / session_slug("work-routine-cli")
    local_root.mkdir(parents=True)
    _save_mixed_factory_workbook(local_root / "source.xlsx")
    _save_template_workbook(local_root / "template.xlsx")

    draft = runner.invoke(
        work_main,
        [
            "--home",
            str(tmp_path),
            "routine",
            "draft",
            "--name",
            "面料统计",
            "--trigger",
            "面料统计",
            "--mapping",
            json.dumps(
                {
                    "面料名称": "品名",
                    "规格": "规格",
                    "数量": "数量",
                    "单位": "单位",
                    "备注": "备注",
                },
                ensure_ascii=False,
            ),
            "--row-filter",
            json.dumps({"field": "类别", "equals": "面料"}, ensure_ascii=False),
            "--aggregation",
            json.dumps(
                {
                    "group_by": ["面料名称", "规格", "单位"],
                    "sum_fields": ["数量"],
                    "merge_text_fields": ["备注"],
                },
                ensure_ascii=False,
            ),
        ],
    )
    assert draft.exit_code == 0
    routine_id = json.loads(draft.output)["routine_id"]
    approve = runner.invoke(work_main, ["--home", str(tmp_path), "routine", "approve", routine_id])
    assert approve.exit_code == 0

    output = local_root / "reports" / "dry_run.xlsx"
    dry_run = runner.invoke(
        work_main,
        [
            "--home",
            str(tmp_path),
            "--profile",
            "office",
            "routine",
            "run",
            routine_id,
            "--source",
            "source.xlsx",
            "--template",
            "template.xlsx",
            "--output",
            "reports/dry_run.xlsx",
            "--dry-run",
        ],
    )

    assert dry_run.exit_code == 0
    payload = json.loads(dry_run.output)
    assert payload["status"] == "passed"
    assert payload["row_count"] == 2
    assert payload["selected_source_row_count"] == 3
    assert payload["excluded_source_row_count"] == 1
    assert output.exists() is False


def test_work_routine_tools_resolve_store_from_runtime_owner(tmp_path: Path) -> None:
    from js.echo.turn_context import RuntimeContext, reset_runtime_context, set_runtime_context
    from js_work.routines import WorkRoutineStore
    from js_work.routines.tools import WorkRoutineTools

    state = tmp_path / "state"
    store_a = WorkRoutineStore(state, owner_key_hash="owner-a")
    store_b = WorkRoutineStore(state, owner_key_hash="owner-b")
    routine_a = store_a.create_draft(
        name="A",
        trigger_phrases=["a"],
        routine_type="spreadsheet_template",
        field_mapping={"面料名称": "品名"},
    )
    store_a.approve(routine_a.routine_id)
    routine_b = store_b.create_draft(
        name="B",
        trigger_phrases=["b"],
        routine_type="spreadsheet_template",
        field_mapping={"面料名称": "品名"},
    )
    store_b.approve(routine_b.routine_id)

    tools = WorkRoutineTools(workspace=tmp_path, state_dir=state)
    context_a = RuntimeContext(
        product_id="js-work",
        channel="test",
        owner_key_hash="owner-a",
        session_id="s-a",
        run_id="r-a",
        role="user",
        profile="office",
        capabilities=("work_routine_preview",),
        workspace=tmp_path,
        state_dir=state,
    )
    token = set_runtime_context(context_a)
    try:
        result = asyncio.run(
            tools.work_routine_preview(
                routine_id=routine_b.routine_id,
                source_path="missing.xlsx",
                template_path="missing.xlsx",
            )
        )
    finally:
        reset_runtime_context(token)

    assert result.success is False
    assert "not found" in (result.error or "").lower()


def test_work_routine_tools_reject_another_owners_file(tmp_path: Path) -> None:
    from js.echo.attachment_gate import owner_slug, session_slug
    from js.echo.turn_context import RuntimeContext, reset_runtime_context, set_runtime_context
    from js_work.routines.tools import WorkRoutineTools

    workspace = tmp_path / "workspace"
    other_relative = Path("owners") / owner_slug("owner-b") / session_slug("s-a") / "template.xlsx"
    other_template = workspace / other_relative
    other_template.parent.mkdir(parents=True)
    _save_template_workbook(other_template)
    tools = WorkRoutineTools(workspace=workspace, state_dir=tmp_path / "state")
    context = RuntimeContext(
        product_id="js-work",
        channel="test",
        owner_key_hash="owner-a",
        session_id="s-a",
        run_id="r-a",
        role="user",
        profile="office",
        capabilities=("excel_template_analyze",),
        workspace=workspace,
        state_dir=tmp_path / "state",
    )

    token = set_runtime_context(context)
    try:
        result = asyncio.run(tools.excel_template_analyze(other_relative.as_posix()))
    finally:
        reset_runtime_context(token)

    assert result.success is False
    assert "denied" in result.error.lower()


def test_work_routine_tool_fails_closed_without_work_runtime_context(
    tmp_path: Path,
) -> None:
    from js_work.routines.tools import WorkRoutineTools

    workspace = tmp_path / "workspace"
    template = workspace / "global-template.xlsx"
    template.parent.mkdir(parents=True)
    _save_template_workbook(template)
    tools = WorkRoutineTools(workspace=workspace, state_dir=tmp_path / "state")

    result = asyncio.run(tools.excel_template_analyze("global-template.xlsx"))

    assert result.success is False
    assert "runtime context" in result.error.lower()


def test_work_routine_tool_sanitizes_unexpected_engine_error(
    tmp_path: Path,
    monkeypatch,
) -> None:
    from js.echo.turn_context import RuntimeContext, reset_runtime_context, set_runtime_context
    from js_work.file_scope import WorkOwnerFileScope
    from js_work.routines.tools import WorkRoutineTools

    private_detail = "/Users/private/Documents/customer.xlsx secret-token"
    workspace = tmp_path / "workspace"
    scope = WorkOwnerFileScope(workspace, owner="owner-a", session_id="s-a")
    scope.private_root.mkdir(parents=True)
    _save_template_workbook(scope.private_root / "safe.xlsx")
    tools = WorkRoutineTools(workspace=workspace, state_dir=tmp_path / "state")
    monkeypatch.setattr(
        tools.engine,
        "analyze_template",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(RuntimeError(private_detail)),
    )
    context = RuntimeContext(
        product_id="js-work",
        channel="test",
        owner_key_hash="owner-a",
        session_id="s-a",
        run_id="r-a",
        role="user",
        profile="office",
        capabilities=("excel_template_analyze",),
        workspace=workspace,
        state_dir=tmp_path / "state",
    )

    token = set_runtime_context(context)
    try:
        result = asyncio.run(tools.excel_template_analyze("safe.xlsx"))
    finally:
        reset_runtime_context(token)

    assert result.success is False
    assert result.error == "Work routine operation failed safely"
    assert private_detail not in str(result)


def test_work_routine_output_is_written_under_the_owner_root(tmp_path: Path) -> None:
    from js.echo.attachment_gate import owner_slug, session_slug
    from js.echo.turn_context import RuntimeContext, reset_runtime_context, set_runtime_context
    from js_work.routines.tools import WorkRoutineTools

    workspace = tmp_path / "workspace"
    private_root = workspace / "owners" / owner_slug("owner-a") / session_slug("s-a")
    template = private_root / "template.xlsx"
    template.parent.mkdir(parents=True)
    _save_template_workbook(template)
    tools = WorkRoutineTools(workspace=workspace, state_dir=tmp_path / "state")
    context = RuntimeContext(
        product_id="js-work",
        channel="test",
        owner_key_hash="owner-a",
        session_id="s-a",
        run_id="r-a",
        role="user",
        profile="office",
        capabilities=("excel_render_from_template",),
        workspace=workspace,
        state_dir=tmp_path / "state",
    )

    token = set_runtime_context(context)
    try:
        result = asyncio.run(
            tools.excel_render_from_template(
                template_path="template.xlsx",
                output_path="reports/result.xlsx",
                rows=json.dumps(
                    [{"面料名称": "全棉", "规格": "40S", "数量": 1, "单位": "米"}],
                    ensure_ascii=False,
                ),
            )
        )
    finally:
        reset_runtime_context(token)

    assert result.success is True, result.error
    assert result.output == "reports/result.xlsx"
    assert result.metadata["path"] == "reports/result.xlsx"
    assert str(workspace) not in str(result)
    assert (private_root / "reports" / "result.xlsx").exists()
    assert not (workspace / "reports" / "result.xlsx").exists()


def test_routine_store_owner_isolation_and_rejects_path_traversal(tmp_path: Path) -> None:
    from js.echo.attachment_gate import owner_slug, session_slug
    from js_work.routines import WorkRoutineStore, validate_routine_id

    state = tmp_path / "state"
    store_a = WorkRoutineStore(state, owner_key_hash="owner-a")
    store_b = WorkRoutineStore(state, owner_key_hash="owner-b")
    routine = store_a.create_draft(
        name="面料统计",
        trigger_phrases=["面料统计"],
        routine_type="spreadsheet_template",
        field_mapping={"面料名称": "品名"},
    )

    assert store_a.get(routine.routine_id).routine_id == routine.routine_id
    assert [item.routine_id for item in store_a.list_routines()] == [routine.routine_id]
    assert store_b.list_routines() == []
    try:
        store_b.get(routine.routine_id)
        raise AssertionError("owner-b must not see owner-a routines")
    except KeyError:
        pass
    try:
        store_b.approve(routine.routine_id)
        raise AssertionError("owner-b must not approve owner-a routines")
    except KeyError:
        pass

    expected_path = (
        state
        / "routines"
        / owner_slug("owner-a")
        / session_slug("default")
        / f"{routine.routine_id}.json"
    )
    assert expected_path.exists()
    assert not (state / "routines" / f"{routine.routine_id}.json").exists()

    for bad_id in (
        "../escape",
        "../../etc/passwd",
        "/tmp/abs",
        "a/b",
        "a\\b",
        "x" * 65,
        "bad id",
        "id with spaces!",
    ):
        try:
            validate_routine_id(bad_id)
            raise AssertionError(f"expected invalid routine_id: {bad_id}")
        except ValueError:
            pass
        try:
            store_a.get(bad_id)
            raise AssertionError(f"store must reject invalid routine_id: {bad_id}")
        except ValueError:
            pass


def test_routine_store_rejects_symlinked_owner_partition(tmp_path: Path) -> None:
    from js.echo.attachment_gate import owner_slug
    from js_work.routines import WorkRoutineStore

    state = tmp_path / "state"
    routines_root = state / "routines"
    routines_root.mkdir(parents=True)
    outside = tmp_path / "outside-routines"
    outside.mkdir()
    (routines_root / owner_slug("owner-a")).symlink_to(
        outside,
        target_is_directory=True,
    )

    with pytest.raises(OSError):
        WorkRoutineStore(state, owner_key_hash="owner-a")

    assert list(outside.iterdir()) == []


def test_routine_store_rejects_symlinked_routine_file(tmp_path: Path) -> None:
    from js_work.routines import WorkRoutineStore

    state = tmp_path / "state"
    store = WorkRoutineStore(state, owner_key_hash="owner-a")
    routine = store.create_draft(
        name="Owned routine",
        trigger_phrases=["owned"],
        routine_type="spreadsheet_template",
    )
    routine_path = store.routines_dir / f"{routine.routine_id}.json"
    outside = tmp_path / "outside-routine.json"
    outside.write_text(routine_path.read_text(encoding="utf-8"), encoding="utf-8")
    routine_path.unlink()
    routine_path.symlink_to(outside)

    with pytest.raises(KeyError):
        store.get(routine.routine_id)
    with pytest.raises(KeyError):
        store.approve(routine.routine_id)

    assert json.loads(outside.read_text(encoding="utf-8"))["status"] == "draft"


def test_routine_store_atomic_publish_failure_preserves_previous_record(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    from js_work.routines import WorkRoutineStore, json_store

    store = WorkRoutineStore(tmp_path / "state", owner_key_hash="owner-a")
    routine = store.create_draft(
        name="Original",
        trigger_phrases=["original"],
        routine_type="spreadsheet_template",
    )

    def fail_replace(_temp_name: str, _filename: str, _directory_fd: int) -> None:
        raise OSError("simulated atomic publish failure")

    monkeypatch.setattr(json_store, "_replace_at", fail_replace)
    routine.name = "Must not publish"

    with pytest.raises(OSError):
        store.save(routine)

    assert store.get(routine.routine_id).name == "Original"


def test_js_work_cli_routine_run_requires_office_profile_and_uses_effect(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    from js.models.providers import ChatMessage
    from js.tools.registry import ToolResult

    runner = CliRunner()
    draft = runner.invoke(
        work_main,
        [
            "--home",
            str(tmp_path),
            "routine",
            "draft",
            "--name",
            "面料统计",
            "--trigger",
            "面料统计",
            "--mapping",
            json.dumps({"面料名称": "品名", "数量": "数量"}, ensure_ascii=False),
        ],
    )
    assert draft.exit_code == 0
    routine_id = json.loads(draft.output)["routine_id"]
    approve = runner.invoke(work_main, ["--home", str(tmp_path), "routine", "approve", routine_id])
    assert approve.exit_code == 0

    rejected = runner.invoke(
        work_main,
        [
            "--home",
            str(tmp_path),
            "--profile",
            "execute",
            "routine",
            "run",
            routine_id,
            "--source",
            "source.xlsx",
            "--template",
            "template.xlsx",
            "--output",
            "reports/out.xlsx",
        ],
    )
    assert rejected.exit_code != 0
    assert "office" in (rejected.output + rejected.stderr).lower()

    captured: dict[str, Any] = {}

    class FakeRuntime:
        def build_context(self, **kwargs: Any) -> Any:
            kwargs["run_id"] = "test-work-run"
            captured["context"] = kwargs
            return type("FakeContext", (), kwargs)()

        async def execute_tool_effect(self, effect: Any, context: Any) -> tuple[Any, ToolResult]:
            captured["effect"] = effect
            captured["effect_context"] = context
            return (
                ChatMessage(role="tool", content="{}", name=effect.tool_name),
                ToolResult(
                    success=True,
                    output=json.dumps(
                        {
                            "status": "passed",
                            "row_count": 1,
                            "report": {"status": "passed", "row_count": 1},
                        },
                        ensure_ascii=False,
                    ),
                ),
            )

    class FakeApprovals:
        def __init__(self) -> None:
            self.callbacks: dict[str, Any] = {}

        def set_callback(
            self,
            session_id: str,
            callback: Any,
            **bindings: Any,
        ) -> None:
            self.callbacks[session_id] = callback
            captured["approval_callback_set"] = session_id
            captured["approval_callback_bindings"] = bindings

        def remove_callback(self, session_id: str) -> None:
            self.callbacks.pop(session_id, None)
            captured["approval_callback_removed"] = session_id

    class FakeAgent:
        def __init__(self) -> None:
            self.echo_runtime = FakeRuntime()
            self.approvals = FakeApprovals()
            self.closed = False

        async def close(self) -> None:
            self.closed = True
            captured["closed"] = True

    monkeypatch.setattr("js_work.cli.create_work_agent", lambda **_kwargs: FakeAgent())

    ok = runner.invoke(
        work_main,
        [
            "--home",
            str(tmp_path),
            "--profile",
            "office",
            "routine",
            "run",
            routine_id,
            "--source",
            "source.xlsx",
            "--template",
            "template.xlsx",
            "--output",
            "reports/out.xlsx",
        ],
    )
    assert ok.exit_code == 0
    assert json.loads(ok.output)["status"] == "passed"
    assert captured["context"]["channel"] == "js_work_routine_cli"
    assert captured["context"]["owner_key_hash"] == "js-work-local"
    assert captured["context"]["capabilities"] == ("work_routine_run",)
    assert captured["effect"].tool_name == "work_routine_run"
    assert captured["approval_callback_set"] == captured["approval_callback_removed"]
    assert captured["context"]["session_id"] == captured["approval_callback_set"]
    assert captured["approval_callback_bindings"] == {
        "owner_key_hash": "js-work-local",
        "run_id": captured["context"]["run_id"],
        "tool_name": "work_routine_run",
        "arguments": json.loads(captured["effect"].arguments_json),
    }
    assert captured["closed"] is True

    dry = runner.invoke(
        work_main,
        [
            "--home",
            str(tmp_path),
            "--profile",
            "office",
            "routine",
            "run",
            routine_id,
            "--source",
            "source.xlsx",
            "--template",
            "template.xlsx",
            "--output",
            "reports/out.xlsx",
            "--dry-run",
        ],
    )
    assert dry.exit_code == 0
    assert captured["context"]["capabilities"] == ("work_routine_preview",)
    assert captured["effect"].tool_name == "work_routine_preview"


def test_work_web_routine_endpoints(tmp_path: Path, monkeypatch: Any) -> None:
    from unittest.mock import AsyncMock

    from js_work.web import create_work_web_app

    config = tmp_path / "config.yaml"
    config.write_text("security:\n  api_key_required: false\nproviders: []\n", encoding="utf-8")
    app = create_work_web_app(config=str(config), home=tmp_path, profile=WorkToolProfile.OFFICE)

    with TestClient(
        app, base_url="http://localhost", headers={"Origin": "http://localhost"}
    ) as client:
        # Admin endpoints reject anonymous guests; authenticate as a work admin.
        from js.web.auth import AuthManager

        client.headers["X-API-Key"] = AuthManager(
            app.state.web_runtime.settings.state_dir
        ).create_key("work-admin", role="admin")
        agent = app.state.web_runtime.agent
        execute_effect = AsyncMock(wraps=agent.echo_runtime.execute_tool_effect)
        monkeypatch.setattr(agent.echo_runtime, "execute_tool_effect", execute_effect)
        draft = client.post(
            "/api/work/routines/draft",
            json={
                "name": "面料统计",
                "trigger_phrases": ["面料统计"],
                "field_mapping": {"面料名称": "品名", "数量": "数量"},
            },
        )
        assert draft.status_code == 200
        routine_id = draft.json()["routine_id"]

        approve = client.post(f"/api/work/routines/{routine_id}/approve")
        assert approve.status_code == 200
        assert approve.json()["status"] == "enabled"

        listed = client.get("/api/work/routines")
        assert listed.status_code == 200
        assert listed.json()["routines"][0]["routine_id"] == routine_id
        effects = [call.args[0] for call in execute_effect.await_args_list]
        assert [effect.tool_name for effect in effects] == [
            "control_work_routine_draft",
            "control_work_routine_approve",
        ]
        assert all(effect.allowed_tools == (effect.tool_name,) for effect in effects)


def test_work_web_routine_run_returns_validation_report(tmp_path: Path) -> None:
    from js.echo.attachment_gate import owner_slug, session_slug
    from js_work.web import create_work_web_app

    config = tmp_path / "config.yaml"
    config.write_text("security:\n  api_key_required: false\nproviders: []\n", encoding="utf-8")
    workspace = tmp_path / ".js-work" / "workspace"
    # Routine definitions are owner+session partitioned: the web control
    # endpoints manage routines in the "web" session, so the run must use
    # the same session (cross-session execution is denied by design).
    # Anonymous guests are read-only; authenticate as a work admin and seed
    # the source files under that key's owner partition.
    from js.web.auth import AuthManager

    state_dir = tmp_path / ".js-work" / "state"
    admin_key = AuthManager(state_dir).create_key("work-admin", role="admin")
    owner = AuthManager(state_dir).verify(admin_key)["key_hash"]
    local_root = workspace / "owners" / owner_slug(owner) / session_slug("web")
    local_root.mkdir(parents=True)
    _save_mixed_factory_workbook(local_root / "source.xlsx")
    _save_template_workbook(local_root / "template.xlsx")
    app = create_work_web_app(config=str(config), home=tmp_path, profile=WorkToolProfile.OFFICE)

    with TestClient(
        app, base_url="http://localhost", headers={"Origin": "http://localhost"}
    ) as client:
        client.headers["X-API-Key"] = admin_key
        draft = client.post(
            "/api/work/routines/draft",
            json={
                "name": "面料统计",
                "trigger_phrases": ["面料统计"],
                "field_mapping": {
                    "面料名称": "品名",
                    "规格": "规格",
                    "数量": "数量",
                    "单位": "单位",
                    "备注": "备注",
                },
                "row_filters": [{"field": "类别", "equals": "面料"}],
                "aggregation_rules": {
                    "group_by": ["面料名称", "规格", "单位"],
                    "sum_fields": ["数量"],
                    "merge_text_fields": ["备注"],
                },
                "validation_rules": {"required_fields": ["面料名称", "规格", "数量", "单位"]},
            },
        )
        assert draft.status_code == 200
        routine_id = draft.json()["routine_id"]
        approve = client.post(f"/api/work/routines/{routine_id}/approve")
        assert approve.status_code == 200

        run = client.post(
            "/api/work/routines/run",
            json={
                "routine_id": routine_id,
                "session_id": "web",
                "source_path": "source.xlsx",
                "template_path": "template.xlsx",
                "output_path": "reports/web.xlsx",
            },
        )

        assert run.status_code == 200
        payload = run.json()
        assert payload["status"] == "passed"
        assert payload["report"]["row_count"] == 2
        assert payload["report"]["excluded_source_row_count"] == 1
        assert payload["report"]["totals"]["数量"] == 21.5


def test_work_web_routine_run_uses_echo_tool_effect_not_direct_runner(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    from unittest.mock import AsyncMock

    from js.models.providers import ChatMessage
    from js.tools.registry import ToolResult
    from js_work.web import create_work_web_app

    config = tmp_path / "config.yaml"
    config.write_text("security:\n  api_key_required: false\nproviders: []\n", encoding="utf-8")
    app = create_work_web_app(config=str(config), home=tmp_path, profile=WorkToolProfile.OFFICE)

    with TestClient(
        app, base_url="http://localhost", headers={"Origin": "http://localhost"}
    ) as client:
        # Admin endpoints reject anonymous guests; authenticate as a work admin.
        from js.web.auth import AuthManager

        client.headers["X-API-Key"] = AuthManager(
            app.state.web_runtime.settings.state_dir
        ).create_key("work-admin", role="admin")
        draft = client.post(
            "/api/work/routines/draft",
            json={
                "name": "面料统计",
                "trigger_phrases": ["面料统计"],
                "field_mapping": {"面料名称": "品名", "数量": "数量"},
            },
        )
        assert draft.status_code == 200
        routine_id = draft.json()["routine_id"]
        assert client.post(f"/api/work/routines/{routine_id}/approve").status_code == 200

        agent = app.state.web_runtime.agent
        captured: dict[str, Any] = {}

        async def fake_execute(
            effect: Any, context: Any, progress_callback: Any = None
        ) -> tuple[Any, ToolResult]:
            captured["effect"] = effect
            captured["context"] = context
            return (
                ChatMessage(role="tool", content="{}", name=effect.tool_name),
                ToolResult(
                    success=True,
                    output=json.dumps(
                        {
                            "status": "passed",
                            "row_count": 2,
                            "report": {
                                "status": "passed",
                                "row_count": 2,
                                "totals": {"数量": 21.5},
                            },
                        },
                        ensure_ascii=False,
                    ),
                ),
            )

        monkeypatch.setattr(
            agent.echo_runtime, "execute_tool_effect", AsyncMock(side_effect=fake_execute)
        )

        run = client.post(
            "/api/work/routines/run",
            json={
                "routine_id": routine_id,
                "session_id": "work-web-session",
                "source_path": "source.xlsx",
                "template_path": "template.xlsx",
                "output_path": "reports/web.xlsx",
            },
        )
        assert run.status_code == 200
        assert run.json()["report"]["row_count"] == 2
        assert captured["effect"].tool_name == "work_routine_run"
        assert captured["context"].channel == "js_work_routine_web"
        expected_owner = AuthManager(app.state.web_runtime.settings.state_dir).verify(
            client.headers["X-API-Key"]
        )["key_hash"]
        assert captured["context"].owner_key_hash == expected_owner
        assert captured["context"].session_id == "work-web-session"
        assert captured["context"].capabilities == ("work_routine_run",)

        preview = client.post(
            "/api/work/routines/run",
            json={
                "routine_id": routine_id,
                "session_id": "work-web-session",
                "source_path": "source.xlsx",
                "template_path": "template.xlsx",
                "dry_run": True,
            },
        )
        assert preview.status_code == 200
        assert captured["effect"].tool_name == "work_routine_preview"
        assert captured["context"].capabilities == ("work_routine_preview",)
