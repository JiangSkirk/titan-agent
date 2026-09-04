from __future__ import annotations

import os
from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook, load_workbook

from js_work.routines import formula_cache
from js_work.routines.precise_edit import PreciseExcelEditEngine
from js_work.safe_output import publish_no_clobber, staged_path


def _save_formula_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet["A1"] = 1
    sheet["A2"] = 2
    sheet["B1"] = "=SUM(A1:A2)"
    workbook.save(path)
    workbook.close()


def test_precise_edit_parent_swap_cannot_stage_bytes_or_cleanup_outside_anchor(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = tmp_path / "source.xlsx"
    output_parent = tmp_path / "reports"
    output = output_parent / "edited.xlsx"
    parked = tmp_path / "reports-parked"
    outside = tmp_path / "outside"
    outside.mkdir()
    _save_formula_workbook(source)
    real_save = Workbook.save
    observed_outside_bytes: list[bytes] = []
    swapped = False

    def swap_parent_before_save(workbook: Workbook, filename: Any) -> None:
        nonlocal swapped
        if not swapped:
            output_parent.rename(parked)
            output_parent.symlink_to(outside, target_is_directory=True)
            swapped = True
        real_save(workbook, filename)
        observed_outside_bytes.extend(
            candidate.read_bytes() for candidate in outside.iterdir() if candidate.is_file()
        )

    monkeypatch.setattr(Workbook, "save", swap_parent_before_save)

    with pytest.raises((OSError, ValueError)):
        PreciseExcelEditEngine(tmp_path).apply(
            source_path=source.name,
            output_path="reports/edited.xlsx",
            operations=[{"op": "set_cell", "sheet": "Data", "cell": "A1", "value": 7}],
        )

    assert swapped is True
    assert observed_outside_bytes == []
    assert list(outside.iterdir()) == []
    assert list(parked.glob(".*.xlsx")) == []
    assert not output.exists()


def test_formula_cache_parent_swap_cannot_replace_outside_hardlink(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output_parent = tmp_path / "reports"
    output_parent.mkdir()
    parked = tmp_path / "reports-parked"
    outside = tmp_path / "outside"
    outside.mkdir()
    target = output_parent / "result.xlsx"
    real_zip_file = formula_cache.zipfile.ZipFile
    swapped = False

    with staged_path(target) as staged:
        _save_formula_workbook(staged)

        def swap_parent_before_zip_open(*args: Any, **kwargs: Any) -> Any:
            nonlocal swapped
            if not swapped:
                output_parent.rename(parked)
                output_parent.symlink_to(outside, target_is_directory=True)
                os.link(parked / staged.name, outside / staged.name)
                swapped = True
            return real_zip_file(*args, **kwargs)

        monkeypatch.setattr(formula_cache.zipfile, "ZipFile", swap_parent_before_zip_open)

        result = formula_cache.refresh_formula_caches(staged, soffice=None)
        assert result.cached_count == 1
        with pytest.raises((OSError, ValueError)):
            publish_no_clobber(staged, target, "output already exists")

    assert swapped is True
    outside_workbook = load_workbook(outside / staged.name, data_only=True)
    try:
        assert outside_workbook["Data"]["B1"].value is None
    finally:
        outside_workbook.close()
    assert list(outside.glob(".*.xlsx")) == [outside / staged.name]
    assert list(parked.glob(".*.xlsx")) == []
    assert not target.exists()
