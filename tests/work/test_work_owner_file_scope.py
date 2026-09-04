from __future__ import annotations

import os
import shutil
from collections.abc import Callable
from dataclasses import replace
from hashlib import sha256
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from js.echo.attachment_gate import owner_slug, session_slug
from js_work.file_scope import LOCAL_WORK_OWNER, WorkFileScopeError, WorkOwnerFileScope


def test_authenticated_owners_are_isolated_to_private_roots(tmp_path: Path) -> None:
    workspace = tmp_path / "workspace"
    owner_a = WorkOwnerFileScope(workspace, owner="owner-a", session_id="session-a")
    owner_b = WorkOwnerFileScope(workspace, owner="owner-b", session_id="session-a")

    path_a = owner_a.resolve_private_read("reports/quarterly.md")
    path_b = owner_b.resolve_private_read("reports/quarterly.md")

    assert path_a == (
        workspace
        / "owners"
        / owner_slug("owner-a")
        / session_slug("session-a")
        / "reports/quarterly.md"
    ).resolve()
    assert path_b == (
        workspace
        / "owners"
        / owner_slug("owner-b")
        / session_slug("session-a")
        / "reports/quarterly.md"
    ).resolve()
    assert path_a != path_b
    assert owner_a.to_registry_path(path_a) == (
        f"owners/{owner_slug('owner-a')}/{session_slug('session-a')}/reports/quarterly.md"
    )
    with pytest.raises(WorkFileScopeError) as exc_info:
        owner_a.resolve_private_read(
            f"owners/{owner_slug('owner-b')}/reports/quarterly.md"
        )
    assert exc_info.value.status_code == 403


def test_local_owner_keeps_workspace_relative_compatibility(tmp_path: Path) -> None:
    workspace = tmp_path / "workspace"
    scope = WorkOwnerFileScope(
        workspace,
        owner=LOCAL_WORK_OWNER,
        session_id="session-a",
    )

    resolved = scope.resolve_private_read("legacy/project/input.txt")

    assert scope.private_root == (workspace / "local" / session_slug("session-a")).resolve()
    assert resolved == (
        workspace / "local" / session_slug("session-a") / "legacy/project/input.txt"
    ).resolve()
    assert scope.to_registry_path(resolved) == (
        f"local/{session_slug('session-a')}/legacy/project/input.txt"
    )


def test_local_owner_cannot_enter_authenticated_owner_or_upload_roots(
    tmp_path: Path,
) -> None:
    workspace = tmp_path / "workspace"
    scope = WorkOwnerFileScope(
        workspace,
        owner=LOCAL_WORK_OWNER,
        session_id="session-a",
    )

    for path in (
        f"owners/{owner_slug('owner-a')}/{session_slug('session-a')}/secret.txt",
        f"uploads/{owner_slug('owner-a')}/{session_slug('session-a')}/secret.txt",
    ):
        with pytest.raises(WorkFileScopeError) as exc_info:
            scope.resolve_routine_input(path)
        assert exc_info.value.status_code == 403


def test_owner_cannot_enter_another_session_root(tmp_path: Path) -> None:
    workspace = tmp_path / "workspace"
    scope = WorkOwnerFileScope(
        workspace,
        owner="owner-a",
        session_id="session-a",
    )

    for path in (
        f"owners/{owner_slug('owner-a')}/{session_slug('session-b')}/secret.txt",
        f"uploads/{owner_slug('owner-a')}/{session_slug('session-b')}/secret.txt",
    ):
        with pytest.raises(WorkFileScopeError) as exc_info:
            scope.resolve_routine_input(path)
        assert exc_info.value.status_code == 403


@pytest.mark.parametrize(
    ("owner", "upload_owner"),
    [("owner-a", "owner-a"), (LOCAL_WORK_OWNER, None)],
)
def test_routine_input_accepts_only_owned_upload(
    tmp_path: Path,
    owner: str,
    upload_owner: str | None,
) -> None:
    workspace = tmp_path / "workspace"
    scope = WorkOwnerFileScope(workspace, owner=owner, session_id="session-a")
    owned = (
        f"uploads/{owner_slug(owner)}/{session_slug('session-a')}/input.csv"
    )
    other = (
        f"uploads/{owner_slug('owner-b')}/{session_slug('session-a')}/input.csv"
    )

    assert scope.resolve_routine_input(owned) == (workspace / owned).resolve()
    assert scope.to_registry_path(owned) == owned
    with pytest.raises(WorkFileScopeError) as exc_info:
        scope.resolve_routine_input(other)
    assert exc_info.value.status_code == 403


@pytest.mark.parametrize("owner", ["owner-a", LOCAL_WORK_OWNER])
def test_output_is_private_and_never_an_upload(tmp_path: Path, owner: str) -> None:
    workspace = tmp_path / "workspace"
    scope = WorkOwnerFileScope(workspace, owner=owner, session_id="session-a")
    expected_root = (
        workspace / "local" / session_slug("session-a")
        if owner == LOCAL_WORK_OWNER
        else workspace / "owners" / owner_slug(owner) / session_slug("session-a")
    )

    assert scope.resolve_output("exports/result.xlsx") == (
        expected_root / "exports/result.xlsx"
    ).resolve()
    with pytest.raises(WorkFileScopeError) as exc_info:
        scope.resolve_output(
            f"uploads/{owner_slug(owner)}/{session_slug('session-a')}/result.xlsx"
        )
    assert exc_info.value.status_code == 403


@pytest.mark.parametrize("logical", ["../outside.txt", "reports/../secret.txt"])
def test_parent_traversal_is_rejected(tmp_path: Path, logical: str) -> None:
    scope = WorkOwnerFileScope(
        tmp_path / "workspace",
        owner="owner-a",
        session_id="session-a",
    )
    with pytest.raises(WorkFileScopeError) as exc_info:
        scope.resolve_routine_input(logical)
    assert exc_info.value.status_code == 400


def test_absolute_path_outside_workspace_is_rejected(tmp_path: Path) -> None:
    scope = WorkOwnerFileScope(
        tmp_path / "workspace",
        owner="owner-a",
        session_id="session-a",
    )
    with pytest.raises(WorkFileScopeError) as exc_info:
        scope.resolve_private_read(tmp_path / "outside.txt")
    assert exc_info.value.status_code == 400


@pytest.mark.parametrize(
    "resolver_name",
    ["resolve_private_read", "resolve_routine_input", "resolve_output"],
)
def test_symlink_escape_is_rejected(
    tmp_path: Path,
    resolver_name: str,
) -> None:
    workspace = tmp_path / "workspace"
    outside = tmp_path / "outside"
    outside.mkdir()
    scope = WorkOwnerFileScope(
        workspace,
        owner="owner-a",
        session_id="session-a",
    )
    scope.private_root.mkdir(parents=True)
    (scope.private_root / "escape").symlink_to(outside, target_is_directory=True)
    resolver: Callable[[str | Path], Path] = getattr(scope, resolver_name)

    with pytest.raises(WorkFileScopeError) as exc_info:
        resolver("escape/secret.txt")
    assert exc_info.value.status_code in {400, 403}


def test_raw_absolute_paths_are_rejected_even_inside_owned_root(tmp_path: Path) -> None:
    workspace = tmp_path / "workspace"
    scope = WorkOwnerFileScope(
        workspace,
        owner="owner-a",
        session_id="session-a",
    )
    owned = scope.private_root / "input.xlsx"

    with pytest.raises(WorkFileScopeError) as exc_info:
        scope.resolve_routine_input(owned)

    assert exc_info.value.status_code == 400
    assert "absolute" in exc_info.value.detail.lower()


def test_secure_routine_read_rejects_symlinked_file(tmp_path: Path) -> None:
    workspace = tmp_path / "workspace"
    outside = tmp_path / "outside.txt"
    outside.write_text("outside secret", encoding="utf-8")
    scope = WorkOwnerFileScope(
        workspace,
        owner="owner-a",
        session_id="session-a",
    )
    scope.private_root.mkdir(parents=True)
    (scope.private_root / "input.txt").symlink_to(outside)

    with pytest.raises(WorkFileScopeError) as exc_info:
        scope.read_routine_input("input.txt")

    assert exc_info.value.status_code in {400, 403, 409}


def test_routine_path_resolver_rejects_symlink_even_within_owner_root(tmp_path: Path) -> None:
    workspace = tmp_path / "workspace"
    scope = WorkOwnerFileScope(
        workspace,
        owner="owner-a",
        session_id="session-a",
    )
    scope.private_root.mkdir(parents=True)
    target = scope.private_root / "target.xlsx"
    target.write_bytes(b"synthetic")
    (scope.private_root / "alias.xlsx").symlink_to(target.name)

    with pytest.raises(WorkFileScopeError, match="symlink"):
        scope.resolve_routine_input("alias.xlsx")


def test_secure_routine_read_survives_parent_directory_swap_without_escape(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    workspace = tmp_path / "workspace"
    outside = tmp_path / "outside"
    outside.mkdir()
    (outside / "secret.txt").write_text("outside secret", encoding="utf-8")
    scope = WorkOwnerFileScope(
        workspace,
        owner="owner-a",
        session_id="session-a",
    )
    scope.private_root.mkdir(parents=True)
    (scope.private_root / "secret.txt").write_text("owned data", encoding="utf-8")
    parked = scope.private_root.with_name(scope.private_root.name + "-parked")

    real_open = os.open
    swapped = False

    def swap_before_final_open(
        path: str | bytes | os.PathLike[str] | os.PathLike[bytes],
        flags: int,
        mode: int = 0o777,
        *,
        dir_fd: int | None = None,
    ) -> int:
        nonlocal swapped
        if path == "secret.txt" and dir_fd is not None and not swapped:
            scope.private_root.rename(parked)
            scope.private_root.symlink_to(outside, target_is_directory=True)
            swapped = True
        return real_open(path, flags, mode, dir_fd=dir_fd)

    monkeypatch.setattr("js_work.file_scope.os.open", swap_before_final_open)
    monkeypatch.setattr(os, "supports_dir_fd", {*os.supports_dir_fd, swap_before_final_open})

    snapshot = scope.read_routine_input("secret.txt")

    assert swapped
    assert snapshot.data == b"owned data"
    assert snapshot.data != b"outside secret"


def test_secure_routine_snapshot_binds_sha256_to_exact_bytes(tmp_path: Path) -> None:
    workspace = tmp_path / "workspace"
    scope = WorkOwnerFileScope(
        workspace,
        owner="owner-a",
        session_id="session-a",
    )
    scope.private_root.mkdir(parents=True)
    payload = b"authorized synthetic bytes"
    (scope.private_root / "input.bin").write_bytes(payload)

    snapshot = scope.read_routine_input("input.bin")

    assert snapshot.size == len(payload)
    assert snapshot.sha256 == sha256(payload).hexdigest()


def test_materialized_snapshot_cleanup_is_anchored_to_open_directory(
    tmp_path: Path,
) -> None:
    workspace = tmp_path / "workspace"
    outside = tmp_path / "outside"
    outside.mkdir()
    scope = WorkOwnerFileScope(
        workspace,
        owner="owner-a",
        session_id="session-a",
    )
    scope.private_root.mkdir(parents=True)
    (scope.private_root / "input.xlsx").write_bytes(b"synthetic snapshot")
    snapshot = scope.read_routine_input("input.xlsx")
    parked = scope.private_root.with_name(scope.private_root.name + "-parked")

    with (
        pytest.raises(WorkFileScopeError) as exc_info,
        scope.materialize_snapshot(snapshot) as staged,
    ):
        assert staged.name == "input.xlsx"
        assert staged.read_bytes() == snapshot.data
        scope.private_root.rename(parked)
        scope.private_root.symlink_to(outside, target_is_directory=True)

    assert exc_info.value.status_code == 409
    assert list(parked.glob(".work-input-snapshot-*")) == []
    assert list(outside.iterdir()) == []


def test_materialized_snapshot_parser_cannot_consume_swapped_parent_decoy(
    tmp_path: Path,
) -> None:
    workspace = tmp_path / "workspace"
    scope = WorkOwnerFileScope(
        workspace,
        owner="owner-a",
        session_id="session-a",
    )
    scope.private_root.mkdir(parents=True)
    source = scope.private_root / "input.xlsx"
    decoy = scope.private_root / "decoy.xlsx"
    original = Workbook()
    original.active["A1"] = "authorized"
    original.save(source)
    original.close()
    replacement = Workbook()
    replacement.active["A1"] = "decoy"
    replacement.save(decoy)
    replacement.close()
    snapshot = scope.read_routine_input("input.xlsx")
    observed: list[str] = []

    with (
        pytest.raises(WorkFileScopeError) as exc_info,
        scope.materialize_snapshot(snapshot) as staged,
    ):
        physical_parent = next(scope.private_root.glob(".work-input-snapshot-*"))
        parked = physical_parent.with_name(physical_parent.name + "-parked")
        physical_parent.rename(parked)
        physical_parent.mkdir()
        shutil.copy2(decoy, physical_parent / snapshot.name)
        workbook = load_workbook(staged, data_only=False, read_only=True)
        try:
            observed.append(workbook.active["A1"].value)
        finally:
            workbook.close()

    assert exc_info.value.status_code == 409
    assert observed == ["authorized"]


def test_secure_routine_snapshot_rejects_oversize_input(tmp_path: Path) -> None:
    scope = WorkOwnerFileScope(
        tmp_path / "workspace",
        owner="owner-a",
        session_id="session-a",
    )
    scope.private_root.mkdir(parents=True)
    (scope.private_root / "input.bin").write_bytes(b"1234")

    with pytest.raises(WorkFileScopeError) as exc_info:
        scope.read_routine_input("input.bin", max_bytes=3)

    assert exc_info.value.status_code == 413


def test_secure_routine_snapshot_rejects_special_file(tmp_path: Path) -> None:
    scope = WorkOwnerFileScope(
        tmp_path / "workspace",
        owner="owner-a",
        session_id="session-a",
    )
    scope.private_root.mkdir(parents=True)
    (scope.private_root / "input.bin").mkdir()

    with pytest.raises(WorkFileScopeError) as exc_info:
        scope.read_routine_input("input.bin")

    assert exc_info.value.status_code == 403


def test_secure_routine_snapshot_rejects_in_read_mutation(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    scope = WorkOwnerFileScope(
        tmp_path / "workspace",
        owner="owner-a",
        session_id="session-a",
    )
    scope.private_root.mkdir(parents=True)
    source = scope.private_root / "input.bin"
    source.write_bytes(b"authorized bytes")
    real_read = os.read
    mutated = False

    def mutate_after_first_read(descriptor: int, size: int) -> bytes:
        nonlocal mutated
        chunk = real_read(descriptor, size)
        if chunk and not mutated:
            source.write_bytes(b"replacement bytes")
            mutated = True
        return chunk

    monkeypatch.setattr("js_work.file_scope.os.read", mutate_after_first_read)

    with pytest.raises(WorkFileScopeError) as exc_info:
        scope.read_routine_input("input.bin")

    assert mutated is True
    assert exc_info.value.status_code == 409


def test_materialized_snapshot_rejects_tampered_payload(tmp_path: Path) -> None:
    scope = WorkOwnerFileScope(
        tmp_path / "workspace",
        owner="owner-a",
        session_id="session-a",
    )
    scope.private_root.mkdir(parents=True)
    (scope.private_root / "input.bin").write_bytes(b"authorized bytes")
    snapshot = scope.read_routine_input("input.bin")

    with (
        pytest.raises(WorkFileScopeError) as exc_info,
        scope.materialize_snapshot(replace(snapshot, data=b"replacement bytes")),
    ):
        raise AssertionError("tampered snapshot must not be published")

    assert exc_info.value.status_code == 409


def test_snapshot_bytes_fail_closed_when_integrity_metadata_is_tampered(
    tmp_path: Path,
) -> None:
    scope = WorkOwnerFileScope(
        tmp_path / "workspace",
        owner="owner-a",
        session_id="session-a",
    )
    scope.private_root.mkdir(parents=True)
    (scope.private_root / "input.bin").write_bytes(b"authorized bytes")
    snapshot = scope.read_routine_input("input.bin")

    with pytest.raises(WorkFileScopeError) as exc_info:
        replace(snapshot, data=b"replacement bytes").verified_data()

    assert exc_info.value.status_code == 409


def test_materialized_snapshot_never_clobbers_existing_stage_directory(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    scope = WorkOwnerFileScope(
        tmp_path / "workspace",
        owner="owner-a",
        session_id="session-a",
    )
    scope.private_root.mkdir(parents=True)
    (scope.private_root / "input.bin").write_bytes(b"authorized bytes")
    snapshot = scope.read_routine_input("input.bin")
    token = "a" * 32
    existing = scope.private_root / f".work-input-snapshot-{token}"
    existing.mkdir()
    sentinel = existing / "sentinel"
    sentinel.write_text("keep", encoding="utf-8")
    monkeypatch.setattr("js_work.file_scope.secrets.token_hex", lambda _size: token)

    with (
        pytest.raises(WorkFileScopeError) as exc_info,
        scope.materialize_snapshot(snapshot),
    ):
        raise AssertionError("existing staging directory must block publication")

    assert exc_info.value.status_code == 409
    assert sentinel.read_text(encoding="utf-8") == "keep"


def test_materialized_snapshot_preserves_existing_empty_stage_directory(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    scope = WorkOwnerFileScope(
        tmp_path / "workspace",
        owner="owner-a",
        session_id="session-a",
    )
    scope.private_root.mkdir(parents=True)
    (scope.private_root / "input.bin").write_bytes(b"authorized bytes")
    snapshot = scope.read_routine_input("input.bin")
    token = "b" * 32
    existing = scope.private_root / f".work-input-snapshot-{token}"
    existing.mkdir()
    monkeypatch.setattr("js_work.file_scope.secrets.token_hex", lambda _size: token)

    with (
        pytest.raises(WorkFileScopeError) as exc_info,
        scope.materialize_snapshot(snapshot),
    ):
        raise AssertionError("existing staging directory must block publication")

    assert exc_info.value.status_code == 409
    assert existing.is_dir()
