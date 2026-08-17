from __future__ import annotations

import asyncio
import json
import zipfile
from pathlib import Path
from typing import Any

from js.echo.attachment_gate import owner_slug, session_slug
from js.echo.turn_context import RuntimeContext, reset_runtime_context, set_runtime_context
from js_work.agent_factory import create_work_agent
from js_work.config import load_work_settings
from js_work.tools import WorkToolProfile


def _context(workspace: Path, state_dir: Path, owner: str) -> RuntimeContext:
    return RuntimeContext(
        product_id="js-work",
        channel="test",
        owner_key_hash=owner,
        session_id="session-a",
        run_id="run-a",
        role="user",
        profile="office",
        capabilities=(
            "csv_write",
            "excel_merge",
            "excel_read",
            "excel_write",
            "file_list",
            "file_write",
        ),
        workspace=workspace,
        state_dir=state_dir,
    )


def _execute(
    agent: Any,
    echo_tool_context: Any,
    *,
    tool_name: str,
    arguments: dict[str, Any],
    owner_root: Path,
) -> Any:
    return asyncio.run(
        agent.registry.execute(
            "run-a",
            tool_name,
            arguments,
            execution_context=echo_tool_context(
                run_id="run-a",
                tool_name=tool_name,
                arguments=arguments,
                fs_roots=(str(owner_root),),
                registry=agent.registry,
            ),
        )
    )


def _save_xlsx(path: Path, rows: list[list[Any]]) -> None:
    from openpyxl import Workbook

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()


def _add_external_relationship(path: Path) -> None:
    patched = path.with_name(f"{path.stem}.patched.xlsx")
    external = (
        b'<Relationship Id="rIdSynthetic" '
        b'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/externalLink" '
        b'Target="https://example.invalid/private.xlsx" TargetMode="External"/>'
    )
    with zipfile.ZipFile(path) as source, zipfile.ZipFile(patched, "w") as output:
        for info in source.infolist():
            payload = source.read(info.filename)
            if info.filename == "xl/_rels/workbook.xml.rels":
                payload = payload.replace(b"</Relationships>", external + b"</Relationships>")
            output.writestr(info, payload)
    patched.replace(path)


def test_work_file_handlers_are_scoped_to_runtime_owner(
    tmp_path: Path,
    echo_tool_context: Any,
) -> None:
    settings = load_work_settings(home=tmp_path)
    workspace = settings.workspace
    owner_a_root = workspace / "owners" / owner_slug("owner-a") / session_slug("session-a")
    owner_b_root = workspace / "owners" / owner_slug("owner-b") / session_slug("session-a")
    owner_a_root.mkdir(parents=True)
    owner_b_root.mkdir(parents=True)
    (owner_a_root / "a-only.txt").write_text("a", encoding="utf-8")
    (owner_b_root / "b-secret.txt").write_text("b", encoding="utf-8")
    agent = create_work_agent(settings=settings, profile=WorkToolProfile.OFFICE)

    token = set_runtime_context(_context(workspace, settings.state_dir, "owner-a"))
    try:
        result = _execute(
            agent,
            echo_tool_context,
            tool_name="file_list",
            arguments={},
            owner_root=owner_a_root,
        )
    finally:
        reset_runtime_context(token)
        asyncio.run(agent.close())

    assert result.success is True
    assert "a-only.txt" in result.output
    assert "b-secret.txt" not in result.output
    assert "owners" not in result.output


def test_work_write_handler_maps_relative_output_to_owner_root(
    tmp_path: Path,
    echo_tool_context: Any,
) -> None:
    settings = load_work_settings(home=tmp_path)
    agent = create_work_agent(settings=settings, profile=WorkToolProfile.OFFICE)
    owner_root = settings.workspace / "owners" / owner_slug("owner-a") / session_slug("session-a")

    token = set_runtime_context(_context(settings.workspace, settings.state_dir, "owner-a"))
    try:
        result = _execute(
            agent,
            echo_tool_context,
            tool_name="file_write",
            arguments={"path": "reports/result.txt", "content": "private"},
            owner_root=owner_root,
        )
    finally:
        reset_runtime_context(token)
        asyncio.run(agent.close())

    assert result.success is True, result.error
    assert result.metadata["path"] == "reports/result.txt"
    assert "reports/result.txt" in result.output
    assert str(settings.workspace) not in str(result)
    assert (
        settings.workspace
        / "owners"
        / owner_slug("owner-a")
        / session_slug("session-a")
        / "reports"
        / "result.txt"
    ).read_text(encoding="utf-8") == "private"
    assert not (settings.workspace / "reports" / "result.txt").exists()


def test_work_shared_office_tools_reject_overwrite_formula_and_unsafe_ooxml(
    tmp_path: Path,
    echo_tool_context: Any,
) -> None:
    settings = load_work_settings(home=tmp_path)
    agent = create_work_agent(settings=settings, profile=WorkToolProfile.OFFICE)
    owner_root = settings.workspace / "owners" / owner_slug("owner-a") / session_slug("session-a")
    existing = owner_root / "existing.xlsx"
    unsafe = owner_root / "unsafe.xlsx"
    _save_xlsx(existing, [["keep"], [1]])
    original = existing.read_bytes()
    _save_xlsx(unsafe, [["unsafe"]])
    _add_external_relationship(unsafe)

    token = set_runtime_context(_context(settings.workspace, settings.state_dir, "owner-a"))
    try:
        overwrite = _execute(
            agent,
            echo_tool_context,
            tool_name="excel_write",
            arguments={"path": "existing.xlsx", "data": json.dumps([["replace"]])},
            owner_root=owner_root,
        )
        formula = _execute(
            agent,
            echo_tool_context,
            tool_name="excel_write",
            arguments={"path": "formula.xlsx", "data": json.dumps([["=2+2"]])},
            owner_root=owner_root,
        )
        unsafe_read = _execute(
            agent,
            echo_tool_context,
            tool_name="excel_read",
            arguments={"path": "unsafe.xlsx"},
            owner_root=owner_root,
        )
        unsafe_csv = _execute(
            agent,
            echo_tool_context,
            tool_name="csv_write",
            arguments={"path": "formula.csv", "data": json.dumps([["=2+2"]])},
            owner_root=owner_root,
        )
    finally:
        reset_runtime_context(token)
        asyncio.run(agent.close())

    assert overwrite.success is False
    assert "already exists" in overwrite.error
    assert existing.read_bytes() == original
    assert formula.success is True
    assert (owner_root / "formula.xlsx").exists()
    from openpyxl import load_workbook

    written = load_workbook(owner_root / "formula.xlsx", data_only=False)
    try:
        cell = written.active["A1"]
        assert cell.value == "=2+2"
        assert cell.data_type == "s"
    finally:
        written.close()
    assert unsafe_read.success is False
    assert "external OOXML relationship" in unsafe_read.error
    assert unsafe_csv.success is True
    assert (owner_root / "formula.csv").read_text(encoding="utf-8").startswith("'=2+2")


def test_work_excel_merge_requires_and_publishes_a_new_output(
    tmp_path: Path,
    echo_tool_context: Any,
) -> None:
    settings = load_work_settings(home=tmp_path)
    agent = create_work_agent(settings=settings, profile=WorkToolProfile.OFFICE)
    owner_root = settings.workspace / "owners" / owner_slug("owner-a") / session_slug("session-a")
    _save_xlsx(owner_root / "source.xlsx", [["A", "B"], [1, 2]])
    _save_xlsx(owner_root / "template.xlsx", [["Template"]])
    template_before = (owner_root / "template.xlsx").read_bytes()

    token = set_runtime_context(_context(settings.workspace, settings.state_dir, "owner-a"))
    try:
        result = _execute(
            agent,
            echo_tool_context,
            tool_name="excel_merge",
            arguments={
                "source_path": "source.xlsx",
                "target_path": "template.xlsx",
                "output_path": "reports/merged.xlsx",
                "target_start_cell": "A2",
            },
            owner_root=owner_root,
        )
    finally:
        reset_runtime_context(token)
        asyncio.run(agent.close())

    assert result.success is True, result.error
    assert (owner_root / "reports" / "merged.xlsx").is_file()
    assert (owner_root / "template.xlsx").read_bytes() == template_before
    assert str(settings.workspace) not in str(result)


def test_work_office_handler_rejects_another_owner_input(
    tmp_path: Path,
    echo_tool_context: Any,
) -> None:
    from openpyxl import Workbook

    settings = load_work_settings(home=tmp_path)
    other_relative = (
        Path("owners") / owner_slug("owner-b") / session_slug("session-a") / "secret.xlsx"
    )
    other = settings.workspace / other_relative
    other.parent.mkdir(parents=True)
    workbook = Workbook()
    workbook.save(other)
    workbook.close()
    agent = create_work_agent(settings=settings, profile=WorkToolProfile.OFFICE)
    # Grant the registry the whole Work workspace so this test reaches the
    # Work owner-scope wrapper instead of being rejected by registry fs_roots.
    owner_root = settings.workspace

    token = set_runtime_context(_context(settings.workspace, settings.state_dir, "owner-a"))
    try:
        result = _execute(
            agent,
            echo_tool_context,
            tool_name="excel_read",
            arguments={"path": other_relative.as_posix()},
            owner_root=owner_root,
        )
    finally:
        reset_runtime_context(token)
        asyncio.run(agent.close())

    assert result.success is False
    assert result.error == "Routine input access denied"


def test_work_file_handler_fails_closed_without_work_runtime_context(
    tmp_path: Path,
    echo_tool_context: Any,
) -> None:
    settings = load_work_settings(home=tmp_path)
    (settings.workspace / "global-secret.txt").write_text("secret", encoding="utf-8")
    agent = create_work_agent(settings=settings, profile=WorkToolProfile.OFFICE)
    owner_root = settings.workspace / "owners" / owner_slug("owner-a") / session_slug("session-a")

    try:
        result = _execute(
            agent,
            echo_tool_context,
            tool_name="file_list",
            arguments={},
            owner_root=owner_root,
        )
    finally:
        asyncio.run(agent.close())

    assert result.success is False
    assert "runtime context" in result.error.lower()
    assert "global-secret.txt" not in result.output


def test_work_file_handler_fails_closed_for_non_work_runtime_context(
    tmp_path: Path,
    echo_tool_context: Any,
) -> None:
    settings = load_work_settings(home=tmp_path)
    agent = create_work_agent(settings=settings, profile=WorkToolProfile.OFFICE)
    owner_root = settings.workspace / "owners" / owner_slug("owner-a") / session_slug("session-a")
    wrong_context = _context(settings.workspace, settings.state_dir, "owner-a")
    object.__setattr__(wrong_context, "product_id", "js-agent")

    token = set_runtime_context(wrong_context)
    try:
        result = _execute(
            agent,
            echo_tool_context,
            tool_name="file_list",
            arguments={},
            owner_root=owner_root,
        )
    finally:
        reset_runtime_context(token)
        asyncio.run(agent.close())

    assert result.success is False
    assert "runtime context" in result.error.lower()
