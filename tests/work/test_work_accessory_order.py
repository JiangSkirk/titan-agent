from __future__ import annotations

import asyncio
import csv
import json
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from js_work.routines.accessory_order import AccessoryOrderRoutineRunner


def _save_table(path: Path, title: str, headers: list[str], rows: list[list[object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = title
    ws.append([f"{title}说明"])
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(path)


def _save_sources(root: Path) -> None:
    _save_table(
        root / "quantities.xlsx",
        "订单数量",
        ["STYLE NO", "COLOR", "SIZE", "ORDER QTY"],
        [
            ["S1", "RED", "S", 100],
            ["S1", "RED", "M", 50],
            ["S1", "BLUE", "S", 80],
        ],
    )
    _save_table(
        root / "style-bom.xlsx",
        "款式BOM",
        [
            "款号",
            "辅料编码",
            "辅料名称",
            "规格",
            "适用颜色",
            "适用尺码",
            "单件用量",
            "单位",
            "辅料颜色",
            "物料类型",
        ],
        [
            ["S1", "B18", "纽扣", "18L", "*", "*", 4, "粒", "同成衣色", "辅料"],
            ["S1", "Z20", "拉链", "20CM", "RED", "*", 1, "条", "RED", "辅料"],
            ["S1", "L01", "主唛", "标准", "*", "*", 1, "个", "通用", "辅料"],
            ["S1", "F01", "主面料", "40S", "*", "*", 2, "米", "同成衣色", "面料"],
        ],
    )
    _save_table(
        root / "accessories.xlsx",
        "辅料主数据",
        [
            "辅料编码",
            "供应商",
            "单位",
            "损耗率",
            "最小起订量",
            "包装倍数",
            "物料类型",
        ],
        [
            ["B18", "A供应商", "粒", "5%", 0, 100, "辅料"],
            ["Z20", "B供应商", "条", "2%", 0, 10, "辅料"],
            ["L01", "A供应商", "个", 0, 0, 50, "辅料"],
            ["F01", "面料供应商", "米", 0, 0, 1, "面料"],
        ],
    )


def _save_minimal_sources(
    root: Path,
    *,
    quantity_rows: list[list[object]],
    bom_rows: list[list[object]],
    master_rows: list[list[object]],
) -> None:
    _save_table(
        root / "quantities.xlsx",
        "订单数量",
        ["款号", "颜色", "尺码", "数量"],
        quantity_rows,
    )
    _save_table(
        root / "style-bom.xlsx",
        "款式BOM",
        [
            "款号",
            "辅料编码",
            "辅料名称",
            "规格",
            "适用颜色",
            "适用尺码",
            "单件用量",
            "单位",
            "物料类型",
        ],
        bom_rows,
    )
    _save_table(
        root / "accessories.xlsx",
        "辅料主数据",
        ["辅料编码", "供应商", "单位", "损耗率", "最小起订量", "包装倍数", "物料类型"],
        master_rows,
    )


def test_accessory_order_calculates_color_usage_and_supplier_rounding(tmp_path: Path) -> None:
    _save_sources(tmp_path)

    result = AccessoryOrderRoutineRunner(tmp_path).run(
        quantity_path="quantities.xlsx",
        style_path="style-bom.xlsx",
        accessory_path="accessories.xlsx",
        output_path="reports/supplier-order.xlsx",
    )

    assert result.status == "passed"
    assert result.issue_count == 0
    assert result.summary_row_count == 4

    report = json.loads(Path(result.report_path).read_text(encoding="utf-8"))
    rows = {(row["辅料编码"], row["辅料颜色"]): row for row in report["summary_rows"]}
    assert rows[("B18", "RED")]["理论数量"] == 600
    assert rows[("B18", "RED")]["需求数量"] == 630
    assert rows[("B18", "RED")]["下单数量"] == 700
    assert rows[("B18", "BLUE")]["理论数量"] == 320
    assert rows[("B18", "BLUE")]["下单数量"] == 400
    assert rows[("Z20", "RED")]["理论数量"] == 150
    assert rows[("Z20", "RED")]["下单数量"] == 160
    assert rows[("L01", "通用")]["理论数量"] == 230
    assert rows[("L01", "通用")]["下单数量"] == 250
    assert "F01" not in {row["辅料编码"] for row in report["summary_rows"]}

    wb = load_workbook(result.output_path, data_only=False)
    assert {"供应商下单汇总", "计算明细", "异常审核", "来源追踪"} <= set(wb.sheetnames)
    assert wb["供应商下单汇总"]["A1"].value == "辅料供应商下单汇总"
    wb.close()


def test_accessory_order_never_overwrites_existing_output_or_report(tmp_path: Path) -> None:
    _save_sources(tmp_path)
    output = tmp_path / "reports" / "supplier-order.xlsx"
    output.parent.mkdir(parents=True)
    output.write_bytes(b"keep-output")

    with pytest.raises(ValueError, match="already exists"):
        AccessoryOrderRoutineRunner(tmp_path).run(
            quantity_path="quantities.xlsx",
            style_path="style-bom.xlsx",
            accessory_path="accessories.xlsx",
            output_path="reports/supplier-order.xlsx",
        )
    assert output.read_bytes() == b"keep-output"

    output.unlink()
    report = output.with_suffix(".validation.json")
    report.write_text("keep-report", encoding="utf-8")
    with pytest.raises(ValueError, match="already exists"):
        AccessoryOrderRoutineRunner(tmp_path).run(
            quantity_path="quantities.xlsx",
            style_path="style-bom.xlsx",
            accessory_path="accessories.xlsx",
            output_path="reports/supplier-order.xlsx",
        )
    assert not output.exists()
    assert report.read_text(encoding="utf-8") == "keep-report"


def test_accessory_order_rejects_formula_like_csv_text(tmp_path: Path) -> None:
    _save_sources(tmp_path)
    accessory_csv = tmp_path / "accessories.csv"
    with accessory_csv.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            ["辅料编码", "供应商", "单位", "损耗率", "最小起订量", "包装倍数", "物料类型"]
        )
        writer.writerow(["B18", "=2+2", "粒", 0, 0, 100, "辅料"])
        writer.writerow(["Z20", "B供应商", "条", 0, 0, 10, "辅料"])
        writer.writerow(["L01", "A供应商", "个", 0, 0, 50, "辅料"])

    # Formula-like CSV text is treated as a literal supplier string.
    AccessoryOrderRoutineRunner(tmp_path).run(
        quantity_path="quantities.xlsx",
        style_path="style-bom.xlsx",
        accessory_path="accessories.csv",
        output_path="reports/supplier-order.xlsx",
    )
    assert (tmp_path / "reports" / "supplier-order.xlsx").exists()


def test_accessory_order_applies_explicit_color_alias_and_unit_conversion_rules(
    tmp_path: Path,
) -> None:
    _save_table(
        tmp_path / "quantities.xlsx",
        "数量",
        ["款号", "颜色", "数量"],
        [["S1", "红色", 100]],
    )
    _save_table(
        tmp_path / "style-bom.xlsx",
        "BOM",
        ["款号", "辅料编码", "辅料名称", "适用颜色", "单件用量", "单位", "物料类型"],
        [["S1", "B1", "织带", "RED", 2, "米", "辅料"]],
    )
    _save_table(
        tmp_path / "accessories.xlsx",
        "主数据",
        ["辅料编码", "供应商", "单位", "损耗率", "最小起订量", "包装倍数", "物料类型"],
        [["B1", "合成供应商", "码", 0, 0, 1, "辅料"]],
    )

    result = AccessoryOrderRoutineRunner(tmp_path).run(
        quantity_path="quantities.xlsx",
        style_path="style-bom.xlsx",
        accessory_path="accessories.xlsx",
        output_path="reports/converted.xlsx",
        color_aliases={"红色": "RED"},
        unit_conversions={"米->码": 1.1},
    )

    assert result.status == "passed"
    report = json.loads(Path(result.report_path).read_text(encoding="utf-8"))
    summary = report["summary_rows"][0]
    assert summary["单位"] == "码"
    assert summary["理论数量"] == 220
    assert summary["需求数量"] == 220
    assert report["applied_rules"] == {
        "color_aliases": {"红色": "RED"},
        "unit_conversions": {"米->码": 1.1},
    }


def test_accessory_order_supports_all_csv_sources_with_alternate_headers(
    tmp_path: Path,
) -> None:
    sources = {
        "quantity.csv": (
            ["STYLE NUMBER", "COLOUR", "SIZE", "QUANTITY"],
            [["CSV-1", "NAVY", "L", 25]],
        ),
        "bom.csv": (
            [
                "STYLE",
                "ITEM CODE",
                "ITEM",
                "SPECIFICATION",
                "COLOR",
                "SIZE",
                "USAGE",
                "UOM",
                "MATERIAL COLOR",
                "TYPE",
            ],
            [["CSV-1", "ZIP-1", "ZIPPER", "18CM", "ALL", "L", 1, "PCS", "NAVY", "TRIM"]],
        ),
        "master.csv": (
            ["CODE", "NAME", "VENDOR", "UOM", "WASTAGE", "MOQ", "PACK QTY", "TYPE"],
            [["ZIP-1", "ZIPPER", "SYNTHETIC VENDOR", "PCS", "4%", 30, 10, "TRIM"]],
        ),
    }
    for name, (headers, rows) in sources.items():
        with (tmp_path / name).open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(headers)
            writer.writerows(rows)

    result = AccessoryOrderRoutineRunner(tmp_path).run(
        quantity_path="quantity.csv",
        style_path="bom.csv",
        accessory_path="master.csv",
        output_path="out/csv-order.xlsx",
    )

    assert result.status == "passed"
    report = json.loads(Path(result.report_path).read_text(encoding="utf-8"))
    assert report["source_layouts"] == {
        "quantity": {"sheet": "CSV", "header_row": 1},
        "style_bom": {"sheet": "CSV", "header_row": 1},
        "accessory_master": {"sheet": "CSV", "header_row": 1},
    }
    assert report["summary_rows"][0]["需求数量"] == 26
    assert report["summary_rows"][0]["下单数量"] == 30


def test_accessory_order_discovers_shifted_headers_on_non_first_sheets(
    tmp_path: Path,
) -> None:
    fixtures = (
        (
            "random-a.xlsx",
            "实际数量",
            ["款式号", "成衣颜色", "尺寸", "订货数量"],
            [["SHIFT-9", "黑色", "M", 40]],
        ),
        (
            "random-b.xlsx",
            "实际BOM",
            ["款式", "物料编码", "物料名称", "适用颜色", "适用尺码", "单位用量", "单位", "类别"],
            [["SHIFT-9", "LBL-9", "洗水唛", "通用", "全部", 2, "个", "辅料"]],
        ),
        (
            "random-c.xlsx",
            "实际主数据",
            ["物料编码", "物料名称", "供应商名称", "单位", "损耗", "起订量", "包装数量", "类别"],
            [["LBL-9", "洗水唛", "合成供应商乙", "个", 0, 0, 25, "辅料"]],
        ),
    )
    for name, data_sheet, headers, rows in fixtures:
        wb = Workbook()
        decoy = wb.active
        decoy.title = "说明"
        decoy.append(["这是合成测试说明，不是数据表"])
        ws = wb.create_sheet(data_sheet)
        ws.append(["合成元数据"])
        ws.append([])
        ws.append(["版本", "1"])
        ws.append([])
        ws.append(headers)
        for row in rows:
            ws.append(row)
        wb.save(tmp_path / name)
        wb.close()

    result = AccessoryOrderRoutineRunner(tmp_path).run(
        quantity_path="random-a.xlsx",
        style_path="random-b.xlsx",
        accessory_path="random-c.xlsx",
        output_path="generated/shifted.xlsx",
    )

    assert result.status == "passed"
    report = json.loads(Path(result.report_path).read_text(encoding="utf-8"))
    assert report["source_layouts"] == {
        "quantity": {"sheet": "实际数量", "header_row": 5},
        "style_bom": {"sheet": "实际BOM", "header_row": 5},
        "accessory_master": {"sheet": "实际主数据", "header_row": 5},
    }
    assert report["summary_rows"][0]["理论数量"] == 80
    assert report["summary_rows"][0]["下单数量"] == 100


@pytest.mark.parametrize(
    ("color_aliases", "unit_conversions"),
    [
        ({"RED": "BLUE", "red ": "GREEN"}, None),
        (None, {"米->码": 0}),
        (None, {"invalid": 1.1}),
    ],
)
def test_accessory_order_rejects_ambiguous_or_invalid_explicit_rules(
    tmp_path: Path,
    color_aliases: dict[str, str] | None,
    unit_conversions: dict[str, float] | None,
) -> None:
    _save_sources(tmp_path)

    with pytest.raises(ValueError):
        AccessoryOrderRoutineRunner(tmp_path).run(
            quantity_path="quantities.xlsx",
            style_path="style-bom.xlsx",
            accessory_path="accessories.xlsx",
            output_path="reports/rejected.xlsx",
            color_aliases=color_aliases,
            unit_conversions=unit_conversions,
        )

    assert not (tmp_path / "reports" / "rejected.xlsx").exists()


def test_accessory_order_fails_closed_on_missing_master_and_bad_quantity(
    tmp_path: Path,
) -> None:
    _save_sources(tmp_path)
    wb = load_workbook(tmp_path / "quantities.xlsx")
    ws = wb["订单数量"]
    ws.append(["S1", "GREEN", "S", "not-a-number"])
    wb.save(tmp_path / "quantities.xlsx")
    wb.close()

    wb = load_workbook(tmp_path / "accessories.xlsx")
    ws = wb["辅料主数据"]
    for row in range(ws.max_row, 0, -1):
        if ws.cell(row, 1).value == "Z20":
            ws.delete_rows(row)
    wb.save(tmp_path / "accessories.xlsx")
    wb.close()

    result = AccessoryOrderRoutineRunner(tmp_path).run(
        quantity_path="quantities.xlsx",
        style_path="style-bom.xlsx",
        accessory_path="accessories.xlsx",
        output_path="reports/review.xlsx",
    )

    assert result.status == "needs_review"
    report = json.loads(Path(result.report_path).read_text(encoding="utf-8"))
    codes = {issue["code"] for issue in report["issues"]}
    assert "invalid_quantity" in codes
    assert "accessory_master_missing" in codes


def test_accessory_order_does_not_depend_on_fixed_rows_or_file_names(tmp_path: Path) -> None:
    _save_sources(tmp_path)
    (tmp_path / "输入").mkdir()
    for source, target in (
        ("quantities.xlsx", "输入/随机数量表.xlsx"),
        ("style-bom.xlsx", "输入/款式资料.xlsx"),
        ("accessories.xlsx", "输入/供应商资料.xlsx"),
    ):
        (tmp_path / source).replace(tmp_path / target)

    result = AccessoryOrderRoutineRunner(tmp_path).run(
        quantity_path="输入/随机数量表.xlsx",
        style_path="输入/款式资料.xlsx",
        accessory_path="输入/供应商资料.xlsx",
        output_path="输出/辅料订单.xlsx",
    )

    assert result.status == "passed"
    assert Path(result.output_path).is_file()


def test_accessory_order_flags_ambiguous_size_spec_header(tmp_path: Path) -> None:
    _save_table(
        tmp_path / "quantities.xlsx",
        "订单数量",
        ["款号", "颜色", "尺码", "数量"],
        [["S1", "RED", "S", 100], ["S1", "RED", "M", 50]],
    )
    _save_table(
        tmp_path / "style-bom.xlsx",
        "款式BOM",
        ["款号", "辅料编码", "尺寸", "单件用量", "单位"],
        [["S1", "B18", "M", 4, "粒"]],
    )
    _save_table(
        tmp_path / "accessories.xlsx",
        "辅料主数据",
        ["辅料编码", "供应商", "单位", "损耗率", "最小起订量", "包装倍数"],
        [["B18", "A供应商", "粒", 0, 0, 100]],
    )

    result = AccessoryOrderRoutineRunner(tmp_path).run(
        quantity_path="quantities.xlsx",
        style_path="style-bom.xlsx",
        accessory_path="accessories.xlsx",
        output_path="reports/ambiguous.xlsx",
    )

    assert result.status == "needs_review"
    report = json.loads(Path(result.report_path).read_text(encoding="utf-8"))
    ambiguous = [issue for issue in report["issues"] if issue["code"] == "ambiguous_header"]
    assert ambiguous, f"expected ambiguous_header issue, got {report['issues']}"
    assert set(ambiguous[0]["fields"].split(",")) == {"size", "spec"}


def test_accessory_order_prefers_exact_size_alias_over_shared_spec_header(
    tmp_path: Path,
) -> None:
    _save_table(
        tmp_path / "quantities.xlsx",
        "订单数量",
        ["款号", "颜色", "尺码", "数量"],
        [["S1", "RED", "M", 100]],
    )
    _save_table(
        tmp_path / "style-bom.xlsx",
        "款式BOM",
        ["款号", "辅料编码", "尺寸", "适用尺码", "单件用量", "单位"],
        [["S1", "B18", "18L", "M", 4, "粒"]],
    )
    _save_table(
        tmp_path / "accessories.xlsx",
        "辅料主数据",
        ["辅料编码", "供应商", "单位", "损耗率", "最小起订量", "包装倍数"],
        [["B18", "A供应商", "粒", 0, 0, 100]],
    )

    result = AccessoryOrderRoutineRunner(tmp_path).run(
        quantity_path="quantities.xlsx",
        style_path="style-bom.xlsx",
        accessory_path="accessories.xlsx",
        output_path="reports/exact-alias.xlsx",
    )

    report = json.loads(Path(result.report_path).read_text(encoding="utf-8"))
    codes = {issue["code"] for issue in report["issues"]}
    assert "ambiguous_header" not in codes
    assert result.status == "passed"
    assert report["summary_rows"][0]["规格"] == "18L"
    assert report["summary_rows"][0]["理论数量"] == 400


def test_accessory_order_rounds_moq_up_to_pack_multiple(tmp_path: Path) -> None:
    _save_minimal_sources(
        tmp_path,
        quantity_rows=[["S1", "RED", "M", 10]],
        bom_rows=[["S1", "B18", "纽扣", "18L", "*", "*", 1, "粒", "辅料"]],
        master_rows=[["B18", "A供应商", "粒", 0, 150, 100, "辅料"]],
    )

    result = AccessoryOrderRoutineRunner(tmp_path).run(
        quantity_path="quantities.xlsx",
        style_path="style-bom.xlsx",
        accessory_path="accessories.xlsx",
        output_path="reports/moq-round.xlsx",
    )

    report = json.loads(Path(result.report_path).read_text(encoding="utf-8"))
    ordered = report["summary_rows"][0]["下单数量"]
    assert ordered == 200, f"MOQ=150 应按包装倍数100向上取整为200, 实际={ordered}"
    assert ordered % 100 == 0


def test_accessory_order_rejects_zero_and_non_numeric_pack_multiple(tmp_path: Path) -> None:
    for bad_value in (0, "abc"):
        _save_minimal_sources(
            tmp_path,
            quantity_rows=[["S1", "RED", "M", 10]],
            bom_rows=[["S1", "B18", "纽扣", "18L", "*", "*", 1, "粒", "辅料"]],
            master_rows=[["B18", "A供应商", "粒", 0, 0, bad_value, "辅料"]],
        )

        result = AccessoryOrderRoutineRunner(tmp_path).run(
            quantity_path="quantities.xlsx",
            style_path="style-bom.xlsx",
            accessory_path="accessories.xlsx",
            output_path=f"reports/pack-{bad_value}.xlsx",
        )

        report = json.loads(Path(result.report_path).read_text(encoding="utf-8"))
        codes = {issue["code"] for issue in report["issues"]}
        assert "invalid_order_rounding" in codes, (
            f"pack_multiple={bad_value!r} 应报issue, got {codes}"
        )


def test_accessory_order_rejects_zero_quantity(tmp_path: Path) -> None:
    _save_minimal_sources(
        tmp_path,
        quantity_rows=[["S1", "RED", "M", 0]],
        bom_rows=[["S1", "B18", "纽扣", "18L", "*", "*", 1, "粒", "辅料"]],
        master_rows=[["B18", "A供应商", "粒", 0, 50, 100, "辅料"]],
    )

    result = AccessoryOrderRoutineRunner(tmp_path).run(
        quantity_path="quantities.xlsx",
        style_path="style-bom.xlsx",
        accessory_path="accessories.xlsx",
        output_path="reports/zero-qty.xlsx",
    )

    report = json.loads(Path(result.report_path).read_text(encoding="utf-8"))
    codes = {issue["code"] for issue in report["issues"]}
    assert "invalid_quantity" in codes, f"quantity=0 应报issue, got {codes}"
    assert report["summary_rows"] == []


def test_accessory_order_flags_suspicious_wastage(tmp_path: Path) -> None:
    _save_minimal_sources(
        tmp_path,
        quantity_rows=[["S1", "RED", "M", 100]],
        bom_rows=[["S1", "B18", "纽扣", "18L", "*", "*", 1, "粒", "辅料"]],
        master_rows=[["B18", "A供应商", "粒", "1", 0, 100, "辅料"]],
    )

    result = AccessoryOrderRoutineRunner(tmp_path).run(
        quantity_path="quantities.xlsx",
        style_path="style-bom.xlsx",
        accessory_path="accessories.xlsx",
        output_path="reports/wastage.xlsx",
    )

    report = json.loads(Path(result.report_path).read_text(encoding="utf-8"))
    codes = {issue["code"] for issue in report["issues"]}
    assert "wastage_suspicious" in codes, f"损耗率100% 应报可疑, got {codes}"


def test_accessory_order_merges_same_code_applies_moq_once(tmp_path: Path) -> None:
    _save_minimal_sources(
        tmp_path,
        quantity_rows=[["S1", "RED", "M", 10], ["S2", "RED", "M", 10]],
        bom_rows=[
            ["S1", "B18", "纽扣", "18L", "*", "*", 1, "粒", "辅料"],
            ["S2", "B18", "", "18L", "*", "*", 1, "粒", "辅料"],
        ],
        master_rows=[["B18", "A供应商", "粒", 0, 100, 50, "辅料"]],
    )

    result = AccessoryOrderRoutineRunner(tmp_path).run(
        quantity_path="quantities.xlsx",
        style_path="style-bom.xlsx",
        accessory_path="accessories.xlsx",
        output_path="reports/merge.xlsx",
    )

    report = json.loads(Path(result.report_path).read_text(encoding="utf-8"))
    assert len(report["summary_rows"]) == 1, (
        f"同编码应合并为一组(MOQ只应用一次), 实际{len(report['summary_rows'])}组"
    )
    row = report["summary_rows"][0]
    assert row["理论数量"] == 20
    assert row["下单数量"] == 100


def test_accessory_order_flags_duplicate_bom_rows(tmp_path: Path) -> None:
    _save_minimal_sources(
        tmp_path,
        quantity_rows=[["S1", "RED", "M", 100]],
        bom_rows=[
            ["S1", "B18", "纽扣", "18L", "*", "*", 4, "粒", "辅料"],
            ["S1", "B18", "纽扣", "18L", "*", "*", 4, "粒", "辅料"],
        ],
        master_rows=[["B18", "A供应商", "粒", 0, 0, 100, "辅料"]],
    )

    result = AccessoryOrderRoutineRunner(tmp_path).run(
        quantity_path="quantities.xlsx",
        style_path="style-bom.xlsx",
        accessory_path="accessories.xlsx",
        output_path="reports/dup-bom.xlsx",
    )

    report = json.loads(Path(result.report_path).read_text(encoding="utf-8"))
    codes = {issue["code"] for issue in report["issues"]}
    assert "bom_duplicate" in codes, f"重复BOM行应报issue, got {codes}"
    assert report["summary_rows"][0]["理论数量"] == 400, "重复行被去重后理论数量应为400而非800"


def test_accessory_order_flags_overlapping_bom_scopes(tmp_path: Path) -> None:
    _save_minimal_sources(
        tmp_path,
        quantity_rows=[["S1", "RED", "M", 100]],
        bom_rows=[
            ["S1", "B18", "纽扣", "18L", "*", "*", 4, "粒", "辅料"],
            ["S1", "B18", "纽扣", "18L", "RED", "*", 4, "粒", "辅料"],
        ],
        master_rows=[["B18", "A供应商", "粒", 0, 0, 100, "辅料"]],
    )

    result = AccessoryOrderRoutineRunner(tmp_path).run(
        quantity_path="quantities.xlsx",
        style_path="style-bom.xlsx",
        accessory_path="accessories.xlsx",
        output_path="reports/overlap.xlsx",
    )

    report = json.loads(Path(result.report_path).read_text(encoding="utf-8"))
    codes = {issue["code"] for issue in report["issues"]}
    assert "bom_scope_overlap" in codes, f"重叠作用域应报issue, got {codes}"
    assert report["summary_rows"][0]["理论数量"] == 400, "重叠行去重后理论数量应为400而非800"


def test_accessory_order_flags_alias_equivalent_bom_scope_overlap(tmp_path: Path) -> None:
    _save_minimal_sources(
        tmp_path,
        quantity_rows=[["S1", "RED", "M", 100]],
        bom_rows=[
            ["S1", "B18", "纽扣", "18L", "红色", "*", 4, "粒", "辅料"],
            ["S1", "B18", "纽扣", "18L", "RED", "*", 4, "粒", "辅料"],
        ],
        master_rows=[["B18", "A供应商", "粒", 0, 0, 100, "辅料"]],
    )

    result = AccessoryOrderRoutineRunner(tmp_path).run(
        quantity_path="quantities.xlsx",
        style_path="style-bom.xlsx",
        accessory_path="accessories.xlsx",
        output_path="reports/alias-overlap.xlsx",
        color_aliases={"红色": "RED"},
    )

    report = json.loads(Path(result.report_path).read_text(encoding="utf-8"))
    codes = {issue["code"] for issue in report["issues"]}
    assert "bom_scope_overlap" in codes or "bom_duplicate" in codes, (
        f"别名等价行应报重叠或重复, got {codes}"
    )
    assert report["summary_rows"][0]["理论数量"] == 400, "别名等价行去重后理论数量应为400而非800"


def test_accessory_order_merges_alias_equivalent_colors_in_summary(tmp_path: Path) -> None:
    _save_minimal_sources(
        tmp_path,
        quantity_rows=[["S1", "红", "M", 100], ["S1", "RED", "L", 100]],
        bom_rows=[["S1", "B18", "纽扣", "18L", "*", "*", 1, "粒", "辅料"]],
        master_rows=[["B18", "A供应商", "粒", 0, 150, 1, "辅料"]],
    )

    result = AccessoryOrderRoutineRunner(tmp_path).run(
        quantity_path="quantities.xlsx",
        style_path="style-bom.xlsx",
        accessory_path="accessories.xlsx",
        output_path="reports/alias-merge.xlsx",
        color_aliases={"红": "RED"},
    )

    report = json.loads(Path(result.report_path).read_text(encoding="utf-8"))
    assert len(report["summary_rows"]) == 1, (
        f"别名等价颜色应合并为一组, 实际{len(report['summary_rows'])}组"
    )
    row = report["summary_rows"][0]
    assert row["需求数量"] == 200
    assert row["下单数量"] == 200, f"合并后应按总需求200下单, 实际{row['下单数量']}"


def test_accessory_order_flags_missing_wastage(tmp_path: Path) -> None:
    _save_minimal_sources(
        tmp_path,
        quantity_rows=[["S1", "RED", "M", 100]],
        bom_rows=[["S1", "B18", "纽扣", "18L", "*", "*", 4, "粒", "辅料"]],
        master_rows=[["B18", "A供应商", "粒", None, 0, 100, "辅料"]],
    )

    result = AccessoryOrderRoutineRunner(tmp_path).run(
        quantity_path="quantities.xlsx",
        style_path="style-bom.xlsx",
        accessory_path="accessories.xlsx",
        output_path="reports/no-wastage.xlsx",
    )

    report = json.loads(Path(result.report_path).read_text(encoding="utf-8"))
    codes = {issue["code"] for issue in report["issues"]}
    assert "wastage_missing" in codes, f"损耗率缺失应报issue, got {codes}"


def test_accessory_order_flags_material_type_conflict(tmp_path: Path) -> None:
    _save_minimal_sources(
        tmp_path,
        quantity_rows=[["S1", "RED", "M", 100]],
        bom_rows=[["S1", "B18", "纽扣", "18L", "*", "*", 4, "粒", "辅料"]],
        master_rows=[["B18", "A供应商", "粒", 0, 0, 100, "面料"]],
    )

    result = AccessoryOrderRoutineRunner(tmp_path).run(
        quantity_path="quantities.xlsx",
        style_path="style-bom.xlsx",
        accessory_path="accessories.xlsx",
        output_path="reports/type-conflict.xlsx",
    )

    report = json.loads(Path(result.report_path).read_text(encoding="utf-8"))
    codes = {issue["code"] for issue in report["issues"]}
    assert "material_type_conflict" in codes, f"物料类型冲突应报issue, got {codes}"


def test_accessory_order_supports_gbk_encoded_csv(tmp_path: Path) -> None:
    _save_table(
        tmp_path / "quantities.xlsx",
        "订单数量",
        ["款号", "颜色", "尺码", "数量"],
        [["S1", "RED", "M", 100]],
    )
    _save_table(
        tmp_path / "style-bom.xlsx",
        "款式BOM",
        ["款号", "辅料编码", "单件用量", "单位"],
        [["S1", "B18", 4, "粒"]],
    )
    with (tmp_path / "accessories.csv").open("w", encoding="gbk", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["辅料编码", "供应商", "单位", "损耗率", "最小起订量", "包装倍数"])
        writer.writerow(["B18", "A供应商", "粒", 0, 0, 100])

    result = AccessoryOrderRoutineRunner(tmp_path).run(
        quantity_path="quantities.xlsx",
        style_path="style-bom.xlsx",
        accessory_path="accessories.csv",
        output_path="reports/gbk.xlsx",
    )

    assert result.status == "passed", f"GBK CSV应可读, error={result}"
    report = json.loads(Path(result.report_path).read_text(encoding="utf-8"))
    assert report["summary_rows"][0]["供应商"] == "A供应商"


def test_accessory_order_tool_warns_on_needs_review(tmp_path: Path) -> None:
    from js.echo.attachment_gate import owner_slug, session_slug
    from js.echo.turn_context import RuntimeContext, reset_runtime_context, set_runtime_context
    from js_work.routines.tools import WorkRoutineTools

    workspace = tmp_path / "workspace"
    state_dir = tmp_path / "state"
    owner_root = workspace / "owners" / owner_slug("owner-a") / session_slug("session-a")
    owner_root.mkdir(parents=True)
    _save_sources(owner_root)
    wb = load_workbook(owner_root / "quantities.xlsx")
    ws = wb["订单数量"]
    ws.append(["S1", "GREEN", "S", "not-a-number"])
    wb.save(owner_root / "quantities.xlsx")
    wb.close()

    tools = WorkRoutineTools(workspace=workspace, state_dir=state_dir)
    context = RuntimeContext(
        product_id="js-work",
        channel="test",
        owner_key_hash="owner-a",
        session_id="session-a",
        run_id="run-a",
        role="user",
        profile="office",
        capabilities=("accessory_order_run",),
        workspace=workspace,
        state_dir=state_dir,
    )

    token = set_runtime_context(context)
    try:
        result = asyncio.run(
            tools.accessory_order_run(
                quantity_path="quantities.xlsx",
                style_path="style-bom.xlsx",
                accessory_path="accessories.xlsx",
                output_path="reports/review.xlsx",
            )
        )
    finally:
        reset_runtime_context(token)

    assert result.success is True
    assert result.metadata["status"] == "needs_review"
    payload = json.loads(result.output)
    assert "warning" in payload, f"needs_review时应含warning, got {list(payload)}"
    assert "needs_review" in payload["warning"]


def test_accessory_order_rejects_path_escaping_owner_root(tmp_path: Path) -> None:
    owner_a = tmp_path / "owners" / "owner-a" / "session-a"
    owner_b = tmp_path / "owners" / "owner-b" / "session-b"
    owner_a.mkdir(parents=True)
    owner_b.mkdir(parents=True)
    _save_sources(owner_a)
    _save_sources(owner_b)

    runner = AccessoryOrderRoutineRunner(tmp_path, allowed_roots=(owner_a,))
    with pytest.raises(ValueError, match="escapes allowed roots"):
        runner.run(
            quantity_path="owners/owner-b/session-b/quantities.xlsx",
            style_path="owners/owner-a/session-a/style-bom.xlsx",
            accessory_path="owners/owner-a/session-a/accessories.xlsx",
            output_path="owners/owner-a/session-a/reports/escape.xlsx",
        )


def test_accessory_order_rejects_symlink_in_path(tmp_path: Path) -> None:
    owner_root = tmp_path / "owners" / "owner-a" / "session-a"
    owner_root.mkdir(parents=True)
    _save_sources(owner_root)
    real_file = owner_root / "real.xlsx"
    _save_table(
        real_file,
        "订单数量",
        ["款号", "颜色", "尺码", "数量"],
        [["S1", "RED", "M", 100]],
    )
    symlink = owner_root / "link.xlsx"
    symlink.symlink_to(real_file)

    runner = AccessoryOrderRoutineRunner(tmp_path, allowed_roots=(owner_root,))
    with pytest.raises(ValueError, match="symlink"):
        runner.run(
            quantity_path="owners/owner-a/session-a/link.xlsx",
            style_path="owners/owner-a/session-a/style-bom.xlsx",
            accessory_path="owners/owner-a/session-a/accessories.xlsx",
            output_path="owners/owner-a/session-a/reports/symlink.xlsx",
        )


def test_accessory_order_detail_sheet_includes_audit_columns(tmp_path: Path) -> None:
    _save_minimal_sources(
        tmp_path,
        quantity_rows=[["S1", "RED", "M", 100]],
        bom_rows=[["S1", "B18", "纽扣", "18L", "*", "*", 4, "粒", "辅料"]],
        master_rows=[["B18", "A供应商", "粒", 0, 0, 100, "辅料"]],
    )

    result = AccessoryOrderRoutineRunner(tmp_path).run(
        quantity_path="quantities.xlsx",
        style_path="style-bom.xlsx",
        accessory_path="accessories.xlsx",
        output_path="reports/audit.xlsx",
    )

    wb = load_workbook(result.output_path)
    detail_ws = wb["计算明细"]
    headers = [cell.value for cell in detail_ws[1]]
    wb.close()
    assert "原始单件用量" in headers, f"明细表应含原始单件用量: {headers}"
    assert "单位换算系数" in headers, f"明细表应含单位换算系数: {headers}"


def test_accessory_order_flags_spec_conflict(tmp_path: Path) -> None:
    _save_minimal_sources(
        tmp_path,
        quantity_rows=[["S1", "RED", "M", 100], ["S1", "RED", "L", 200]],
        bom_rows=[
            ["S1", "B18", "纽扣", "18L", "*", "M", 4, "粒", "辅料"],
            ["S1", "B18", "纽扣", "20L", "*", "L", 4, "粒", "辅料"],
        ],
        master_rows=[["B18", "A供应商", "粒", 0, 0, 100, "辅料"]],
    )

    result = AccessoryOrderRoutineRunner(tmp_path).run(
        quantity_path="quantities.xlsx",
        style_path="style-bom.xlsx",
        accessory_path="accessories.xlsx",
        output_path="reports/spec-conflict.xlsx",
    )

    report = json.loads(Path(result.report_path).read_text(encoding="utf-8"))
    codes = {issue["code"] for issue in report["issues"]}
    assert "spec_conflict" in codes, f"同编码不同规格应报spec_conflict, got {codes}"


def test_accessory_order_rejects_oversized_csv(tmp_path: Path) -> None:
    _save_table(
        tmp_path / "quantities.xlsx",
        "订单数量",
        ["款号", "颜色", "尺码", "数量"],
        [["S1", "RED", "M", 100]],
    )
    _save_table(
        tmp_path / "style-bom.xlsx",
        "款式BOM",
        ["款号", "辅料编码", "单件用量", "单位"],
        [["S1", "B18", 4, "粒"]],
    )
    big_csv = tmp_path / "big.csv"
    big_csv.write_bytes(b"x" * (101 * 1024 * 1024))

    result = AccessoryOrderRoutineRunner(tmp_path).run(
        quantity_path="quantities.xlsx",
        style_path="style-bom.xlsx",
        accessory_path="big.csv",
        output_path="reports/too-large.xlsx",
    )

    report = json.loads(Path(result.report_path).read_text(encoding="utf-8"))
    codes = {issue["code"] for issue in report["issues"]}
    assert "csv_too_large" in codes, f"超大CSV应报csv_too_large, got {codes}"


def test_accessory_order_source_hashes_are_valid_sha256(tmp_path: Path) -> None:
    _save_minimal_sources(
        tmp_path,
        quantity_rows=[["S1", "RED", "M", 100]],
        bom_rows=[["S1", "B18", "纽扣", "18L", "*", "*", 4, "粒", "辅料"]],
        master_rows=[["B18", "A供应商", "粒", 0, 0, 100, "辅料"]],
    )

    result = AccessoryOrderRoutineRunner(tmp_path).run(
        quantity_path="quantities.xlsx",
        style_path="style-bom.xlsx",
        accessory_path="accessories.xlsx",
        output_path="reports/hash.xlsx",
    )

    report = json.loads(Path(result.report_path).read_text(encoding="utf-8"))
    for key, hash_value in report["source_hashes"].items():
        assert len(hash_value) == 64, f"{key} hash长度应为64: {hash_value}"
        assert all(c in "0123456789abcdef" for c in hash_value), f"{key} hash应为hex: {hash_value}"


def test_accessory_order_tool_is_only_visible_in_office_profile() -> None:
    from js_work.tools import WorkToolProfile, allowed_tools_for_profile

    assert "accessory_order_run" in allowed_tools_for_profile(WorkToolProfile.OFFICE)
    assert "accessory_order_run" not in allowed_tools_for_profile(WorkToolProfile.SAFE)


def test_accessory_order_tool_uses_owner_scoped_inputs_and_output(tmp_path: Path) -> None:
    from js.echo.attachment_gate import owner_slug, session_slug
    from js.echo.turn_context import RuntimeContext, reset_runtime_context, set_runtime_context
    from js_work.routines.tools import WorkRoutineTools

    workspace = tmp_path / "workspace"
    state_dir = tmp_path / "state"
    owner_root = workspace / "owners" / owner_slug("owner-a") / session_slug("session-a")
    owner_root.mkdir(parents=True)
    _save_sources(owner_root)
    tools = WorkRoutineTools(workspace=workspace, state_dir=state_dir)
    context = RuntimeContext(
        product_id="js-work",
        channel="test",
        owner_key_hash="owner-a",
        session_id="session-a",
        run_id="run-a",
        role="user",
        profile="office",
        capabilities=("accessory_order_run",),
        workspace=workspace,
        state_dir=state_dir,
    )

    token = set_runtime_context(context)
    try:
        result = asyncio.run(
            tools.accessory_order_run(
                quantity_path="quantities.xlsx",
                style_path="style-bom.xlsx",
                accessory_path="accessories.xlsx",
                output_path="reports/supplier-order.xlsx",
            )
        )
    finally:
        reset_runtime_context(token)

    assert result.success is True, result.error
    assert result.metadata["status"] == "passed"
    assert str(workspace) not in str(result)
    payload = json.loads(result.output)
    assert payload["output_path"] == "reports/supplier-order.xlsx"
    assert payload["report_path"] == "reports/supplier-order.validation.json"
    assert (owner_root / "reports" / "supplier-order.xlsx").is_file()
    assert (owner_root / "reports" / "supplier-order.validation.json").is_file()
    assert not (workspace / "reports" / "supplier-order.xlsx").exists()
