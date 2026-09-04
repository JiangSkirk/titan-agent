"""Round 8.2 D: packing details formula safety via apply_work_cell_value."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from js_work.routines.office_safety import Formula, validate_restricted_formula
from js_work.routines.packing_details import PackingDetailsRoutineRunner


def _save_roll_source_with_formula_literal(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "发货"
    ws.append(["FABRICS", "PON.", "COLOR", "卷号", "米数", "卷号", "米数"])
    ws.append(["A", "P1", "=2+2", 1, 10, 2, 40])
    ws.append(["小计", None, None, "1卷", "=SUM(E2:E2)", None, None])
    wb.save(path)
    wb.close()


def _save_shipment_source_with_string_literals(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Shipment"
    ws.append(["Synthetic shipment source"])
    ws.append(["NOTE", "STYLE NO", "COLOUR", "ORDER QTY", "CARTONS", "N.W.", "G.W.", "CBM"])
    ws.append(["first", "+123", "RED", 10, 1, 1, 1, 0.1])
    ws.append(["second", "@name", "BLUE", 20, 2, 2, 2, 0.2])
    ws.append(["third", "-9", "GREEN", 30, 3, 3, 3, 0.3])
    wb.save(path)
    wb.close()


def _save_template(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "PACKING DETAILS"
    headers = ["FABRICS", "PON.", "COLOR", "ROLL NO", "QTY(M)", "ROLL NO", "QTY(M)"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    ws.merge_cells("A3:C3")
    ws["A3"] = "TOTAL "
    ws["D3"] = 1
    ws["E3"] = "=SUM(E2:E2)"
    wb.save(path)
    wb.close()


def _save_shipment_template(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "PACKING SUMMARY"
    ws.merge_cells("A1:G1")
    ws["A1"] = "PACKING SUMMARY"
    headers = ["款号", "颜色", "数量", "箱数", "净重", "毛重", "体积"]
    for column, header in enumerate(headers, start=1):
        ws.cell(2, column, header)
    ws["A4"] = "TOTAL"
    wb.save(path)
    wb.close()


def _three_templates(tmp_path: Path) -> tuple[str, str, str]:
    names: list[str] = []
    for index, title in enumerate(("PACKING DETAILS", "ROLL PACK", "SHIP LIST")):
        path = tmp_path / f"template_{index}.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = title
        ws["A1"] = "FABRICS"
        ws["B1"] = "PON."
        ws["C1"] = "COLOR"
        ws["D1"] = "ROLL NO"
        ws["E1"] = "QTY(M)"
        ws["F1"] = "ROLL NO"
        ws["G1"] = "QTY(M)"
        ws.merge_cells("A3:C3")
        ws["A3"] = "TOTAL "
        ws["D3"] = 1
        ws["E3"] = "=SUM(E2:E2)"
        wb.save(path)
        wb.close()
        names.append(path.name)
    return names[0], names[1], names[2]


def test_roll_manifest_formula_like_qty_stays_string_literal(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    template = tmp_path / "template.xlsx"
    output = tmp_path / "out.xlsx"
    _save_roll_source_with_formula_literal(source)
    _save_template(template)

    result = PackingDetailsRoutineRunner(tmp_path).run(
        source_path="source.xlsx",
        template_path="template.xlsx",
        output_path="out.xlsx",
    )
    assert result.mode == "roll_manifest"

    wb = load_workbook(output, data_only=False)
    try:
        cell = wb.active["C2"]
        assert cell.value == "=2+2"
        assert cell.data_type == "s"
    finally:
        wb.close()


def test_shipment_string_literals_stay_strings(tmp_path: Path) -> None:
    source = tmp_path / "ship_source.xlsx"
    template = tmp_path / "ship_template.xlsx"
    output = tmp_path / "ship_out.xlsx"
    _save_shipment_source_with_string_literals(source)
    _save_shipment_template(template)

    result = PackingDetailsRoutineRunner(tmp_path).run(
        source_path="ship_source.xlsx",
        template_path="ship_template.xlsx",
        output_path="ship_out.xlsx",
    )
    assert result.mode == "shipment_summary"

    wb = load_workbook(output, data_only=False)
    try:
        ws = wb.active
        assert ws is not None
        styles = {
            str(ws.cell(row, 1).value): ws.cell(row, 1)
            for row in range(3, ws.max_row)
            if ws.cell(row, 1).value not in (None, "TOTAL")
        }
        assert styles["+123"].data_type == "s"
        assert styles["@name"].data_type == "s"
        assert styles["-9"].data_type == "s"
    finally:
        wb.close()


@pytest.mark.parametrize(
    "expression",
    [
        '=IMAGE("http://example.test/x.png")',
        '=WEBSERVICE("http://example.test")',
        '=HYPERLINK("http://example.test","x")',
        "=UNKNOWNFUNC(1)",
    ],
)
def test_restricted_formula_rejects_unsafe_functions(expression: str) -> None:
    with pytest.raises(ValueError):
        validate_restricted_formula(expression)


def test_internal_sum_via_formula_type_allowed() -> None:
    Formula("=SUM(A1:A3)")


@pytest.mark.parametrize("template_name", ["template_0.xlsx", "template_1.xlsx", "template_2.xlsx"])
def test_roll_total_sum_is_real_formula_after_reload(tmp_path: Path, template_name: str) -> None:
    _three_templates(tmp_path)
    source = tmp_path / "source.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "发货"
    ws.append(["FABRICS", "PON.", "COLOR", "卷号", "米数", "卷号", "米数"])
    ws.append(["A", "P1", "WHITE", 1, 10, 2, 40])
    ws.append(["小计", None, None, "1卷", "=SUM(E2:E2)", None, None])
    wb.save(source)
    wb.close()
    output = tmp_path / f"out_{template_name}"

    result = PackingDetailsRoutineRunner(tmp_path).run(
        source_path="source.xlsx",
        template_path=template_name,
        output_path=output.name,
    )
    assert result.mode == "roll_manifest"

    wb = load_workbook(output, data_only=False)
    try:
        ws = wb.active
        assert ws is not None
        total_row = None
        for row in range(1, ws.max_row + 1):
            if ws.cell(row, 1).value and str(ws.cell(row, 1).value).strip().startswith("TOTAL"):
                total_row = row
                break
        assert total_row is not None
        formula_cell = ws.cell(total_row, 5)
        assert isinstance(formula_cell.value, str)
        assert formula_cell.value.startswith("=SUM(")
        assert formula_cell.data_type == "f"
    finally:
        wb.close()
