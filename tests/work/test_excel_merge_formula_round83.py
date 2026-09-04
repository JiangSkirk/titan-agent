"""Round 8.3 D: excel_merge must preserve formula-like string literals under Work."""

from __future__ import annotations

import asyncio
import json
from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook, load_workbook

from js.echo.attachment_gate import owner_slug, session_slug
from js.echo.turn_context import RuntimeContext, reset_runtime_context, set_runtime_context
from js_work.agent_factory import create_work_agent
from js_work.config import load_work_settings
from js_work.tools import WorkToolProfile


def _context(workspace: Path, state_dir: Path, owner: str) -> RuntimeContext:
    return RuntimeContext(
        product_id="js-work",
        channel="test",
        owner_key_hash=owner,
        session_id="session-a",
        run_id="run-a",
        role="user",
        profile="office",
        capabilities=(
            "csv_write",
            "excel_merge",
            "excel_read",
            "excel_write",
            "file_list",
            "file_write",
        ),
        workspace=workspace,
        state_dir=state_dir,
    )


def _save_literal_workbook(path: Path, rows: list[list[Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    for row_index, row in enumerate(rows, start=1):
        for col_index, value in enumerate(row, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            if isinstance(value, str) and value.lstrip()[:1] in {"=", "+", "-", "@"}:
                cell.data_type = "s"
    wb.save(path)
    wb.close()


def _execute(
    agent: Any,
    echo_tool_context: Any,
    *,
    tool_name: str,
    arguments: dict[str, Any],
    owner_root: Path,
) -> Any:
    return asyncio.run(
        agent.registry.execute(
            "run-a",
            tool_name,
            arguments,
            echo_mode="on",
            execution_context=echo_tool_context(
                run_id="run-a",
                tool_name=tool_name,
                arguments=arguments,
                owner_key_hash="owner-a",
                fs_roots=(str(owner_root),),
                registry=agent.registry,
            ),
        )
    )


def _merge_literals(
    tmp_path: Path,
    echo_tool_context: Any,
    *,
    source_rows: list[list[Any]],
    output_name: str = "reports/merged.xlsx",
) -> Path:
    settings = load_work_settings(home=tmp_path)
    agent = create_work_agent(settings=settings, profile=WorkToolProfile.OFFICE)
    owner_root = settings.workspace / "owners" / owner_slug("owner-a") / session_slug("session-a")
    _save_literal_workbook(owner_root / "source.xlsx", source_rows)
    _save_literal_workbook(owner_root / "template.xlsx", [["Template"]])
    arguments = {
        "source_path": "source.xlsx",
        "target_path": "template.xlsx",
        "output_path": output_name,
        "target_start_cell": "A1",
    }
    token = set_runtime_context(_context(settings.workspace, settings.state_dir, "owner-a"))
    try:
        result = _execute(
            agent,
            echo_tool_context,
            tool_name="excel_merge",
            arguments=arguments,
            owner_root=owner_root,
        )
    finally:
        reset_runtime_context(token)
        asyncio.run(agent.close())
    assert result.success is True, result.error
    return owner_root / output_name


def test_excel_merge_formula_like_string_stays_literal(
    tmp_path: Path,
    echo_tool_context: Any,
) -> None:
    output = _merge_literals(
        tmp_path,
        echo_tool_context,
        source_rows=[["label", "qty"], ["item", "=2+2"]],
    )
    wb = load_workbook(output, data_only=False)
    try:
        cell = wb.active["B2"]
        assert cell.value == "=2+2"
        assert cell.data_type == "s"
    finally:
        wb.close()


@pytest.mark.parametrize(
    ("row", "column", "expected"),
    [
        (2, 2, "+123"),
        (3, 2, "@name"),
        (4, 2, "-9"),
    ],
)
def test_excel_merge_leading_trigger_chars_stay_strings(
    tmp_path: Path,
    echo_tool_context: Any,
    row: int,
    column: int,
    expected: str,
) -> None:
    output = _merge_literals(
        tmp_path,
        echo_tool_context,
        source_rows=[
            ["style", "note"],
            ["first", "+123"],
            ["second", "@name"],
            ["third", "-9"],
        ],
        output_name=f"reports/literals_{expected.lstrip('@').replace('+', 'plus')}.xlsx",
    )
    wb = load_workbook(output, data_only=False)
    try:
        cell = wb.active.cell(row=row, column=column)
        assert cell.value == expected
        assert cell.data_type == "s"
    finally:
        wb.close()


@pytest.mark.parametrize(
    "expression",
    [
        '=IMAGE("http://example.test/x.png")',
        '=WEBSERVICE("http://example.test")',
        '=HYPERLINK("http://example.test","x")',
        "=UNKNOWNFUNC(1)",
    ],
)
def test_excel_merge_unsafe_formula_text_stays_literal_string(
    tmp_path: Path,
    echo_tool_context: Any,
    expression: str,
) -> None:
    output = _merge_literals(
        tmp_path,
        echo_tool_context,
        source_rows=[["payload"], [expression]],
        output_name=f"reports/unsafe_{abs(hash(expression))}.xlsx",
    )
    wb = load_workbook(output, data_only=False)
    try:
        cell = wb.active["A2"]
        assert cell.value == expression
        assert cell.data_type == "s"
    finally:
        wb.close()


@pytest.mark.parametrize(
    "expression",
    [
        '=IMAGE("http://example.test/x.png")',
        '=WEBSERVICE("http://example.test")',
        '=HYPERLINK("http://example.test","x")',
        "=UNKNOWNFUNC(1)",
    ],
)
def test_excel_write_rejects_typed_unsafe_formula(
    tmp_path: Path,
    echo_tool_context: Any,
    expression: str,
) -> None:
    settings = load_work_settings(home=tmp_path)
    agent = create_work_agent(settings=settings, profile=WorkToolProfile.OFFICE)
    owner_root = settings.workspace / "owners" / owner_slug("owner-a") / session_slug("session-a")
    arguments = {
        "path": "unsafe.xlsx",
        "data": json.dumps([[{"__work_formula__": expression}]]),
    }
    token = set_runtime_context(_context(settings.workspace, settings.state_dir, "owner-a"))
    try:
        result = _execute(
            agent,
            echo_tool_context,
            tool_name="excel_write",
            arguments=arguments,
            owner_root=owner_root,
        )
    finally:
        reset_runtime_context(token)
        asyncio.run(agent.close())
    assert result.success is False


def test_excel_write_typed_sum_formula_still_works(
    tmp_path: Path,
    echo_tool_context: Any,
) -> None:
    settings = load_work_settings(home=tmp_path)
    agent = create_work_agent(settings=settings, profile=WorkToolProfile.OFFICE)
    owner_root = settings.workspace / "owners" / owner_slug("owner-a") / session_slug("session-a")
    arguments = {
        "path": "sum.xlsx",
        "data": json.dumps([[2, 3, {"__work_formula__": "=SUM(A1:B1)"}]]),
    }
    token = set_runtime_context(_context(settings.workspace, settings.state_dir, "owner-a"))
    try:
        result = _execute(
            agent,
            echo_tool_context,
            tool_name="excel_write",
            arguments=arguments,
            owner_root=owner_root,
        )
    finally:
        reset_runtime_context(token)
        asyncio.run(agent.close())
    assert result.success is True, result.error
    wb = load_workbook(owner_root / "sum.xlsx", data_only=False)
    try:
        cell = wb.active["C1"]
        assert cell.value == "=SUM(A1:B1)"
        assert cell.data_type == "f"
    finally:
        wb.close()


@pytest.mark.asyncio
async def test_excel_merge_requires_signed_echo_lease(
    tmp_path: Path,
    echo_tool_context: Any,
) -> None:
    settings = load_work_settings(home=tmp_path)
    agent = create_work_agent(settings=settings, profile=WorkToolProfile.OFFICE)
    context = agent.echo_runtime.build_context(
        channel="test",
        owner_key_hash="owner-a",
        session_id="session-a",
        run_id="run-a",
    )
    owner_root = settings.workspace / "owners" / owner_slug("owner-a") / session_slug("session-a")
    _save_literal_workbook(owner_root / "source.xlsx", [["=2+2"]])
    _save_literal_workbook(owner_root / "template.xlsx", [["Template"]])
    arguments = {
        "source_path": "source.xlsx",
        "target_path": "template.xlsx",
        "output_path": "reports/leased.xlsx",
        "target_start_cell": "A1",
    }

    token = set_runtime_context(context)
    try:
        denied = await agent.registry.execute(
            "run-a",
            "excel_merge",
            arguments,
            echo_mode="on",
        )
        signed = echo_tool_context(
            run_id="run-a",
            tool_name="excel_merge",
            arguments=arguments,
            owner_key_hash="owner-a",
            fs_roots=tuple(str(root) for root in context.fs_roots),
            registry=agent.registry,
        )
        allowed = await agent.registry.execute(
            "run-a",
            "excel_merge",
            arguments,
            echo_mode="on",
            execution_context=signed,
        )
        replay = await agent.registry.execute(
            "run-a",
            "excel_merge",
            arguments,
            echo_mode="on",
            execution_context=signed,
        )
    finally:
        reset_runtime_context(token)
        await agent.close()

    assert denied.success is False
    assert "context required" in (denied.error or "").lower()
    assert allowed.success is True, allowed.error
    wb = load_workbook(owner_root / "reports" / "leased.xlsx", data_only=False)
    try:
        cell = wb.active["A1"]
        assert cell.value == "=2+2"
        assert cell.data_type == "s"
    finally:
        wb.close()
    assert replay.success is False
    assert "lease denied" in (replay.error or "").lower()
