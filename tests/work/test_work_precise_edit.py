from __future__ import annotations

import asyncio
import json
import os
import zipfile
from hashlib import sha256
from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from js.echo.attachment_gate import owner_slug, session_slug
from js.echo.turn_context import RuntimeContext, reset_runtime_context, set_runtime_context
from js_work.agent_factory import create_work_agent
from js_work.config import load_work_settings
from js_work.routines import precise_edit as precise_edit_module
from js_work.routines.precise_edit import PreciseExcelEditEngine
from js_work.routines.tools import WorkRoutineTools
from js_work.tools import WorkToolProfile, allowed_tools_for_profile


def _save_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "订单"
    sheet["A1"] = "款号"
    sheet["B1"] = "数量"
    sheet["A2"] = "SRT-001"
    sheet["B2"] = 12
    sheet["C2"] = "=B2*2"
    sheet["A4"] = "保持不变"
    sheet["A4"].font = Font(name="Arial", bold=True, color="00FF0000")
    sheet["A4"].fill = PatternFill("solid", fgColor="00FFF2CC")
    sheet.merge_cells("D1:E1")
    sheet["D1"] = "原合并标题"
    workbook.save(path)
    workbook.close()


def _digest(path: Path) -> str:
    return sha256(path.read_bytes()).hexdigest()


def _inject_zip_member(path: Path, name: str, payload: bytes) -> None:
    replacement = path.with_suffix(".replacement.xlsx")
    with (
        zipfile.ZipFile(path, "r") as source,
        zipfile.ZipFile(
            replacement,
            "w",
            zipfile.ZIP_DEFLATED,
        ) as output,
    ):
        for item in source.infolist():
            output.writestr(item, source.read(item.filename))
        output.writestr(name, payload)
    os.replace(replacement, path)


def _replace_zip_member(path: Path, name: str, transform: Any) -> None:
    replacement = path.with_suffix(".replacement.xlsx")
    with (
        zipfile.ZipFile(path, "r") as source,
        zipfile.ZipFile(
            replacement,
            "w",
            zipfile.ZIP_DEFLATED,
        ) as output,
    ):
        for item in source.infolist():
            payload = source.read(item.filename)
            output.writestr(item, transform(payload) if item.filename == name else payload)
    os.replace(replacement, path)


def test_precise_edit_writes_new_workbook_and_preserves_untouched_content(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "reports" / "edited.xlsx"
    _save_workbook(source)
    source_digest = _digest(source)

    report = PreciseExcelEditEngine(tmp_path).apply(
        source_path="source.xlsx",
        output_path="reports/edited.xlsx",
        operations=[
            {"op": "set_cell", "sheet": "订单", "cell": "B2", "value": 25},
            {
                "op": "copy_style",
                "sheet": "订单",
                "source_cell": "A4",
                "target_cell": "A5",
            },
            {
                "op": "set_number_format",
                "sheet": "订单",
                "cell": "B2",
                "number_format": "#,##0",
            },
            {"op": "set_row_height", "sheet": "订单", "row": 2, "height": 24},
            {
                "op": "set_column_width",
                "sheet": "订单",
                "column": "A",
                "width": 18,
            },
            {"op": "unmerge_cells", "sheet": "订单", "range": "D1:E1"},
            {"op": "merge_cells", "sheet": "订单", "range": "D2:E2"},
            {"op": "clear_cell", "sheet": "订单", "cell": "A2"},
        ],
    )

    assert source_digest == _digest(source)
    assert output.exists()
    assert report["status"] == "passed"
    assert report["operation_count"] == 8
    assert Path(report["validation_path"]).exists()

    source_wb = load_workbook(source, data_only=False)
    output_wb = load_workbook(output, data_only=False)
    source_sheet = source_wb["订单"]
    output_sheet = output_wb["订单"]
    assert source_sheet["A2"].value == "SRT-001"
    assert output_sheet["A2"].value is None
    assert output_sheet["B2"].value == 25
    assert output_sheet["B2"].number_format == "#,##0"
    assert output_sheet["C2"].value == "=B2*2"
    assert output_sheet["A4"].value == "保持不变"
    assert output_sheet["A4"]._style == source_sheet["A4"]._style
    assert output_sheet["A5"]._style == output_sheet["A4"]._style
    assert output_sheet.row_dimensions[2].height == 24
    assert output_sheet.column_dimensions["A"].width == 18
    assert "D1:E1" not in {str(item) for item in output_sheet.merged_cells.ranges}
    assert "D2:E2" in {str(item) for item in output_sheet.merged_cells.ranges}
    source_wb.close()
    output_wb.close()


def test_precise_edit_rejects_overwrite_and_external_formula(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    _save_workbook(source)
    engine = PreciseExcelEditEngine(tmp_path)

    with pytest.raises(ValueError, match="must not overwrite"):
        engine.apply(
            source_path="source.xlsx",
            output_path="source.xlsx",
            operations=[{"op": "set_cell", "sheet": "订单", "cell": "A1", "value": "x"}],
        )

    from js_work.routines.office_safety import Formula

    with pytest.raises(ValueError, match="external or executable formula|unsupported"):
        engine.apply(
            source_path="source.xlsx",
            output_path="out.xlsx",
            operations=[
                {
                    "op": "set_cell",
                    "sheet": "订单",
                    "cell": "A1",
                    "value": Formula('=WEBSERVICE("https://example.com")'),
                }
            ],
        )
    assert not (tmp_path / "out.xlsx").exists()

    # Plain strings remain literals even when they look like formulas.
    engine.apply(
        source_path="source.xlsx",
        output_path="safe-formula.xlsx",
        operations=[{"op": "set_cell", "sheet": "订单", "cell": "A1", "value": "=SUM(B1:B2)"}],
    )
    from openpyxl import load_workbook as _load

    wb = _load(tmp_path / "safe-formula.xlsx", data_only=False)
    try:
        cell = wb["订单"]["A1"]
        assert cell.value == "=SUM(B1:B2)"
        assert cell.data_type == "s"
    finally:
        wb.close()


def test_precise_edit_rejects_hardlink_alias_and_stale_source_hash(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    alias = tmp_path / "alias.xlsx"
    _save_workbook(source)
    os.link(source, alias)
    engine = PreciseExcelEditEngine(tmp_path)

    with pytest.raises(ValueError, match="must not overwrite"):
        engine.apply(
            source_path="source.xlsx",
            output_path="alias.xlsx",
            operations=[{"op": "set_cell", "sheet": "订单", "cell": "A1", "value": "x"}],
        )

    with pytest.raises(ValueError, match="source hash"):
        engine.apply(
            source_path="source.xlsx",
            output_path="new.xlsx",
            expected_source_sha256="0" * 64,
            operations=[{"op": "set_cell", "sheet": "订单", "cell": "A1", "value": "x"}],
        )
    assert not (tmp_path / "new.xlsx").exists()


def test_precise_edit_rejects_existing_output_and_symlink_source(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    _save_workbook(source)
    _save_workbook(output)
    output_digest = _digest(output)
    engine = PreciseExcelEditEngine(tmp_path)

    with pytest.raises(ValueError, match="already exists"):
        engine.apply(
            source_path="source.xlsx",
            output_path="output.xlsx",
            operations=[{"op": "set_cell", "sheet": "订单", "cell": "A1", "value": "x"}],
        )
    assert _digest(output) == output_digest

    alias = tmp_path / "alias.xlsx"
    alias.symlink_to(source.name)
    with pytest.raises(ValueError, match="symlink"):
        engine.apply(
            source_path="alias.xlsx",
            output_path="new.xlsx",
            operations=[{"op": "set_cell", "sheet": "订单", "cell": "A1", "value": "x"}],
        )
    assert not (tmp_path / "new.xlsx").exists()


def test_precise_edit_never_overwrites_existing_validation_report(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    validation = output.with_suffix(".validation.json")
    _save_workbook(source)
    validation.write_text('{"owner":"existing"}\n', encoding="utf-8")
    original_validation = validation.read_bytes()

    with pytest.raises(ValueError, match="validation report already exists"):
        PreciseExcelEditEngine(tmp_path).apply(
            source_path="source.xlsx",
            output_path="output.xlsx",
            operations=[{"op": "set_cell", "sheet": "订单", "cell": "A1", "value": "x"}],
        )

    assert validation.read_bytes() == original_validation
    assert not output.exists()


def test_precise_edit_does_not_clobber_output_created_during_publish(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    _save_workbook(source)
    real_link = os.link

    def competing_link(
        source_path: str | Path,
        target_path: str | Path,
        *,
        src_dir_fd: int | None = None,
        dst_dir_fd: int | None = None,
        follow_symlinks: bool = True,
    ) -> None:
        # The publish is descriptor-relative now: the competitor races the
        # atomic linkat() by pre-creating the same target *name*.
        if Path(target_path).name == output.name and not output.exists():
            output.write_bytes(b"concurrent-owner")
        real_link(
            source_path,
            target_path,
            src_dir_fd=src_dir_fd,
            dst_dir_fd=dst_dir_fd,
            follow_symlinks=follow_symlinks,
        )

    monkeypatch.setattr(precise_edit_module.os, "link", competing_link)

    with pytest.raises(ValueError, match="output workbook already exists"):
        PreciseExcelEditEngine(tmp_path).apply(
            source_path="source.xlsx",
            output_path="output.xlsx",
            operations=[{"op": "set_cell", "sheet": "订单", "cell": "A1", "value": "x"}],
        )

    assert output.read_bytes() == b"concurrent-owner"
    assert not output.with_suffix(".validation.json").exists()


def test_precise_edit_removes_staged_output_when_source_changes(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    _save_workbook(source)
    real_file_hash = precise_edit_module.file_hash
    source_hash_calls = 0

    def changed_source_hash(path: Path) -> str:
        nonlocal source_hash_calls
        if path == source:
            source_hash_calls += 1
            if source_hash_calls > 1:
                return "1" * 64
        return real_file_hash(path)

    monkeypatch.setattr(precise_edit_module, "file_hash", changed_source_hash)
    with pytest.raises(RuntimeError, match="source workbook changed"):
        PreciseExcelEditEngine(tmp_path).apply(
            source_path="source.xlsx",
            output_path="output.xlsx",
            operations=[{"op": "set_cell", "sheet": "订单", "cell": "A1", "value": "x"}],
        )

    assert not output.exists()
    assert not output.with_suffix(".validation.json").exists()


@pytest.mark.parametrize(
    ("member", "payload", "message"),
    [
        ("xl/externalLinks/externalLink1.xml", b"<externalLink/>", "external"),
        ("xl/vbaProject.bin", b"macro", "unsafe"),
        ("xl/media/compression-bomb.bin", b"0" * (5 * 1024 * 1024), "compression ratio"),
    ],
    ids=("external-link", "macro", "compression-ratio"),
)
def test_precise_edit_rejects_unsafe_ooxml_parts(
    tmp_path: Path,
    member: str,
    payload: bytes,
    message: str,
) -> None:
    source = tmp_path / "source.xlsx"
    _save_workbook(source)
    _inject_zip_member(source, member, payload)

    with pytest.raises(ValueError, match=message):
        PreciseExcelEditEngine(tmp_path).apply(
            source_path="source.xlsx",
            output_path="output.xlsx",
            operations=[{"op": "set_cell", "sheet": "订单", "cell": "A1", "value": "x"}],
        )
    assert not (tmp_path / "output.xlsx").exists()


def test_precise_edit_rejects_existing_dangerous_formula(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    _save_workbook(source)
    workbook = load_workbook(source)
    workbook["订单"]["A3"] = '=WEBSERVICE("https://example.com")'
    workbook.save(source)
    workbook.close()

    with pytest.raises(ValueError, match="external or executable formula"):
        PreciseExcelEditEngine(tmp_path).apply(
            source_path="source.xlsx",
            output_path="output.xlsx",
            operations=[{"op": "set_cell", "sheet": "订单", "cell": "A1", "value": "x"}],
        )

    assert not (tmp_path / "output.xlsx").exists()


def test_precise_edit_rejects_complex_ooxml_that_is_not_preserved(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    _save_workbook(source)
    _inject_zip_member(source, "xl/charts/chart1.xml", b"<chart-space />")

    with pytest.raises(ValueError, match="complex OOXML content changed"):
        PreciseExcelEditEngine(tmp_path).apply(
            source_path="source.xlsx",
            output_path="output.xlsx",
            operations=[{"op": "set_cell", "sheet": "订单", "cell": "A1", "value": "x"}],
        )

    assert not (tmp_path / "output.xlsx").exists()


def test_precise_edit_rejects_worksheet_extensions_that_cannot_roundtrip(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source.xlsx"
    _save_workbook(source)

    def add_extension(payload: bytes) -> bytes:
        return payload.replace(
            b"</worksheet>",
            b'<extLst><ext uri="unsupported" /></extLst></worksheet>',
        )

    _replace_zip_member(source, "xl/worksheets/sheet1.xml", add_extension)

    with pytest.raises(ValueError, match="unsupported worksheet OOXML feature"):
        PreciseExcelEditEngine(tmp_path).apply(
            source_path="source.xlsx",
            output_path="output.xlsx",
            operations=[{"op": "set_cell", "sheet": "订单", "cell": "A1", "value": "x"}],
        )

    assert not (tmp_path / "output.xlsx").exists()


def test_precise_edit_rejects_merge_that_discards_or_overlaps_content(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    _save_workbook(source)
    engine = PreciseExcelEditEngine(tmp_path)

    with pytest.raises(ValueError, match="discard"):
        engine.apply(
            source_path="source.xlsx",
            output_path="discard.xlsx",
            operations=[{"op": "merge_cells", "sheet": "订单", "range": "A1:B1"}],
        )
    with pytest.raises(ValueError, match="overlap"):
        engine.apply(
            source_path="source.xlsx",
            output_path="overlap.xlsx",
            operations=[{"op": "merge_cells", "sheet": "订单", "range": "E1:F1"}],
        )
    assert not (tmp_path / "discard.xlsx").exists()
    assert not (tmp_path / "overlap.xlsx").exists()


def test_precise_edit_tool_is_office_only_and_owner_scoped(tmp_path: Path) -> None:
    assert "excel_precise_edit" in allowed_tools_for_profile(WorkToolProfile.OFFICE)
    assert "excel_precise_edit" not in allowed_tools_for_profile(WorkToolProfile.SAFE)

    workspace = tmp_path / "workspace"
    private_root = workspace / "owners" / owner_slug("owner-a") / session_slug("session-a")
    source = private_root / "source.xlsx"
    _save_workbook(source)
    tools = WorkRoutineTools(workspace=workspace, state_dir=tmp_path / "state")
    context = RuntimeContext(
        product_id="js-work",
        channel="test",
        owner_key_hash="owner-a",
        session_id="session-a",
        run_id="run-a",
        role="user",
        profile="office",
        capabilities=("excel_precise_edit",),
        workspace=workspace,
        state_dir=tmp_path / "state",
    )
    token = set_runtime_context(context)
    try:
        result = asyncio.run(
            tools.excel_precise_edit(
                source_path="source.xlsx",
                output_path="reports/result.xlsx",
                operations=json.dumps(
                    [{"op": "set_cell", "sheet": "订单", "cell": "B2", "value": 99}],
                    ensure_ascii=False,
                ),
            )
        )
    finally:
        reset_runtime_context(token)

    assert result.success is True, result.error
    assert (private_root / "reports" / "result.xlsx").exists()
    assert not (workspace / "reports" / "result.xlsx").exists()


def test_precise_edit_tool_schema_documents_every_supported_operation(tmp_path: Path) -> None:
    specs = WorkRoutineTools(workspace=tmp_path, state_dir=tmp_path / "state").get_specs()
    spec = next(item for item in specs if item.name == "excel_precise_edit")
    operations = next(item for item in spec.parameters if item.name == "operations")

    for operation in (
        "set_cell",
        "clear_cell",
        "copy_style",
        "set_number_format",
        "set_row_height",
        "set_column_width",
        "merge_cells",
        "unmerge_cells",
    ):
        assert operation in operations.description
    assert "formulas" in operations.description


def test_precise_edit_tool_rejects_another_owners_source(tmp_path: Path) -> None:
    workspace = tmp_path / "workspace"
    other_relative = (
        Path("owners") / owner_slug("owner-b") / session_slug("session-a") / "source.xlsx"
    )
    other = workspace / other_relative
    _save_workbook(other)
    tools = WorkRoutineTools(workspace=workspace, state_dir=tmp_path / "state")
    context = RuntimeContext(
        product_id="js-work",
        channel="test",
        owner_key_hash="owner-a",
        session_id="session-a",
        run_id="run-a",
        role="user",
        profile="office",
        capabilities=("excel_precise_edit",),
        workspace=workspace,
        state_dir=tmp_path / "state",
    )
    token = set_runtime_context(context)
    try:
        result = asyncio.run(
            tools.excel_precise_edit(
                source_path=other_relative.as_posix(),
                output_path="reports/result.xlsx",
                operations=json.dumps(
                    [{"op": "set_cell", "sheet": "订单", "cell": "B2", "value": 99}]
                ),
            )
        )
    finally:
        reset_runtime_context(token)

    assert result.success is False
    assert "denied" in (result.error or "").lower()


@pytest.mark.asyncio
async def test_precise_edit_executes_only_with_single_use_signed_echo_lease(
    tmp_path: Path,
    echo_tool_context: Any,
) -> None:
    settings = load_work_settings(home=tmp_path)
    agent = create_work_agent(settings=settings, profile=WorkToolProfile.OFFICE)
    context = agent.echo_runtime.build_context(
        channel="test",
        owner_key_hash="owner-a",
        session_id="session-a",
        run_id="run-a",
    )
    owner_root = settings.workspace / "owners" / owner_slug("owner-a") / session_slug("session-a")
    _save_workbook(owner_root / "source.xlsx")
    arguments = {
        "source_path": "source.xlsx",
        "output_path": "reports/result.xlsx",
        "operations": json.dumps(
            [{"op": "set_cell", "sheet": "订单", "cell": "B2", "value": 88}],
            ensure_ascii=False,
        ),
    }

    token = set_runtime_context(context)
    try:
        denied = await agent.registry.execute(
            "run-a",
            "excel_precise_edit",
            arguments,
            echo_mode="on",
        )
        signed = echo_tool_context(
            run_id="run-a",
            tool_name="excel_precise_edit",
            arguments=arguments,
            owner_key_hash="owner-a",
            fs_roots=tuple(str(root) for root in context.fs_roots),
            registry=agent.registry,
        )
        allowed = await agent.registry.execute(
            "run-a",
            "excel_precise_edit",
            arguments,
            echo_mode="on",
            execution_context=signed,
        )
        replay = await agent.registry.execute(
            "run-a",
            "excel_precise_edit",
            arguments,
            echo_mode="on",
            execution_context=signed,
        )
    finally:
        reset_runtime_context(token)
        await agent.close()

    assert denied.success is False
    assert "context required" in (denied.error or "").lower()
    assert allowed.success is True, allowed.error
    assert replay.success is False
    assert "lease denied" in (replay.error or "").lower()
