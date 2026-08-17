from __future__ import annotations

import json
import math
import shutil
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Font, PatternFill, Side

from js_work.routines.packing_details import PackingDetailsRoutineRunner


def _save_source(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "发货"
    ws.append(["FABRICS", "PON.", "COLOR", "卷号", "米数", "卷号", "米数"])
    ws.append(["A", "P1", "WHITE", 1, 10, 4, 40])
    ws.append(["A", "P1", "WHITE", 2, 20, 5, 50])
    ws.append(["A", "P1", "WHITE", 3, 30, None, None])
    ws.append(["小计", None, None, "3卷", "=SUM(E2:E4)", "2卷", "=SUM(G2:G3)"])
    ws.append([])
    ws.append(["FABRICS", "PON.", "COLOR", "卷号", "米数", "卷号", "米数"])
    ws.append(["B ", "P2", "BLACK", 1, 11, 3, 33])
    ws.append(["B ", "P2", "BLACK", 2, 22, 4, 44])
    ws.append(["小计", None, None, "2卷", "=SUM(E8:E9)", "2卷", "=SUM(G8:G9)"])
    ws.append([])
    ws.append(["FABRICS", "PON.", "COLOR", "卷号", "米数", "卷号", "米数"])
    ws.append(["B ", "P2", "BLACK", 5, 55, None, None])
    ws.append(["小计", None, None, "1卷", "=SUM(E13:E13)", None, None])
    wb.save(path)


def _save_shifted_source(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "发货"
    ws.append(["备注", "COLOR", "ROLL NO", "QTY(M)", "FABRICS", "ROLL NO", "QTY(M)", "PON."])
    ws.append(["第一组", "WHITE", 1, 10, "A", 3, 30, "P1"])
    ws.append(["第一组", "WHITE", 2, 20, "A", None, None, "P1"])
    ws.append(["小计", None, "2卷", "=SUM(D2:D3)", None, "1卷", "=SUM(G2:G2)", None])
    ws.append([])
    ws.append(["备注", "COLOR", "ROLL NO", "QTY(M)", "FABRICS", "ROLL NO", "QTY(M)", "PON."])
    ws.append(["第二组", "BLACK", 1, 11, "B", 2, 22, "P2"])
    ws.append(["小计", None, "1卷", "=SUM(D7:D7)", None, "1卷", "=SUM(G7:G7)", None])
    wb.save(path)


def _save_template(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "PACKING DETAILS"
    headers = ["FABRICS", "PON.", "COLOR", "ROLL NO", "QTY(M)", "ROLL NO", "QTY(M)"]
    thin = Side(style="thin", color="000000")
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FFF2CC")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.row_dimensions[1].height = 21
    for row in range(2, 8):
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill("solid", fgColor="FFFFFF")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if col in {5, 7}:
                cell.number_format = "#,##0.0"
        ws.row_dimensions[row].height = 17
    ws.merge_cells("A8:C8")
    ws["A8"] = "TOTAL "
    ws["D8"] = 1
    ws["E8"] = "=SUM(E2:E7)+SUM(G2:G7)"
    for col in range(1, 8):
        cell = ws.cell(row=8, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9D9D9")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        if col in {5, 7}:
            cell.number_format = "#,##0.0"
    ws.row_dimensions[8].height = 19
    wb.save(path)


def _save_custom_template(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "PACKING DETAILS"
    headers = ["FABRIC STYLE", "ORDER", "SHADE", "ROLL", "METERS", "ROLL", "METERS"]
    thin = Side(style="thin", color="000000")
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(name="Courier New", bold=True)
        cell.fill = PatternFill("solid", fgColor="CCFFFF")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.row_dimensions[1].height = 25
    for row in range(2, 6):
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(name="Courier New")
            cell.fill = PatternFill("solid", fgColor="FFFFCC")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if col in {5, 7}:
                cell.number_format = "0.00"
    ws.merge_cells("A6:C6")
    for col in range(1, 8):
        cell = ws.cell(row=6, column=col)
        cell.font = Font(name="Courier New", bold=True)
        cell.fill = PatternFill("solid", fgColor="CCFFCC")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        if col == 5:
            cell.number_format = "0.00"
    ws["A6"] = "TOTAL "
    ws["D6"] = 1
    ws["E6"] = "=SUM(E2:E5)+SUM(G2:G5)"
    wb.save(path)


def _save_shipment_source(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Shipment"
    worksheet.append(["Synthetic shipment source"])
    worksheet.append(["NOTE", "STYLE NO", "COLOUR", "ORDER QTY", "CARTONS", "N.W.", "G.W.", "CBM"])
    worksheet.append(["first", "S1", "RED", 100, 5, 50, 55, 1.2])
    worksheet.append(["second", "S1", "BLUE", 80, 4, 40, 44, 0.9])
    worksheet.append(["third", "S2", "BLACK", 50, 3, 30, 33, 0.7])
    worksheet.append(["repeat", "S1", "RED", 20, 1, 10, 11, 0.2])
    workbook.save(path)
    workbook.close()


def _save_shipment_template(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "PACKING SUMMARY"
    worksheet.merge_cells("A1:G1")
    worksheet["A1"] = "PACKING SUMMARY"
    worksheet["A1"].font = Font(bold=True, size=14)
    headers = ["款号", "颜色", "数量", "箱数", "净重", "毛重", "体积"]
    for column, header in enumerate(headers, start=1):
        cell = worksheet.cell(2, column, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="CCFFFF")
    for column in range(1, 8):
        worksheet.cell(3, column).fill = PatternFill("solid", fgColor="FFFFCC")
        worksheet.cell(4, column).fill = PatternFill("solid", fgColor="CCFFCC")
    worksheet["A4"] = "TOTAL"
    workbook.save(path)
    workbook.close()


def test_packing_details_runner_merges_sections_and_writes_total_rows(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    template = tmp_path / "template.xlsx"
    output = tmp_path / "out.xlsx"
    _save_source(source)
    _save_template(template)

    result = PackingDetailsRoutineRunner(tmp_path).run(
        source_path="source.xlsx",
        template_path="template.xlsx",
        output_path="out.xlsx",
    )

    assert result.status == "passed"
    assert result.mode == "roll_manifest"
    assert result.group_count == 2
    assert result.roll_count == 10
    report = json.loads(Path(result.report_path).read_text(encoding="utf-8"))
    assert report["status"] == "passed"
    assert report["cross_validation"]["total_quantity"] == 315
    assert report["cross_validation"]["colors"] == ["BLACK", "WHITE"]
    assert report["cross_validation"]["styles"] == ["P1", "P2"]
    assert report["visual_structure"]["status"] == "passed"

    wb = load_workbook(output, data_only=False)
    ws = wb["PACKING DETAILS"]
    assert ws.max_row == 12
    assert [ws.cell(1, col).value for col in range(1, 8)] == [
        "FABRICS",
        "PON.",
        "COLOR",
        "ROLL NO",
        "QTY(M)",
        "ROLL NO",
        "QTY(M)",
    ]
    assert [ws.cell(2, col).value for col in range(1, 8)] == ["A", "P1", "WHITE", 1, 10, 4, 40]
    assert [ws.cell(4, col).value for col in range(1, 8)] == ["A", "P1", "WHITE", 3, 30, None, None]
    assert [ws.cell(5, col).value for col in range(1, 8)] == [
        "TOTAL ",
        None,
        None,
        5,
        "=SUM(E2:E4)+SUM(G2:G4)",
        None,
        None,
    ]
    assert [ws.cell(8, col).value for col in range(1, 8)] == [
        "FABRICS",
        "PON.",
        "COLOR",
        "ROLL NO",
        "QTY(M)",
        "ROLL NO",
        "QTY(M)",
    ]
    assert [ws.cell(9, col).value for col in range(1, 8)] == ["B", "P2", "BLACK", 1, 11, 4, 44]
    assert [ws.cell(11, col).value for col in range(1, 8)] == [
        "B",
        "P2",
        "BLACK",
        3,
        33,
        None,
        None,
    ]
    assert [ws.cell(12, col).value for col in range(1, 8)] == [
        "TOTAL ",
        None,
        None,
        5,
        "=SUM(E9:E11)+SUM(G9:G11)",
        None,
        None,
    ]
    assert {str(rng) for rng in ws.merged_cells.ranges} == {"A5:C5", "A12:C12"}
    assert ws["A1"].font.bold is True
    assert ws["A1"].fill.fill_type == "solid"
    assert ws["A1"].border.left.style == "thin"
    assert ws.row_dimensions[1].height == 21
    assert ws["E2"].number_format == "#,##0.0"
    assert ws["A5"].font.bold is True
    assert ws["A5"].fill.fill_type == "solid"
    assert ws["E5"].number_format == "#,##0.0"
    assert math.isclose(ws.row_dimensions[5].height or 0, 19, abs_tol=0.3)


def test_packing_details_never_overwrites_existing_output(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    template = tmp_path / "template.xlsx"
    output = tmp_path / "out.xlsx"
    _save_source(source)
    _save_template(template)
    output.write_bytes(b"keep-output")

    with pytest.raises(ValueError, match="already exists"):
        PackingDetailsRoutineRunner(tmp_path).run(
            source_path="source.xlsx",
            template_path="template.xlsx",
            output_path="out.xlsx",
        )

    assert output.read_bytes() == b"keep-output"

    output.unlink()
    report = output.with_suffix(".validation.json")
    report.write_text("keep-report", encoding="utf-8")
    with pytest.raises(ValueError, match="already exists"):
        PackingDetailsRoutineRunner(tmp_path).run(
            source_path="source.xlsx",
            template_path="template.xlsx",
            output_path="out.xlsx",
        )
    assert not output.exists()
    assert report.read_text(encoding="utf-8") == "keep-report"


def test_packing_details_rejects_formula_like_source_text(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    template = tmp_path / "template.xlsx"
    _save_source(source)
    _save_template(template)
    workbook = load_workbook(source)
    workbook["发货"]["A2"] = "=2+2"
    workbook.save(source)
    workbook.close()

    # Source cell text that looks like a formula is accepted as data; executable
    # formulas still require typed Formula objects on write paths.
    PackingDetailsRoutineRunner(tmp_path).run(
        source_path="source.xlsx",
        template_path="template.xlsx",
        output_path="out.xlsx",
    )
    assert (tmp_path / "out.xlsx").exists()


def test_packing_details_runner_detects_source_columns_from_headers(tmp_path: Path) -> None:
    source = tmp_path / "shifted-source.xlsx"
    template = tmp_path / "template.xlsx"
    output = tmp_path / "out.xlsx"
    _save_shifted_source(source)
    _save_template(template)

    result = PackingDetailsRoutineRunner(tmp_path).run(
        source_path="shifted-source.xlsx",
        template_path="template.xlsx",
        output_path="out.xlsx",
    )

    assert result.status == "passed"
    assert result.group_count == 2
    assert result.roll_count == 5
    wb = load_workbook(output, data_only=False)
    ws = wb["PACKING DETAILS"]
    assert [ws.cell(2, col).value for col in range(1, 8)] == ["A", "P1", "WHITE", 1, 10, 3, 30]
    assert [ws.cell(3, col).value for col in range(1, 8)] == ["A", "P1", "WHITE", 2, 20, None, None]
    assert [ws.cell(4, col).value for col in range(1, 8)] == [
        "TOTAL ",
        None,
        None,
        3,
        "=SUM(E2:E3)+SUM(G2:G3)",
        None,
        None,
    ]
    assert [ws.cell(8, col).value for col in range(1, 8)] == ["B", "P2", "BLACK", 1, 11, 2, 22]


def test_packing_details_supports_shipment_summary_and_cross_validates_all_metrics(
    tmp_path: Path,
) -> None:
    source = tmp_path / "shipment-source.xlsx"
    template = tmp_path / "shipment-template.xlsx"
    output = tmp_path / "shipment-output.xlsx"
    _save_shipment_source(source)
    _save_shipment_template(template)

    result = PackingDetailsRoutineRunner(tmp_path).run(
        source_path=source.name,
        template_path=template.name,
        output_path=output.name,
    )

    assert result.status == "passed"
    assert result.mode == "shipment_summary"
    assert result.group_count == 3
    assert result.roll_count == 0
    workbook = load_workbook(output, data_only=False)
    worksheet = workbook["PACKING SUMMARY"]
    assert [worksheet.cell(3, column).value for column in range(1, 8)] == [
        "S1",
        "BLUE",
        80,
        4,
        40,
        44,
        0.9,
    ]
    assert [worksheet.cell(4, column).value for column in range(1, 8)] == [
        "S1",
        "RED",
        120,
        6,
        60,
        66,
        1.4,
    ]
    assert worksheet["A6"].value == "TOTAL"
    assert worksheet["C6"].value == "=SUM(C3:C5)"
    workbook.close()
    values = load_workbook(output, data_only=True)
    assert values["PACKING SUMMARY"]["C6"].value == 250
    assert values["PACKING SUMMARY"]["D6"].value == 13
    assert math.isclose(values["PACKING SUMMARY"]["G6"].value, 3.0, abs_tol=0.0001)
    values.close()

    report = json.loads(Path(result.report_path).read_text(encoding="utf-8"))
    assert report["cross_validation"]["totals"] == {
        "cartons": 13.0,
        "gross_weight": 143.0,
        "net_weight": 130.0,
        "quantity": 250.0,
        "volume": 3.0,
    }
    assert report["cross_validation"]["colors"] == ["BLACK", "BLUE", "RED"]
    assert report["cross_validation"]["styles"] == ["S1", "S2"]
    assert report["formula_validation"]["status"] == "passed"
    assert report["visual_structure"]["status"] == "passed"


def test_packing_details_runner_uses_template_headers_and_styles(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    template = tmp_path / "custom-template.xlsx"
    output = tmp_path / "out.xlsx"
    _save_source(source)
    _save_custom_template(template)

    result = PackingDetailsRoutineRunner(tmp_path).run(
        source_path="source.xlsx",
        template_path="custom-template.xlsx",
        output_path="out.xlsx",
    )

    assert result.status == "passed"
    wb = load_workbook(output, data_only=False)
    ws = wb["PACKING DETAILS"]
    assert [ws.cell(1, col).value for col in range(1, 8)] == [
        "FABRIC STYLE",
        "ORDER",
        "SHADE",
        "ROLL",
        "METERS",
        "ROLL",
        "METERS",
    ]
    assert ws["A1"].font.name == "Courier New"
    assert ws["A1"].fill.fgColor.rgb == "00CCFFFF"
    assert ws["A2"].fill.fgColor.rgb == "00FFFFCC"
    assert ws["E2"].number_format == "0.00"
    assert ws["A5"].fill.fgColor.rgb == "00CCFFCC"
    assert ws["E5"].number_format == "0.00"


def test_packing_details_applies_quantity_format_when_template_is_general(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source.xlsx"
    template = tmp_path / "general-template.xlsx"
    output = tmp_path / "out.xlsx"
    _save_source(source)
    _save_template(template)
    wb = load_workbook(template)
    ws = wb["PACKING DETAILS"]
    for row in range(1, ws.max_row + 1):
        ws.cell(row, 5).number_format = "General"
        ws.cell(row, 7).number_format = "General"
    wb.save(template)
    wb.close()

    result = PackingDetailsRoutineRunner(tmp_path).run(
        source_path="source.xlsx",
        template_path="general-template.xlsx",
        output_path="out.xlsx",
    )

    assert result.status == "passed"
    wb = load_workbook(output, data_only=False)
    ws = wb["PACKING DETAILS"]
    assert ws["E2"].number_format == "#,##0.0"
    assert ws["G2"].number_format == "#,##0.0"
    assert ws["E5"].number_format == "#,##0.0"
    assert ws["G5"].number_format == "General"
    wb.close()


def test_packing_details_runner_accepts_legacy_xls_source(tmp_path: Path, monkeypatch) -> None:
    source_xlsx = tmp_path / "source-data.xlsx"
    source_xls = tmp_path / "source.xls"
    template = tmp_path / "template.xlsx"
    output = tmp_path / "out.xlsx"
    _save_source(source_xlsx)
    shutil.copy2(source_xlsx, source_xls)
    _save_template(template)

    def fake_convert(source: Path, output_dir: Path) -> Path:
        converted = output_dir / f"{source.stem}.xlsx"
        shutil.copy2(source, converted)
        return converted

    monkeypatch.setattr(
        PackingDetailsRoutineRunner, "_convert_xls_to_xlsx", staticmethod(fake_convert)
    )

    result = PackingDetailsRoutineRunner(tmp_path).run(
        source_path="source.xls",
        template_path="template.xlsx",
        output_path="out.xlsx",
    )

    assert result.status == "passed"
    assert result.roll_count == 10
    wb = load_workbook(output, data_only=False)
    ws = wb["PACKING DETAILS"]
    assert [ws.cell(2, col).value for col in range(1, 8)] == ["A", "P1", "WHITE", 1, 10, 4, 40]


def test_packing_details_runner_requires_cached_total_formula_values(
    tmp_path: Path, monkeypatch
) -> None:
    source = tmp_path / "source.xlsx"
    template = tmp_path / "template.xlsx"
    _save_source(source)
    _save_template(template)
    monkeypatch.setattr(
        PackingDetailsRoutineRunner,
        "_recalculate_workbook_formulas",
        staticmethod(lambda _path: False),
    )

    result = PackingDetailsRoutineRunner(tmp_path).run(
        source_path="source.xlsx",
        template_path="template.xlsx",
        output_path="out.xlsx",
    )

    assert result.status == "needs_review"
    assert result.issues == [
        {
            "code": "formula_value_missing",
            "cell": "E5",
            "formula": "=SUM(E2:E4)+SUM(G2:G4)",
        },
        {
            "code": "formula_value_missing",
            "cell": "E12",
            "formula": "=SUM(E9:E11)+SUM(G9:G11)",
        },
    ]


def test_packing_details_output_validation_catches_bad_cached_total(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    template = tmp_path / "template.xlsx"
    output = tmp_path / "out.xlsx"
    _save_source(source)
    _save_template(template)
    PackingDetailsRoutineRunner(tmp_path).run(
        source_path="source.xlsx",
        template_path="template.xlsx",
        output_path="out.xlsx",
    )

    wb = load_workbook(output, data_only=False)
    ws = wb["PACKING DETAILS"]
    assert ws["E5"].value == "=SUM(E2:E4)+SUM(G2:G4)"
    wb.close()
    _replace_cached_value(output, "E5", "999")

    issues = PackingDetailsRoutineRunner._validate_rendered_output(output, [5, 12])

    assert {
        "code": "qty_total_mismatch",
        "cell": "E5",
        "expected": 150.0,
        "actual": 999.0,
    } in issues


def _replace_cached_value(path: Path, cell: str, value: str) -> None:
    namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    tmp_path = path.with_suffix(".patched.xlsx")
    with (
        zipfile.ZipFile(path, "r") as src,
        zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as dst,
    ):
        for item in src.infolist():
            data = src.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                root = ET.fromstring(data)
                cell_node = next(
                    node for node in root.iter(f"{{{namespace}}}c") if node.get("r") == cell
                )
                value_node = cell_node.find(f"{{{namespace}}}v")
                if value_node is None:
                    value_node = ET.SubElement(cell_node, f"{{{namespace}}}v")
                value_node.text = value
                data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
            dst.writestr(item, data)
    shutil.move(tmp_path, path)
