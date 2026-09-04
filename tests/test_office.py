"""Tests for office document tools (Excel + PDF + CSV)."""

from pathlib import Path

import pytest
from openpyxl import load_workbook

from js.config import SecurityConfig, ToolLimits
from js.security.guard import BehaviorGuard
from js.tools.office import OfficeTools


class TestOfficeTools:
    @pytest.fixture
    def office(self, tmp_path: Path) -> OfficeTools:
        limits = ToolLimits()
        guard = BehaviorGuard(SecurityConfig(allow_workspace_delete=True), tmp_path)
        return OfficeTools(tmp_path, limits, guard)

    # ------------------------------------------------------------------
    # CSV
    # ------------------------------------------------------------------

    @pytest.mark.asyncio
    async def test_csv_write_and_read(self, office: OfficeTools, tmp_path: Path) -> None:
        result = await office.csv_write(
            "test.csv", data='[["Name","Age","City"],["Alice",30,"NYC"],["Bob",25,"LA"]]'
        )
        assert result.success

        result = await office.csv_read("test.csv")
        assert result.success
        assert "Alice" in result.output
        assert "NYC" in result.output
        assert result.metadata is not None
        assert result.metadata.get("rows") == 3
        assert result.metadata.get("columns") == 3

    @pytest.mark.asyncio
    async def test_csv_custom_delimiter(self, office: OfficeTools, tmp_path: Path) -> None:
        result = await office.csv_write("test.tsv", data='[["A","B"],["1","2"]]', delimiter="\t")
        assert result.success

        result = await office.csv_read("test.tsv", delimiter="\t")
        assert result.success
        assert "Alice" not in result.output  # Just sanity check
        rows = __import__("json").loads(result.output)
        assert rows[1] == ["1", "2"]

    @pytest.mark.asyncio
    async def test_csv_encoding_gbk(self, office: OfficeTools, tmp_path: Path) -> None:
        """Generic JS Agent keeps Python codec support for legacy Chinese CSVs."""
        (tmp_path / "chinese.csv").write_bytes("姓名,年龄\n张三,30\n".encode("gbk"))
        result = await office.csv_read("chinese.csv", encoding="gbk")
        assert result.success
        assert "张三" in result.output

    @pytest.mark.asyncio
    async def test_csv_read_missing_file(self, office: OfficeTools) -> None:
        result = await office.csv_read("missing.csv")
        assert not result.success
        assert "not found" in result.error.lower()

    @pytest.mark.asyncio
    async def test_csv_read_bad_encoding(self, office: OfficeTools, tmp_path: Path) -> None:
        # Write UTF-8 bytes then try to read as latin-1 (should work), then as ascii (should fail)
        (tmp_path / "mixed.csv").write_bytes("姓名,年龄\n张三,30\n".encode())
        result = await office.csv_read("mixed.csv", encoding="ascii")
        assert not result.success
        assert "encoding error" in result.error.lower()

    # ------------------------------------------------------------------
    # Excel
    # ------------------------------------------------------------------

    @pytest.mark.asyncio
    async def test_excel_create_and_read(self, office: OfficeTools, tmp_path: Path) -> None:
        result = await office.excel_create(
            "test.xlsx", sheet_name="Data", headers='["Name","Age","City"]'
        )
        assert result.success

        result = await office.excel_read("test.xlsx", sheet="Data")
        assert result.success
        assert "Name" in result.output
        assert "Age" in result.output

    @pytest.mark.asyncio
    async def test_excel_write_and_read(self, office: OfficeTools, tmp_path: Path) -> None:
        await office.excel_create("write_test.xlsx")
        data = '[["Alice", 30, "NYC"], ["Bob", 25, "LA"]]'
        result = await office.excel_write("write_test.xlsx", data=data, start_cell="A1")
        assert result.success

        result = await office.excel_read("write_test.xlsx")
        assert result.success
        assert "Alice" in result.output
        assert "Bob" in result.output

    @pytest.mark.asyncio
    async def test_excel_merge(self, office: OfficeTools, tmp_path: Path) -> None:
        # Create source file with data
        await office.excel_create("source.xlsx", headers='["ID","Value"]')
        await office.excel_write(
            "source.xlsx", data='[[1, "A"], [2, "B"], [3, "C"]]', start_cell="A2"
        )

        # Create target file with existing structure
        await office.excel_create("target.xlsx", headers='["X","Y","Z","A","B","C","D"]')
        await office.excel_write(
            "target.xlsx", data="[[10, 20, 30, 40, 50, 60, 70]]", start_cell="A2"
        )

        # Merge source data into target at column E (E2)
        result = await office.excel_merge(
            source_path="source.xlsx",
            target_path="target.xlsx",
            source_range="A2:B4",
            target_start_cell="E2",
            include_header=False,
        )
        assert result.success
        assert result.metadata is not None
        assert result.metadata.get("rows_copied") == 3

        # Verify target contents
        result = await office.excel_read("target.xlsx")
        assert result.success
        rows = __import__("json").loads(result.output)
        assert "1" in str(rows)
        assert "A" in str(rows)

    @pytest.mark.asyncio
    async def test_excel_read_missing_sheet(self, office: OfficeTools, tmp_path: Path) -> None:
        await office.excel_create("single.xlsx", sheet_name="Sheet1")
        result = await office.excel_read("single.xlsx", sheet="NonExistent")
        assert not result.success
        assert "sheet not found" in result.error.lower()

    @pytest.mark.asyncio
    async def test_excel_write_append_mode(self, office: OfficeTools, tmp_path: Path) -> None:
        await office.excel_create("append.xlsx")
        await office.excel_write("append.xlsx", data='[["row1"]]', start_cell="A1")
        result = await office.excel_write(
            "append.xlsx", data='[["row2"]]', start_cell="A1", append=True
        )
        assert result.success

        result = await office.excel_read("append.xlsx")
        rows = __import__("json").loads(result.output)
        assert len(rows) == 2

    # ------------------------------------------------------------------
    # PDF
    # ------------------------------------------------------------------

    @pytest.mark.asyncio
    async def test_pdf_generate(self, office: OfficeTools, tmp_path: Path) -> None:
        data = '[["Product", "Price", "Qty"], ["Apple", 1.5, 10], ["Banana", 0.8, 20]]'
        result = await office.pdf_generate("report.pdf", title="Sales Report", data=data)
        assert result.success
        assert (tmp_path / "report.pdf").exists()

    @pytest.mark.asyncio
    async def test_pdf_generate_empty_data(self, office: OfficeTools) -> None:
        result = await office.pdf_generate("empty.pdf", data="")
        assert not result.success
        assert "no data" in result.error.lower()

    # ------------------------------------------------------------------
    # Security
    # ------------------------------------------------------------------

    @pytest.mark.asyncio
    async def test_excel_path_escape_blocked(self, office: OfficeTools) -> None:
        result = await office.excel_read("../../../etc/passwd")
        assert not result.success
        assert "escapes workspace" in result.error or "blocked" in result.error.lower()

    @pytest.mark.asyncio
    async def test_csv_path_escape_blocked(self, office: OfficeTools) -> None:
        result = await office.csv_read("../../../etc/passwd")
        assert not result.success
        assert "escapes workspace" in result.error.lower()

    @pytest.mark.asyncio
    async def test_csv_write_escapes_formula_injection(
        self, office: OfficeTools, tmp_path: Path
    ) -> None:
        data = (
            '[["=cmd|\'/c calc\'!A1", "=HYPERLINK(\\"http://evil\\",\\"x\\")"],'
            ' ["+1", "-2", "@sum", "safe", "\'=already"], [1, 2.5, true]]'
        )
        result = await office.csv_write("inject.csv", data=data)
        assert result.success

        content = (tmp_path / "inject.csv").read_text(encoding="utf-8")
        assert "'=cmd|'/c calc'!A1" in content
        assert "'=HYPERLINK" in content
        assert "'+1" in content
        assert "'-2" in content
        assert "'@sum" in content
        # Already-escaped text must not be double-escaped.
        assert "''=already" not in content
        # Numbers pass through untouched.
        assert "1,2.5" in content

    @pytest.mark.asyncio
    async def test_excel_write_escapes_formula_injection(
        self, office: OfficeTools, tmp_path: Path
    ) -> None:
        data = '[["=cmd|\'/c calc\'!A1", "=HYPERLINK(\\"http://evil\\",\\"x\\")", -5]]'
        result = await office.excel_write("inject_write.xlsx", data=data, start_cell="A1")
        assert result.success

        wb = load_workbook(tmp_path / "inject_write.xlsx")
        ws = wb.active
        assert ws is not None
        assert ws.cell(row=1, column=1).value == "'=cmd|'/c calc'!A1"
        assert ws.cell(row=1, column=2).value == '\'=HYPERLINK("http://evil","x")'
        # Numeric cells keep their native type.
        assert ws.cell(row=1, column=3).value == -5
        wb.close()

    @pytest.mark.asyncio
    async def test_excel_create_escapes_formula_headers(
        self, office: OfficeTools, tmp_path: Path
    ) -> None:
        result = await office.excel_create(
            "inject_create.xlsx", headers='["=HYPERLINK(\\"http://evil\\",\\"x\\")", "ok"]'
        )
        assert result.success

        wb = load_workbook(tmp_path / "inject_create.xlsx")
        ws = wb.active
        assert ws is not None
        assert ws.cell(row=1, column=1).value == '\'=HYPERLINK("http://evil","x")'
        assert ws.cell(row=1, column=2).value == "ok"
        wb.close()

    @pytest.mark.asyncio
    async def test_excel_merge_escapes_formula_values(
        self, office: OfficeTools, tmp_path: Path
    ) -> None:
        # Seed the source workbook directly so it carries a raw formula-like string.
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.cell(row=1, column=1).value = "Name"
        # Force literal text; openpyxl would otherwise store "=..." as a formula.
        source_cell = ws.cell(row=2, column=1)
        source_cell.value = "=cmd|'/c calc'!A1"
        source_cell.data_type = "s"
        wb.save(tmp_path / "merge_source.xlsx")
        wb.close()

        await office.excel_create("merge_target.xlsx", headers='["Name"]')
        result = await office.excel_merge(
            source_path="merge_source.xlsx",
            target_path="merge_target.xlsx",
            source_range="A2:A2",
            target_start_cell="A2",
            include_header=False,
        )
        assert result.success

        wb = load_workbook(tmp_path / "merge_target.xlsx")
        ws = wb.active
        assert ws is not None
        assert ws.cell(row=2, column=1).value == "'=cmd|'/c calc'!A1"
        wb.close()

    @pytest.mark.asyncio
    async def test_csv_write_rejects_symlink_final(
        self, office: OfficeTools, tmp_path: Path
    ) -> None:
        victim = tmp_path / "victim.csv"
        victim.write_text("keep\n", encoding="utf-8")
        (tmp_path / "alias.csv").symlink_to(victim)
        result = await office.csv_write("alias.csv", data='[["hijacked"]]')
        assert not result.success
        assert victim.read_text(encoding="utf-8") == "keep\n"
