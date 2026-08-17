#!/usr/bin/env python3
"""Offline isolated-venv install E2E for JS Agent + JS Agent Work.

Builds wheel+sdist from the current source tree, installs each artifact into a
brand-new venv outside the repository using a local wheelhouse only
(``--no-index --find-links``), and verifies imports, CLI, tokenizer, pip check,
and a synthetic server E2E (HTTP + WebSocket + attachments + Work office).

Gate mode requires ``--source-digest HEX`` to match the current tree digest and
``--wheelhouse PATH`` with dependency wheels (see ``--prepare-wheelhouse``).

Usage:
    .venv/bin/python scripts/isolated_venv_e2e.py \\
        --source-digest "$(.venv/bin/python -c 'from pathlib import Path; \\
        from js.echo.ledger.release_gates import release_source_digest; \\
        print(release_source_digest(Path(\".\").resolve()))')" \\
        --wheelhouse /tmp/js-agent-wheelhouse \\
        --json dist/isolated_venv_e2e.json

    .venv/bin/python scripts/isolated_venv_e2e.py --prepare-wheelhouse /tmp/js-agent-wheelhouse

Exit code 0 only when every check passes.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import shutil
import signal
import subprocess
import sys
import tempfile
import time
from collections.abc import Iterator, Mapping, Sequence
from datetime import UTC, datetime
from importlib.metadata import distributions
from io import TextIOBase
from pathlib import Path
from typing import Any

REPO_ROOT = Path(__file__).resolve().parents[1]
ISOLATED_VENV_E2E_SCHEMA_VERSION = "isolated-venv-e2e-v8"
ISO_E2E_EVIDENCE_DIR_ENV = "JS_ISO_E2E_EVIDENCE_DIR"
ISO_E2E_SANDBOX_ENV = "JS_ISO_E2E_SANDBOX"
ISO_E2E_HOME_ENV = "JS_ISO_E2E_HOME"
ISO_E2E_PROVIDER_FAIL_ENV = "JS_ISO_E2E_PROVIDER_FAIL"
ISO_E2E_KIND_ENV = "JS_ISO_E2E_KIND"
SERVER_E2E_STEP_SUFFIX = "server HTTP+WS+attachment+work E2E"
VENV_PYTHON = REPO_ROOT / ".venv" / "bin" / "python"
SITE_PACKAGES = (
    REPO_ROOT
    / ".venv"
    / "lib"
    / f"python{sys.version_info.major}.{sys.version_info.minor}"
    / "site-packages"
)

IMPORT_CHECK = r"""
import hashlib, json, pathlib, sys, sysconfig
import js, js_work
errors = []
modules = {}
for module in (js, js_work):
    path = str(pathlib.Path(module.__file__).resolve())
    if "site-packages" not in path:
        errors.append(f"{module.__name__} loaded from non-site-packages: {path}")
    modules[module.__name__] = {
        "file": path,
        "file_sha256": hashlib.sha256(pathlib.Path(path).read_bytes()).hexdigest(),
        "site_packages": str(pathlib.Path(sysconfig.get_paths()["purelib"]).resolve()),
    }
import js.agent, js.echo.turn_runtime, js.echo.ledger.service
import js.models.router, js.models.permit
import js.security.approvals, js.web.server, js.tools.registry
import js_work.cli, js_work.routines.store, js_work.safe_output, js_work.routines
from js.echo.turn_runtime import EchoPulseRuntime  # noqa: F401
from js.security.approvals import ApprovalQueue  # noqa: F401
print(json.dumps({"modules": modules, "errors": errors}))
sys.exit(1 if errors else 0)
"""

TOKENIZER_CHECK = r"""
import json, os, sys
os.environ.pop("TIKTOKEN_CACHE_DIR", None)
from js.echo.context_tokenizer import tiktoken_counter_factory
counter = tiktoken_counter_factory("cl100k_base")
count = counter(b"hello offline world")
cache_dir = os.environ.get("TIKTOKEN_CACHE_DIR", "")
ok = count > 0 and "site-packages" in cache_dir
print(json.dumps({"count": count, "cache_dir": cache_dir, "ok": ok}))
sys.exit(0 if ok else 1)
"""

_PREPARE_SKIP_PREFIXES = ("js-agent", "js_agent", "-e")
_WHEEL_FILENAME_RE = re.compile(
    r"^(?P<name>[^-]+(?:-[^-]+)*?)-(?P<version>\d[^-]*(?:\.\d[^-]*)*?)"
    r"(?:-(?:cp\d+|py3|py2|py|abi3|none|any|universal).*?)?\.whl$",
    re.IGNORECASE,
)


def _utc_now() -> str:
    return datetime.now(tz=UTC).strftime("%Y-%m-%dT%H:%M:%SZ")


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _canonical_name(name: str) -> str:
    return re.sub(r"[-_.]+", "-", name).lower()


def _pip_cache_roots() -> list[Path]:
    roots: list[Path] = []
    if sys.platform == "darwin":
        roots.append(Path.home() / "Library/Caches/pip")
    else:
        roots.append(Path.home() / ".cache/pip")
    uv_cache = Path.home() / ".cache/uv"
    if uv_cache.is_dir():
        roots.append(uv_cache / "sdists-v9")
        roots.append(uv_cache / "archive-v0")
    try:
        proc = subprocess.run(
            [sys.executable, "-m", "pip", "cache", "dir"],
            check=True,
            capture_output=True,
            text=True,
        )
        cache_dir = proc.stdout.strip()
        if cache_dir:
            roots.append(Path(cache_dir))
    except (subprocess.CalledProcessError, FileNotFoundError):
        pass
    deduped: list[Path] = []
    seen: set[str] = set()
    for root in roots:
        key = str(root.resolve())
        if key not in seen and root.is_dir():
            seen.add(key)
            deduped.append(root)
    return deduped


def _iter_cached_wheels() -> Iterator[Path]:
    for root in _pip_cache_roots():
        for sub in ("wheels", "http-v2", "http"):
            candidate = root / sub
            if candidate.is_dir():
                yield from candidate.rglob("*.whl")


def _match_wheel_filename(path: Path, *, name: str, version: str) -> bool:
    match = _WHEEL_FILENAME_RE.match(path.name)
    if match is None:
        return False
    return (
        _canonical_name(match.group("name")) == _canonical_name(name)
        and match.group("version") == version
    )


def _is_product_wheel(path: Path) -> bool:
    match = _WHEEL_FILENAME_RE.match(path.name)
    return match is not None and _canonical_name(match.group("name")) == "js-agent"


def _find_cached_wheel(name: str, version: str) -> Path | None:
    matches = [
        path
        for path in _iter_cached_wheels()
        if _match_wheel_filename(path, name=name, version=version)
    ]
    if not matches:
        return None
    return max(matches, key=lambda item: item.stat().st_mtime)


def _installed_distributions() -> list[tuple[str, str]]:
    if not SITE_PACKAGES.is_dir():
        raise RuntimeError(f"Repository venv site-packages not found: {SITE_PACKAGES}")
    original_path = sys.path[:]
    try:
        sys.path.insert(0, str(SITE_PACKAGES))
        items: list[tuple[str, str]] = []
        for dist in distributions(path=[str(SITE_PACKAGES)]):
            name = dist.metadata.get("Name") or dist.name
            if not name:
                continue
            canonical = _canonical_name(name)
            if any(canonical.startswith(prefix) for prefix in _PREPARE_SKIP_PREFIXES):
                continue
            version = dist.version
            if not version:
                continue
            items.append((name, version))
        return sorted({(n, v) for n, v in items}, key=lambda item: _canonical_name(item[0]))
    finally:
        sys.path[:] = original_path


def _critical_dependency_names() -> set[str]:
    try:
        import tomllib
    except ModuleNotFoundError:  # pragma: no cover - py312+
        import tomli as tomllib  # type: ignore[no-redef]

    data = tomllib.loads((REPO_ROOT / "pyproject.toml").read_text(encoding="utf-8"))
    project = data.get("project", {})
    names: set[str] = set()
    for dep in project.get("dependencies", []):
        token = str(dep).split(";", 1)[0].strip()
        match = re.match(r"^([A-Za-z0-9_.-]+)", token)
        if match:
            names.add(_canonical_name(match.group(1)))
    optional = project.get("optional-dependencies", {})
    for extra in ("echo-tokenizer", "office"):
        for dep in optional.get(extra, []):
            token = str(dep).split(";", 1)[0].strip()
            match = re.match(r"^([A-Za-z0-9_.-]+)", token)
            if match:
                names.add(_canonical_name(match.group(1)))
    names.add("hatchling")
    names.add("build")
    return names


def prepare_wheelhouse(target: Path) -> dict[str, Any]:
    """Copy cached wheels for packages installed in the repo ``.venv``."""
    target.mkdir(parents=True, exist_ok=True)
    historical_dir = target / "historical"
    archived: list[dict[str, str]] = []
    for stale in tuple(target.glob("*.whl")):
        if _is_product_wheel(stale):
            historical_dir.mkdir(parents=True, exist_ok=True)
            destination = historical_dir / stale.name
            if destination.exists():
                destination.unlink()
            shutil.move(stale, destination)
            archived.append({"wheel": destination.name, "sha256": _sha256(destination)})
    copied: list[dict[str, str]] = []
    missing: list[str] = []
    critical = _critical_dependency_names()
    missing_critical: list[str] = []

    for name, version in _installed_distributions():
        wheel = _find_cached_wheel(name, version)
        if wheel is None:
            label = f"{name}=={version}"
            missing.append(label)
            if _canonical_name(name) in critical:
                missing_critical.append(label)
            continue
        destination = target / wheel.name
        if not destination.exists():
            shutil.copy2(wheel, destination)
        copied.append(
            {
                "name": name,
                "version": version,
                "wheel": wheel.name,
                "sha256": _sha256(destination),
            }
        )

    ok = not missing_critical
    report = {
        "wheelhouse": str(target.resolve()),
        "ok": ok,
        "copied_count": len(copied),
        "missing_count": len(missing),
        "missing_critical": missing_critical,
        "missing": missing,
        "copied": copied,
        "archived_same_version": archived,
    }
    (target / "MANIFEST.json").write_text(
        json.dumps({"schema_version": "wheelhouse-manifest-v1", "wheels": copied}, indent=2) + "\n",
        encoding="utf-8",
    )
    if archived:
        (historical_dir / "MANIFEST.json").write_text(
            json.dumps(
                {"schema_version": "wheelhouse-historical-manifest-v1", "wheels": archived},
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )
    if not ok:
        print(
            "WHEELHOUSE_PREPARE FAILED: missing critical dependency wheels:",
            ", ".join(missing_critical),
            file=sys.stderr,
        )
        print(
            "Populate pip/uv wheel caches (e.g. pip download -d <wheelhouse> with deps once), "
            "then retry.",
            file=sys.stderr,
        )
    else:
        print(f"WHEELHOUSE_PREPARE OK: copied {len(copied)} wheels to {target}")
    return report


def _validate_wheelhouse(wheelhouse: Path) -> None:
    if not wheelhouse.is_dir():
        raise SystemExit(f"wheelhouse does not exist: {wheelhouse}")
    wheels = list(wheelhouse.glob("*.whl"))
    if not wheels:
        raise SystemExit(f"wheelhouse contains no .whl files: {wheelhouse}")
    product_wheels = sorted(path.name for path in wheels if _is_product_wheel(path))
    if product_wheels:
        raise SystemExit(
            "wheelhouse must contain dependencies only; remove product wheels: "
            + ", ".join(product_wheels)
        )


class _Tee(TextIOBase):
    def __init__(self, *streams: TextIOBase) -> None:
        self._streams = streams

    def write(self, data: str) -> int:
        for stream in self._streams:
            stream.write(data)
        return len(data)

    def flush(self) -> None:
        for stream in self._streams:
            stream.flush()


def _base_env(
    home: Path,
    *,
    wheelhouse: Path | None = None,
    config_path: Path | None = None,
    evidence_dir: Path | None = None,
) -> dict[str, str]:
    tmpdir = home / "tmp"
    xdg_config = home / ".config"
    xdg_cache = home / ".cache"
    xdg_state = home / ".local" / "state"
    for directory in (home, tmpdir, xdg_config, xdg_cache, xdg_state):
        directory.mkdir(parents=True, exist_ok=True)
    env = {
        "PATH": os.environ.get("PATH", "/usr/bin:/bin"),
        "HOME": str(home.resolve()),
        "TMPDIR": str(tmpdir.resolve()),
        "XDG_CONFIG_HOME": str(xdg_config.resolve()),
        "XDG_CACHE_HOME": str(xdg_cache.resolve()),
        "XDG_STATE_HOME": str(xdg_state.resolve()),
        "PYTHONPATH": "",
        "PYTHONDONTWRITEBYTECODE": "1",
        "JS_ECHO_ENGINE": "on",
        "JS_ALLOWED_ORIGINS": "http://localhost",
        ISO_E2E_SANDBOX_ENV: "1",
        ISO_E2E_HOME_ENV: str(home.resolve()),
        "PIP_DISABLE_PIP_VERSION_CHECK": "1",
        "NO_PROXY": "127.0.0.1,localhost,::1",
        "no_proxy": "127.0.0.1,localhost,::1",
    }
    if config_path is not None:
        env["JS_CONFIG_PATH"] = str(config_path.resolve())
    if wheelhouse is not None:
        env["PIP_NO_INDEX"] = "1"
        env["PIP_FIND_LINKS"] = str(wheelhouse.resolve())
    if evidence_dir is not None:
        resolved_evidence = evidence_dir.resolve()
        resolved_evidence.mkdir(parents=True, exist_ok=True)
        env[ISO_E2E_EVIDENCE_DIR_ENV] = str(resolved_evidence)
    # Point isolated server E2E at the real repo / ephemeral private key (not sandbox cwd).
    env["JS_ISO_E2E_REPO_ROOT"] = str(REPO_ROOT)
    try:
        from js.echo.ledger.e2e_signing import E2E_PRIVATE_ENV, resolve_private_key_path

        env[E2E_PRIVATE_ENV] = str(resolve_private_key_path(REPO_ROOT))
    except Exception:
        # Signing step will fail-closed if the key is required and missing.
        pass
    return env


def _parse_json_payload(stdout: str) -> dict[str, Any] | None:
    text = stdout.strip()
    if not text:
        return None
    # Prefer the last JSON object line (server E2E prints a single summary object).
    for line in reversed(text.splitlines()):
        line = line.strip()
        if not line.startswith("{"):
            continue
        try:
            payload = json.loads(line)
        except json.JSONDecodeError:
            continue
        if isinstance(payload, dict):
            return payload
    try:
        payload = json.loads(text)
    except json.JSONDecodeError:
        return None
    return payload if isinstance(payload, dict) else None


def _run(
    step: str,
    cmd: Sequence[str],
    *,
    cwd: Path,
    env: Mapping[str, str],
    results: list[dict[str, Any]],
    source_digest: str | None = None,
    timeout: int = 600,
) -> bool:
    started = _utc_now()
    start_mono = time.monotonic()
    proc = subprocess.run(
        list(cmd),
        cwd=cwd,
        env=dict(env),
        capture_output=True,
        text=True,
        timeout=timeout,
    )
    finished = _utc_now()
    duration = round(time.monotonic() - start_mono, 3)
    ok = proc.returncode == 0
    detail: dict[str, Any] = {}
    payload = _parse_json_payload(proc.stdout)
    if payload is not None:
        nested = payload.get("results")
        if isinstance(nested, dict):
            detail.update(nested)
        detail["payload_ok"] = payload.get("ok")
        chat_status = detail.get("chat_status")
        if chat_status is not None and chat_status != 200:
            ok = False
        provider_calls = detail.get("provider_calls")
        if provider_calls is not None and int(provider_calls) <= 0:
            ok = False
        if detail.get("attachment_consumed") is False:
            ok = False
        work_receipt = detail.get("work_receipt")
        if isinstance(work_receipt, dict) and work_receipt.get("status") != "ok":
            ok = False
        if detail.get("ws_terminal_ok") is False:
            ok = False
        if payload.get("ok") is False:
            ok = False
    stdout_tail = proc.stdout[-4000:]
    stderr_tail = proc.stderr[-4000:]
    evidence_raw = env.get(ISO_E2E_EVIDENCE_DIR_ENV)
    capture_fields: dict[str, str] = {}
    if evidence_raw:
        evidence_root = Path(evidence_raw).resolve()
        steps_dir = evidence_root / "e2e" / "steps"
        steps_dir.mkdir(parents=True, exist_ok=True)
        index = len(results) + 1
        slug = re.sub(r"[^a-z0-9]+", "_", step.lower()).strip("_")
        prefix = steps_dir / f"{index:02d}_{slug}"
        stdout_path = prefix.with_suffix(".stdout.txt")
        stderr_path = prefix.with_suffix(".stderr.txt")
        receipt_path = prefix.with_suffix(".receipt.json")
        stdout_path.write_text(proc.stdout, encoding="utf-8")
        stderr_path.write_text(proc.stderr, encoding="utf-8")
        capture_fields = {
            "stdout_path": str(stdout_path),
            "stderr_path": str(stderr_path),
            "step_receipt_path": str(receipt_path),
        }
    entry: dict[str, Any] = {
        "step": step,
        "argv": list(cmd),
        "cwd": str(cwd),
        "started_utc": started,
        "finished_utc": finished,
        "exit_code": proc.returncode,
        "duration_seconds": duration,
        "stdout_tail": stdout_tail,
        "stderr_tail": stderr_tail,
        "stdout_sha256": hashlib.sha256(proc.stdout.encode()).hexdigest(),
        "stderr_sha256": hashlib.sha256(proc.stderr.encode()).hexdigest(),
        "ok": ok,
        **capture_fields,
    }
    if source_digest is not None:
        entry["source_digest"] = source_digest
    if step.endswith("import js/js_work from venv site-packages") and payload is not None:
        entry["import_evidence"] = {
            "modules": payload.get("modules"),
            "errors": payload.get("errors"),
        }
    if detail:
        entry["detail"] = detail
    if capture_fields:
        receipt_path = Path(capture_fields["step_receipt_path"])
        receipt_path.write_text(
            json.dumps(
                {
                    "argv": list(cmd),
                    "cwd": str(cwd.resolve()),
                    "exit_code": proc.returncode,
                    "stdout_sha256": entry["stdout_sha256"],
                    "stderr_sha256": entry["stderr_sha256"],
                    **capture_fields,
                },
                indent=2,
                sort_keys=True,
            )
            + "\n",
            encoding="utf-8",
        )
    results.append(entry)
    status = "OK" if ok else "FAIL"
    print(f"[{status}] {step} (exit={proc.returncode}, {duration}s)", flush=True)
    if not ok:
        print(f"  stderr: {proc.stderr[-800:]}", flush=True)
    return bool(ok)


def _write_server_e2e_script(path: Path) -> None:
    path.write_text(
        '''#!/usr/bin/env python3
"""Real Echo server E2E executed inside an isolated install venv."""
from __future__ import annotations

import asyncio
import base64
import hashlib
import json
import os
import re
import shutil
import socket
import sys
import tempfile
import threading
import time
from pathlib import Path
from typing import Any

import httpx
import yaml
from fastapi import FastAPI
from fastapi.responses import StreamingResponse
from fastapi.testclient import TestClient

ISO_E2E_ATTACHMENT_MARKER = "iso-e2e-attachment-marker-v4"
ISO_E2E_TOOL_MARKER = "__iso_tool__"
ISO_E2E_ATTACHMENT_PROMPT = "__iso_attachment__"
ISO_E2E_CHAT_MARKER = "__iso_chat__"
ISO_E2E_SANDBOX_ENV = "JS_ISO_E2E_SANDBOX"
ISO_E2E_HOME_ENV = "JS_ISO_E2E_HOME"
ISO_E2E_PROVIDER_FAIL_ENV = "JS_ISO_E2E_PROVIDER_FAIL"
LOOPBACK = "127.0.0.1"
FORBIDDEN_PROVIDERS = frozenset({"xai-oauth", "lmstudio"})


def _install_loopback_socket_guard() -> None:
    import socket as _socket

    original_connect = _socket.socket.connect
    original_getaddrinfo = _socket.getaddrinfo

    def _host_allowed(host: object) -> bool:
        if host is None:
            return True
        if not isinstance(host, str):
            return False
        normalized = host.strip().lower().rstrip(".")
        if not normalized:
            return True
        if normalized == "localhost" or normalized.endswith(".localhost"):
            return True
        if normalized == "::1":
            return True
        if normalized.startswith("127."):
            return True
        try:
            import ipaddress

            address = ipaddress.ip_address(normalized)
        except ValueError:
            return False
        return address.is_loopback

    def guarded_getaddrinfo(host, *args, **kwargs):
        if not _host_allowed(host):
            raise OSError(f"non-loopback host blocked by iso E2E guard: {host!r}")
        return original_getaddrinfo(host, *args, **kwargs)

    def guarded_connect(self, address):
        if self.family == _socket.AF_UNIX:
            return original_connect(self, address)
        if isinstance(address, tuple) and len(address) >= 1:
            if not _host_allowed(address[0]):
                raise OSError(f"non-loopback connect blocked by iso E2E guard: {address[0]!r}")
        return original_connect(self, address)

    _socket.socket.connect = guarded_connect  # type: ignore[method-assign]
    _socket.getaddrinfo = guarded_getaddrinfo  # type: ignore[assignment]


def _assert_isolation() -> None:
    if os.environ.get(ISO_E2E_SANDBOX_ENV) != "1":
        raise RuntimeError("server E2E requires isolation sandbox marker")
    expected_home = os.environ.get(ISO_E2E_HOME_ENV, "")
    actual_home = str(Path.home())
    if not expected_home or actual_home != expected_home:
        raise RuntimeError(
            f"server E2E HOME mismatch: expected {expected_home!r}, got {actual_home!r}"
        )


def _assert_settings_loopback_only() -> None:
    from js.config import JSSettings

    settings = JSSettings.from_file()
    provider_names = {provider.name for provider in settings.providers}
    if provider_names != {"iso-local"}:
        raise RuntimeError(f"unexpected providers in isolated E2E: {sorted(provider_names)}")
    forbidden = sorted(provider_names & FORBIDDEN_PROVIDERS)
    if forbidden:
        raise RuntimeError(f"forbidden providers present: {forbidden}")
    for provider in settings.providers:
        base_url = provider.base_url.lower()
        if not (
            base_url.startswith("http://127.0.0.1")
            or base_url.startswith("http://localhost")
            or base_url.startswith("http://[::1]")
        ):
            raise RuntimeError(f"provider base_url is not loopback: {provider.base_url}")
    from js.models.router import ModelRouter

    router = ModelRouter(settings)
    router_names = set(getattr(router, "_providers", {}).keys())
    unexpected = sorted(router_names - {"iso-local"} | (router_names & FORBIDDEN_PROVIDERS))
    if unexpected:
        raise RuntimeError(f"router exposes unexpected providers: {unexpected}")


def _free_port() -> int:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
        sock.bind((LOOPBACK, 0))
        return int(sock.getsockname()[1])


def _latest_user_text(messages: list[Any]) -> str:
    for message in reversed(messages):
        if isinstance(message, dict) and message.get("role") == "user":
            content = message.get("content")
            if isinstance(content, str):
                return content
            if isinstance(content, list):
                parts: list[str] = []
                for block in content:
                    if isinstance(block, dict) and block.get("type") == "text":
                        parts.append(str(block.get("text", "")))
                return "".join(parts)
    return ""


def _build_fake_provider_app() -> tuple[FastAPI, dict[str, Any]]:
    app = FastAPI()
    state: dict[str, Any] = {
        "chat_calls": 0,
        "attachment_observed": False,
        "attachment_marker_in_messages": False,
        "tool_rounds": 0,
        "scenarios": {},
        "unexpected_provider_calls": [],
        "all_provider_hosts_loopback": True,
        "force_error": os.environ.get(ISO_E2E_PROVIDER_FAIL_ENV) == "1",
    }

    def _scenario(text: str) -> str:
        if ISO_E2E_ATTACHMENT_PROMPT in text or ISO_E2E_ATTACHMENT_MARKER in text:
            return "attachment"
        if ISO_E2E_TOOL_MARKER in text:
            return "tool"
        if ISO_E2E_CHAT_MARKER in text:
            return "chat"
        return "default"

    @app.get("/v1/models")
    async def models() -> dict[str, Any]:
        return {"object": "list", "data": [{"id": "iso-local-model", "object": "model"}]}

    @app.get("/__iso__/stats")
    async def stats() -> dict[str, Any]:
        return state

    @app.post("/v1/chat/completions")
    async def completions(body: dict[str, Any]) -> Any:
        if state["force_error"]:
            from fastapi import HTTPException

            raise HTTPException(status_code=503, detail="iso-local forced provider failure")
        messages = body.get("messages", [])
        text = _latest_user_text(messages)
        if ISO_E2E_ATTACHMENT_MARKER in text:
            state["attachment_marker_in_messages"] = True
        scenario = _scenario(text)
        state["chat_calls"] = int(state["chat_calls"]) + 1
        state["scenarios"][scenario] = int(state["scenarios"].get(scenario, 0)) + 1
        if ISO_E2E_ATTACHMENT_MARKER in text:
            state["attachment_observed"] = True

        has_tool_result = any(
            isinstance(message, dict) and message.get("role") == "tool" for message in messages
        )
        if scenario == "tool":
            state["tool_rounds"] = int(state["tool_rounds"]) + (1 if has_tool_result else 0)

        if body.get("stream"):

            async def events() -> Any:
                thinking = {
                    "id": "iso-stream",
                    "object": "chat.completion.chunk",
                    "model": "iso-local-model",
                    "choices": [
                        {
                            "index": 0,
                            "delta": {"reasoning_content": "iso thinking"},
                            "finish_reason": None,
                        }
                    ],
                }
                yield f"data: {json.dumps(thinking)}\\n\\n"
                if scenario == "tool" and not has_tool_result:
                    tool_chunk = {
                        "id": "iso-stream",
                        "object": "chat.completion.chunk",
                        "model": "iso-local-model",
                        "choices": [
                            {
                                "index": 0,
                                "delta": {
                                    "tool_calls": [
                                        {
                                            "index": 0,
                                            "id": "iso-tool-call",
                                            "type": "function",
                                            "function": {"name": "file_list", "arguments": "{}"},
                                        }
                                    ]
                                },
                                "finish_reason": None,
                            }
                        ],
                    }
                    yield f"data: {json.dumps(tool_chunk)}\\n\\n"
                    final = {
                        "id": "iso-stream",
                        "object": "chat.completion.chunk",
                        "model": "iso-local-model",
                        "choices": [{"index": 0, "delta": {}, "finish_reason": "tool_calls"}],
                    }
                    yield f"data: {json.dumps(final)}\\n\\n"
                    yield "data: [DONE]\\n\\n"
                    return
                for token in ("hello ", "world"):
                    chunk = {
                        "id": "iso-stream",
                        "object": "chat.completion.chunk",
                        "model": "iso-local-model",
                        "choices": [
                            {"index": 0, "delta": {"content": token}, "finish_reason": None}
                        ],
                    }
                    yield f"data: {json.dumps(chunk)}\\n\\n"
                    await asyncio.sleep(0.01)
                final = {
                    "id": "iso-stream",
                    "object": "chat.completion.chunk",
                    "model": "iso-local-model",
                    "choices": [{"index": 0, "delta": {}, "finish_reason": "stop"}],
                    "usage": {"prompt_tokens": 3, "completion_tokens": 2, "total_tokens": 5},
                }
                yield f"data: {json.dumps(final)}\\n\\n"
                yield "data: [DONE]\\n\\n"

            return StreamingResponse(events(), media_type="text/event-stream")

        if scenario == "tool" and not has_tool_result:
            message = {
                "role": "assistant",
                "content": "",
                "tool_calls": [
                    {
                        "id": "iso-tool-call",
                        "type": "function",
                        "function": {"name": "file_list", "arguments": "{}"},
                    }
                ],
            }
            finish_reason = "tool_calls"
        elif has_tool_result:
            message = {"role": "assistant", "content": "tool continuation complete"}
            finish_reason = "stop"
        elif scenario == "attachment" and ISO_E2E_ATTACHMENT_MARKER in text:
            message = {
                "role": "assistant",
                "content": f"attachment acknowledged: {ISO_E2E_ATTACHMENT_MARKER}",
            }
            finish_reason = "stop"
        else:
            message = {
                "role": "assistant",
                "content": f"iso e2e complete {ISO_E2E_CHAT_MARKER}".strip(),
            }
            finish_reason = "stop"
        return {
            "id": "iso-completion",
            "object": "chat.completion",
            "created": 0,
            "model": "iso-local-model",
            "choices": [{"index": 0, "message": message, "finish_reason": finish_reason}],
            "usage": {"prompt_tokens": 3, "completion_tokens": 2, "total_tokens": 5},
        }

    return app, state


def _start_fake_provider(port: int) -> None:
    import uvicorn

    app, _state = _build_fake_provider_app()
    uvicorn.run(app, host=LOOPBACK, port=port, log_level="error", access_log=False)


def _wait_provider_ready(port: int) -> None:
    deadline = time.monotonic() + 20
    url = f"http://{LOOPBACK}:{port}/v1/models"
    while time.monotonic() < deadline:
        try:
            response = httpx.get(url, timeout=1.0, trust_env=False)
            if response.status_code < 500:
                return
        except httpx.HTTPError:
            pass
        time.sleep(0.1)
    raise RuntimeError(f"fake provider did not become ready on port {port}")


def _provider_stats(port: int) -> dict[str, Any]:
    response = httpx.get(
        f"http://{LOOPBACK}:{port}/__iso__/stats",
        timeout=5.0,
        trust_env=False,
    )
    response.raise_for_status()
    payload = response.json()
    return payload if isinstance(payload, dict) else {}


def _write_config(path: Path, *, provider_port: int, workspace: Path, state_dir: Path) -> None:
    config = {
        "workspace": str(workspace),
        "state_dir": str(state_dir),
        "max_turns": 4,
        "echo_engine": "on",
        "providers": [
            {
                "name": "iso-local",
                "base_url": f"http://{LOOPBACK}:{provider_port}/v1",
                "api_key": "local-only",
                "timeout": 12,
                "max_retries": 1,
                "default_model": "iso-local-model",
                "models": [
                    {
                        "id": "iso-local-model",
                        "name": "Isolated E2E local model",
                        "provider": "iso-local",
                        "supports_tools": True,
                        "supports_streaming": True,
                    }
                ],
            }
        ],
        "security": {"api_key_required": True, "defense_mode": "enforce"},
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(yaml.safe_dump(config, sort_keys=False), encoding="utf-8")


async def _work_office_via_lease(base: Path) -> dict[str, Any]:
    import hashlib

    from js.echo.attachment_gate import owner_slug, session_slug
    from js.echo.effect_interpreter import ToolEffect
    from js.echo.ledger.journal import _read_file_records
    from js_work.agent_factory import create_work_agent
    from js_work.config import WorkSettings
    from js_work.tools import WorkToolProfile

    work_home = Path.home() / ".js-work"
    settings = WorkSettings(
        work_home=work_home,
        workspace=work_home / "workspace",
        state_dir=work_home / "state",
        echo_engine="on",
        providers=[],
        models=[],
    )
    agent = create_work_agent(settings=settings, profile=WorkToolProfile.OFFICE)
    owner = "iso-e2e-owner"
    session = "iso-e2e-work"
    run_id = "iso-e2e-work-run"
    owner_root = settings.workspace / "owners" / owner_slug(owner) / session_slug(session)
    owner_root.mkdir(parents=True, exist_ok=True)
    arguments = {
        "path": "outputs/iso-e2e.xlsx",
        "data": '[["iso", "e2e", "leased"]]',
        "start_cell": "A1",
    }

    from js.security.approvals import ApprovalDecision, ApprovalDecisionType

    def _approve(_request: object) -> ApprovalDecision:
        request_id = getattr(_request, "id", "")
        return ApprovalDecision(
            ApprovalDecisionType.APPROVE,
            request_id=str(request_id),
            reason="iso-e2e",
        )

    agent.approvals.set_callback(
        session,
        _approve,
        owner_key_hash=owner,
        run_id=run_id,
        tool_name="excel_write",
        arguments=arguments,
    )
    context = agent.echo_runtime.build_context(
        channel="web",
        owner_key_hash=owner,
        session_id=session,
        run_id=run_id,
    )
    effect = ToolEffect.from_arguments(
        "excel_write",
        arguments,
        allowed_tools=("excel_write",),
    )
    before_records = agent.echo_safety_service.health().record_count
    _message, result = await agent.echo_runtime.execute_tool_effect(effect, context)
    after_records = agent.echo_safety_service.health().record_count
    output_path = owner_root / "outputs" / "iso-e2e.xlsx"
    journal_path = agent.echo_safety_service.journal_path_for_scope(
        owner,
        product_id="js-work",
        session_id=session,
    )
    mac_key = agent.echo_safety_service.journal_key_for_scope(
        owner,
        product_id="js-work",
        session_id=session,
    )
    from js.echo.ledger.release_gates import build_work_ledger_receipt_binding

    binding = build_work_ledger_receipt_binding(
        journal_path=journal_path,
        mac_key=mac_key,
        state_dir=settings.state_dir,
        owner=owner,
        session=session,
        product_id="js-work",
        run_id=run_id,
        tool_name="excel_write",
    )
    lease_id = ""
    lease_consumed = False
    if binding is not None:
        lease_id = str(binding.get("lease_id") or "")
        lease_consumed = binding.get("lease_consumed") is True
    output_cells: list[list[str]] = []
    if output_path.is_file():
        from openpyxl import load_workbook

        workbook = load_workbook(output_path, read_only=True, data_only=True)
        sheet = workbook.active
        output_cells = [[str(cell.value) for cell in row] for row in sheet.iter_rows(min_row=1, max_row=1)]
        workbook.close()
    evidence_root_raw = os.environ.get("JS_ISO_E2E_EVIDENCE_DIR", "")
    evidence_kind = os.environ.get("JS_ISO_E2E_KIND", "")
    archived_output = output_path
    if binding is not None and evidence_root_raw and evidence_kind in {"wheel", "sdist"}:
        archive_dir = Path(evidence_root_raw) / "e2e" / "work" / evidence_kind
        archive_dir.mkdir(parents=True, exist_ok=True)
        archived_output = archive_dir / "iso-e2e.xlsx"
        archived_journal = archive_dir / "ledger.journal"
        shutil.copy2(output_path, archived_output)
        shutil.copy2(journal_path, archived_journal)
        keys_dir = Path(evidence_root_raw) / "e2e" / "keys"
        # Public material only under evidence — never write private key material here.
        keys_dir.mkdir(parents=True, exist_ok=True)
        from js.echo.ledger.e2e_signing import (
            assert_no_private_key_under,
            load_private_key,
            resolve_private_key_path,
        )

        repo_root_env = os.environ.get("JS_ISO_E2E_REPO_ROOT", "").strip()
        repo_root = Path(repo_root_env) if repo_root_env else None
        private_path = resolve_private_key_path(repo_root)
        evidence_resolved = Path(evidence_root_raw).resolve()
        try:
            private_path.resolve().relative_to(evidence_resolved)
        except ValueError:
            pass
        else:
            raise RuntimeError("E2E private key must not reside under evidence root")
        private_key = load_private_key(private_path)
        arguments_sha256 = str(
            binding["ledger_chain"][0]["payload"]["tool_effect"]["args_hash"]
        )
        output_sha256 = hashlib.sha256(archived_output.read_bytes()).hexdigest()
        signature_payload = {
            "journal_sha256": hashlib.sha256(archived_journal.read_bytes()).hexdigest(),
            "arguments_sha256": arguments_sha256,
            "output_sha256": output_sha256,
            "product_id": "js-work",
            "owner": owner,
            "session": session,
            "run_id": run_id,
            "effect_id": str(binding["effect_id"]),
        }
        canonical = json.dumps(
            signature_payload, sort_keys=True, separators=(",", ":")
        ).encode()
        public_raw = private_key.public_key().public_bytes_raw()
        signature_b64 = base64.b64encode(private_key.sign(canonical)).decode()
        assert_no_private_key_under(evidence_resolved)
        binding.update(
            {
                "journal_evidence_path": f"e2e/work/{evidence_kind}/ledger.journal",
                "arguments_sha256": arguments_sha256,
                "ledger_signature_b64": signature_b64,
                "pubkey_fingerprint": hashlib.sha256(public_raw).hexdigest(),
                "signature_purpose": "ephemeral-e2e-ledger-consistency-v1",
                "not_a_third_party_signature": True,
            }
        )
        (keys_dir / "ledger.ed25519.public.b64").write_text(
            base64.b64encode(public_raw).decode("ascii") + "\\n",
            encoding="utf-8",
        )
    terminal_status = str(binding.get("terminal_status") or "") if binding is not None else ""
    receipt = {
        "product_id": "js-work",
        "owner": owner,
        "session": session,
        "run_id": run_id,
        "tool_name": "excel_write",
        "lease_id": lease_id,
        "lease_consumed": lease_consumed,
        "status": terminal_status or "failed",
        "terminal": terminal_status == "ok",
        "output_path": str(output_path),
        "output_exists": output_path.is_file(),
        "output_sha256": hashlib.sha256(archived_output.read_bytes()).hexdigest()
        if archived_output.is_file()
        else "",
        "output_cells": output_cells,
        "journal_records_added": after_records > before_records,
    }
    if binding is not None:
        for field in (
            "journal_relative_path",
            "journal_sha256",
            "ledger_sequence",
            "record_hash",
            "effect_id",
            "terminal_status",
            "ledger_chain",
            "journal_evidence_path",
            "arguments_sha256",
            "ledger_signature_b64",
            "pubkey_fingerprint",
        ):
            if field in binding:
                receipt[field] = binding[field]
    await agent.close()
    return receipt


def main() -> int:
    try:
        _install_loopback_socket_guard()
        _assert_isolation()
    except Exception as exc:
        print(json.dumps({"ok": False, "error": str(exc)}))
        return 1

    base = Path(tempfile.mkdtemp(prefix="iso-e2e-server-"))
    provider_port = _free_port()
    provider_thread = threading.Thread(
        target=_start_fake_provider,
        args=(provider_port,),
        daemon=True,
    )
    provider_thread.start()
    _wait_provider_ready(provider_port)

    workspace = base / "workspace"
    state_dir = base / "state"
    config_path = base / "config.yaml"
    _write_config(
        config_path,
        provider_port=provider_port,
        workspace=workspace,
        state_dir=state_dir,
    )

    os.environ["JS_CONFIG_PATH"] = str(config_path)
    os.environ["JS_ECHO_ENGINE"] = "on"
    os.environ["JS_ALLOWED_ORIGINS"] = "http://localhost"
    os.environ["NO_PROXY"] = "127.0.0.1,localhost,::1"
    os.environ["no_proxy"] = "127.0.0.1,localhost,::1"

    try:
        _assert_settings_loopback_only()
    except Exception as exc:
        print(json.dumps({"ok": False, "error": str(exc)}))
        return 1

    import js.web.auth as auth_mod

    auth_mod._ALLOWED_ORIGINS = None

    from js.web.auth import AuthManager
    from js.web.server import create_app

    app = create_app()
    admin_key = AuthManager(state_dir).ensure_bootstrap_admin_key()
    headers = {
        "Origin": "http://localhost",
        "X-API-Key": admin_key,
    }
    results: dict[str, Any] = {"offline": True, "schema_version": "isolated-venv-e2e-v6"}

    with TestClient(app, base_url="http://localhost", headers=headers) as client:
        results["health_status"] = client.get("/api/health").status_code
        results["status_status"] = client.get("/api/status").status_code
        chat = client.post(
            "/api/chat",
            json={
                "message": f"{ISO_E2E_CHAT_MARKER} iso e2e chat",
                "session_id": "iso-e2e-chat",
            },
        )
        results["chat_status"] = chat.status_code
        results["chat_response_ok"] = (
            chat.status_code == 200
            and ISO_E2E_CHAT_MARKER in str(chat.json().get("response", ""))
        )

        attachment_session = "iso-e2e-attachment"
        upload = client.post(
            "/api/upload",
            data={"session_id": attachment_session},
            files={"file": ("note.txt", ISO_E2E_ATTACHMENT_MARKER.encode(), "text/plain")},
        )
        results["upload_status"] = upload.status_code
        attachment_path = upload.json().get("path") if upload.status_code == 200 else None
        results["attachment_path_ok"] = bool(attachment_path)
        if attachment_path:
            attachment_chat = client.post(
                "/api/chat",
                json={
                    "message": f"{ISO_E2E_ATTACHMENT_PROMPT} summarize the note",
                    "session_id": attachment_session,
                    "attachments": [attachment_path],
                },
            )
            results["attachment_chat_status"] = attachment_chat.status_code
            results["attachment_chat_ok"] = (
                attachment_chat.status_code == 200
                and f"attachment acknowledged: {ISO_E2E_ATTACHMENT_MARKER}"
                in str(attachment_chat.json().get("response", ""))
            )
        else:
            results["attachment_chat_status"] = 0
            results["attachment_chat_ok"] = False

        with client.websocket_connect("/ws", headers=headers) as ws:
            payload: dict[str, Any] = {
                "type": "stream",
                "content": f"{ISO_E2E_TOOL_MARKER} stream iso e2e",
                "session_id": "iso-e2e-ws",
            }
            ws.send_json(payload)
            frame_types: list[str] = []
            terminal_ok = False
            saw_token = False
            saw_thinking = False
            saw_tool_call = False
            saw_done = False
            for _ in range(40):
                frame = ws.receive_json()
                frame_type = str(frame.get("type", ""))
                frame_types.append(frame_type)
                if frame_type == "error":
                    terminal_ok = False
                    break
                if frame_type in {"token", "response"}:
                    content = frame.get("content")
                    if isinstance(content, str) and content.strip():
                        saw_token = True
                if frame_type == "thinking" and frame.get("content"):
                    saw_thinking = True
                if frame_type == "tool_call" and frame.get("tool_call"):
                    saw_tool_call = True
                if frame_type == "done":
                    saw_done = True
                    terminal_ok = bool(frame.get("terminal")) or frame.get("status") == "completed"
                    break
            results["ws_frame_types"] = frame_types
            results["ws_terminal_ok"] = terminal_ok
            results["ws_saw_token"] = saw_token
            results["ws_saw_thinking"] = saw_thinking
            results["ws_saw_tool_call"] = saw_tool_call
            results["ws_saw_done"] = saw_done

    stats = _provider_stats(provider_port)
    results["provider_calls"] = int(stats.get("chat_calls", 0))
    results["provider_scenarios"] = dict(stats.get("scenarios", {}))
    results["attachment_consumed"] = bool(stats.get("attachment_observed"))
    results["attachment_marker_in_messages"] = bool(stats.get("attachment_marker_in_messages"))
    results["unexpected_provider_calls"] = list(stats.get("unexpected_provider_calls", []))
    results["all_provider_hosts_loopback"] = bool(stats.get("all_provider_hosts_loopback", True))
    results["tool_rounds"] = int(stats.get("tool_rounds", 0))
    work_receipt = asyncio.run(_work_office_via_lease(base))
    results["work_receipt"] = work_receipt

    ok = (
        results.get("health_status") == 200
        and results.get("status_status") == 200
        and results.get("chat_status") == 200
        and results.get("chat_response_ok") is True
        and results.get("upload_status") == 200
        and results.get("attachment_path_ok") is True
        and results.get("attachment_chat_ok") is True
        and results.get("ws_terminal_ok") is True
        and results.get("ws_saw_token") is True
        and results.get("ws_saw_thinking") is True
        and results.get("ws_saw_tool_call") is True
        and results.get("ws_saw_done") is True
        and results.get("provider_calls", 0) > 0
        and results.get("attachment_consumed") is True
        and results.get("attachment_marker_in_messages") is True
        and results.get("unexpected_provider_calls") == []
        and results.get("all_provider_hosts_loopback") is True
        and int(results.get("provider_scenarios", {}).get("chat", 0)) >= 1
        and int(results.get("provider_scenarios", {}).get("attachment", 0)) >= 1
        and int(results.get("provider_scenarios", {}).get("tool", 0)) >= 1
        and results.get("tool_rounds", 0) >= 1
        and isinstance(results.get("work_receipt"), dict)
        and results["work_receipt"].get("product_id") == "js-work"
        and results["work_receipt"].get("status") == "ok"
        and results["work_receipt"].get("lease_consumed") is True
        and results["work_receipt"].get("effect_id")
        and results["work_receipt"].get("ledger_chain")
        and results["work_receipt"].get("record_hash")
        and results["work_receipt"].get("output_exists") is True
        and results["work_receipt"].get("journal_records_added") is True
        and results["work_receipt"].get("output_cells") == [["iso", "e2e", "leased"]]
    )
    print(json.dumps({"ok": ok, "results": results}))
    return 0 if ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
''',
        encoding="utf-8",
    )
    path.chmod(0o755)


def _check_artifact(
    artifact: Path,
    *,
    kind: str,
    work_root: Path,
    wheelhouse: Path,
    source_digest: str,
    results: list[dict[str, Any]],
    pip_checks: dict[str, dict[str, Any]],
    evidence_dir: Path | None = None,
) -> bool:
    ok = True
    sandbox = work_root / f"install-{kind}"
    sandbox.mkdir(parents=True)
    home = sandbox / "home"
    home.mkdir()
    env = _base_env(home, wheelhouse=wheelhouse, evidence_dir=evidence_dir)
    env[ISO_E2E_KIND_ENV] = kind
    venv_dir = sandbox / "venv"

    ok &= _run(
        f"{kind}: create venv",
        [str(VENV_PYTHON), "-m", "venv", str(venv_dir)],
        cwd=sandbox,
        env=env,
        results=results,
        source_digest=source_digest,
    )
    if not ok:
        return False
    venv_python = venv_dir / "bin" / "python"
    find_links = str(wheelhouse.resolve())
    # Offline sdist builds need the hatchling backend already present.
    if kind == "sdist":
        ok &= _run(
            f"{kind}: pip install build backends offline",
            [
                str(venv_python),
                "-m",
                "pip",
                "install",
                "--no-index",
                "--find-links",
                find_links,
                "--no-input",
                "hatchling",
                "pathspec",
                "packaging",
                "trove-classifiers",
                "pluggy",
            ],
            cwd=sandbox,
            env=env,
            results=results,
            source_digest=source_digest,
            timeout=600,
        )
        if not ok:
            return False
    install_cmd = [
        str(venv_python),
        "-m",
        "pip",
        "install",
        "--no-index",
        "--find-links",
        find_links,
        "--no-input",
    ]
    pip_report: Path | None = None
    if evidence_dir is not None:
        pip_report = evidence_dir / "e2e" / "pip" / f"{kind}.install-report.json"
        pip_report.parent.mkdir(parents=True, exist_ok=True)
        install_cmd.extend(["--report", str(pip_report)])
    if kind == "sdist":
        install_cmd.append("--no-build-isolation")
    install_cmd.append(f"{artifact}[echo-tokenizer,office]")
    ok &= _run(
        f"{kind}: pip install artifact offline (echo-tokenizer,office)",
        install_cmd,
        cwd=sandbox,
        env=env,
        results=results,
        source_digest=source_digest,
        timeout=900,
    )
    if not ok:
        return False
    if pip_report is not None and pip_report.is_file():
        report_payload = pip_report.read_bytes()
        install_step = results[-1]
        install_step["pip_report"] = {
            "path": str(pip_report.resolve()),
            "sha256": hashlib.sha256(report_payload).hexdigest(),
            "packages": [
                {
                    "name": str(item.get("metadata", {}).get("name", "")),
                    "version": str(item.get("metadata", {}).get("version", "")),
                    "archive_hash": str(
                        item.get("download_info", {}).get("archive_info", {}).get("hash", "")
                    ),
                }
                for item in json.loads(report_payload).get("install", [])
                if isinstance(item, dict)
            ],
        }

    pip_ok = _run(
        f"{kind}: pip check",
        [str(venv_python), "-m", "pip", "check"],
        cwd=sandbox,
        env=env,
        results=results,
        source_digest=source_digest,
        timeout=120,
    )
    pip_step = results[-1]
    pip_checks[kind] = {
        "ok": pip_ok,
        "exit_code": pip_step["exit_code"],
        "stdout_tail": pip_step["stdout_tail"],
        "stderr_tail": pip_step["stderr_tail"],
    }
    ok &= pip_ok
    if not ok:
        print(f"[FAIL] {kind}: pip check (exit={pip_step['exit_code']})", flush=True)
        return False

    ok &= _run(
        f"{kind}: import js/js_work from venv site-packages",
        [str(venv_python), "-c", IMPORT_CHECK],
        cwd=sandbox,
        env=env,
        results=results,
        source_digest=source_digest,
    )
    ok &= _run(
        f"{kind}: tokenizer loads offline from vendored cache",
        [str(venv_python), "-c", TOKENIZER_CHECK],
        cwd=sandbox,
        env=env,
        results=results,
        source_digest=source_digest,
    )
    ok &= _run(
        f"{kind}: CLI js --help",
        [str(venv_dir / "bin" / "js"), "--help"],
        cwd=sandbox,
        env=env,
        results=results,
        source_digest=source_digest,
    )
    ok &= _run(
        f"{kind}: CLI js work --help",
        [str(venv_dir / "bin" / "js"), "work", "--help"],
        cwd=sandbox,
        env=env,
        results=results,
        source_digest=source_digest,
    )
    ok &= _run(
        f"{kind}: CLI js-work --help",
        [str(venv_dir / "bin" / "js-work"), "--help"],
        cwd=sandbox,
        env=env,
        results=results,
        source_digest=source_digest,
    )
    ok &= _run(
        f"{kind}: CLI python -m js_work --help",
        [str(venv_python), "-m", "js_work", "--help"],
        cwd=sandbox,
        env=env,
        results=results,
        source_digest=source_digest,
    )

    server_script = sandbox / "server_e2e.py"
    _write_server_e2e_script(server_script)
    ok &= _run(
        f"{kind}: server HTTP+WS+attachment+work E2E",
        [str(venv_python), str(server_script)],
        cwd=sandbox,
        env=env,
        results=results,
        source_digest=source_digest,
        timeout=900,
    )
    return bool(ok)


def _expected_source_digest() -> str:
    from js.echo.ledger.release_gates import release_source_digest

    return release_source_digest(REPO_ROOT)


def _validate_summary_schema(summary: Mapping[str, Any]) -> list[str]:
    errors: list[str] = []
    if summary.get("schema_version") != ISOLATED_VENV_E2E_SCHEMA_VERSION:
        errors.append(f"schema_version must be {ISOLATED_VENV_E2E_SCHEMA_VERSION!r}")
    if summary.get("offline") is not True:
        errors.append("offline must be true")
    if not isinstance(summary.get("source_digest"), str) or not summary["source_digest"]:
        errors.append("source_digest must be a non-empty string")
    if not isinstance(summary.get("evidence_root"), str) or not summary["evidence_root"]:
        errors.append("evidence_root must be a non-empty string")
    if not isinstance(summary.get("ok"), bool):
        errors.append("ok must be a boolean")
    artifacts = summary.get("artifacts")
    if not isinstance(artifacts, dict):
        errors.append("artifacts must be an object")
    else:
        for kind in ("wheel", "sdist"):
            meta = artifacts.get(kind)
            if not isinstance(meta, dict):
                errors.append(f"artifacts.{kind} must be an object")
                continue
            if not isinstance(meta.get("path"), str) or not meta.get("path"):
                errors.append(f"artifacts.{kind}.path must be a non-empty relative path")
            if not isinstance(meta.get("sha256"), str) or len(meta.get("sha256", "")) != 64:
                errors.append(f"artifacts.{kind}.sha256 must be a 64-char hex digest")
            if isinstance(meta.get("bytes"), bool) or not isinstance(meta.get("bytes"), int):
                errors.append(f"artifacts.{kind}.bytes must be an integer")
    work_output = summary.get("work_output")
    if not isinstance(work_output, dict):
        errors.append("work_output must be an object")
    else:
        for key in ("path", "sha256", "bytes", "cells"):
            if key not in work_output:
                errors.append(f"work_output missing {key}")
    work_outputs = summary.get("work_outputs")
    if not isinstance(work_outputs, dict) or set(work_outputs) != {"wheel", "sdist"}:
        errors.append("work_outputs must contain wheel and sdist")
    manifest = summary.get("manifest")
    if not isinstance(manifest, list) or not manifest:
        errors.append("manifest must be a non-empty list")
    pip_check = summary.get("pip_check")
    if not isinstance(pip_check, dict):
        errors.append("pip_check must be an object")
    else:
        for kind in ("wheel", "sdist"):
            entry = pip_check.get(kind)
            if (
                not isinstance(entry, dict)
                or entry.get("ok") is not True
                or entry.get("exit_code") != 0
            ):
                errors.append(f"pip_check.{kind} must include ok=true and exit_code=0")
    results = summary.get("results")
    if not isinstance(results, list):
        errors.append("results must be a list")
    elif len(results) != 22:
        errors.append("results must contain exactly 22 steps")
    else:
        for idx, step in enumerate(results):
            if not isinstance(step, dict):
                errors.append(f"results[{idx}] must be an object")
                continue
            for key in ("argv", "cwd", "started_utc", "finished_utc", "exit_code", "ok"):
                if key not in step:
                    errors.append(f"results[{idx}] missing {key}")
    return errors


def _manifest_entry(relative_path: str, *, root: Path) -> dict[str, Any]:
    path = (root / relative_path).resolve()
    payload = path.read_bytes()
    return {
        "path": relative_path,
        "sha256": hashlib.sha256(payload).hexdigest(),
        "bytes": len(payload),
    }


def _path_aliases(path: Path) -> tuple[str, ...]:
    """Return path spellings that may appear before/after ``Path.resolve()``.

    macOS exposes ``/var`` as a symlink to ``/private/var``. Ephemeral sandboxes
    often record the unresolved ``/var/...`` form in argv/cwd while
    ``Path.resolve()`` yields ``/private/var/...`` — both must normalize.
    """
    aliases = {str(path), str(path.resolve())}
    expanded: set[str] = set()
    for alias in aliases:
        expanded.add(alias)
        if alias.startswith("/private/var/"):
            expanded.add("/var/" + alias.removeprefix("/private/var/"))
        elif alias.startswith("/var/"):
            expanded.add("/private/var/" + alias.removeprefix("/var/"))
    return tuple(sorted(expanded, key=len, reverse=True))


def _rewrite_path_aliases(value: Any, *, aliases: Sequence[str], target: str) -> Any:
    """Recursively rewrite ephemeral sandbox path spellings to durable evidence paths."""
    if isinstance(value, str):
        rewritten = value
        for alias in aliases:
            rewritten = rewritten.replace(alias, target)
        return rewritten
    if isinstance(value, list):
        return [_rewrite_path_aliases(item, aliases=aliases, target=target) for item in value]
    if isinstance(value, dict):
        return {
            key: _rewrite_path_aliases(item, aliases=aliases, target=target)
            for key, item in value.items()
        }
    return value


def _refresh_import_evidence_shas(import_evidence: dict[str, Any]) -> None:
    """Re-bind module file digests after durable path normalization."""
    modules = import_evidence.get("modules")
    if not isinstance(modules, dict):
        return
    for module_evidence in modules.values():
        if not isinstance(module_evidence, dict):
            continue
        file_raw = module_evidence.get("file")
        if not isinstance(file_raw, str) or not file_raw.strip():
            continue
        module_file = Path(file_raw)
        if not module_file.is_file():
            continue
        module_evidence["file"] = str(module_file.resolve())
        module_evidence["file_sha256"] = hashlib.sha256(module_file.read_bytes()).hexdigest()
        site_raw = module_evidence.get("site_packages")
        if isinstance(site_raw, str) and site_raw.strip():
            module_evidence["site_packages"] = str(Path(site_raw).resolve())


def _normalize_step_contracts(
    results: list[dict[str, Any]],
    *,
    work_root: Path,
    evidence_dir: Path,
) -> None:
    """Replace ephemeral install roots with durable normalized evidence roots."""
    runtime_root = (evidence_dir / "e2e" / "runtime").resolve()

    def _ignore_heavy(directory: str, names: list[str]) -> set[str]:
        # Skip venv executables/headers; keep lib/site-packages for import evidence.
        # Sanitized export still excludes e2e/runtime entirely.
        base = Path(directory).name
        ignored = {name for name in names if name in {"__pycache__", ".pytest_cache"}}
        if base == "venv":
            ignored.update(name for name in names if name in {"bin", "include", "share"})
        return ignored

    for kind in ("wheel", "sdist"):
        source_path = work_root / f"install-{kind}"
        if not source_path.exists():
            continue
        aliases = _path_aliases(source_path)
        target_path = runtime_root / f"install-{kind}"
        if target_path.exists():
            shutil.rmtree(target_path)
        shutil.copytree(source_path, target_path, symlinks=False, ignore=_ignore_heavy)
        target = str(target_path)
        for step in results:
            if not str(step.get("step", "")).startswith(f"{kind}:"):
                continue
            cwd = str(step.get("cwd", ""))
            argv = list(step.get("argv", []))
            for alias in aliases:
                cwd = cwd.replace(alias, target)
                argv = [
                    argument.replace(alias, target) if isinstance(argument, str) else argument
                    for argument in argv
                ]
            step["cwd"] = cwd
            step["argv"] = argv
            if "import_evidence" in step and isinstance(step["import_evidence"], dict):
                step["import_evidence"] = _rewrite_path_aliases(
                    step["import_evidence"],
                    aliases=aliases,
                    target=target,
                )
                _refresh_import_evidence_shas(step["import_evidence"])
            if "detail" in step:
                step["detail"] = _rewrite_path_aliases(
                    step["detail"],
                    aliases=aliases,
                    target=target,
                )
            for path_field, tail_field, digest_field in (
                ("stdout_path", "stdout_tail", "stdout_sha256"),
                ("stderr_path", "stderr_tail", "stderr_sha256"),
            ):
                raw_capture = step.get(path_field)
                if not isinstance(raw_capture, str):
                    continue
                capture_path = Path(raw_capture)
                if not capture_path.is_file():
                    continue
                rewritten = _rewrite_path_aliases(
                    capture_path.read_text(encoding="utf-8"),
                    aliases=aliases,
                    target=target,
                )
                if not isinstance(rewritten, str):
                    continue
                capture_path.write_text(rewritten, encoding="utf-8")
                step[digest_field] = hashlib.sha256(rewritten.encode()).hexdigest()
                # Keep tails bound to the rewritten capture files (not independently mutated).
                step[tail_field] = rewritten[-4000:]
            receipt_raw = step.get("step_receipt_path")
            if isinstance(receipt_raw, str):
                receipt_path = Path(receipt_raw)
                receipt = json.loads(receipt_path.read_text(encoding="utf-8"))
                receipt["cwd"] = cwd
                receipt["argv"] = argv
                receipt["stdout_sha256"] = step.get("stdout_sha256", receipt.get("stdout_sha256"))
                receipt["stderr_sha256"] = step.get("stderr_sha256", receipt.get("stderr_sha256"))
                receipt_path.write_text(
                    json.dumps(receipt, indent=2, sort_keys=True) + "\n",
                    encoding="utf-8",
                )


def _copy_artifact_to_evidence(
    artifact: Path,
    *,
    evidence_dir: Path,
    relative_dir: str,
) -> tuple[str, dict[str, Any]]:
    target_dir = evidence_dir / relative_dir
    target_dir.mkdir(parents=True, exist_ok=True)
    destination = target_dir / artifact.name
    shutil.copy2(artifact, destination)
    relative_path = f"{relative_dir}/{artifact.name}".replace("\\", "/")
    payload = destination.read_bytes()
    return relative_path, {
        "path": relative_path,
        "sha256": hashlib.sha256(payload).hexdigest(),
        "bytes": len(payload),
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--source-digest",
        required=False,
        help="required gate digest; must match current release_source_digest()",
    )
    parser.add_argument(
        "--wheelhouse",
        type=Path,
        help="local find-links directory containing dependency wheels",
    )
    parser.add_argument("--prepare-wheelhouse", type=Path, help="populate wheelhouse offline")
    parser.add_argument("--json", type=Path, default=None, help="write machine-readable results")
    parser.add_argument(
        "--evidence-dir",
        type=Path,
        default=None,
        help="evidence bundle root; artifacts and work output are copied here",
    )
    parser.add_argument("--log", type=Path, default=None, help="tee full log to this .out file")
    parser.add_argument("--keep", action="store_true", help="keep temp sandboxes")
    args = parser.parse_args()

    if args.prepare_wheelhouse is not None:
        report = prepare_wheelhouse(args.prepare_wheelhouse.resolve())
        if args.json is not None:
            args.json.write_text(json.dumps(report, indent=2), encoding="utf-8")
        return 0 if report["ok"] else 1

    if args.wheelhouse is None:
        parser.error("--wheelhouse is required unless --prepare-wheelhouse is used")
    if args.source_digest is None:
        parser.error("--source-digest is required for gate mode")

    wheelhouse = args.wheelhouse.resolve()
    _validate_wheelhouse(wheelhouse)
    evidence_dir = args.evidence_dir.resolve() if args.evidence_dir is not None else None
    if evidence_dir is not None:
        evidence_dir.mkdir(parents=True, exist_ok=True)
        os.environ[ISO_E2E_EVIDENCE_DIR_ENV] = str(evidence_dir)

    expected_digest = _expected_source_digest()
    if args.source_digest != expected_digest:
        print(
            "SOURCE_DIGEST_MISMATCH",
            f"expected={expected_digest}",
            f"got={args.source_digest}",
            sep="\n",
            file=sys.stderr,
        )
        return 1

    from js.echo.ledger.e2e_signing import (
        E2E_PARENT_OWNS_CLEANUP_ENV,
        assert_no_private_key_under,
        resolve_private_key_path,
    )
    from js.echo.ledger.strict_json import StrictJSONError, strict_load_path

    # Keypair must already be prepared by the freeze orchestrator (pubkey on digest
    # surface). Refuse to invent a key mid-gate after digest freeze.
    # Parent process owns EphemeralKeyHandle destroy; child must not Path-destroy.
    parent_owns_cleanup = os.environ.get(E2E_PARENT_OWNS_CLEANUP_ENV, "").strip() == "1"
    try:
        private_key_path = resolve_private_key_path(REPO_ROOT)
    except Exception as exc:
        print(f"[FAIL] E2E private key not prepared by orchestrator: {exc}", flush=True)
        return 1
    _ = private_key_path

    provenance_path = (
        evidence_dir / "e2e" / "E2E_KEY_PROVENANCE.json" if evidence_dir is not None else None
    )
    provenance: dict[str, object] | None = None
    if provenance_path is not None and provenance_path.is_file():
        try:
            loaded = strict_load_path(provenance_path)
            if isinstance(loaded, dict):
                provenance = loaded
        except (OSError, StrictJSONError, ValueError):
            provenance = None
    if provenance is None:
        print(
            "[FAIL] E2E_KEY_PROVENANCE.json missing; orchestrator must write it before freeze",
            flush=True,
        )
        return 1
    if not parent_owns_cleanup:
        print(
            f"[FAIL] {E2E_PARENT_OWNS_CLEANUP_ENV}=1 required; "
            "child must not destroy via bare Path",
            flush=True,
        )
        return 1

    log_path = args.log
    if log_path is None and args.json is not None:
        log_path = args.json.with_suffix(".out")

    log_fp: TextIOBase | None = None
    original_stdout = sys.stdout
    if log_path is not None:
        log_path.parent.mkdir(parents=True, exist_ok=True)
        log_fp = log_path.open("w", encoding="utf-8")
        sys.stdout = _Tee(original_stdout, log_fp)  # type: ignore[arg-type,assignment]

    results: list[dict[str, Any]] = []
    pip_checks: dict[str, dict[str, Any]] = {}
    started = _utc_now()
    work_root = Path(tempfile.mkdtemp(prefix="iso-venv-e2e-"))
    home = work_root / "build-home"
    home.mkdir()
    env = _base_env(home, wheelhouse=wheelhouse, evidence_dir=evidence_dir)
    dist = work_root / "dist"

    overall = True
    exit_code = 1
    summary: dict[str, Any] = {}
    cleanup_done = False

    def _cleanup_private_key(*, reason: str) -> None:
        """Child never destroys; parent retains EphemeralKeyHandle lifecycle."""
        nonlocal cleanup_done, exit_code
        if cleanup_done:
            return
        cleanup_done = True
        try:
            if evidence_dir is not None:
                assert_no_private_key_under(evidence_dir)
            print(
                f"[OK] e2e private key cleanup deferred to parent ({reason})",
                flush=True,
            )
        except Exception as exc:
            print(f"[FAIL] e2e private key cleanup check: {exc}", flush=True)
            exit_code = 1

    def _on_signal(signum: int, _frame: object) -> None:
        _cleanup_private_key(reason=f"signal:{signum}")
        raise SystemExit(128 + int(signum))

    previous_sigterm = signal.getsignal(signal.SIGTERM)
    previous_sigint = signal.getsignal(signal.SIGINT)
    signal.signal(signal.SIGTERM, _on_signal)
    signal.signal(signal.SIGINT, _on_signal)
    try:
        build_python = VENV_PYTHON if VENV_PYTHON.is_file() else Path(sys.executable)
        overall &= _run(
            "build: python -m build --no-isolation (wheel+sdist)",
            [
                str(build_python),
                "-m",
                "build",
                "--outdir",
                str(dist),
                "--no-isolation",
            ],
            cwd=REPO_ROOT,
            env=env,
            results=results,
            source_digest=expected_digest,
            timeout=900,
        )
        if not overall:
            exit_code = 1
        else:
            wheel = next(dist.glob("*.whl"))
            sdist = next(dist.glob("*.tar.gz"))
            artifacts = {"wheel": wheel, "sdist": sdist}
            artifact_meta: dict[str, dict[str, Any]] = {}
            if evidence_dir is not None:
                for kind, artifact in artifacts.items():
                    relative_path, meta = _copy_artifact_to_evidence(
                        artifact,
                        evidence_dir=evidence_dir,
                        relative_dir="e2e/artifacts",
                    )
                    artifact_meta[kind] = meta
            else:
                for kind, artifact in artifacts.items():
                    payload = artifact.read_bytes()
                    artifact_meta[kind] = {
                        "path": artifact.name,
                        "sha256": hashlib.sha256(payload).hexdigest(),
                        "bytes": len(payload),
                    }
            check_artifacts = dict(artifacts)
            if evidence_dir is not None:
                check_artifacts = {
                    kind: (evidence_dir / str(meta["path"])).resolve()
                    for kind, meta in artifact_meta.items()
                }
            for kind, artifact in check_artifacts.items():
                overall &= _check_artifact(
                    artifact,
                    kind=kind,
                    work_root=work_root,
                    wheelhouse=wheelhouse,
                    source_digest=expected_digest,
                    results=results,
                    pip_checks=pip_checks,
                    evidence_dir=evidence_dir,
                )
            if evidence_dir is not None:
                _normalize_step_contracts(
                    results,
                    work_root=work_root,
                    evidence_dir=evidence_dir,
                )

            work_output: dict[str, Any] | None = None
            work_outputs: dict[str, dict[str, Any]] = {}
            manifest: list[dict[str, Any]] = []
            if evidence_dir is not None:
                manifest.extend(
                    _manifest_entry(entry["path"], root=evidence_dir)
                    for entry in artifact_meta.values()
                )
                for kind in ("wheel", "sdist"):
                    relative_work = f"e2e/work/{kind}/iso-e2e.xlsx"
                    required_paths = (
                        relative_work,
                        f"e2e/work/{kind}/ledger.journal",
                    )
                    if not all((evidence_dir / relative).is_file() for relative in required_paths):
                        overall = False
                        continue
                    payload = (evidence_dir / relative_work).read_bytes()
                    work_outputs[kind] = {
                        "path": relative_work,
                        "sha256": hashlib.sha256(payload).hexdigest(),
                        "bytes": len(payload),
                        "cells": [["iso", "e2e", "leased"]],
                    }
                    manifest.extend(
                        _manifest_entry(relative, root=evidence_dir) for relative in required_paths
                    )
                work_output = work_outputs.get("sdist")

            try:
                evidence_root = (
                    evidence_dir.relative_to(REPO_ROOT).as_posix() if evidence_dir is not None else "."
                )
            except ValueError:
                evidence_root = str(evidence_dir) if evidence_dir is not None else "."
            summary = {
                "schema_version": ISOLATED_VENV_E2E_SCHEMA_VERSION,
                "offline": True,
                "source_digest": expected_digest,
                "evidence_root": evidence_root,
                "started_utc": started,
                "finished_utc": _utc_now(),
                "ok": overall,
                "wheelhouse": str(wheelhouse),
                "artifacts": artifact_meta,
                "work_output": work_output or {},
                "work_outputs": work_outputs,
                "manifest": manifest,
                "pip_check": pip_checks,
                "results": results,
            }
            schema_errors = _validate_summary_schema(summary)
            if schema_errors:
                overall = False
                summary["ok"] = False
                summary["schema_errors"] = schema_errors
                for err in schema_errors:
                    print(f"[FAIL] schema: {err}", flush=True)
            if args.json is not None:
                args.json.write_text(json.dumps(summary, indent=2), encoding="utf-8")
            print(f"ISOLATED_VENV_E2E {'OK' if overall else 'FAILED'}")
            exit_code = 0 if overall else 1
    finally:
        sys.stdout = original_stdout
        if log_fp is not None:
            log_fp.close()
        if not args.keep:
            shutil.rmtree(work_root, ignore_errors=True)
        _cleanup_private_key(reason="finally")
        signal.signal(signal.SIGTERM, previous_sigterm)
        signal.signal(signal.SIGINT, previous_sigint)

    return exit_code


if __name__ == "__main__":
    sys.exit(main())
