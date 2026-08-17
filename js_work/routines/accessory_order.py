"""Deterministic garment-accessory demand and supplier-order workflow."""

from __future__ import annotations

import csv
import io
import json
import re
import tempfile
from dataclasses import asdict, dataclass
from decimal import ROUND_CEILING, Decimal, InvalidOperation
from hashlib import sha256
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from js_work.file_scope import MaterializedSnapshotPath, WorkFileSnapshot
from js_work.routines.legacy_xls import LegacyXlsError, convert_legacy_xls_to_xlsx
from js_work.routines.office_safety import (
    apply_work_cell_value,
    reject_formula_like_text,
    validate_safe_xlsx,
)
from js_work.safe_output import (
    ensure_absent,
    open_artifact,
    publish_no_clobber,
    remove_published_link,
    staged_path,
    write_json_no_clobber,
)


@dataclass(frozen=True)
class AccessoryOrderRunResult:
    status: str
    output_path: str
    report_path: str
    summary_row_count: int
    detail_row_count: int
    issue_count: int

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True)
class _SourceTable:
    rows: list[dict[str, Any]]
    issues: list[dict[str, Any]]
    sheet: str
    header_row: int


_QUANTITY_ALIASES = {
    "style": ("款号", "款式", "款式号", "STYLE", "STYLE NO", "STYLE NUMBER"),
    "garment_color": ("颜色", "色号", "成衣颜色", "COLOR", "COLOUR", "SHADE"),
    "size": ("尺码", "SIZE", "尺寸"),
    "quantity": ("数量", "订单数量", "订货数量", "QTY", "ORDER QTY", "QUANTITY"),
}
_BOM_ALIASES = {
    "style": _QUANTITY_ALIASES["style"],
    "accessory_code": ("辅料编码", "物料编码", "ITEM CODE", "MATERIAL CODE", "CODE"),
    "accessory_name": ("辅料名称", "物料名称", "品名", "ITEM", "MATERIAL", "NAME"),
    "spec": ("规格", "型号", "尺寸", "SPEC", "SPECIFICATION"),
    "garment_color": ("适用颜色", "成衣颜色", "颜色", "COLOR", "COLOUR"),
    "size": ("适用尺码", "尺码", "SIZE", "尺寸"),
    "usage": ("单件用量", "单位用量", "用量", "CONSUMPTION", "USAGE", "QTY/PC"),
    "unit": ("单位", "UNIT", "UOM"),
    "accessory_color": ("辅料颜色", "物料颜色", "ACCESSORY COLOR", "MATERIAL COLOR"),
    "material_type": ("物料类型", "类别", "MATERIAL TYPE", "TYPE"),
    "wastage": ("损耗率", "损耗", "WASTAGE", "WASTE RATE"),
}
_MASTER_ALIASES = {
    "accessory_code": _BOM_ALIASES["accessory_code"],
    "accessory_name": _BOM_ALIASES["accessory_name"],
    "supplier": ("供应商", "供应商名称", "SUPPLIER", "VENDOR"),
    "unit": _BOM_ALIASES["unit"],
    "wastage": _BOM_ALIASES["wastage"],
    "moq": ("最小起订量", "起订量", "MOQ", "MIN ORDER"),
    "pack_multiple": ("包装倍数", "包装数量", "PACK MULTIPLE", "PACK QTY"),
    "material_type": _BOM_ALIASES["material_type"],
    "accessory_color": _BOM_ALIASES["accessory_color"],
}
_FABRIC_TYPES = frozenset({"FABRIC", "SHELLFABRIC", "面料", "主面料"})
_ALL_MARKERS = frozenset({"", "*", "ALL", "全部", "通用"})
_MATCH_GARMENT_COLOR = frozenset({"同成衣色", "同色", "MATCH", "SELF", "SAMEASCOLOR"})
_MAX_REASONABLE_WASTAGE = Decimal("0.2")
_MAX_CSV_BYTES = 100 * 1024 * 1024
_CSV_ENCODINGS = ("utf-8-sig", "gbk", "gb18030")


class AccessoryOrderRoutineRunner:
    """Join quantity, style/BOM and accessory-master files without LLM arithmetic."""

    def __init__(self, workspace: Path, *, allowed_roots: tuple[Path, ...] | None = None) -> None:
        self.workspace = workspace.resolve()
        self.allowed_roots = (
            tuple(root.resolve() for root in allowed_roots) if allowed_roots is not None else None
        )

    def run(
        self,
        *,
        quantity_path: str | WorkFileSnapshot,
        style_path: str | WorkFileSnapshot,
        accessory_path: str | WorkFileSnapshot,
        output_path: str,
        color_aliases: dict[str, str] | None = None,
        unit_conversions: dict[str, float] | None = None,
    ) -> AccessoryOrderRunResult:
        normalized_color_aliases, public_color_aliases = _normalize_color_aliases(color_aliases)
        normalized_unit_conversions, public_unit_conversions = _normalize_unit_conversions(
            unit_conversions
        )
        quantity_source = self._resolve_source(quantity_path)
        style_source = self._resolve_source(style_path)
        accessory_source = self._resolve_source(accessory_path)
        output = self._resolve(output_path, must_exist=False)
        report_path = output.with_suffix(".validation.json")
        ensure_absent(output, "accessory order output already exists")
        ensure_absent(report_path, "accessory order validation report already exists")
        output.parent.mkdir(parents=True, exist_ok=True)
        source_hashes = {
            "quantity": _file_hash(quantity_source),
            "style_bom": _file_hash(style_source),
            "accessory_master": _file_hash(accessory_source),
        }

        quantities = self._read_source(
            quantity_source,
            aliases=_QUANTITY_ALIASES,
            required=("style", "garment_color", "quantity"),
            source_kind="quantity",
        )
        bom = self._read_source(
            style_source,
            aliases=_BOM_ALIASES,
            required=("style", "accessory_code", "usage"),
            source_kind="style_bom",
        )
        master = self._read_source(
            accessory_source,
            aliases=_MASTER_ALIASES,
            required=("accessory_code", "supplier", "unit"),
            source_kind="accessory_master",
        )

        issues: list[dict[str, Any]] = [*quantities.issues, *bom.issues, *master.issues]
        issue_keys = {_issue_key(issue) for issue in issues}

        master_by_code: dict[str, dict[str, Any]] = {}
        for row in master.rows:
            code = _key(row.get("accessory_code"))
            if not code:
                self._add_issue(
                    issues,
                    issue_keys,
                    "accessory_code_missing",
                    source=row.get("__source"),
                )
                continue
            if code in master_by_code:
                self._add_issue(
                    issues,
                    issue_keys,
                    "accessory_master_duplicate",
                    accessory_code=str(row.get("accessory_code", "")),
                    existing_code=str(master_by_code[code].get("accessory_code", "")),
                )
                continue
            master_by_code[code] = row

        bom_by_style: dict[str, list[dict[str, Any]]] = {}
        bom_seen: set[tuple[str, str, str, str]] = set()
        bom_scope_index: dict[tuple[str, str], list[dict[str, Any]]] = {}
        for row in bom.rows:
            style = _key(row.get("style"))
            if not style:
                self._add_issue(
                    issues,
                    issue_keys,
                    "style_missing",
                    source=row.get("__source"),
                )
                continue
            bom_key = (
                style,
                _key(row.get("accessory_code")),
                _canonical_alias(_key(row.get("garment_color")), normalized_color_aliases),
                _key(row.get("size")),
            )
            if bom_key in bom_seen:
                self._add_issue(
                    issues,
                    issue_keys,
                    "bom_duplicate",
                    accessory_code=_text(row.get("accessory_code")),
                    source=row.get("__source"),
                )
                continue
            scope_group = (style, _key(row.get("accessory_code")))
            if any(
                _scopes_overlap(row, existing, color_aliases=normalized_color_aliases)
                for existing in bom_scope_index.get(scope_group, [])
            ):
                self._add_issue(
                    issues,
                    issue_keys,
                    "bom_scope_overlap",
                    accessory_code=_text(row.get("accessory_code")),
                    source=row.get("__source"),
                )
                continue
            bom_seen.add(bom_key)
            bom_scope_index.setdefault(scope_group, []).append(row)
            bom_by_style.setdefault(style, []).append(row)

        details: list[dict[str, Any]] = []
        for quantity_row in quantities.rows:
            style_display = _text(quantity_row.get("style"))
            style = _key(style_display)
            garment_color = _text(quantity_row.get("garment_color"))
            size = _text(quantity_row.get("size"))
            quantity = _number(quantity_row.get("quantity"))
            if quantity is None or quantity <= 0:
                self._add_issue(
                    issues,
                    issue_keys,
                    "invalid_quantity",
                    source=quantity_row.get("__source"),
                    value=_text(quantity_row.get("quantity")),
                )
                continue
            matching_bom = [
                row
                for row in bom_by_style.get(style, [])
                if _matches_scope(
                    row.get("garment_color"),
                    garment_color,
                    aliases=normalized_color_aliases,
                )
                and _matches_scope(row.get("size"), size)
            ]
            if not matching_bom:
                self._add_issue(
                    issues,
                    issue_keys,
                    "style_bom_missing",
                    style=style_display,
                    garment_color=garment_color,
                    size=size,
                )
                continue

            for bom_row in matching_bom:
                code_display = _text(bom_row.get("accessory_code"))
                code = _key(code_display)
                material_type = _key(bom_row.get("material_type"))
                if material_type in _FABRIC_TYPES:
                    continue
                master_row = master_by_code.get(code)
                if master_row is None:
                    self._add_issue(
                        issues,
                        issue_keys,
                        "accessory_master_missing",
                        accessory_code=code_display,
                    )
                    continue
                if _key(master_row.get("material_type")) in _FABRIC_TYPES:
                    self._add_issue(
                        issues,
                        issue_keys,
                        "material_type_conflict",
                        accessory_code=code_display,
                        bom_material_type=_text(bom_row.get("material_type")),
                        master_material_type=_text(master_row.get("material_type")),
                    )
                    continue

                usage = _number(bom_row.get("usage"))
                if usage is None or usage <= 0:
                    self._add_issue(
                        issues,
                        issue_keys,
                        "invalid_usage",
                        accessory_code=code_display,
                        source=bom_row.get("__source"),
                    )
                    continue
                supplier = _text(master_row.get("supplier"))
                if not supplier:
                    self._add_issue(
                        issues,
                        issue_keys,
                        "supplier_missing",
                        accessory_code=code_display,
                    )
                    continue
                bom_unit = _text(bom_row.get("unit"))
                master_unit = _text(master_row.get("unit"))
                unit_conversion = Decimal(1)
                if bom_unit and master_unit and _key(bom_unit) != _key(master_unit):
                    unit_conversion = normalized_unit_conversions.get(
                        (_key(bom_unit), _key(master_unit)),
                        Decimal(0),
                    )
                    if unit_conversion <= 0:
                        self._add_issue(
                            issues,
                            issue_keys,
                            "unit_conflict",
                            accessory_code=code_display,
                            bom_unit=bom_unit,
                            master_unit=master_unit,
                        )
                        continue
                unit = master_unit or bom_unit
                bom_wastage_text = _text(bom_row.get("wastage"))
                master_wastage_text = _text(master_row.get("wastage"))
                if not bom_wastage_text and not master_wastage_text:
                    self._add_issue(
                        issues,
                        issue_keys,
                        "wastage_missing",
                        accessory_code=code_display,
                    )
                    continue
                wastage = _rate(
                    bom_row.get("wastage") if bom_wastage_text else master_row.get("wastage")
                )
                if wastage is None or wastage < 0:
                    self._add_issue(
                        issues,
                        issue_keys,
                        "invalid_wastage",
                        accessory_code=code_display,
                    )
                    continue
                if wastage > _MAX_REASONABLE_WASTAGE:
                    self._add_issue(
                        issues,
                        issue_keys,
                        "wastage_suspicious",
                        accessory_code=code_display,
                        wastage=str(wastage),
                    )
                    continue
                moq = _order_param(master_row.get("moq"), Decimal(0))
                pack_multiple = _order_param(master_row.get("pack_multiple"), Decimal(1))
                if moq is None or pack_multiple is None or moq < 0 or pack_multiple <= 0:
                    self._add_issue(
                        issues,
                        issue_keys,
                        "invalid_order_rounding",
                        accessory_code=code_display,
                    )
                    continue
                accessory_color = _resolve_accessory_color(
                    bom_row.get("accessory_color") or master_row.get("accessory_color"),
                    garment_color,
                )
                effective_usage = usage * unit_conversion
                theoretical = quantity * effective_usage
                required = theoretical * (Decimal(1) + wastage)
                details.append(
                    {
                        "供应商": supplier,
                        "款号": style_display,
                        "成衣颜色": garment_color,
                        "尺码": size,
                        "成衣数量": quantity,
                        "辅料编码": code_display,
                        "辅料名称": _text(bom_row.get("accessory_name"))
                        or _text(master_row.get("accessory_name")),
                        "规格": _text(bom_row.get("spec")),
                        "辅料颜色": accessory_color,
                        "单位": unit,
                        "单件用量": effective_usage,
                        "原始单件用量": usage,
                        "单位换算系数": unit_conversion,
                        "损耗率": wastage,
                        "理论数量": theoretical,
                        "需求数量": required,
                        "MOQ": moq,
                        "包装倍数": pack_multiple,
                        "数量来源": quantity_row.get("__source"),
                        "BOM来源": bom_row.get("__source"),
                        "主数据来源": master_row.get("__source"),
                    }
                )

        summary, aggregate_conflicts = _aggregate_details(
            details, color_aliases=normalized_color_aliases
        )
        for conflict in aggregate_conflicts:
            self._add_issue(
                issues,
                issue_keys,
                conflict["code"],
                **{key: value for key, value in conflict.items() if key != "code"},
            )
        status = "passed" if not issues else "needs_review"
        with staged_path(output) as staged:
            with open_artifact(staged, "w+b") as handle:
                self._write_workbook(
                    handle,
                    summary=summary,
                    details=details,
                    issues=issues,
                    status=status,
                )
            validate_safe_xlsx(staged)
            report = {
                "status": status,
                "source_hashes": source_hashes,
                "source_layouts": {
                    "quantity": {"sheet": quantities.sheet, "header_row": quantities.header_row},
                    "style_bom": {"sheet": bom.sheet, "header_row": bom.header_row},
                    "accessory_master": {"sheet": master.sheet, "header_row": master.header_row},
                },
                "applied_rules": {
                    "color_aliases": public_color_aliases,
                    "unit_conversions": public_unit_conversions,
                },
                "summary_rows": [_json_row(row) for row in summary],
                "detail_rows": [_json_row(row) for row in details],
                "issues": issues,
                "output_hash": _file_hash(staged),
            }
            publish_no_clobber(staged, output, "accessory order output already exists")
            try:
                write_json_no_clobber(
                    report_path,
                    report,
                    "accessory order validation report already exists",
                    anchor=staged,
                )
            except Exception:
                remove_published_link(staged, output)
                raise
        return AccessoryOrderRunResult(
            status=status,
            output_path=str(output),
            report_path=str(report_path),
            summary_row_count=len(summary),
            detail_row_count=len(details),
            issue_count=len(issues),
        )

    def _read_source(
        self,
        path: Path | WorkFileSnapshot,
        *,
        aliases: dict[str, tuple[str, ...]],
        required: tuple[str, ...],
        source_kind: str,
    ) -> _SourceTable:
        converted_tmp: tempfile.TemporaryDirectory[str] | None = None
        workbook_path = path
        try:
            if isinstance(path, WorkFileSnapshot) and path.suffix != ".csv":
                raise ValueError("Excel snapshots must be materialized before validation")
            if path.suffix.lower() == ".xls":
                if isinstance(path, WorkFileSnapshot):
                    raise ValueError("legacy Excel snapshots must be materialized")
                converted_tmp = tempfile.TemporaryDirectory(
                    prefix=".js_work_accessory_xls_",
                    dir=self.workspace,
                )
                try:
                    workbook_path = convert_legacy_xls_to_xlsx(
                        path,
                        Path(converted_tmp.name) / f"{path.stem}.xlsx",
                    )
                except LegacyXlsError as exc:
                    return _SourceTable(
                        rows=[],
                        issues=[
                            {
                                "code": "unsafe_legacy_xls",
                                "source": path.name,
                                "detail": str(exc),
                            }
                        ],
                        sheet="",
                        header_row=0,
                    )
            if workbook_path.suffix.lower() == ".csv":
                return _read_csv(workbook_path, aliases, required, source_kind)
            if isinstance(workbook_path, WorkFileSnapshot):
                raise ValueError("Excel snapshots must be materialized before validation")
            validate_safe_xlsx(workbook_path)
            return _read_workbook(workbook_path, aliases, required, source_kind, path.name)
        finally:
            if converted_tmp is not None:
                converted_tmp.cleanup()

    def _resolve_source(self, path: str | WorkFileSnapshot) -> Path | WorkFileSnapshot:
        if isinstance(path, WorkFileSnapshot):
            path.verified_data()
            return path
        return self._resolve(path)

    def _resolve(self, path: str | Path, *, must_exist: bool = True) -> Path:
        if isinstance(path, MaterializedSnapshotPath):
            if not must_exist:
                raise ValueError("materialized snapshots are input-only")
            return path
        raw = Path(path)
        candidate = raw if raw.is_absolute() else self.workspace / raw
        self._reject_symlink_components(candidate)
        resolved = candidate.resolve()
        try:
            resolved.relative_to(self.workspace)
        except ValueError as exc:
            raise ValueError(f"Path escapes workspace: {path}") from exc
        if self.allowed_roots is not None and not any(
            _is_relative_to(resolved, root) for root in self.allowed_roots
        ):
            raise ValueError(f"Path escapes allowed roots: {path}")
        if must_exist and not resolved.is_file():
            raise FileNotFoundError(str(path))
        return resolved

    def _reject_symlink_components(self, candidate: Path) -> None:
        from js_work.safe_output import reject_symlink_components

        reject_symlink_components(self.workspace, candidate)

    @staticmethod
    def _add_issue(
        issues: list[dict[str, Any]],
        issue_keys: set[str],
        code: str,
        **context: Any,
    ) -> None:
        issue = {"code": code, **context}
        key = _issue_key(issue)
        if key not in issue_keys:
            issue_keys.add(key)
            issues.append(issue)

    @staticmethod
    def _write_workbook(
        path: Any,
        *,
        summary: list[dict[str, Any]],
        details: list[dict[str, Any]],
        issues: list[dict[str, Any]],
        status: str,
    ) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "供应商下单汇总"
        ws.merge_cells("A1:Q1")
        ws["A1"] = "辅料供应商下单汇总"
        ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A2"] = "状态"
        ws["B2"] = status
        ws["D2"] = "说明"
        ws["E2"] = "needs_review 时不得直接作为供应商正式订单"
        summary_headers = [
            "供应商",
            "款号",
            "成衣颜色",
            "辅料编码",
            "辅料名称",
            "规格",
            "辅料颜色",
            "单位",
            "成衣数量",
            "理论数量",
            "损耗率",
            "需求数量",
            "MOQ",
            "包装倍数",
            "下单数量",
            "计算公式",
            "状态",
        ]
        _write_table(ws, summary_headers, summary, start_row=4)

        detail_ws = wb.create_sheet("计算明细")
        detail_headers = [
            "供应商",
            "款号",
            "成衣颜色",
            "尺码",
            "成衣数量",
            "辅料编码",
            "辅料名称",
            "规格",
            "辅料颜色",
            "单位",
            "单件用量",
            "原始单件用量",
            "单位换算系数",
            "损耗率",
            "理论数量",
            "需求数量",
            "MOQ",
            "包装倍数",
        ]
        _write_table(detail_ws, detail_headers, details, start_row=1)

        issue_ws = wb.create_sheet("异常审核")
        issue_headers = sorted({key for issue in issues for key in issue} | {"code"})
        _write_table(issue_ws, issue_headers, issues, start_row=1)

        trace_ws = wb.create_sheet("来源追踪")
        trace_headers = ["辅料编码", "款号", "成衣颜色", "数量来源", "BOM来源", "主数据来源"]
        _write_table(trace_ws, trace_headers, details, start_row=1)

        for sheet in wb.worksheets:
            sheet.freeze_panes = "A5" if sheet.title == "供应商下单汇总" else "A2"
            sheet.sheet_view.showGridLines = False
        wb.save(path)
        wb.close()


def _read_workbook(
    path: Path,
    aliases: dict[str, tuple[str, ...]],
    required: tuple[str, ...],
    source_kind: str,
    source_name: str,
) -> _SourceTable:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        best: tuple[int, Any, int, list[Any], dict[str, int]] | None = None
        alias_priority = {
            canonical: tuple(_header(alias) for alias in names)
            for canonical, names in aliases.items()
        }
        for ws in wb.worksheets:
            for row_number, values in enumerate(
                ws.iter_rows(min_row=1, max_row=min(ws.max_row, 50), values_only=True),
                start=1,
            ):
                normalized = [_header(value) for value in values]
                column_of: dict[str, int] = {}
                for index, value in enumerate(normalized):
                    if value and value not in column_of:
                        column_of[value] = index
                mapping: dict[str, int] = {}
                for canonical, names in alias_priority.items():
                    for normalized_alias in names:
                        column = column_of.get(normalized_alias)
                        if column is not None:
                            mapping[canonical] = column
                            break
                score = sum(1 for field in required if field in mapping)
                candidate = (score, ws, row_number, list(values), mapping)
                if best is None or candidate[0] > best[0]:
                    best = candidate
        if best is None:
            return _SourceTable(
                rows=[],
                issues=[{"code": "header_not_found", "source_kind": source_kind}],
                sheet="",
                header_row=0,
            )
        _score, ws, header_row, _headers, mapping = best
        issues = _ambiguous_header_issues(mapping, source_kind)
        missing = [field for field in required if field not in mapping]
        issues.extend(
            {"code": "required_header_missing", "source_kind": source_kind, "field": field}
            for field in missing
        )
        rows: list[dict[str, Any]] = []
        for row_number, values in enumerate(
            ws.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            if not values or all(value in (None, "") for value in values):
                continue
            row = {
                canonical: values[index] if index < len(values) else None
                for canonical, index in mapping.items()
            }
            row["__source"] = f"{source_name}:{ws.title}!{row_number}"
            rows.append(row)
        return _SourceTable(rows=rows, issues=issues, sheet=ws.title, header_row=header_row)
    finally:
        wb.close()


def _ambiguous_header_issues(
    mapping: dict[str, Any],
    source_kind: str,
) -> list[dict[str, Any]]:
    owners: dict[Any, list[str]] = {}
    for canonical, column in mapping.items():
        owners.setdefault(column, []).append(canonical)
    return [
        {
            "code": "ambiguous_header",
            "source_kind": source_kind,
            "fields": ",".join(sorted(canonicals)),
        }
        for column, canonicals in owners.items()
        if len(canonicals) > 1
    ]


def _read_csv(
    path: Path | WorkFileSnapshot,
    aliases: dict[str, tuple[str, ...]],
    required: tuple[str, ...],
    source_kind: str,
) -> _SourceTable:
    size = path.size if isinstance(path, WorkFileSnapshot) else path.stat().st_size
    if size > _MAX_CSV_BYTES:
        return _SourceTable(
            rows=[],
            issues=[
                {
                    "code": "csv_too_large",
                    "source_kind": source_kind,
                    "size_bytes": size,
                }
            ],
            sheet="",
            header_row=0,
        )
    raw = path.verified_data() if isinstance(path, WorkFileSnapshot) else path.read_bytes()
    text: str | None = None
    for encoding in _CSV_ENCODINGS:
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        return _SourceTable(
            rows=[],
            issues=[{"code": "csv_encoding_unsupported", "source_kind": source_kind}],
            sheet="",
            header_row=0,
        )
    reader = csv.DictReader(io.StringIO(text))
    headers = reader.fieldnames or []
    index = {_header(value): value for value in headers}
    mapping = {
        canonical: next((index[_header(alias)] for alias in names if _header(alias) in index), "")
        for canonical, names in aliases.items()
    }
    mapping = {canonical: header for canonical, header in mapping.items() if header}
    issues = _ambiguous_header_issues(mapping, source_kind)
    issues.extend(
        {"code": "required_header_missing", "source_kind": source_kind, "field": field}
        for field in required
        if field not in mapping
    )
    rows = []
    for row_number, raw_row in enumerate(reader, start=2):
        row = {canonical: raw_row.get(header) for canonical, header in mapping.items()}
        row["__source"] = f"{path.name}:CSV!{row_number}"
        rows.append(row)
    return _SourceTable(rows=rows, issues=issues, sheet="CSV", header_row=1)


def _aggregate_details(
    details: list[dict[str, Any]],
    *,
    color_aliases: dict[str, str] | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    aliases = color_aliases or {}
    groups: dict[tuple[str, ...], dict[str, Any]] = {}
    for detail in details:
        key = (
            _key(detail.get("供应商")),
            _key(detail.get("辅料编码")),
            _canonical_alias(_key(detail.get("辅料颜色")), aliases),
            _key(detail.get("单位")),
        )
        group = groups.setdefault(
            key,
            {
                "供应商": _text(detail.get("供应商")),
                "辅料编码": _text(detail.get("辅料编码")),
                "辅料颜色": _text(detail.get("辅料颜色")),
                "单位": _text(detail.get("单位")),
                "辅料名称": "",
                "规格": "",
                "名称集合": set(),
                "规格集合": set(),
                "款号集合": set(),
                "成衣颜色集合": set(),
                "成衣数量": Decimal(0),
                "理论数量": Decimal(0),
                "需求数量": Decimal(0),
                "损耗率集合": set(),
                "MOQ": detail["MOQ"],
                "包装倍数": detail["包装倍数"],
            },
        )
        name = _text(detail.get("辅料名称"))
        spec = _text(detail.get("规格"))
        group["名称集合"].add(name)
        group["规格集合"].add(spec)
        if not group["辅料名称"] and name:
            group["辅料名称"] = name
        if not group["规格"] and spec:
            group["规格"] = spec
        group["款号集合"].add(detail["款号"])
        group["成衣颜色集合"].add(detail["成衣颜色"])
        group["成衣数量"] += detail["成衣数量"]
        group["理论数量"] += detail["理论数量"]
        group["需求数量"] += detail["需求数量"]
        group["损耗率集合"].add(detail["损耗率"])

    conflicts: list[dict[str, Any]] = []
    rows: list[dict[str, Any]] = []
    for group in groups.values():
        names = {name for name in group["名称集合"] if name}
        specs = {spec for spec in group["规格集合"] if spec}
        if len(names) > 1 or len(specs) > 1:
            conflicts.append(
                {
                    "code": "spec_conflict",
                    "accessory_code": group["辅料编码"],
                    "names": ",".join(sorted(names)),
                    "specs": ",".join(sorted(specs)),
                }
            )
        multiple = group["包装倍数"]
        required = group["需求数量"]
        rounded = (required / multiple).to_integral_value(rounding=ROUND_CEILING) * multiple
        ordered = max(group["MOQ"], rounded)
        ordered = (ordered / multiple).to_integral_value(rounding=ROUND_CEILING) * multiple
        rates = sorted(group["损耗率集合"])
        rows.append(
            {
                "供应商": group["供应商"],
                "款号": ", ".join(sorted(group["款号集合"])),
                "成衣颜色": ", ".join(sorted(group["成衣颜色集合"])),
                "辅料编码": group["辅料编码"],
                "辅料名称": group["辅料名称"],
                "规格": group["规格"],
                "辅料颜色": group["辅料颜色"],
                "单位": group["单位"],
                "成衣数量": group["成衣数量"],
                "理论数量": group["理论数量"],
                "损耗率": rates[0] if len(rates) == 1 else "mixed",
                "需求数量": required,
                "MOQ": group["MOQ"],
                "包装倍数": multiple,
                "下单数量": ordered,
                "计算公式": "Σ(成衣数量×单件用量)×(1+损耗率)，再按MOQ/包装倍数向上取整",
                "状态": "待确认",
            }
        )
    rows.sort(
        key=lambda row: tuple(_text(row[field]) for field in ("供应商", "辅料编码", "辅料颜色"))
    )
    return rows, conflicts


def _write_table(
    ws: Any, headers: list[str], rows: list[dict[str, Any]], *, start_row: int
) -> None:
    if not headers:
        headers = ["code"]
    for column, header in enumerate(headers, start=1):
        cell = ws.cell(start_row, column, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_offset, row in enumerate(rows, start=1):
        for column, header in enumerate(headers, start=1):
            value = row.get(header)
            if isinstance(value, Decimal):
                value = _excel_number(value)
            reject_formula_like_text(value, label="accessory order value")
            cell = ws.cell(start_row + row_offset, column)
            if isinstance(value, str):
                apply_work_cell_value(cell, value)
            else:
                cell.value = value
    if rows:
        ws.auto_filter.ref = f"A{start_row}:{_column_letter(len(headers))}{start_row + len(rows)}"
    for index, header in enumerate(headers, start=1):
        width = min(42, max(12, len(str(header)) * 2 + 2))
        ws.column_dimensions[_column_letter(index)].width = width


def _column_letter(index: int) -> str:
    result = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _scopes_overlap(
    row_a: dict[str, Any],
    row_b: dict[str, Any],
    *,
    color_aliases: dict[str, str] | None = None,
) -> bool:
    aliases = color_aliases or {}
    color_a = _canonical_alias(_key(row_a.get("garment_color")), aliases)
    color_b = _canonical_alias(_key(row_b.get("garment_color")), aliases)
    return _scope_keys_overlap(color_a, color_b) and _scope_keys_overlap(
        _key(row_a.get("size")), _key(row_b.get("size"))
    )


def _scope_keys_overlap(key_a: str, key_b: str) -> bool:
    return key_a in _ALL_MARKERS or key_b in _ALL_MARKERS or key_a == key_b


def _matches_scope(
    rule: Any,
    actual: Any,
    *,
    aliases: dict[str, str] | None = None,
) -> bool:
    normalized_rule = _canonical_alias(_key(rule), aliases or {})
    normalized_actual = _canonical_alias(_key(actual), aliases or {})
    return normalized_rule in _ALL_MARKERS or normalized_rule == normalized_actual


def _canonical_alias(value: str, aliases: dict[str, str]) -> str:
    seen: set[str] = set()
    current = value
    while current in aliases:
        if current in seen:
            raise ValueError("Color alias cycle is not allowed")
        seen.add(current)
        current = aliases[current]
    return current


def _normalize_color_aliases(
    aliases: dict[str, str] | None,
) -> tuple[dict[str, str], dict[str, str]]:
    if aliases is None:
        return {}, {}
    if not isinstance(aliases, dict) or len(aliases) > 256:
        raise ValueError("Color aliases must be an object with at most 256 entries")
    normalized: dict[str, str] = {}
    public: dict[str, str] = {}
    for raw_source, raw_target in aliases.items():
        if not isinstance(raw_source, str) or not isinstance(raw_target, str):
            raise ValueError("Color aliases must map strings to strings")
        source = raw_source.strip()
        target = raw_target.strip()
        reject_formula_like_text(source, label="color alias")
        reject_formula_like_text(target, label="color alias")
        source_key = _key(source)
        target_key = _key(target)
        if not source_key or not target_key:
            raise ValueError("Color aliases cannot contain empty values")
        existing = normalized.get(source_key)
        if existing is not None and existing != target_key:
            raise ValueError("Conflicting normalized color aliases are not allowed")
        normalized[source_key] = target_key
        public[source] = target
    for source_key in normalized:
        _canonical_alias(source_key, normalized)
    return normalized, public


def _normalize_unit_conversions(
    conversions: dict[str, float] | None,
) -> tuple[dict[tuple[str, str], Decimal], dict[str, int | float]]:
    if conversions is None:
        return {}, {}
    if not isinstance(conversions, dict) or len(conversions) > 256:
        raise ValueError("Unit conversions must be an object with at most 256 entries")
    normalized: dict[tuple[str, str], Decimal] = {}
    public: dict[str, int | float] = {}
    for raw_rule, raw_factor in conversions.items():
        if not isinstance(raw_rule, str):
            raise ValueError("Unit conversion keys must be strings")
        separator = "->" if "->" in raw_rule else ""
        parts = raw_rule.split(separator) if separator else []
        if len(parts) != 2:
            raise ValueError("Unit conversions must use the 'source->target' form")
        source = parts[0].strip()
        target = parts[1].strip()
        source_key = _key(source)
        target_key = _key(target)
        if not source_key or not target_key:
            raise ValueError("Unit conversion units cannot be empty")
        factor = _number(raw_factor)
        if factor is None or factor <= 0:
            raise ValueError("Unit conversion factors must be positive finite numbers")
        pair = (source_key, target_key)
        existing = normalized.get(pair)
        if existing is not None and existing != factor:
            raise ValueError("Conflicting normalized unit conversions are not allowed")
        normalized[pair] = factor
        public[f"{source}->{target}"] = _decimal_json(factor)
    return normalized, public


def _resolve_accessory_color(value: Any, garment_color: str) -> str:
    text = _text(value)
    return garment_color if _key(text) in _MATCH_GARMENT_COLOR or not text else text


def _number(value: Any) -> Decimal | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, Decimal):
        return value
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        result = Decimal(text)
    except InvalidOperation:
        return None
    return result if result.is_finite() else None


def _order_param(value: Any, default: Decimal) -> Decimal | None:
    if isinstance(value, Decimal):
        return value if value.is_finite() else None
    if not _text(value):
        return default
    return _number(value)


def _rate(value: Any) -> Decimal | None:
    text = _text(value)
    if not text:
        return Decimal(0)
    percent = text.endswith("%")
    number = _number(text[:-1] if percent else text)
    if number is None:
        return None
    if percent or number > 1:
        number /= Decimal(100)
    return number


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _header(value: Any) -> str:
    return re.sub(r"[^0-9A-Z\u4e00-\u9fff]+", "", _text(value).upper())


def _key(value: Any) -> str:
    return _header(value)


def _issue_key(issue: dict[str, Any]) -> str:
    return json.dumps(issue, ensure_ascii=False, sort_keys=True, default=str)


def _json_row(row: dict[str, Any]) -> dict[str, Any]:
    return {
        key: _decimal_json(value) if isinstance(value, Decimal) else value
        for key, value in row.items()
    }


def _decimal_json(value: Decimal) -> int | float:
    return int(value) if value == value.to_integral_value() else float(value)


def _excel_number(value: Decimal) -> int | float:
    return _decimal_json(value)


def _is_relative_to(path: Path, root: Path) -> bool:
    try:
        path.relative_to(root)
    except ValueError:
        return False
    return True


def _file_hash(path: Path | WorkFileSnapshot) -> str:
    if isinstance(path, WorkFileSnapshot):
        path.verified_data()
        return path.sha256
    digest = sha256()
    with open_artifact(path) as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()
