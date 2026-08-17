"""Deterministic spreadsheet template routines for JS Agent Work."""

from __future__ import annotations

import csv
import io
import json
import shutil
from copy import copy
from dataclasses import asdict, dataclass
from hashlib import sha256
from pathlib import Path
from typing import Any

from js.echo.ledger.service import EchoSafetyService
from js.echo.turn_context import current_runtime_context
from js_work.file_scope import MaterializedSnapshotPath, WorkFileSnapshot
from js_work.routines.formula_cache import (
    evaluate_formula,
    referenced_ranges,
    refresh_formula_caches,
)
from js_work.routines.models import RoutineRunResult, WorkRoutine
from js_work.routines.office_safety import (
    apply_work_cell_value,
    reject_formula_like_text,
    validate_safe_xlsx,
)
from js_work.routines.store import DEFAULT_WORK_OWNER_KEY_HASH
from js_work.safe_output import (
    StagedArtifact,
    create_staged,
    discard_staged,
    ensure_absent,
    open_artifact,
    publish_no_clobber,
    reject_symlink_components,
    remove_published_link,
    write_json_no_clobber,
)


@dataclass(frozen=True)
class TemplateAnalysis:
    sheet_name: str
    header_row: int
    data_start_row: int
    headers: list[str]
    merged_cells: list[str]
    column_widths: dict[str, float]
    row_heights: dict[str, float]
    formulas: dict[str, str]
    number_formats: dict[str, str]
    template_hash: str

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True)
class ExtractionResult:
    rows: list[dict[str, Any]]
    source_row_count: int
    selected_source_row_count: int
    excluded_rows: list[dict[str, Any]]
    unmapped_fields: list[str]


class SpreadsheetTemplateEngine:
    """Analyze, render, extract, and validate Excel templates inside a workspace."""

    def __init__(self, workspace: Path) -> None:
        self.workspace = workspace.resolve()

    def analyze_template(self, path: str) -> TemplateAnalysis:
        target = self._resolve(path)
        validate_safe_xlsx(target)
        wb = None
        try:
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter

            wb = load_workbook(target, data_only=False)
            ws = wb.active
            if ws is None:
                raise ValueError("template workbook has no active sheet")
            header_row, headers = self._detect_header(ws)
            formulas: dict[str, str] = {}
            number_formats: dict[str, str] = {}
            for row in ws.iter_rows():
                for cell in row:
                    value = cell.value
                    if isinstance(value, str) and value.startswith("="):
                        formulas[cell.coordinate] = value
                    if cell.number_format and cell.number_format != "General":
                        number_formats[cell.coordinate] = cell.number_format

            column_widths: dict[str, float] = {}
            for idx in range(1, ws.max_column + 1):
                letter = get_column_letter(idx)
                width = ws.column_dimensions[letter].width
                if width is not None:
                    column_widths[letter] = width

            row_heights = {
                str(index): float(dim.height)
                for index, dim in ws.row_dimensions.items()
                if dim.height is not None
            }
            return TemplateAnalysis(
                sheet_name=ws.title,
                header_row=header_row,
                data_start_row=header_row + 1,
                headers=headers,
                merged_cells=[str(rng) for rng in ws.merged_cells.ranges],
                column_widths=column_widths,
                row_heights=row_heights,
                formulas=formulas,
                number_formats=number_formats,
                template_hash=file_hash(target),
            )
        finally:
            if wb is not None:
                wb.close()

    def extract_rows(
        self,
        source_path: str | WorkFileSnapshot,
        *,
        field_mapping: dict[str, str],
        sheet: str | None = None,
        row_filters: list[dict[str, Any]] | None = None,
        header_aliases: dict[str, list[str]] | None = None,
    ) -> list[dict[str, Any]]:
        return self.extract_table(
            source_path,
            field_mapping=field_mapping,
            sheet=sheet,
            row_filters=row_filters,
            header_aliases=header_aliases,
        ).rows

    def extract_table(
        self,
        source_path: str | WorkFileSnapshot,
        *,
        field_mapping: dict[str, str],
        sheet: str | None = None,
        row_filters: list[dict[str, Any]] | None = None,
        header_aliases: dict[str, list[str]] | None = None,
    ) -> ExtractionResult:
        target = (
            source_path
            if isinstance(source_path, WorkFileSnapshot)
            else self._resolve(source_path)
        )
        if target.suffix.lower() == ".csv":
            return self._extract_csv_table(
                target,
                field_mapping=field_mapping,
                row_filters=row_filters or [],
                header_aliases=header_aliases or {},
            )
        if isinstance(target, WorkFileSnapshot):
            raise ValueError("Excel snapshots must be materialized before validation")
        return self._extract_excel_table(
            target,
            field_mapping=field_mapping,
            sheet=sheet,
            row_filters=row_filters or [],
            header_aliases=header_aliases or {},
        )

    def _extract_excel_table(
        self,
        target: Path,
        *,
        field_mapping: dict[str, str],
        sheet: str | None,
        row_filters: list[dict[str, Any]],
        header_aliases: dict[str, list[str]],
    ) -> ExtractionResult:
        validate_safe_xlsx(target)
        wb = None
        try:
            from openpyxl import load_workbook

            wb = load_workbook(target, data_only=True, read_only=True)
            ws = wb[sheet] if sheet else wb.active
            if ws is None:
                raise ValueError("source sheet not found")
            header_row, source_headers = self._detect_header(ws)
            records: list[dict[str, Any]] = []
            for source_row_number, raw_row in enumerate(
                ws.iter_rows(min_row=header_row + 1, values_only=True),
                start=header_row + 1,
            ):
                if not raw_row or all(value in (None, "") for value in raw_row):
                    continue
                records.append(
                    {
                        "__row_number": source_row_number,
                        **{
                            header: raw_row[idx] if idx < len(raw_row) else ""
                            for idx, header in enumerate(source_headers)
                            if header
                        },
                    }
                )
            return self._map_and_filter_records(
                records,
                field_mapping=field_mapping,
                row_filters=row_filters,
                header_aliases=header_aliases,
            )
        finally:
            if wb is not None:
                wb.close()

    def _extract_csv_table(
        self,
        target: Path | WorkFileSnapshot,
        *,
        field_mapping: dict[str, str],
        row_filters: list[dict[str, Any]],
        header_aliases: dict[str, list[str]],
    ) -> ExtractionResult:
        handle_context = (
            io.StringIO(target.verified_data().decode("utf-8-sig"), newline="")
            if isinstance(target, WorkFileSnapshot)
            else target.open("r", encoding="utf-8-sig", newline="")
        )
        with handle_context as handle:
            reader = csv.DictReader(handle)
            records = [
                {"__row_number": row_number, **{k: _coerce_scalar(v) for k, v in row.items()}}
                for row_number, row in enumerate(reader, start=2)
            ]
        return self._map_and_filter_records(
            records,
            field_mapping=field_mapping,
            row_filters=row_filters,
            header_aliases=header_aliases,
        )

    def _map_and_filter_records(
        self,
        records: list[dict[str, Any]],
        *,
        field_mapping: dict[str, str],
        row_filters: list[dict[str, Any]],
        header_aliases: dict[str, list[str]],
    ) -> ExtractionResult:
        source_headers = [key for key in records[0] if key != "__row_number"] if records else []
        source_index = {normalize_header(header): header for header in source_headers if header}
        selected_rows: list[dict[str, Any]] = []
        excluded_rows: list[dict[str, Any]] = []
        unmapped_fields: list[str] = []
        mapped_headers: dict[str, str] = {}
        for output_field, source_field in field_mapping.items():
            header = _find_source_header(
                source_field,
                source_index,
                header_aliases=header_aliases,
                output_field=output_field,
            )
            if header is None:
                unmapped_fields.append(output_field)
            else:
                mapped_headers[output_field] = header

        for record in records:
            matched, reason = _record_matches_filters(record, row_filters, header_aliases)
            if not matched:
                excluded_rows.append(
                    {
                        "source_row": record.get("__row_number"),
                        "reason": reason or "row_filter_not_matched",
                    }
                )
                continue
            item = {
                output_field: record.get(source_header, "")
                for output_field, source_header in mapped_headers.items()
            }
            if any(value not in (None, "") for value in item.values()):
                selected_rows.append(item)

        return ExtractionResult(
            rows=selected_rows,
            source_row_count=len(records),
            selected_source_row_count=len(selected_rows),
            excluded_rows=excluded_rows,
            unmapped_fields=unmapped_fields,
        )

    def render_from_template(
        self,
        *,
        template_path: str,
        output_path: str,
        rows: list[dict[str, Any]],
        data_start_row: int | None = None,
    ) -> Path:
        output, staged = self._render_from_template_with_anchor(
            template_path=template_path,
            output_path=output_path,
            rows=rows,
            data_start_row=data_start_row,
        )
        try:
            return output
        finally:
            discard_staged(staged)

    def _render_from_template_with_anchor(
        self,
        *,
        template_path: str,
        output_path: str,
        rows: list[dict[str, Any]],
        data_start_row: int | None = None,
    ) -> tuple[Path, StagedArtifact]:
        template = self._resolve(template_path)
        output = self._resolve(output_path, must_exist=False)
        ensure_absent(output, "output workbook already exists")
        output.parent.mkdir(parents=True, exist_ok=True)
        staged = create_staged(output)
        try:
            with open_artifact(template) as source, open_artifact(staged, "w+b") as target:
                shutil.copyfileobj(source, target)
            wb = None
            try:
                from openpyxl import load_workbook

                analysis = self.analyze_template(template_path)
                with open_artifact(staged) as handle:
                    wb = load_workbook(handle, data_only=False)
                ws = wb[analysis.sheet_name]
                start_row = data_start_row or analysis.data_start_row
                header_to_col = {
                    normalize_header(header): index + 1
                    for index, header in enumerate(analysis.headers)
                    if header
                }
                anchor_row = start_row
                for row_offset, row_data in enumerate(rows):
                    row_index = start_row + row_offset
                    self._copy_row_style(ws, source_row=anchor_row, target_row=row_index)
                    for header, value in row_data.items():
                        col_index = header_to_col.get(normalize_header(header))
                        if col_index is None:
                            continue
                        reject_formula_like_text(value, label="generated spreadsheet value")
                        apply_work_cell_value(
                            ws.cell(row=row_index, column=col_index),
                            value,
                        )
                with open_artifact(staged, "w+b") as handle:
                    wb.save(handle)
                validate_safe_xlsx(staged)
                _recalculate_workbook_formulas(staged)
                validate_safe_xlsx(staged)
            finally:
                if wb is not None:
                    wb.close()
            publish_no_clobber(staged, output, "output workbook already exists")
            return output, staged
        except BaseException:
            discard_staged(staged)
            raise

    def validate_output(
        self,
        *,
        source_path: str | WorkFileSnapshot,
        template_path: str,
        output_path: str,
        rows: list[dict[str, Any]],
        required_fields: list[str] | None = None,
        source_row_count: int | None = None,
        selected_source_row_count: int | None = None,
        excluded_rows: list[dict[str, Any]] | None = None,
        unmapped_fields: list[str] | None = None,
        write_report: bool = True,
    ) -> dict[str, Any]:
        source = (
            source_path
            if isinstance(source_path, WorkFileSnapshot)
            else self._resolve(source_path)
        )
        template = self._resolve(template_path)
        output = self._resolve(output_path)
        if source.suffix.lower() == ".xlsx":
            if isinstance(source, WorkFileSnapshot):
                raise ValueError("Excel snapshots must be materialized before validation")
            validate_safe_xlsx(source)
        validate_safe_xlsx(template)
        validate_safe_xlsx(output)
        analysis = self.analyze_template(template_path)
        issues: list[dict[str, Any]] = []
        required = required_fields or []
        for idx, row in enumerate(rows, start=1):
            for field in required:
                if row.get(field) in (None, ""):
                    issues.append({"code": "required_missing", "row": idx, "field": field})

        seen_units: dict[tuple[Any, Any], Any] = {}
        seen_rows: set[tuple[tuple[str, str], ...]] = set()
        totals: dict[str, float] = {}
        for idx, row in enumerate(rows, start=1):
            canonical = tuple(sorted((str(k), str(v)) for k, v in row.items()))
            if canonical in seen_rows:
                issues.append({"code": "duplicate_row", "row": idx})
            seen_rows.add(canonical)

            key = (row.get("面料名称") or row.get("品名"), row.get("规格"))
            unit = row.get("单位")
            if key[0] and key[1] and unit:
                previous = seen_units.get(key)
                if previous is not None and previous != unit:
                    issues.append(
                        {
                            "code": "unit_conflict",
                            "row": idx,
                            "key": list(key),
                            "units": [previous, unit],
                        }
                    )
                seen_units[key] = unit

            for field, value in row.items():
                if isinstance(value, (int, float)) and not isinstance(value, bool):
                    totals[field] = round(totals.get(field, 0.0) + float(value), 6)

        if source_row_count and selected_source_row_count == 0:
            issues.append({"code": "no_rows_selected"})
        for field in unmapped_fields or []:
            issues.append({"code": "field_unmapped", "field": field})

        formula_checks = self._check_formulas(output, analysis, row_count=len(rows))
        for coordinate, check in formula_checks.items():
            if check["status"] != "passed":
                issues.append(
                    {
                        "code": check["status"],
                        "cell": coordinate,
                        "formula": check["formula"],
                    }
                )

        template_fidelity = self._check_template_fidelity(template_path, output_path)
        for issue in template_fidelity["issues"]:
            issues.append(issue)

        report = {
            "status": "passed" if not issues else "needs_review",
            "source_hash": file_hash(source),
            "template_hash": file_hash(template),
            "output_hash": file_hash(output),
            "sheet": analysis.sheet_name,
            "row_count": len(rows),
            "source_row_count": source_row_count if source_row_count is not None else len(rows),
            "selected_source_row_count": (
                selected_source_row_count if selected_source_row_count is not None else len(rows)
            ),
            "excluded_source_row_count": len(excluded_rows or []),
            "excluded_rows": excluded_rows or [],
            "unmapped_fields": unmapped_fields or [],
            "headers": analysis.headers,
            "totals": totals,
            "issues": issues,
            "formulas": analysis.formulas,
            "formula_checks": formula_checks,
            "template_fidelity": template_fidelity,
        }
        report_path = output.with_suffix(".validation.json")
        if write_report:
            write_json_no_clobber(
                report_path,
                report,
                "validation report already exists",
            )
        return report

    def _resolve(self, path: str | Path, *, must_exist: bool = True) -> Path:
        if isinstance(path, MaterializedSnapshotPath):
            if not must_exist:
                raise ValueError("materialized snapshots are input-only")
            return path
        raw = Path(path)
        target = raw if raw.is_absolute() else self.workspace / raw
        reject_symlink_components(self.workspace, target)
        resolved = target.resolve()
        try:
            resolved.relative_to(self.workspace)
        except ValueError as e:
            raise ValueError(f"Path escapes workspace: {path}") from e
        if must_exist and not resolved.exists():
            raise FileNotFoundError(str(path))
        return resolved

    def _check_formulas(
        self,
        output: Path,
        analysis: TemplateAnalysis,
        *,
        row_count: int,
    ) -> dict[str, dict[str, Any]]:
        if not analysis.formulas:
            return {}
        from openpyxl import load_workbook

        wb = None
        values_wb = None
        checks: dict[str, dict[str, Any]] = {}
        try:
            values_wb = load_workbook(output, data_only=True)
            wb = load_workbook(output, data_only=False)
            ws = wb[analysis.sheet_name]
            values_ws = values_wb[analysis.sheet_name]
            for coordinate, formula in analysis.formulas.items():
                computed = _evaluate_simple_formula(ws, formula)
                if computed is None:
                    checks[coordinate] = {
                        "formula": formula,
                        "status": "unsupported_formula",
                    }
                    continue
                formula_range_issue = _formula_range_issue(
                    formula,
                    data_start_row=analysis.data_start_row,
                    row_count=row_count,
                )
                cached = values_ws[coordinate].value
                if formula_range_issue is not None:
                    checks[coordinate] = {
                        "formula": formula,
                        "status": formula_range_issue,
                        "computed": computed,
                        "cached": cached,
                    }
                    continue
                if cached in (None, ""):
                    checks[coordinate] = {
                        "formula": formula,
                        "status": "formula_value_missing",
                        "computed": computed,
                        "cached": cached,
                    }
                    continue
                if not _numbers_close(cached, computed):
                    checks[coordinate] = {
                        "formula": formula,
                        "status": "formula_value_mismatch",
                        "computed": computed,
                        "cached": cached,
                    }
                    continue
                checks[coordinate] = {
                    "formula": formula,
                    "status": "passed",
                    "computed": computed,
                    "cached": cached,
                }
            return checks
        finally:
            if wb is not None:
                wb.close()
            if values_wb is not None:
                values_wb.close()

    def _check_template_fidelity(
        self,
        template_path: str,
        output_path: str,
    ) -> dict[str, Any]:
        template = self.analyze_template(template_path)
        output = self.analyze_template(output_path)
        issues: list[dict[str, Any]] = []
        if template.headers != output.headers:
            issues.append({"code": "template_headers_changed"})
        if set(template.merged_cells) != set(output.merged_cells):
            issues.append({"code": "template_merged_cells_changed"})
        for column, width in template.column_widths.items():
            if output.column_widths.get(column) != width:
                issues.append({"code": "template_column_width_changed", "column": column})
        return {
            "status": "passed" if not issues else "needs_review",
            "issues": issues,
        }

    @staticmethod
    def _detect_header(ws: Any) -> tuple[int, list[str]]:
        best_row = 1
        best_headers: list[str] = []
        best_score = -1
        max_row = min(ws.max_row or 1, 20)
        for row_index in range(1, max_row + 1):
            values = [
                "" if cell.value is None else str(cell.value).strip() for cell in ws[row_index]
            ]
            non_empty = [value for value in values if value]
            score = len(non_empty)
            if any(value in {"面料名称", "品名", "规格", "数量", "单位"} for value in non_empty):
                score += 10
            if score > best_score and len(non_empty) >= 2:
                best_score = score
                best_row = row_index
                best_headers = values
        return best_row, best_headers

    @staticmethod
    def _copy_row_style(ws: Any, *, source_row: int, target_row: int) -> None:
        if source_row == target_row:
            return
        for col_index in range(1, ws.max_column + 1):
            source = ws.cell(row=source_row, column=col_index)
            target = ws.cell(row=target_row, column=col_index)
            if source.has_style:
                target._style = copy(source._style)
            if source.number_format:
                target.number_format = source.number_format


class DeterministicRoutineReviewer:
    """Cheap reviewer that only sees validation summaries."""

    def review(self, summary: dict[str, Any]) -> dict[str, Any]:
        issues = summary.get("issues") or []
        return {"status": "passed" if not issues else "needs_review", "issues": issues}


class WorkSpreadsheetRoutineRunner:
    """End-to-end runner for spreadsheet template routines."""

    def __init__(self, workspace: Path, state_dir: Path) -> None:
        self.workspace = workspace
        self.state_dir = state_dir
        self.engine = SpreadsheetTemplateEngine(workspace)
        self.echo_ledger = EchoSafetyService(state_dir=state_dir)

    def run(
        self,
        *,
        routine: WorkRoutine,
        source_path: str | WorkFileSnapshot,
        template_path: str,
        output_path: str,
        reviewer: Any | None = None,
    ) -> RoutineRunResult:
        if not routine.enabled:
            raise PermissionError("routine must be approved before it can run")
        output_candidate = self.engine._resolve(output_path, must_exist=False)
        report_candidate = output_candidate.with_suffix(".validation.json")
        ensure_absent(output_candidate, "output workbook already exists")
        ensure_absent(report_candidate, "validation report already exists")
        row_filters = _effective_row_filters(routine)
        extraction = self.engine.extract_table(
            source_path,
            field_mapping=routine.field_mapping,
            sheet=routine.source_sheet or None,
            row_filters=row_filters,
            header_aliases=routine.header_aliases,
        )
        rows = _aggregate_rows(
            extraction.rows,
            _effective_aggregation_rules(routine, extraction.rows),
        )
        output, staged_output = self.engine._render_from_template_with_anchor(
            template_path=template_path,
            output_path=output_path,
            rows=rows,
        )
        report_committed = False
        try:
            required = list(routine.validation_rules.get("required_fields", []))
            report = self.engine.validate_output(
                source_path=source_path,
                template_path=template_path,
                output_path=output_path,
                rows=rows,
                required_fields=required,
                source_row_count=extraction.source_row_count,
                selected_source_row_count=extraction.selected_source_row_count,
                excluded_rows=extraction.excluded_rows,
                unmapped_fields=extraction.unmapped_fields,
                write_report=False,
            )
            review_summary = {
                "routine_id": routine.routine_id,
                "routine_name": routine.name,
                "row_count": len(rows),
                "source_row_count": report["source_row_count"],
                "selected_source_row_count": report["selected_source_row_count"],
                "excluded_source_row_count": report["excluded_source_row_count"],
                "headers": report["headers"],
                "totals": report["totals"],
                "issues": report["issues"],
                "source_hash": report["source_hash"],
                "template_hash": report["template_hash"],
                "output_hash": report["output_hash"],
                "formula_checks": report["formula_checks"],
            }
            reviewer_result = (reviewer or DeterministicRoutineReviewer()).review(
                review_summary
            )
            issues = list(report["issues"])
            for issue in reviewer_result.get("issues", []):
                if issue not in issues:
                    issues.append(issue)
            status = (
                "passed"
                if report["status"] == "passed"
                and reviewer_result.get("status") == "passed"
                else "needs_review"
            )
            final_report = {
                **report,
                "status": status,
                "issues": issues,
                "reviewer": reviewer_result,
            }
            report_path = output.with_suffix(".validation.json")
            write_json_no_clobber(
                report_path,
                final_report,
                "validation report already exists",
                anchor=staged_output,
            )
            report_committed = True
            self._record_echo_ledger(routine=routine, report=final_report)
            return RoutineRunResult(
                status=status,
                output_path=str(output),
                report_path=str(report_path),
                row_count=len(rows),
                issues=issues,
                reviewer=reviewer_result,
            )
        except BaseException:
            if not report_committed:
                try:
                    remove_published_link(staged_output, output)
                except Exception as rollback_error:
                    raise RuntimeError(
                        "spreadsheet output rollback could not be confirmed"
                    ) from rollback_error
            raise
        finally:
            discard_staged(staged_output)

    def preview(
        self,
        *,
        routine: WorkRoutine,
        source_path: str | WorkFileSnapshot,
        template_path: str,
    ) -> dict[str, Any]:
        _ = template_path
        if not routine.enabled:
            raise PermissionError("routine must be approved before it can run")
        extraction = self.engine.extract_table(
            source_path,
            field_mapping=routine.field_mapping,
            sheet=routine.source_sheet or None,
            row_filters=_effective_row_filters(routine),
            header_aliases=routine.header_aliases,
        )
        rows = _aggregate_rows(
            extraction.rows,
            _effective_aggregation_rules(routine, extraction.rows),
        )
        issues = _row_validation_issues(
            rows,
            required_fields=list(routine.validation_rules.get("required_fields", [])),
            source_row_count=extraction.source_row_count,
            selected_source_row_count=extraction.selected_source_row_count,
            unmapped_fields=extraction.unmapped_fields,
        )
        return {
            "status": "passed" if not issues else "needs_review",
            "dry_run": True,
            "row_count": len(rows),
            "source_row_count": extraction.source_row_count,
            "selected_source_row_count": extraction.selected_source_row_count,
            "excluded_source_row_count": len(extraction.excluded_rows),
            "excluded_rows": extraction.excluded_rows,
            "unmapped_fields": extraction.unmapped_fields,
            "totals": _totals(rows),
            "issues": issues,
            "rows_preview": rows[:20],
        }

    def _record_echo_ledger(self, *, routine: WorkRoutine, report: dict[str, Any]) -> None:
        """Record a compact, auditable routine result without storing table rows."""
        context = current_runtime_context()
        tenant_id = (
            context.owner_key_hash
            if context is not None and context.owner_key_hash
            else DEFAULT_WORK_OWNER_KEY_HASH
        )
        run_id = (
            context.run_id
            if context is not None and context.run_id
            else f"{routine.routine_id}:{report['output_hash']}"
        )
        user_text = json.dumps(
            {
                "kind": "work_routine_run",
                "routine_id": routine.routine_id,
                "routine_version": routine.version,
                "source_hash": report["source_hash"],
                "template_hash": report["template_hash"],
                "output_hash": report["output_hash"],
                "session_id": context.session_id if context is not None else "",
            },
            ensure_ascii=False,
            sort_keys=True,
        )
        assistant_text = json.dumps(
            {
                "status": report["status"],
                "row_count": report["row_count"],
                "totals": report["totals"],
                "issue_count": len(report["issues"]),
                "reviewer_status": (report.get("reviewer") or {}).get("status"),
            },
            ensure_ascii=False,
            sort_keys=True,
        )
        self.echo_ledger.record_chat_turn(
            tenant_id=tenant_id,
            run_id=run_id,
            user_text=user_text,
            assistant_text=assistant_text,
            status=report["status"],
            token_totals={"input": 0, "output": 0},
        )


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    return "".join(ch for ch in text if not ch.isspace())


def _find_source_header(
    source_field: str,
    source_index: dict[str, str],
    *,
    header_aliases: dict[str, list[str]],
    output_field: str,
) -> str | None:
    candidates = [source_field, output_field, _source_synonym_key(source_field)]
    candidates.extend(header_aliases.get(source_field, []))
    candidates.extend(header_aliases.get(output_field, []))
    for candidate in candidates:
        key = normalize_header(candidate)
        if key in source_index:
            return source_index[key]
    return None


def _record_matches_filters(
    record: dict[str, Any],
    row_filters: list[dict[str, Any]],
    header_aliases: dict[str, list[str]],
) -> tuple[bool, str | None]:
    if not row_filters:
        return True, None
    source_index = {
        normalize_header(header): header for header in record if header != "__row_number"
    }
    for rule in row_filters:
        field = str(rule.get("field") or "")
        header = _find_source_header(
            field,
            source_index,
            header_aliases=header_aliases,
            output_field=field,
        )
        value = "" if header is None else str(record.get(header, "") or "")
        if header is None:
            if rule.get("optional_if_missing"):
                continue
            return False, "row_filter_field_missing"
        if "equals" in rule and value != str(rule["equals"]):
            return False, "row_filter_not_matched"
        if "not_equals" in rule and value == str(rule["not_equals"]):
            return False, "row_filter_not_matched"
        if "contains" in rule and str(rule["contains"]) not in value:
            return False, "row_filter_not_matched"
        if "not_contains" in rule and str(rule["not_contains"]) in value:
            return False, "row_filter_not_matched"
        include_contains = [str(item) for item in rule.get("include_contains", [])]
        if include_contains and not any(item in value for item in include_contains):
            return False, "row_filter_not_matched"
        exclude_contains = [str(item) for item in rule.get("exclude_contains", [])]
        if exclude_contains and any(item in value for item in exclude_contains):
            return False, "row_filter_not_matched"
        if rule.get("required") and value == "":
            return False, "row_filter_not_matched"
    return True, None


def _effective_row_filters(routine: WorkRoutine) -> list[dict[str, Any]]:
    if routine.row_filters:
        return routine.row_filters
    haystack = " ".join([routine.name, *routine.trigger_phrases])
    if "面料" in haystack:
        return [{"field": "类别", "equals": "面料", "optional_if_missing": True}]
    return []


def _effective_aggregation_rules(
    routine: WorkRoutine,
    rows: list[dict[str, Any]],
) -> dict[str, Any]:
    if routine.aggregation_rules:
        return routine.aggregation_rules
    if not rows:
        return {}
    fields = set().union(*(row.keys() for row in rows))
    if {"面料名称", "规格", "单位", "数量"} <= fields:
        return {
            "group_by": ["面料名称", "规格", "单位"],
            "sum_fields": ["数量"],
            "merge_text_fields": ["备注"] if "备注" in fields else [],
        }
    return {}


def _aggregate_rows(
    rows: list[dict[str, Any]],
    aggregation_rules: dict[str, Any],
) -> list[dict[str, Any]]:
    group_by = [str(field) for field in aggregation_rules.get("group_by", [])]
    if not group_by:
        return rows
    sum_fields = {str(field) for field in aggregation_rules.get("sum_fields", [])}
    merge_text_fields = {str(field) for field in aggregation_rules.get("merge_text_fields", [])}
    grouped: dict[tuple[Any, ...], dict[str, Any]] = {}
    for row in rows:
        key = tuple(row.get(field, "") for field in group_by)
        if key not in grouped:
            grouped[key] = dict(row)
            continue
        target = grouped[key]
        for field, value in row.items():
            if field in sum_fields:
                target[field] = _numeric_sum(target.get(field), value)
            elif field in merge_text_fields:
                target[field] = _merge_text(target.get(field), value)
            elif field not in target or target[field] in (None, ""):
                target[field] = value
    return list(grouped.values())


def _row_validation_issues(
    rows: list[dict[str, Any]],
    *,
    required_fields: list[str],
    source_row_count: int | None,
    selected_source_row_count: int | None,
    unmapped_fields: list[str],
) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []
    for idx, row in enumerate(rows, start=1):
        for field in required_fields:
            if row.get(field) in (None, ""):
                issues.append({"code": "required_missing", "row": idx, "field": field})
    if source_row_count and selected_source_row_count == 0:
        issues.append({"code": "no_rows_selected"})
    for field in unmapped_fields:
        issues.append({"code": "field_unmapped", "field": field})

    seen_units: dict[tuple[Any, Any], Any] = {}
    seen_rows: set[tuple[tuple[str, str], ...]] = set()
    for idx, row in enumerate(rows, start=1):
        canonical = tuple(sorted((str(k), str(v)) for k, v in row.items()))
        if canonical in seen_rows:
            issues.append({"code": "duplicate_row", "row": idx})
        seen_rows.add(canonical)

        key = (row.get("面料名称") or row.get("品名"), row.get("规格"))
        unit = row.get("单位")
        if key[0] and key[1] and unit:
            previous = seen_units.get(key)
            if previous is not None and previous != unit:
                issues.append(
                    {
                        "code": "unit_conflict",
                        "row": idx,
                        "key": list(key),
                        "units": [previous, unit],
                    }
                )
            seen_units[key] = unit
    return issues


def _totals(rows: list[dict[str, Any]]) -> dict[str, float]:
    totals: dict[str, float] = {}
    for row in rows:
        for field, value in row.items():
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                totals[field] = round(totals.get(field, 0.0) + float(value), 6)
    return totals


def _numeric_sum(left: Any, right: Any) -> float | int:
    total = float(left or 0) + float(right or 0)
    return int(total) if total.is_integer() else round(total, 6)


def _numbers_close(left: Any, right: Any) -> bool:
    if not isinstance(left, (int, float)) or isinstance(left, bool):
        return bool(left == right)
    if not isinstance(right, (int, float)) or isinstance(right, bool):
        return False
    return math_isclose(float(left), float(right))


def math_isclose(left: float, right: float) -> bool:
    return abs(left - right) <= 0.0001


def _recalculate_workbook_formulas(path: Path) -> None:
    refresh_formula_caches(path, soffice=_find_soffice())


def _find_soffice() -> Path | None:
    found = shutil.which("soffice")
    if found:
        return Path(found)
    for candidate in (
        Path("/Applications/LibreOffice.app/Contents/MacOS/soffice"),
        Path("/opt/homebrew/bin/soffice"),
        Path("/usr/local/bin/soffice"),
    ):
        if candidate.exists():
            return candidate
    return None


def _merge_text(left: Any, right: Any) -> str:
    values: list[str] = []
    for value in (left, right):
        text = str(value or "").strip()
        if text and text not in values:
            values.append(text)
    return "；".join(values)


def _coerce_scalar(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    text = value.strip()
    if text == "":
        return ""
    try:
        number = float(text)
    except ValueError:
        return text
    return int(number) if number.is_integer() else number


def _evaluate_simple_formula(ws: Any, formula: str) -> float | int | None:
    return evaluate_formula(ws, formula)


def _formula_range_issue(
    formula: str,
    *,
    data_start_row: int,
    row_count: int,
) -> str | None:
    if row_count <= 0:
        return None
    generated_last_row = data_start_row + row_count - 1
    for formula_range in referenced_ranges(formula):
        if formula_range.min_row <= data_start_row and formula_range.max_row < generated_last_row:
            return "formula_range_excludes_data"
    return None


def _source_synonym_key(value: str) -> str:
    synonyms = {
        "面料名称": "品名",
        "面料": "品名",
        "材质": "品名",
        "类别": "类别",
    }
    return normalize_header(synonyms.get(value, value))


def file_hash(path: Path | WorkFileSnapshot) -> str:
    if isinstance(path, WorkFileSnapshot):
        path.verified_data()
        return f"sha256:{path.sha256}"
    digest = sha256()
    with open_artifact(path) as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            digest.update(chunk)
    return f"sha256:{digest.hexdigest()}"
