"""Deterministic formula evaluation and cached-value updates for XLSX files."""

from __future__ import annotations

import fcntl
import math
import os
import posixpath
import shutil
import stat
import subprocess
import sys
import tempfile
import threading
import xml.etree.ElementTree as ET
import zipfile
from collections.abc import Iterator, Mapping
from contextlib import contextmanager
from dataclasses import dataclass
from functools import lru_cache
from hashlib import sha256
from pathlib import Path, PurePosixPath
from typing import Any

from js.utils.log import get_logger
from js_work.safe_output import StagedArtifact, open_artifact, rewrite_artifact

logger = get_logger("js_work.routines.formula_cache")

type Number = int | float

_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_DOCUMENT_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
_PROCESS_LOCKS: dict[str, threading.Lock] = {}
_PROCESS_LOCKS_GUARD = threading.Lock()
MAX_FORMULA_RANGE_CELLS = 100_000
MAX_FORMULA_TOTAL_CELL_VISITS = 250_000
_MACOS_NETWORK_DENY_PROFILE = "(version 1)(allow default)(deny network*)"
_NETWORK_SANDBOX_PROBE = """\
import socket
import sys

sock = socket.socket()
try:
    sock.bind(("127.0.0.1", 0))
except PermissionError:
    sys.exit(0)
except OSError:
    sys.exit(2)
else:
    sys.exit(3)
"""


@dataclass(frozen=True)
class FormulaCacheResult:
    cached_values: dict[str, dict[str, Number]]
    unsupported_formulas: dict[str, dict[str, str]]
    libreoffice_used: bool

    @property
    def cached_count(self) -> int:
        return sum(len(values) for values in self.cached_values.values())

    @property
    def complete(self) -> bool:
        return not any(self.unsupported_formulas.values())


@dataclass(frozen=True)
class LibreOfficeConversionResult:
    output_path: Path | None
    detail: str
    timed_out: bool = False


@dataclass(frozen=True)
class FormulaRange:
    min_col: int
    min_row: int
    max_col: int
    max_row: int


@dataclass(frozen=True)
class _Token:
    kind: str
    value: str


@dataclass(frozen=True)
class _NumberNode:
    value: Number


@dataclass(frozen=True)
class _CellNode:
    reference: str


@dataclass(frozen=True)
class _RangeNode:
    start: str
    end: str


@dataclass(frozen=True)
class _UnaryNode:
    operator: str
    operand: _Node


@dataclass(frozen=True)
class _BinaryNode:
    operator: str
    left: _Node
    right: _Node


@dataclass(frozen=True)
class _SumNode:
    arguments: tuple[_Node, ...]


type _Node = _NumberNode | _CellNode | _RangeNode | _UnaryNode | _BinaryNode | _SumNode


class _UnsupportedFormulaError(ValueError):
    pass


class _FormulaParser:
    def __init__(self, formula: str) -> None:
        self.tokens = _tokenize(formula)
        self.position = 0

    def parse(self) -> _Node:
        expression = self._parse_expression()
        if self._current.kind != "EOF":
            raise _UnsupportedFormulaError(f"unexpected token: {self._current.value}")
        return expression

    @property
    def _current(self) -> _Token:
        return self.tokens[self.position]

    def _advance(self) -> _Token:
        token = self._current
        self.position += 1
        return token

    def _accept(self, kind: str) -> bool:
        if self._current.kind != kind:
            return False
        self._advance()
        return True

    def _require(self, kind: str) -> _Token:
        if self._current.kind != kind:
            raise _UnsupportedFormulaError(f"expected {kind}, got {self._current.value}")
        return self._advance()

    def _parse_expression(self) -> _Node:
        node = self._parse_term()
        while self._current.kind in {"+", "-"}:
            operator = self._advance().kind
            node = _BinaryNode(operator, node, self._parse_term())
        return node

    def _parse_term(self) -> _Node:
        node = self._parse_unary()
        while self._current.kind in {"*", "/"}:
            operator = self._advance().kind
            node = _BinaryNode(operator, node, self._parse_unary())
        return node

    def _parse_unary(self) -> _Node:
        if self._current.kind in {"+", "-"}:
            operator = self._advance().kind
            return _UnaryNode(operator, self._parse_unary())
        return self._parse_primary()

    def _parse_primary(self) -> _Node:
        token = self._current
        if token.kind == "NUMBER":
            self._advance()
            return _NumberNode(_parse_number(token.value))
        if token.kind == "CELL":
            start = _normalize_reference(self._advance().value)
            if self._accept(":"):
                end = _normalize_reference(self._require("CELL").value)
                return _RangeNode(start, end)
            return _CellNode(start)
        if token.kind == "NAME":
            function = self._advance().value.upper()
            if function != "SUM":
                raise _UnsupportedFormulaError(f"unsupported function: {function}")
            self._require("(")
            arguments: list[_Node] = []
            if self._current.kind != ")":
                arguments.append(self._parse_expression())
                while self._accept(","):
                    arguments.append(self._parse_expression())
            self._require(")")
            return _SumNode(tuple(arguments))
        if self._accept("("):
            expression = self._parse_expression()
            self._require(")")
            return expression
        raise _UnsupportedFormulaError(f"unexpected token: {token.value}")


class _SheetFormulaEvaluator:
    def __init__(self, worksheet: Any) -> None:
        self.worksheet = worksheet
        self._cache: dict[str, Number] = {}
        self._visiting: set[str] = set()
        self._range_cell_visits = 0

    def evaluate_cell(self, coordinate: str) -> Number:
        normalized = _normalize_reference(coordinate)
        if normalized in self._cache:
            return self._cache[normalized]
        if normalized in self._visiting:
            raise _UnsupportedFormulaError("cyclic formula reference")
        value = self.worksheet[normalized].value
        if not isinstance(value, str) or not value.startswith("="):
            numeric = _direct_numeric(value)
            if numeric is None:
                raise _UnsupportedFormulaError(f"non-numeric cell: {normalized}")
            return numeric
        self._visiting.add(normalized)
        try:
            result = self._evaluate(_FormulaParser(value).parse())
            self._cache[normalized] = _normalize_number(result)
            return self._cache[normalized]
        finally:
            self._visiting.remove(normalized)

    def evaluate_formula(self, formula: str) -> Number:
        return _normalize_number(self._evaluate(_FormulaParser(formula).parse()))

    def _evaluate(self, node: _Node) -> Number:
        if isinstance(node, _NumberNode):
            return node.value
        if isinstance(node, _CellNode):
            return self._evaluate_cell_reference(node.reference, ignore_text=False) or 0
        if isinstance(node, _RangeNode):
            raise _UnsupportedFormulaError("range is only supported inside SUM")
        if isinstance(node, _UnaryNode):
            value = self._evaluate(node.operand)
            return value if node.operator == "+" else -value
        if isinstance(node, _BinaryNode):
            left = self._evaluate(node.left)
            right = self._evaluate(node.right)
            if node.operator == "+":
                return left + right
            if node.operator == "-":
                return left - right
            if node.operator == "*":
                return left * right
            if right == 0:
                raise _UnsupportedFormulaError("division by zero")
            return left / right
        if isinstance(node, _SumNode):
            total = 0.0
            for argument in node.arguments:
                if isinstance(argument, _RangeNode):
                    total += self._sum_range(argument)
                elif isinstance(argument, _CellNode):
                    total += self._evaluate_cell_reference(
                        argument.reference,
                        ignore_text=True,
                    ) or 0
                else:
                    total += self._evaluate(argument)
            return total
        raise _UnsupportedFormulaError("unknown formula node")

    def _sum_range(self, node: _RangeNode) -> Number:
        from openpyxl.utils.cell import range_boundaries

        min_col, min_row, max_col, max_row = range_boundaries(f"{node.start}:{node.end}")
        if min_col is None or min_row is None or max_col is None or max_row is None:
            raise _UnsupportedFormulaError("invalid SUM range")
        width = max_col - min_col + 1
        height = max_row - min_row + 1
        cell_count = width * height
        if cell_count > MAX_FORMULA_RANGE_CELLS:
            raise _UnsupportedFormulaError(
                f"SUM range exceeds {MAX_FORMULA_RANGE_CELLS} cells"
            )
        if self._range_cell_visits + cell_count > MAX_FORMULA_TOTAL_CELL_VISITS:
            raise _UnsupportedFormulaError(
                f"formula exceeds {MAX_FORMULA_TOTAL_CELL_VISITS} cell visits"
            )
        self._range_cell_visits += cell_count
        total = 0.0
        for row in range(min_row, max_row + 1):
            for column in range(min_col, max_col + 1):
                coordinate = self.worksheet.cell(row=row, column=column).coordinate
                total += self._evaluate_cell_reference(coordinate, ignore_text=True) or 0
        return total

    def _evaluate_cell_reference(self, reference: str, *, ignore_text: bool) -> Number | None:
        value = self.worksheet[_normalize_reference(reference)].value
        if isinstance(value, str) and value.startswith("="):
            return self.evaluate_cell(reference)
        numeric = _direct_numeric(value)
        if numeric is not None:
            return numeric
        if value in (None, ""):
            return None
        if ignore_text:
            return None
        raise _UnsupportedFormulaError(f"non-numeric cell: {reference}")


def evaluate_formula(worksheet: Any, formula: str) -> Number | None:
    """Evaluate a numeric formula from the deliberately restricted project grammar."""
    try:
        return _SheetFormulaEvaluator(worksheet).evaluate_formula(formula)
    except (ArithmeticError, ValueError, OverflowError):
        return None


def referenced_ranges(formula: str) -> tuple[FormulaRange, ...]:
    """Return cell ranges from a supported formula without evaluating it."""
    from openpyxl.utils.cell import range_boundaries

    try:
        root = _FormulaParser(formula).parse()
    except ValueError:
        return ()
    ranges: list[FormulaRange] = []
    for node in _walk_nodes(root):
        if not isinstance(node, _RangeNode):
            continue
        min_col, min_row, max_col, max_row = range_boundaries(f"{node.start}:{node.end}")
        if min_col is None or min_row is None or max_col is None or max_row is None:
            continue
        ranges.append(FormulaRange(min_col, min_row, max_col, max_row))
    return tuple(ranges)


def refresh_formula_caches(
    path: Path,
    *,
    soffice: Path | None,
    timeout: float = 30.0,
) -> FormulaCacheResult:
    """Write supported formula caches and optionally use LibreOffice for unresolved cells."""
    target = path if isinstance(path, StagedArtifact) else path.absolute()
    with _output_lock(target):
        cached_values, unsupported = _evaluate_workbook(target)
        if cached_values:
            _write_cached_values(target, cached_values)
        libreoffice_used = False
        if soffice is not None and any(unsupported.values()):
            fallback_values, libreoffice_used = _read_libreoffice_fallback(
                target,
                unresolved=unsupported,
                soffice=soffice,
                timeout=timeout,
            )
            if fallback_values:
                _write_cached_values(target, fallback_values)
        return FormulaCacheResult(
            cached_values=cached_values,
            unsupported_formulas=unsupported,
            libreoffice_used=libreoffice_used,
        )


def run_libreoffice_conversion(
    source: Path,
    output_dir: Path,
    *,
    soffice: Path,
    timeout: float = 30.0,
) -> LibreOfficeConversionResult:
    """Convert one workbook only behind a verified OS-level network sandbox."""
    sandbox_prefix = _libreoffice_network_sandbox_prefix()
    if sandbox_prefix is None:
        return LibreOfficeConversionResult(
            output_path=None,
            detail=(
                "LibreOffice disabled: a verified OS network sandbox is unavailable"
            ),
        )
    output_dir.mkdir(parents=True, exist_ok=True)
    expected = output_dir / f"{source.stem}.xlsx"
    with _output_lock(expected), tempfile.TemporaryDirectory(
        prefix=".js_work_soffice_"
    ) as tmp:
        expected.unlink(missing_ok=True)
        root = Path(tmp)
        profile_dir = root / "profile"
        home_dir = root / "home"
        runtime_dir = root / "runtime"
        for directory in (profile_dir, home_dir, runtime_dir):
            directory.mkdir()
        command = [
            *sandbox_prefix,
            str(soffice),
            f"-env:UserInstallation={profile_dir.resolve().as_uri()}",
            "--headless",
            "--nologo",
            "--nodefault",
            "--norestore",
            "--invisible",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(output_dir),
            str(source),
        ]
        try:
            completed = subprocess.run(
                command,
                capture_output=True,
                text=True,
                check=False,
                timeout=timeout,
                env=_offline_environment(
                    home_dir=home_dir,
                    runtime_dir=runtime_dir,
                    temp_dir=root,
                ),
            )
        except subprocess.TimeoutExpired:
            return LibreOfficeConversionResult(
                output_path=None,
                detail=f"LibreOffice conversion timed out after {timeout:g}s",
                timed_out=True,
            )
        except OSError as exc:
            logger.warning("LibreOffice conversion failed: %s", type(exc).__name__)
            return LibreOfficeConversionResult(
                output_path=None,
                detail="LibreOffice conversion failed safely",
            )
        if completed.returncode != 0 or not expected.exists():
            logger.warning(
                "LibreOffice conversion process failed with return code %s",
                completed.returncode,
            )
            return LibreOfficeConversionResult(
                output_path=None,
                detail="LibreOffice conversion failed safely",
            )
        return LibreOfficeConversionResult(output_path=expected, detail="")


@lru_cache(maxsize=1)
def _libreoffice_network_sandbox_prefix() -> tuple[str, ...] | None:
    """Return a locally verified command prefix that denies all networking.

    Clearing proxy variables is not a network boundary.  Work therefore only
    enables LibreOffice where a root-owned OS sandbox executable both exists
    and demonstrably denies a loopback bind in a fresh subprocess.
    """
    if sys.platform != "darwin":
        return None
    sandbox = Path("/usr/bin/sandbox-exec")
    try:
        metadata = sandbox.stat()
    except OSError:
        return None
    if (
        not stat.S_ISREG(metadata.st_mode)
        or metadata.st_uid != 0
        or metadata.st_mode & (stat.S_IWGRP | stat.S_IWOTH)
    ):
        return None
    prefix = (str(sandbox), "-p", _MACOS_NETWORK_DENY_PROFILE)
    try:
        completed = subprocess.run(
            [*prefix, sys.executable, "-c", _NETWORK_SANDBOX_PROBE],
            capture_output=True,
            text=True,
            check=False,
            timeout=5.0,
            env=_sandbox_probe_environment(),
        )
    except (OSError, subprocess.TimeoutExpired):
        return None
    return prefix if completed.returncode == 0 else None


@contextmanager
def _output_lock(path: Path) -> Iterator[None]:
    if isinstance(path, StagedArtifact):
        parent = os.fstat(path._parent_fd)
        key = f"fd:{parent.st_dev}:{parent.st_ino}:{path._staged_name}"
    else:
        key = str(path.absolute())
    with _PROCESS_LOCKS_GUARD:
        process_lock = _PROCESS_LOCKS.setdefault(key, threading.Lock())
    with process_lock:
        lock_dir = Path(tempfile.gettempdir()) / "js_work_formula_locks"
        lock_dir.mkdir(parents=True, exist_ok=True)
        lock_name = sha256(key.encode("utf-8")).hexdigest() + ".lock"
        with (lock_dir / lock_name).open("a+b") as handle:
            fcntl.flock(handle.fileno(), fcntl.LOCK_EX)
            try:
                yield
            finally:
                fcntl.flock(handle.fileno(), fcntl.LOCK_UN)


def _evaluate_workbook(
    path: Path,
) -> tuple[dict[str, dict[str, Number]], dict[str, dict[str, str]]]:
    from openpyxl import load_workbook

    with open_artifact(path) as handle:
        workbook = load_workbook(handle, data_only=False)
        cached_values: dict[str, dict[str, Number]] = {}
        unsupported: dict[str, dict[str, str]] = {}
        try:
            for worksheet in workbook.worksheets:
                evaluator = _SheetFormulaEvaluator(worksheet)
                sheet_values: dict[str, Number] = {}
                sheet_unsupported: dict[str, str] = {}
                for row in worksheet.iter_rows():
                    for cell in row:
                        formula = cell.value
                        if not isinstance(formula, str) or not formula.startswith("="):
                            continue
                        try:
                            sheet_values[cell.coordinate] = evaluator.evaluate_cell(
                                cell.coordinate
                            )
                        except (ArithmeticError, ValueError, OverflowError):
                            sheet_unsupported[cell.coordinate] = formula
                if sheet_values:
                    cached_values[worksheet.title] = sheet_values
                if sheet_unsupported:
                    unsupported[worksheet.title] = sheet_unsupported
            return cached_values, unsupported
        finally:
            workbook.close()


def _write_cached_values(path: Path, values: Mapping[str, Mapping[str, Number]]) -> None:
    with (
        rewrite_artifact(path) as (source_handle, temporary_handle),
        zipfile.ZipFile(source_handle, "r") as source,
    ):
        sheet_parts = _sheet_parts(source)
        parts_to_values = {
            sheet_parts[sheet_name]: sheet_values
            for sheet_name, sheet_values in values.items()
            if sheet_name in sheet_parts and sheet_values
        }
        with zipfile.ZipFile(temporary_handle, "w") as destination:
            for item in source.infolist():
                data = source.read(item.filename)
                sheet_values = parts_to_values.get(item.filename)
                if sheet_values:
                    data = _update_sheet_cached_values(data, sheet_values)
                destination.writestr(item, data)


def _sheet_parts(workbook: zipfile.ZipFile) -> dict[str, str]:
    workbook_root = ET.fromstring(workbook.read("xl/workbook.xml"))
    relationships_root = ET.fromstring(workbook.read("xl/_rels/workbook.xml.rels"))
    relationships = {
        relation.get("Id"): relation.get("Target")
        for relation in relationships_root.findall(f"{{{_PACKAGE_REL_NS}}}Relationship")
    }
    result: dict[str, str] = {}
    for sheet in workbook_root.findall(f".//{{{_MAIN_NS}}}sheet"):
        name = sheet.get("name")
        relation_id = sheet.get(f"{{{_DOCUMENT_REL_NS}}}id")
        target = relationships.get(relation_id)
        if not name or not target:
            continue
        if target.startswith("/"):
            part = target.lstrip("/")
        else:
            part = posixpath.normpath(str(PurePosixPath("xl") / target))
        result[name] = part
    return result


def _update_sheet_cached_values(data: bytes, values: Mapping[str, Number]) -> bytes:
    ET.register_namespace("", _MAIN_NS)
    root = ET.fromstring(data)
    for cell in root.iter(f"{{{_MAIN_NS}}}c"):
        coordinate = cell.get("r")
        if (
            coordinate is None
            or coordinate not in values
            or cell.find(f"{{{_MAIN_NS}}}f") is None
        ):
            continue
        cell.attrib.pop("t", None)
        value_node = cell.find(f"{{{_MAIN_NS}}}v")
        if value_node is None:
            value_node = ET.SubElement(cell, f"{{{_MAIN_NS}}}v")
        value_node.text = _format_number(values[coordinate])
    rendered = ET.tostring(root, encoding="utf-8", xml_declaration=True)
    return rendered if isinstance(rendered, bytes) else rendered.encode("utf-8")


def _read_libreoffice_fallback(
    path: Path,
    *,
    unresolved: Mapping[str, Mapping[str, str]],
    soffice: Path,
    timeout: float,
) -> tuple[dict[str, dict[str, Number]], bool]:
    with tempfile.TemporaryDirectory(prefix=".js_work_recalc_") as tmp:
        root = Path(tmp)
        input_dir = root / "input"
        output_dir = root / "out"
        input_dir.mkdir()
        output_dir.mkdir()
        source_copy = input_dir / path.name
        with open_artifact(path) as source_handle, source_copy.open("wb") as destination:
            shutil.copyfileobj(source_handle, destination)
            destination.flush()
            os.fsync(destination.fileno())
        conversion = run_libreoffice_conversion(
            source_copy,
            output_dir,
            soffice=soffice,
            timeout=timeout,
        )
        if conversion.output_path is None:
            return {}, True
        return _read_cached_values(conversion.output_path, unresolved), True


def _read_cached_values(
    path: Path,
    requested: Mapping[str, Mapping[str, Any]],
) -> dict[str, dict[str, Number]]:
    values: dict[str, dict[str, Number]] = {}
    with open_artifact(path) as handle, zipfile.ZipFile(handle, "r") as workbook:
        parts = _sheet_parts(workbook)
        for sheet_name, coordinates in requested.items():
            part = parts.get(sheet_name)
            if part is None:
                continue
            root = ET.fromstring(workbook.read(part))
            sheet_values: dict[str, Number] = {}
            for cell in root.iter(f"{{{_MAIN_NS}}}c"):
                coordinate = cell.get("r")
                if coordinate is None or coordinate not in coordinates:
                    continue
                value_node = cell.find(f"{{{_MAIN_NS}}}v")
                if value_node is None or value_node.text in (None, ""):
                    continue
                value_text = value_node.text
                if value_text is None:
                    continue
                try:
                    sheet_values[coordinate] = _parse_number(value_text)
                except ValueError:
                    continue
            if sheet_values:
                values[sheet_name] = sheet_values
    return values


def _offline_environment(
    *,
    home_dir: Path,
    runtime_dir: Path,
    temp_dir: Path,
) -> dict[str, str]:
    environment = dict(os.environ)
    for key in (
        "HTTP_PROXY",
        "HTTPS_PROXY",
        "ALL_PROXY",
        "http_proxy",
        "https_proxy",
        "all_proxy",
    ):
        environment.pop(key, None)
    environment.update(
        {
            "HOME": str(home_dir),
            "XDG_CONFIG_HOME": str(home_dir / ".config"),
            "XDG_CACHE_HOME": str(home_dir / ".cache"),
            "XDG_RUNTIME_DIR": str(runtime_dir),
            "TMPDIR": str(temp_dir),
            "NO_PROXY": "*",
            "no_proxy": "*",
            "SAL_USE_VCLPLUGIN": "svp",
        }
    )
    return environment


def _sandbox_probe_environment() -> dict[str, str]:
    """Use a deterministic probe environment without inherited proxy or loader hooks."""
    return {
        "HOME": "/var/empty",
        "LANG": "C",
        "LC_ALL": "C",
        "PATH": "/usr/bin:/bin:/usr/sbin:/sbin",
        "PYTHONNOUSERSITE": "1",
        "TMPDIR": "/tmp",
    }


def _tokenize(formula: str) -> list[_Token]:
    if not isinstance(formula, str) or not formula.startswith("="):
        raise _UnsupportedFormulaError("formula must start with =")
    text = formula[1:]
    tokens: list[_Token] = []
    index = 0
    while index < len(text):
        character = text[index]
        if character.isspace():
            index += 1
            continue
        if character in "+-*/(),:":
            tokens.append(_Token(character, character))
            index += 1
            continue
        if character.isdigit() or (
            character == "." and index + 1 < len(text) and text[index + 1].isdigit()
        ):
            end = _scan_number(text, index)
            tokens.append(_Token("NUMBER", text[index:end]))
            index = end
            continue
        if character.isalpha() or character in {"$", "_"}:
            end = index + 1
            while end < len(text) and (text[end].isalnum() or text[end] in {"$", "_", "."}):
                end += 1
            value = text[index:end]
            kind = "CELL" if _is_cell_reference(value) else "NAME"
            tokens.append(_Token(kind, value))
            index = end
            continue
        raise _UnsupportedFormulaError(f"unsupported character: {character}")
    tokens.append(_Token("EOF", ""))
    return tokens


def _scan_number(text: str, start: int) -> int:
    index = start
    seen_decimal = False
    while index < len(text):
        if text[index].isdigit():
            index += 1
            continue
        if text[index] == "." and not seen_decimal:
            seen_decimal = True
            index += 1
            continue
        break
    if index < len(text) and text[index] in {"e", "E"}:
        exponent = index + 1
        if exponent < len(text) and text[exponent] in {"+", "-"}:
            exponent += 1
        digits = exponent
        while exponent < len(text) and text[exponent].isdigit():
            exponent += 1
        if exponent == digits:
            raise _UnsupportedFormulaError("invalid numeric exponent")
        index = exponent
    return index


def _is_cell_reference(value: str) -> bool:
    normalized = value.replace("$", "")
    split = 0
    while split < len(normalized) and normalized[split].isalpha():
        split += 1
    return (
        1 <= split <= 3
        and split < len(normalized)
        and normalized[:split].isascii()
        and normalized[split:].isdigit()
        and int(normalized[split:]) > 0
    )


def _normalize_reference(value: str) -> str:
    return value.replace("$", "").upper()


def _parse_number(value: str) -> Number:
    number = float(value)
    if not math.isfinite(number):
        raise ValueError("formula result must be finite")
    return int(number) if number.is_integer() else number


def _normalize_number(value: Number) -> Number:
    number = float(value)
    if not math.isfinite(number):
        raise _UnsupportedFormulaError("formula result must be finite")
    return int(number) if number.is_integer() else round(number, 12)


def _direct_numeric(value: Any) -> Number | None:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return None
    return _normalize_number(value)


def _format_number(value: Number) -> str:
    normalized = _normalize_number(value)
    return str(normalized) if isinstance(normalized, int) else format(normalized, ".15g")


def _walk_nodes(root: _Node) -> Iterator[_Node]:
    yield root
    if isinstance(root, _UnaryNode):
        yield from _walk_nodes(root.operand)
    elif isinstance(root, _BinaryNode):
        yield from _walk_nodes(root.left)
        yield from _walk_nodes(root.right)
    elif isinstance(root, _SumNode):
        for argument in root.arguments:
            yield from _walk_nodes(argument)
