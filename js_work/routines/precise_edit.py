"""Bounded, owner-scoped precise edits for Excel workbooks."""

from __future__ import annotations

import math
import os
import re
import zipfile
from copy import copy
from hashlib import sha256
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, range_boundaries

from js_work.file_scope import MaterializedSnapshotPath
from js_work.safe_output import StagedArtifact

MAX_OPERATIONS = 256
MAX_COMPRESSED_BYTES = 32 * 1024 * 1024
MAX_UNCOMPRESSED_BYTES = 128 * 1024 * 1024
MAX_ARCHIVE_MEMBERS = 1_000
MAX_COMPRESSION_RATIO = 100
MAX_SHEETS = 64
MAX_SHEET_ROWS = 200_000
MAX_SHEET_COLUMNS = 1_000
MAX_RANGE_CELLS = 10_000
MAX_CELL_VALUE_LENGTH = 32_767
MAX_NUMBER_FORMAT_LENGTH = 512
MAX_ROW_HEIGHT = 409.0
MAX_COLUMN_WIDTH = 255.0
_MAX_EXCEL_ROWS = 1_048_576
_MAX_EXCEL_COLUMNS = 16_384
_CELL_RE = re.compile(r"^[A-Z]{1,3}[1-9][0-9]*$")
_RANGE_RE = re.compile(r"^[A-Z]{1,3}[1-9][0-9]*:[A-Z]{1,3}[1-9][0-9]*$")
_COLUMN_RE = re.compile(r"^[A-Z]{1,3}$")
_SHA256_RE = re.compile(r"^[0-9a-fA-F]{64}$")
_UNSAFE_FORMULA_RE = re.compile(
    r"\b(?:CALL|DDE|EXEC|FILTERXML|HYPERLINK|REGISTER(?:\.ID)?|RTD|SHELL|URLDOWNLOADTOFILE|WEBSERVICE)\s*\(",
    re.IGNORECASE,
)
_MUTABLE_OOXML_PREFIXES = (
    "docprops/",
    "xl/theme/",
)
_MUTABLE_OOXML_PARTS = {
    "[content_types].xml",
    "_rels/.rels",
    "xl/_rels/workbook.xml.rels",
    "xl/calcchain.xml",
    "xl/sharedstrings.xml",
    "xl/styles.xml",
    "xl/workbook.xml",
}
_WORKSHEET_PART_RE = re.compile(r"^xl/worksheets/sheet[1-9][0-9]*\.xml$", re.IGNORECASE)
_SPREADSHEETML_NAMESPACE = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_RELATIONSHIPS_NAMESPACE = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_XML_NAMESPACE = "http://www.w3.org/XML/1998/namespace"
_ALLOWED_WORKSHEET_NAMESPACES = {
    "",
    _SPREADSHEETML_NAMESPACE,
    _RELATIONSHIPS_NAMESPACE,
    _XML_NAMESPACE,
}
_UNSUPPORTED_WORKSHEET_ELEMENTS = {
    "controls",
    "customSheetViews",
    "extLst",
    "legacyDrawing",
    "legacyDrawingHF",
    "oleObjects",
    "smartTags",
    "webPublishItems",
}
_EXTERNAL_RELATIONSHIP_RE = re.compile(rb"TargetMode\s*=\s*['\"]External['\"]")
_UNSAFE_OOXML_PREFIXES = (
    "customui/",
    "xl/activex/",
    "xl/embeddings/",
    "xl/externallinks/",
    "xl/macrosheets/",
    "xl/persons/",
    "xl/querytables/",
    "xl/slicers/",
    "xl/threadedcomments/",
    "xl/timelines/",
)
_UNSAFE_OOXML_PARTS = {
    "xl/cellmetadata.xml",
    "xl/connections.xml",
    "xl/metadata.xml",
    "xl/vbadata.xml",
    "xl/vbaproject.bin",
}


class PreciseExcelEditEngine:
    """Apply a small, validated edit set to a copied XLSX workbook."""

    def __init__(self, workspace: Path) -> None:
        self.workspace = workspace.expanduser().resolve()

    def apply(
        self,
        *,
        source_path: str | Path,
        output_path: str,
        operations: list[dict[str, Any]],
        expected_source_sha256: str | None = None,
    ) -> dict[str, Any]:
        self._reject_source_symlinks(source_path)
        source = self._resolve(source_path, must_exist=True)
        output = self._resolve(output_path, must_exist=False)
        validation_path = output.with_suffix(".validation.json")
        self._validate_workbook_paths(source, output, output_path, validation_path)
        source_hash_before = file_hash(source)
        self._validate_expected_source_hash(expected_source_sha256, source_hash_before)
        self._validate_operations(operations)
        self._validate_archive(source)

        workbook = None
        staged_output: Path | None = None
        try:
            with self._open_artifact(source) as source_handle:
                workbook = load_workbook(source_handle, data_only=False, read_only=False)
            self._validate_workbook_shape(workbook)
            self._validate_existing_formulas(workbook)
            for operation in operations:
                self._apply_operation(workbook, operation)
            staged_output = self._stage_workbook(workbook, output)
            self._validate_archive(staged_output)
            self._validate_complex_part_preservation(source, staged_output)
            source_hash_after = file_hash(source)
            if source_hash_after != source_hash_before:
                raise RuntimeError("source workbook changed during precise edit")
            report = {
                "status": "passed",
                "operation_count": len(operations),
                "source_path": str(source),
                "output_path": str(output),
                "source_hash": source_hash_before,
                "source_hash_after": source_hash_after,
                "output_hash": file_hash(staged_output),
                "validation_path": str(validation_path),
            }
            output_published = False
            try:
                self._publish_no_clobber(
                    staged_output,
                    output,
                    "output workbook already exists",
                )
                output_published = True
                self._atomic_write_json(validation_path, report, anchor=staged_output)
            except Exception:
                if output_published:
                    from js_work.safe_output import remove_published_link

                    remove_published_link(staged_output, output)
                raise
        finally:
            if workbook is not None:
                workbook.close()
            if staged_output is not None:
                from js_work.safe_output import discard_staged

                discard_staged(staged_output)

        return report

    def _resolve(self, path: str | Path, *, must_exist: bool) -> Path:
        if isinstance(path, MaterializedSnapshotPath):
            if not must_exist:
                raise ValueError("materialized snapshots are input-only")
            return path
        if not isinstance(path, str) or not path:
            raise ValueError("workbook path is required")
        logical = Path(path)
        if logical.is_absolute() or ".." in logical.parts:
            raise ValueError("workbook path must be workspace-relative")
        resolved = (self.workspace / logical).resolve()
        try:
            resolved.relative_to(self.workspace)
        except ValueError as exc:
            raise ValueError("workbook path escapes workspace") from exc
        if must_exist and (not resolved.is_file() or resolved.is_symlink()):
            raise ValueError("source workbook does not exist")
        return resolved

    def _reject_source_symlinks(self, path: str | Path) -> None:
        if isinstance(path, MaterializedSnapshotPath):
            return
        if not isinstance(path, str) or not path:
            raise ValueError("workbook path is required")
        logical = Path(path)
        if logical.is_absolute() or ".." in logical.parts:
            return
        current = self.workspace
        for component in logical.parts:
            current = current / component
            if current.is_symlink():
                raise ValueError("source path contains a symlink")

    def _validate_workbook_paths(
        self,
        source: Path,
        output: Path,
        output_path: str,
        validation_path: Path,
    ) -> None:
        if source.suffix.lower() != ".xlsx" or output.suffix.lower() != ".xlsx":
            raise ValueError("source and output must be .xlsx files")
        if source == output:
            raise ValueError("output must not overwrite source workbook")
        if output.exists() and os.path.samefile(source, output):
            raise ValueError("output must not overwrite source workbook")
        if os.path.lexists(self.workspace / output_path):
            raise ValueError("output workbook already exists")
        if os.path.lexists(validation_path):
            raise ValueError("validation report already exists")

    @staticmethod
    def _validate_expected_source_hash(expected: str | None, actual: str) -> None:
        if expected is None:
            return
        if not isinstance(expected, str) or not _SHA256_RE.fullmatch(expected):
            raise ValueError("expected source hash must be a SHA-256 digest")
        if expected.lower() != actual:
            raise ValueError("expected source hash does not match source workbook")

    @staticmethod
    def _validate_operations(operations: list[dict[str, Any]]) -> None:
        if not isinstance(operations, list) or not operations:
            raise ValueError("operations must be a non-empty JSON array")
        if len(operations) > MAX_OPERATIONS:
            raise ValueError(f"operations exceed limit of {MAX_OPERATIONS}")
        if not all(isinstance(operation, dict) for operation in operations):
            raise ValueError("each operation must be an object")

    @staticmethod
    def _validate_archive(source: Path) -> None:
        with PreciseExcelEditEngine._open_artifact(source) as source_handle:
            source_handle.seek(0, os.SEEK_END)
            if source_handle.tell() > MAX_COMPRESSED_BYTES:
                raise ValueError("source workbook exceeds compressed size limit")
            source_handle.seek(0)
            if not zipfile.is_zipfile(source_handle):
                raise ValueError("source is not a valid .xlsx archive")
            source_handle.seek(0)
            archive = zipfile.ZipFile(source_handle)
            try:
                members = archive.infolist()
                if len(members) > MAX_ARCHIVE_MEMBERS:
                    raise ValueError("source workbook has too many OOXML parts")
                if sum(member.file_size for member in members) > MAX_UNCOMPRESSED_BYTES:
                    raise ValueError("source workbook exceeds uncompressed size limit")
                part_names = [member.filename for member in members]
                if len(set(part_names)) != len(part_names):
                    raise ValueError("source workbook has duplicate OOXML parts")
                for member in members:
                    lower_name = member.filename.lower()
                    if member.filename.startswith("/") or ".." in Path(member.filename).parts:
                        raise ValueError("source workbook has an unsafe OOXML path")
                    if (
                        member.file_size
                        and member.file_size / max(member.compress_size, 1)
                        > MAX_COMPRESSION_RATIO
                    ):
                        raise ValueError("OOXML compression ratio exceeds limit")
                    if (
                        lower_name.startswith(_UNSAFE_OOXML_PREFIXES)
                        or lower_name in _UNSAFE_OOXML_PARTS
                    ):
                        raise ValueError(
                            "source workbook contains an unsafe external or executable OOXML part"
                        )
                    if lower_name.endswith(".rels") and _EXTERNAL_RELATIONSHIP_RE.search(
                        archive.read(member)
                    ):
                        raise ValueError("source workbook contains an external OOXML relationship")
                    if _WORKSHEET_PART_RE.fullmatch(lower_name):
                        PreciseExcelEditEngine._validate_worksheet_xml(archive.read(member))
                part_name_set = set(part_names)
                if (
                    "[Content_Types].xml" not in part_name_set
                    or "xl/workbook.xml" not in part_name_set
                ):
                    raise ValueError("source is not a valid .xlsx workbook")
                content_types = archive.read("[Content_Types].xml").lower()
                if b"macroenabled" in content_types or b"activex" in content_types:
                    raise ValueError("source workbook contains macro or ActiveX OOXML content")
            finally:
                archive.close()

    @staticmethod
    def _validate_workbook_shape(workbook: Any) -> None:
        if not workbook.worksheets or len(workbook.worksheets) > MAX_SHEETS:
            raise ValueError("workbook sheet count is outside supported limits")
        for sheet in workbook.worksheets:
            if sheet.max_row > MAX_SHEET_ROWS or sheet.max_column > MAX_SHEET_COLUMNS:
                raise ValueError("workbook sheet dimensions exceed supported limits")

    @classmethod
    def _validate_existing_formulas(cls, workbook: Any) -> None:
        for sheet in workbook.worksheets:
            for cell in sheet._cells.values():
                value = cell.value
                if cell.data_type == "f" and isinstance(value, str):
                    cls._validate_formula(value if value.startswith("=") else f"={value}")
        for defined_name in workbook.defined_names.values():
            value = getattr(defined_name, "attr_text", None)
            if isinstance(value, str):
                cls._validate_formula(value if value.startswith("=") else f"={value}")

    @staticmethod
    def _validate_complex_part_preservation(source: Path, staged: Path) -> None:
        with (
            PreciseExcelEditEngine._open_artifact(source) as source_handle,
            PreciseExcelEditEngine._open_artifact(staged) as staged_handle,
            zipfile.ZipFile(source_handle) as source_archive,
            zipfile.ZipFile(staged_handle) as staged_archive,
        ):
            source_names = set(source_archive.namelist())
            staged_names = set(staged_archive.namelist())
            protected_names = {
                name
                for name in source_names | staged_names
                if not PreciseExcelEditEngine._is_mutable_ooxml_part(name)
            }
            for name in protected_names:
                if name not in source_names or name not in staged_names:
                    raise ValueError("complex OOXML content changed during precise edit")
                if source_archive.read(name) != staged_archive.read(name):
                    raise ValueError("complex OOXML content changed during precise edit")

    @staticmethod
    def _is_mutable_ooxml_part(name: str) -> bool:
        lower_name = name.lower()
        return (
            lower_name in _MUTABLE_OOXML_PARTS
            or lower_name.startswith(_MUTABLE_OOXML_PREFIXES)
            or _WORKSHEET_PART_RE.fullmatch(lower_name) is not None
        )

    @staticmethod
    def _validate_worksheet_xml(payload: bytes) -> None:
        upper_payload = payload.upper()
        if b"<!DOCTYPE" in upper_payload or b"<!ENTITY" in upper_payload:
            raise ValueError("unsupported worksheet OOXML feature")
        try:
            root = ElementTree.fromstring(payload)
        except ElementTree.ParseError as exc:
            raise ValueError("worksheet OOXML is malformed") from exc
        for element in root.iter():
            namespace, local_name = PreciseExcelEditEngine._xml_name(element.tag)
            if (
                namespace not in _ALLOWED_WORKSHEET_NAMESPACES
                or local_name in _UNSUPPORTED_WORKSHEET_ELEMENTS
            ):
                raise ValueError("unsupported worksheet OOXML feature")
            for attribute in element.attrib:
                attribute_namespace, _ = PreciseExcelEditEngine._xml_name(attribute)
                if attribute_namespace not in _ALLOWED_WORKSHEET_NAMESPACES:
                    raise ValueError("unsupported worksheet OOXML feature")

    @staticmethod
    def _xml_name(name: str) -> tuple[str, str]:
        if name.startswith("{") and "}" in name:
            namespace, local_name = name[1:].split("}", 1)
            return namespace, local_name
        return "", name

    def _apply_operation(self, workbook: Any, operation: dict[str, Any]) -> None:
        name = operation.get("op")
        if not isinstance(name, str):
            raise ValueError("operation name is required")
        handlers = {
            "set_cell": self._set_cell,
            "clear_cell": self._clear_cell,
            "copy_style": self._copy_style,
            "set_number_format": self._set_number_format,
            "set_row_height": self._set_row_height,
            "set_column_width": self._set_column_width,
            "merge_cells": self._merge_cells,
            "unmerge_cells": self._unmerge_cells,
        }
        handler = handlers.get(name)
        if handler is None:
            raise ValueError(f"unknown precise edit operation: {name}")
        sheet_name = operation.get("sheet")
        if not isinstance(sheet_name, str) or sheet_name not in workbook.sheetnames:
            raise ValueError("unknown worksheet")
        handler(workbook[sheet_name], operation)

    def _set_cell(self, sheet: Any, operation: dict[str, Any]) -> None:
        from js_work.routines.office_safety import Formula, apply_work_cell_value

        self._require_keys(operation, {"op", "sheet", "cell", "value"})
        cell = self._cell(operation["cell"])
        value = operation["value"]
        self._validate_cell_value(value)
        if isinstance(value, (Formula, str, dict)):
            apply_work_cell_value(sheet[cell], value)
            return
        sheet[cell].value = value

    def _clear_cell(self, sheet: Any, operation: dict[str, Any]) -> None:
        self._require_keys(operation, {"op", "sheet", "cell"})
        sheet[self._cell(operation["cell"])].value = None

    def _copy_style(self, sheet: Any, operation: dict[str, Any]) -> None:
        self._require_keys(operation, {"op", "sheet", "source_cell", "target_cell"})
        source = sheet[self._cell(operation["source_cell"])]
        target = sheet[self._cell(operation["target_cell"])]
        target._style = copy(source._style)
        target.number_format = source.number_format
        target.protection = copy(source.protection)
        target.alignment = copy(source.alignment)

    def _set_number_format(self, sheet: Any, operation: dict[str, Any]) -> None:
        self._require_keys(operation, {"op", "sheet", "cell", "number_format"})
        number_format = operation["number_format"]
        if not isinstance(number_format, str) or len(number_format) > MAX_NUMBER_FORMAT_LENGTH:
            raise ValueError("number_format must be a bounded string")
        sheet[self._cell(operation["cell"])].number_format = number_format

    def _set_row_height(self, sheet: Any, operation: dict[str, Any]) -> None:
        self._require_keys(operation, {"op", "sheet", "row", "height"})
        row = self._row(operation["row"])
        height = self._finite_number(operation["height"], "height", maximum=MAX_ROW_HEIGHT)
        sheet.row_dimensions[row].height = height

    def _set_column_width(self, sheet: Any, operation: dict[str, Any]) -> None:
        self._require_keys(operation, {"op", "sheet", "column", "width"})
        column = self._column(operation["column"])
        width = self._finite_number(operation["width"], "width", maximum=MAX_COLUMN_WIDTH)
        sheet.column_dimensions[column].width = width

    def _merge_cells(self, sheet: Any, operation: dict[str, Any]) -> None:
        self._require_keys(operation, {"op", "sheet", "range"})
        cell_range = self._range(operation["range"])
        min_column, min_row, max_column, max_row = range_boundaries(cell_range)
        for existing_range in sheet.merged_cells.ranges:
            existing_bounds = range_boundaries(str(existing_range))
            if self._ranges_overlap(
                (min_column, min_row, max_column, max_row),
                existing_bounds,
            ):
                raise ValueError("merge range overlaps existing merged cells")
        for row in sheet.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_column,
            max_col=max_column,
        ):
            for cell in row:
                if (cell.row != min_row or cell.column != min_column) and cell.value is not None:
                    raise ValueError("merge would discard non-top-left cell content")
        sheet.merge_cells(cell_range)

    def _unmerge_cells(self, sheet: Any, operation: dict[str, Any]) -> None:
        self._require_keys(operation, {"op", "sheet", "range"})
        cell_range = self._range(operation["range"])
        if cell_range not in {str(item) for item in sheet.merged_cells.ranges}:
            raise ValueError("range is not merged")
        sheet.unmerge_cells(cell_range)

    @staticmethod
    def _require_keys(operation: dict[str, Any], expected: set[str]) -> None:
        if set(operation) != expected:
            raise ValueError("operation has missing or unsupported fields")

    @staticmethod
    def _cell(value: Any) -> str:
        if not isinstance(value, str) or not _CELL_RE.fullmatch(value):
            raise ValueError("invalid cell reference")
        column = re.match(r"[A-Z]+", value)
        assert column is not None
        if column_index_from_string(column.group()) > _MAX_EXCEL_COLUMNS:
            raise ValueError("cell column exceeds Excel limit")
        if int(value[len(column.group()) :]) > _MAX_EXCEL_ROWS:
            raise ValueError("cell row exceeds Excel limit")
        return value

    @staticmethod
    def _column(value: Any) -> str:
        if not isinstance(value, str) or not _COLUMN_RE.fullmatch(value):
            raise ValueError("invalid column reference")
        if column_index_from_string(value) > _MAX_EXCEL_COLUMNS:
            raise ValueError("column exceeds Excel limit")
        return value

    @staticmethod
    def _row(value: Any) -> int:
        if (
            isinstance(value, bool)
            or not isinstance(value, int)
            or not 1 <= value <= _MAX_EXCEL_ROWS
        ):
            raise ValueError("invalid row reference")
        return value

    @staticmethod
    def _range(value: Any) -> str:
        if not isinstance(value, str) or not _RANGE_RE.fullmatch(value):
            raise ValueError("invalid cell range")
        min_column, min_row, max_column, max_row = range_boundaries(value)
        if max_column > _MAX_EXCEL_COLUMNS or max_row > _MAX_EXCEL_ROWS:
            raise ValueError("range exceeds Excel limits")
        if (max_column - min_column + 1) * (max_row - min_row + 1) > MAX_RANGE_CELLS:
            raise ValueError("range exceeds supported size")
        return value

    @staticmethod
    def _finite_number(value: Any, name: str, *, maximum: float) -> float:
        if isinstance(value, bool) or not isinstance(value, int | float):
            raise ValueError(f"{name} must be a number")
        number = float(value)
        if not math.isfinite(number) or not 0 <= number <= maximum:
            raise ValueError(f"{name} is outside supported limits")
        return number

    @staticmethod
    def _validate_cell_value(value: Any) -> None:
        from js_work.routines.office_safety import Formula

        if isinstance(value, Formula):
            return
        if isinstance(value, dict) and set(value.keys()) == {"__work_formula__"}:
            expression = value.get("__work_formula__")
            if not isinstance(expression, str):
                raise ValueError("Work formula payload must be a string")
            Formula(expression)
            return
        if isinstance(value, (list, dict)) or not isinstance(
            value, str | int | float | bool | type(None)
        ):
            raise ValueError("set_cell value must be a scalar")
        if isinstance(value, float) and not math.isfinite(value):
            raise ValueError("set_cell value must be finite")
        if isinstance(value, str) and len(value) > MAX_CELL_VALUE_LENGTH:
            raise ValueError("set_cell value exceeds Excel length limit")

    @staticmethod
    def _validate_formula(formula: str) -> None:
        from js_work.routines.office_safety import validate_restricted_formula

        validate_restricted_formula(formula)

    @staticmethod
    def _ranges_overlap(
        first: tuple[int, int, int, int],
        second: tuple[int, int, int, int],
    ) -> bool:
        first_min_column, first_min_row, first_max_column, first_max_row = first
        second_min_column, second_min_row, second_max_column, second_max_row = second
        return not (
            first_max_column < second_min_column
            or second_max_column < first_min_column
            or first_max_row < second_min_row
            or second_max_row < first_min_row
        )

    @staticmethod
    def _stage_workbook(workbook: Any, output: Path) -> StagedArtifact:
        from js_work.safe_output import create_staged, discard_staged, open_artifact

        output.parent.mkdir(parents=True, exist_ok=True)
        PreciseExcelEditEngine._fsync_directory(output.parent)
        staged = create_staged(output)
        try:
            with open_artifact(staged, "w+b") as handle:
                workbook.save(handle)
            with open_artifact(staged) as handle:
                staged_workbook = load_workbook(handle, data_only=False, read_only=True)
                staged_workbook.close()
            return staged
        except Exception:
            discard_staged(staged)
            raise

    @staticmethod
    def _open_artifact(path: Path) -> Any:
        from js_work.safe_output import open_artifact

        return open_artifact(path)

    @staticmethod
    def _atomic_write_json(
        path: Path,
        report: dict[str, Any],
        *,
        anchor: Path,
    ) -> None:
        from js_work.safe_output import StagedArtifact, write_json_no_clobber

        path.parent.mkdir(parents=True, exist_ok=True)
        if not isinstance(anchor, StagedArtifact):
            raise RuntimeError("validation report anchor is unavailable")
        write_json_no_clobber(
            path,
            report,
            "validation report already exists",
            anchor=anchor,
        )

    @staticmethod
    def _publish_no_clobber(source: Path, target: Path, message: str) -> None:
        from js_work.safe_output import publish_no_clobber

        publish_no_clobber(source, target, message)

    @staticmethod
    def _fsync_file(path: Path) -> None:
        from js_work.safe_output import fsync_file

        fsync_file(path)

    @staticmethod
    def _fsync_directory(path: Path) -> None:
        from js_work.safe_output import fsync_directory

        fsync_directory(path)


def file_hash(path: Path) -> str:
    """Return the SHA-256 digest for a workbook or validation artifact."""
    digest = sha256()
    with PreciseExcelEditEngine._open_artifact(path) as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()
