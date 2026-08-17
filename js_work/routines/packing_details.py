"""PACKING DETAILS routine for repeated two-column roll manifests."""

from __future__ import annotations

import math
import re
import shutil
import tempfile
from copy import copy
from dataclasses import asdict, dataclass
from hashlib import sha256
from pathlib import Path
from typing import Any

from js_work.file_scope import MaterializedSnapshotPath
from js_work.routines.formula_cache import (
    evaluate_formula,
    refresh_formula_caches,
)
from js_work.routines.legacy_xls import convert_legacy_xls_to_xlsx
from js_work.routines.office_safety import (
    Formula,
    apply_work_cell_value,
    reject_formula_like_text,
    validate_safe_xlsx,
)
from js_work.safe_output import (
    ensure_absent,
    open_artifact,
    publish_no_clobber,
    reject_symlink_components,
    remove_published_link,
    staged_path,
    write_json_no_clobber,
)


@dataclass(frozen=True)
class PackingDetailsRunResult:
    status: str
    mode: str
    output_path: str
    report_path: str
    group_count: int
    roll_count: int
    issues: list[dict[str, Any]]

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True)
class _RollGroup:
    fabric: str
    pon: str
    color: str
    rolls: list[tuple[int, Any]]

    @property
    def key(self) -> tuple[str, str, str]:
        return (self.fabric, self.pon, self.color)


@dataclass(frozen=True)
class _RowStyleSnapshot:
    height: float | None
    styles: tuple[Any | None, ...]


@dataclass(frozen=True)
class _SourceLayout:
    fabric_col: int
    pon_col: int
    color_col: int
    roll_qty_pairs: tuple[tuple[int, int], ...]


@dataclass(frozen=True)
class _ShipmentLayout:
    header_row: int
    columns: dict[str, int]


_SHIPMENT_FIELDS = (
    "style",
    "color",
    "quantity",
    "cartons",
    "net_weight",
    "gross_weight",
    "volume",
)
_SHIPMENT_NUMERIC_FIELDS = _SHIPMENT_FIELDS[2:]
_SHIPMENT_HEADER_ALIASES = {
    "style": {"STYLE", "STYLENO", "SKU", "款号", "款式", "款式号"},
    "color": {"COLOR", "COLOUR", "SHADE", "颜色", "色号"},
    "quantity": {"QTY", "QUANTITY", "ORDERQTY", "订单数量", "数量"},
    "cartons": {"CARTON", "CARTONS", "CTN", "CTNS", "箱数"},
    "net_weight": {"NETWEIGHT", "NW", "净重"},
    "gross_weight": {"GROSSWEIGHT", "GW", "毛重"},
    "volume": {"CBM", "VOLUME", "体积", "立方"},
}


class PackingDetailsRoutineRunner:
    """Render factory packing details from repeated source sections."""

    headers = ["FABRICS", "PON.", "COLOR", "ROLL NO", "QTY(M)", "ROLL NO", "QTY(M)"]

    def __init__(self, workspace: Path) -> None:
        self.workspace = workspace.resolve()

    def run(
        self,
        *,
        source_path: str,
        template_path: str,
        output_path: str,
    ) -> PackingDetailsRunResult:
        source = self._resolve(source_path)
        template = self._resolve(template_path)
        output = self._resolve(output_path, must_exist=False)
        report_path = output.with_suffix(".validation.json")
        ensure_absent(output, "packing output already exists")
        ensure_absent(report_path, "packing validation report already exists")
        validate_safe_xlsx(template)
        if source.suffix.lower() == ".xlsx":
            validate_safe_xlsx(source)
        shipment = self._extract_shipment_summary(source)
        issues: list[dict[str, Any]] = []
        groups: list[_RollGroup] = []
        shipment_rows: list[dict[str, Any]] = []
        if shipment is None:
            mode = "roll_manifest"
            groups = self._extract_groups(source)
            if not groups:
                issues.append({"code": "no_packing_groups"})
        else:
            mode = "shipment_summary"
            shipment_rows, shipment_issues = shipment
            issues.extend(shipment_issues)
            if not shipment_rows:
                issues.append({"code": "no_shipment_rows"})
        output.parent.mkdir(parents=True, exist_ok=True)
        with staged_path(output) as staged:
            with open_artifact(template) as source_handle, open_artifact(
                staged, "w+b"
            ) as target_handle:
                shutil.copyfileobj(source_handle, target_handle)
            if mode == "roll_manifest":
                if groups:
                    issues.extend(self._render(staged, groups))
                cross_validation = self._roll_cross_validation(groups)
                formula_validation = {
                    "status": (
                        "needs_review"
                        if any("formula" in str(issue.get("code", "")) for issue in issues)
                        else "passed"
                    ),
                    "issues": [
                        issue for issue in issues if "formula" in str(issue.get("code", ""))
                    ],
                }
                header_row = 1
                group_count = len(groups)
                roll_count = sum(len(group.rolls) for group in groups)
            else:
                render_issues, formula_validation, header_row = self._render_shipment_summary(
                    staged,
                    shipment_rows,
                )
                issues.extend(render_issues)
                cross_validation = self._shipment_cross_validation(shipment_rows)
                group_count = len(shipment_rows)
                roll_count = 0
            validate_safe_xlsx(staged)
            visual_structure = self._validate_visual_structure(staged, header_row=header_row)
            if visual_structure["status"] != "passed":
                issues.extend(visual_structure["issues"])
            status = "passed" if not issues else "needs_review"
            report = {
                "status": status,
                "mode": mode,
                "source_hash": _file_hash(source),
                "template_hash": _file_hash(template),
                "output_hash": _file_hash(staged),
                "cross_validation": cross_validation,
                "formula_validation": formula_validation,
                "visual_structure": visual_structure,
                "issues": issues,
            }
            publish_no_clobber(staged, output, "packing output already exists")
            try:
                write_json_no_clobber(
                    report_path,
                    report,
                    "packing validation report already exists",
                    anchor=staged,
                )
            except Exception:
                remove_published_link(staged, output)
                raise
        return PackingDetailsRunResult(
            status=status,
            mode=mode,
            output_path=str(output),
            report_path=str(report_path),
            group_count=group_count,
            roll_count=roll_count,
            issues=issues,
        )

    def _extract_shipment_summary(
        self,
        source: Path,
    ) -> tuple[list[dict[str, Any]], list[dict[str, Any]]] | None:
        if source.suffix.lower() != ".xlsx":
            return None
        from openpyxl import load_workbook

        workbook = load_workbook(source, data_only=False, read_only=True)
        try:
            worksheet = workbook.active
            if worksheet is None:
                return None
            layout: _ShipmentLayout | None = None
            for row_number, values in enumerate(
                worksheet.iter_rows(
                    min_row=1, max_row=min(worksheet.max_row, 50), values_only=True
                ),
                start=1,
            ):
                columns = _detect_shipment_columns(values)
                if {"style", "color", "quantity"} <= set(columns):
                    layout = _ShipmentLayout(header_row=row_number, columns=columns)
                    break
            if layout is None:
                return None

            issues: list[dict[str, Any]] = [
                {"code": "shipment_field_missing", "field": field}
                for field in _SHIPMENT_FIELDS
                if field not in layout.columns
            ]
            grouped: dict[tuple[str, str], dict[str, Any]] = {}
            for row_number, values in enumerate(
                worksheet.iter_rows(min_row=layout.header_row + 1, values_only=True),
                start=layout.header_row + 1,
            ):
                style = _clean_text(_row_value(values, layout.columns["style"]))
                color = _clean_text(_row_value(values, layout.columns["color"]))
                reject_formula_like_text(style, label="packing style")
                reject_formula_like_text(color, label="packing color")
                if not style and not color:
                    continue
                if _normalize_header(style) in {"TOTAL", "合计", "总计"}:
                    continue
                if not style or not color:
                    issues.append({"code": "shipment_group_key_missing", "row": row_number})
                    continue
                key = (style, color)
                initial_group: dict[str, Any] = {"style": style, "color": color}
                initial_group.update(dict.fromkeys(_SHIPMENT_NUMERIC_FIELDS, 0.0))
                group = grouped.setdefault(key, initial_group)
                for field in _SHIPMENT_NUMERIC_FIELDS:
                    column = layout.columns.get(field)
                    value = _row_value(values, column) if column is not None else None
                    reject_formula_like_text(value, label=f"packing {field}")
                    numeric = _numeric(value)
                    if numeric is None:
                        issues.append(
                            {
                                "code": "shipment_numeric_value_missing",
                                "field": field,
                                "row": row_number,
                            }
                        )
                        continue
                    group[field] = round(float(group[field]) + numeric, 6)
            return sorted(grouped.values(), key=lambda row: (row["style"], row["color"])), issues
        finally:
            workbook.close()

    def _render_shipment_summary(
        self,
        output: Path,
        rows: list[dict[str, Any]],
    ) -> tuple[list[dict[str, Any]], dict[str, Any], int]:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter

        with open_artifact(output) as handle:
            workbook = load_workbook(handle, data_only=False, read_only=False)
        issues: list[dict[str, Any]] = []
        try:
            worksheet = workbook.active
            if worksheet is None:
                return ([{"code": "missing_output_sheet"}], {"status": "needs_review"}, 1)
            layout: _ShipmentLayout | None = None
            for row_number, values in enumerate(
                worksheet.iter_rows(
                    min_row=1, max_row=min(worksheet.max_row, 50), values_only=True
                ),
                start=1,
            ):
                columns = _detect_shipment_columns(values)
                if {"style", "color", "quantity"} <= set(columns):
                    layout = _ShipmentLayout(header_row=row_number, columns=columns)
                    break
            if layout is None:
                raise ValueError("packing shipment template headers are unsupported")
            missing = [field for field in _SHIPMENT_FIELDS if field not in layout.columns]
            if missing:
                raise ValueError("packing shipment template is missing required metric headers")

            data_start = layout.header_row + 1
            width = max(worksheet.max_column, max(layout.columns.values()) + 1)
            data_style = self._snapshot_row_style(worksheet, data_start, width=width)
            total_style_row = self._find_total_style_row(worksheet)
            total_style = self._snapshot_row_style(worksheet, total_style_row, width=width)
            for merged in list(worksheet.merged_cells.ranges):
                if merged.min_row >= data_start:
                    worksheet.unmerge_cells(str(merged))
            if worksheet.max_row >= data_start:
                worksheet.delete_rows(data_start, worksheet.max_row - data_start + 1)

            for offset, row in enumerate(rows):
                target_row = data_start + offset
                self._apply_row_style(worksheet, data_style, target_row=target_row)
                for field in _SHIPMENT_FIELDS:
                    apply_work_cell_value(
                        worksheet.cell(target_row, layout.columns[field] + 1),
                        row[field],
                    )

            total_row = data_start + len(rows)
            self._apply_row_style(worksheet, total_style, target_row=total_row)
            apply_work_cell_value(
                worksheet.cell(total_row, layout.columns["style"] + 1),
                "TOTAL",
            )
            if rows:
                data_end = total_row - 1
                for field in _SHIPMENT_NUMERIC_FIELDS:
                    column = layout.columns[field] + 1
                    letter = get_column_letter(column)
                    apply_work_cell_value(
                        worksheet.cell(total_row, column),
                        Formula(f"=SUM({letter}{data_start}:{letter}{data_end})"),
                    )
            with open_artifact(output, "w+b") as handle:
                workbook.save(handle)
        finally:
            workbook.close()

        self._recalculate_workbook_formulas(output)
        formula_validation, validation_issues = self._validate_shipment_output(
            output,
            rows=rows,
            layout=layout,
        )
        issues.extend(validation_issues)
        return issues, formula_validation, layout.header_row

    @staticmethod
    def _validate_shipment_output(
        path: Path,
        *,
        rows: list[dict[str, Any]],
        layout: _ShipmentLayout,
    ) -> tuple[dict[str, Any], list[dict[str, Any]]]:
        from openpyxl import load_workbook

        with open_artifact(path) as formula_handle:
            formula_workbook = load_workbook(
                formula_handle, data_only=False, read_only=False
            )
        with open_artifact(path) as values_handle:
            values_workbook = load_workbook(values_handle, data_only=True, read_only=False)
        issues: list[dict[str, Any]] = []
        checks: dict[str, dict[str, Any]] = {}
        try:
            formula_sheet = formula_workbook.active
            values_sheet = values_workbook.active
            if formula_sheet is None or values_sheet is None:
                return {"status": "needs_review", "checks": {}}, [{"code": "missing_output_sheet"}]
            data_start = layout.header_row + 1
            total_row = data_start + len(rows)
            for offset, expected_row in enumerate(rows):
                row_number = data_start + offset
                for field in _SHIPMENT_FIELDS:
                    actual = values_sheet.cell(row_number, layout.columns[field] + 1).value
                    expected = expected_row[field]
                    if field in _SHIPMENT_NUMERIC_FIELDS:
                        actual_number = _numeric(actual)
                        if actual_number is None or not math.isclose(
                            actual_number,
                            float(expected),
                            rel_tol=0,
                            abs_tol=0.0001,
                        ):
                            issues.append(
                                {
                                    "code": "shipment_output_mismatch",
                                    "field": field,
                                    "row": row_number,
                                }
                            )
                    elif _clean_text(actual) != _clean_text(expected):
                        issues.append(
                            {"code": "shipment_output_mismatch", "field": field, "row": row_number}
                        )
            for field in _SHIPMENT_NUMERIC_FIELDS:
                column = layout.columns[field] + 1
                formula_cell = formula_sheet.cell(total_row, column)
                cached = _numeric(values_sheet.cell(total_row, column).value)
                expected = round(sum(float(row[field]) for row in rows), 6)
                passed = cached is not None and math.isclose(
                    cached, expected, rel_tol=0, abs_tol=0.0001
                )
                checks[formula_cell.coordinate] = {
                    "field": field,
                    "formula": formula_cell.value,
                    "expected": expected,
                    "actual": cached,
                    "status": "passed" if passed else "mismatch",
                }
                if not passed:
                    issues.append(
                        {
                            "code": "shipment_total_mismatch",
                            "field": field,
                            "cell": formula_cell.coordinate,
                        }
                    )
            return {
                "status": "passed"
                if not any(issue["code"].endswith("mismatch") for issue in issues)
                else "needs_review",
                "checks": checks,
            }, issues
        finally:
            formula_workbook.close()
            values_workbook.close()

    @staticmethod
    def _shipment_cross_validation(rows: list[dict[str, Any]]) -> dict[str, Any]:
        return {
            "group_count": len(rows),
            "styles": sorted({str(row["style"]) for row in rows}),
            "colors": sorted({str(row["color"]) for row in rows}),
            "style_color_summaries": rows,
            "totals": {
                field: round(sum(float(row[field]) for row in rows), 6)
                for field in _SHIPMENT_NUMERIC_FIELDS
            },
        }

    @staticmethod
    def _roll_cross_validation(groups: list[_RollGroup]) -> dict[str, Any]:
        summaries: list[dict[str, Any]] = []
        total_quantity = 0.0
        for group in groups:
            quantity = sum(float(value) for _roll, value in group.rolls)
            total_quantity += quantity
            summaries.append(
                {
                    "fabric": group.fabric,
                    "style": group.pon,
                    "color": group.color,
                    "roll_count": len(group.rolls),
                    "quantity": quantity,
                }
            )
        return {
            "group_summaries": summaries,
            "group_count": len(groups),
            "roll_count": sum(len(group.rolls) for group in groups),
            "total_quantity": total_quantity,
            "styles": sorted({group.pon for group in groups}),
            "colors": sorted({group.color for group in groups}),
        }

    @staticmethod
    def _validate_visual_structure(path: Path, *, header_row: int = 1) -> dict[str, Any]:
        from openpyxl import load_workbook

        with open_artifact(path) as handle:
            workbook = load_workbook(handle, data_only=False, read_only=False)
        issues: list[dict[str, Any]] = []
        try:
            worksheet = workbook.active
            if worksheet is None:
                issues.append({"code": "missing_output_sheet"})
                return {"status": "needs_review", "issues": issues}
            if worksheet.max_column < 7 or worksheet.max_row < 1:
                issues.append({"code": "packing_visual_bounds_invalid"})
            if any(
                worksheet.cell(header_row, column).value in (None, "") for column in range(1, 8)
            ):
                issues.append({"code": "packing_visual_header_missing"})
            if any(not worksheet.cell(header_row, column).has_style for column in range(1, 8)):
                issues.append({"code": "packing_visual_header_style_missing"})
            return {
                "status": "passed" if not issues else "needs_review",
                "sheet": worksheet.title,
                "max_row": worksheet.max_row,
                "max_column": worksheet.max_column,
                "merged_cells": sorted(str(item) for item in worksheet.merged_cells.ranges),
                "issues": issues,
            }
        finally:
            workbook.close()

    def _extract_groups(self, source: Path) -> list[_RollGroup]:
        from openpyxl import load_workbook

        conversion_tmp: tempfile.TemporaryDirectory[str] | None = None
        workbook_path = source
        if source.suffix.lower() == ".xls":
            self.workspace.mkdir(parents=True, exist_ok=True)
            conversion_tmp = tempfile.TemporaryDirectory(prefix=".js_work_xls_", dir=self.workspace)
            workbook_path = self._convert_xls_to_xlsx(source, Path(conversion_tmp.name))
            validate_safe_xlsx(workbook_path)

        with open_artifact(workbook_path) as handle:
            wb = load_workbook(handle, data_only=False)
        try:
            ws = wb.active
            if ws is None:
                raise ValueError("source workbook has no active sheet")
            grouped: dict[tuple[str, str, str], _RollGroup] = {}
            order: list[tuple[str, str, str]] = []
            layout: _SourceLayout | None = None
            for row in ws.iter_rows(values_only=True):
                if not row or all(value in (None, "") for value in row):
                    continue
                detected_layout = _detect_source_layout(row)
                if detected_layout is not None:
                    layout = detected_layout
                    continue
                if layout is None:
                    continue
                fabric = _clean_text(_row_value(row, layout.fabric_col))
                pon = _clean_text(_row_value(row, layout.pon_col))
                color = _clean_text(_row_value(row, layout.color_col))
                reject_formula_like_text(fabric, label="packing fabric")
                reject_formula_like_text(pon, label="packing order")
                reject_formula_like_text(color, label="packing color")
                if not fabric or fabric in {"FABRICS", "小计", "TOTAL"}:
                    continue
                key = (fabric, pon, color)
                if key not in grouped:
                    grouped[key] = _RollGroup(fabric=fabric, pon=pon, color=color, rolls=[])
                    order.append(key)
                for roll_col, qty_col in layout.roll_qty_pairs:
                    roll = _row_value(row, roll_col)
                    qty = _row_value(row, qty_col)
                    reject_formula_like_text(qty, label="packing quantity")
                    if _is_roll_number(roll) and qty not in (None, ""):
                        if _numeric(qty) is None:
                            raise ValueError("packing quantity must be numeric")
                        grouped[key].rolls.append((_roll_to_int(roll), qty))
            result: list[_RollGroup] = []
            for key in order:
                group = grouped[key]
                result.append(
                    _RollGroup(
                        fabric=group.fabric,
                        pon=group.pon,
                        color=group.color,
                        rolls=sorted(group.rolls, key=lambda item: item[0]),
                    )
                )
            return result
        finally:
            wb.close()
            if conversion_tmp is not None:
                conversion_tmp.cleanup()

    def _render(self, output: Path, groups: list[_RollGroup]) -> list[dict[str, Any]]:
        from openpyxl import load_workbook

        total_rows: list[int] = []
        with open_artifact(output) as handle:
            wb = load_workbook(handle)
        try:
            ws = wb.active
            if ws is None:
                raise ValueError("template workbook has no active sheet")
            header_style_row = self._find_header_style_row(ws)
            total_style_row = self._find_total_style_row(ws)
            data_style_row = header_style_row + 1
            header_style = self._snapshot_row_style(ws, header_style_row)
            data_style = self._snapshot_row_style(ws, data_style_row)
            total_style = self._snapshot_row_style(ws, total_style_row)
            header_values = [
                ws.cell(header_style_row, col).value or self.headers[col - 1] for col in range(1, 8)
            ]
            quantity_columns = [
                col
                for col, header in enumerate(header_values, start=1)
                if _normalize_header(header) == "QTY"
            ]
            for merged in list(ws.merged_cells.ranges):
                ws.unmerge_cells(str(merged))
            if ws.max_row > 1:
                ws.delete_rows(1, ws.max_row)
            cursor = 1
            for index, group in enumerate(groups):
                if index:
                    cursor += 2
                self._apply_row_style(ws, header_style, target_row=cursor)
                for col, value in enumerate(header_values, start=1):
                    apply_work_cell_value(ws.cell(cursor, col), value)
                data_start = cursor + 1
                half = math.ceil(len(group.rolls) / 2)
                left = group.rolls[:half]
                right = group.rolls[half:]
                for offset in range(half):
                    row_index = data_start + offset
                    self._apply_row_style(ws, data_style, target_row=row_index)
                    self._apply_default_quantity_formats(
                        ws,
                        target_row=row_index,
                        quantity_columns=quantity_columns,
                    )
                    apply_work_cell_value(ws.cell(row_index, 1), group.fabric)
                    apply_work_cell_value(ws.cell(row_index, 2), group.pon)
                    apply_work_cell_value(ws.cell(row_index, 3), group.color)
                    apply_work_cell_value(ws.cell(row_index, 4), left[offset][0])
                    apply_work_cell_value(ws.cell(row_index, 5), left[offset][1])
                    if offset < len(right):
                        apply_work_cell_value(ws.cell(row_index, 6), right[offset][0])
                        apply_work_cell_value(ws.cell(row_index, 7), right[offset][1])
                total_row = data_start + half
                self._apply_row_style(ws, total_style, target_row=total_row)
                self._apply_default_quantity_formats(
                    ws,
                    target_row=total_row,
                    quantity_columns=[5],
                )
                apply_work_cell_value(ws.cell(total_row, 1), "TOTAL ")
                apply_work_cell_value(ws.cell(total_row, 4), len(group.rolls))
                data_end = total_row - 1
                apply_work_cell_value(
                    ws.cell(total_row, 5),
                    Formula(f"=SUM(E{data_start}:E{data_end})+SUM(G{data_start}:G{data_end})"),
                )
                ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
                total_rows.append(total_row)
                cursor = total_row + 1
            with open_artifact(output, "w+b") as handle:
                wb.save(handle)
            self._recalculate_workbook_formulas(output)
            return self._validate_rendered_output(output, total_rows)
        finally:
            wb.close()

    def _resolve(self, path: str | Path, *, must_exist: bool = True) -> Path:
        if isinstance(path, MaterializedSnapshotPath):
            if not must_exist:
                raise ValueError("materialized snapshots are input-only")
            return path
        raw = Path(path)
        target = raw if raw.is_absolute() else self.workspace / raw
        reject_symlink_components(self.workspace, target)
        resolved = target.resolve()
        try:
            resolved.relative_to(self.workspace)
        except ValueError as e:
            raise ValueError(f"Path escapes workspace: {path}") from e
        if must_exist and not resolved.exists():
            raise FileNotFoundError(str(path))
        return resolved

    @staticmethod
    def _convert_xls_to_xlsx(source: Path, output_dir: Path) -> Path:
        return convert_legacy_xls_to_xlsx(
            source,
            output_dir / f"{source.stem}.xlsx",
        )

    @staticmethod
    def _recalculate_workbook_formulas(path: Path) -> None:
        refresh_formula_caches(path, soffice=_find_soffice())

    @staticmethod
    def _validate_rendered_output(path: Path, total_rows: list[int]) -> list[dict[str, Any]]:
        from openpyxl import load_workbook

        issues: list[dict[str, Any]] = []
        with open_artifact(path) as formula_handle:
            formula_wb = load_workbook(formula_handle, data_only=False)
        with open_artifact(path) as values_handle:
            values_wb = load_workbook(values_handle, data_only=True)
        try:
            formula_ws = formula_wb.active
            values_ws = values_wb.active
            if formula_ws is None or values_ws is None:
                return [{"code": "missing_output_sheet"}]
            for row in total_rows:
                formula = formula_ws.cell(row, 5).value
                cached_value = values_ws.cell(row, 5).value
                if (
                    isinstance(formula, str)
                    and formula.startswith("=")
                    and cached_value in (None, "")
                ):
                    issues.append(
                        {
                            "code": "formula_value_missing",
                            "cell": f"E{row}",
                            "formula": formula,
                        }
                    )
                    continue
                expected_rolls = _numeric(values_ws.cell(row, 4).value)
                if expected_rolls is None:
                    issues.append({"code": "roll_total_missing", "cell": f"D{row}"})
                expected_qty = (
                    evaluate_formula(formula_ws, formula) if isinstance(formula, str) else None
                )
                actual_qty = _numeric(cached_value)
                if actual_qty is None:
                    issues.append({"code": "qty_total_missing", "cell": f"E{row}"})
                elif expected_qty is None:
                    issues.append(
                        {"code": "qty_formula_unsupported", "cell": f"E{row}", "formula": formula}
                    )
                elif not math.isclose(actual_qty, expected_qty, rel_tol=0, abs_tol=0.0001):
                    issues.append(
                        {
                            "code": "qty_total_mismatch",
                            "cell": f"E{row}",
                            "expected": expected_qty,
                            "actual": actual_qty,
                        }
                    )
            return issues
        finally:
            formula_wb.close()
            values_wb.close()

    @staticmethod
    def _snapshot_row_style(
        ws: Any,
        source_row: int,
        *,
        width: int = 7,
    ) -> _RowStyleSnapshot:
        height = ws.row_dimensions[source_row].height
        styles: list[Any | None] = []
        for col in range(1, width + 1):
            source = ws.cell(source_row, col)
            if source.has_style:
                styles.append(copy(source._style))
            else:
                styles.append(None)
        return _RowStyleSnapshot(height=height, styles=tuple(styles))

    @staticmethod
    def _apply_row_style(ws: Any, snapshot: _RowStyleSnapshot, *, target_row: int) -> None:
        if snapshot.height is not None:
            ws.row_dimensions[target_row].height = snapshot.height
        for col, style in enumerate(snapshot.styles, start=1):
            if style is not None:
                ws.cell(target_row, col)._style = copy(style)

    @staticmethod
    def _apply_default_quantity_formats(
        ws: Any,
        *,
        target_row: int,
        quantity_columns: list[int],
    ) -> None:
        for column in quantity_columns:
            cell = ws.cell(target_row, column)
            if cell.number_format == "General":
                cell.number_format = "#,##0.0"

    @staticmethod
    def _find_header_style_row(ws: Any) -> int:
        for row in range(1, ws.max_row + 1):
            if _clean_text(ws.cell(row, 1).value) == "FABRICS":
                return row
        return 1

    @staticmethod
    def _find_total_style_row(ws: Any) -> int:
        for row in range(1, ws.max_row + 1):
            if _clean_text(ws.cell(row, 1).value) == "TOTAL":
                return row
        return 1


def _clean_text(value: Any) -> str:
    return str(value or "").strip()


def _row_value(row: tuple[Any, ...], index: int) -> Any:
    return row[index] if 0 <= index < len(row) else None


def _detect_source_layout(row: tuple[Any, ...]) -> _SourceLayout | None:
    normalized = [_normalize_header(value) for value in row]
    fabric_col = _first_matching_col(normalized, {"FABRICS", "FABRIC", "面料", "面料名称", "品名"})
    pon_col = _first_matching_col(normalized, {"PON", "PO", "订单号"})
    color_col = _first_matching_col(normalized, {"COLOR", "COLOUR", "颜色", "色号"})
    pairs: list[tuple[int, int]] = []
    for index, value in enumerate(normalized):
        if value in {"ROLLNO", "ROLL", "卷号"}:
            qty_col = _next_qty_col(normalized, index + 1)
            if qty_col is not None:
                pairs.append((index, qty_col))
    if fabric_col is None or pon_col is None or color_col is None or not pairs:
        return None
    return _SourceLayout(
        fabric_col=fabric_col,
        pon_col=pon_col,
        color_col=color_col,
        roll_qty_pairs=tuple(pairs),
    )


def _detect_shipment_columns(row: tuple[Any, ...]) -> dict[str, int]:
    normalized = [_normalize_header(value) for value in row]
    columns: dict[str, int] = {}
    for field, aliases in _SHIPMENT_HEADER_ALIASES.items():
        accepted = {_normalize_header(alias) for alias in aliases}
        for index, value in enumerate(normalized):
            if value in accepted:
                columns[field] = index
                break
    return columns


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().upper()
    text = re.sub(r"[\s.()（）/\\_-]+", "", text)
    if text in {"PON", "PONO", "PO NO", "PO号码"}:
        return "PON"
    if text in {"QTYM", "QTY", "米数", "数量"}:
        return "QTY"
    if text in {"ROLLNO", "ROLLNO#", "ROLLNUMBER", "卷号"}:
        return "ROLLNO"
    return text


def _first_matching_col(normalized: list[str], candidates: set[str]) -> int | None:
    normalized_candidates = {_normalize_header(candidate) for candidate in candidates}
    for index, value in enumerate(normalized):
        if value in normalized_candidates:
            return index
    return None


def _next_qty_col(normalized: list[str], start: int) -> int | None:
    for index in range(start, len(normalized)):
        if normalized[index] == "QTY":
            return index
        if normalized[index] == "ROLLNO":
            return None
    return None


def _is_roll_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def _roll_to_int(value: Any) -> int:
    if not _is_roll_number(value):
        raise ValueError(f"invalid roll number: {value}")
    return int(value)


def _numeric(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return None


def _find_soffice() -> Path | None:
    found = shutil.which("soffice")
    if found:
        return Path(found)
    for candidate in (
        Path("/Applications/LibreOffice.app/Contents/MacOS/soffice"),
        Path("/opt/homebrew/bin/soffice"),
        Path("/usr/local/bin/soffice"),
    ):
        if candidate.exists():
            return candidate
    return None


def _file_hash(path: Path) -> str:
    digest = sha256()
    with open_artifact(path) as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()
