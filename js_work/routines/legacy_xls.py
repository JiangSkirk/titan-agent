"""Bounded, pure-Python BIFF ``.xls`` ingestion for JS Agent Work.

Only cached cell values are copied into a new macro-free OOXML workbook. The
reader never starts an office application, preserves no executable objects or
links, and enforces resource limits before materializing cells.
"""

from __future__ import annotations

import io
import math
import os
import re
import stat
from datetime import datetime
from pathlib import Path
from typing import Any

from js_work.safe_output import ensure_absent, publish_no_clobber, staged_path

MAX_LEGACY_XLS_BYTES = 50 * 1024 * 1024
MAX_LEGACY_XLS_SHEETS = 64
MAX_LEGACY_XLS_ROWS_PER_SHEET = 200_000
MAX_LEGACY_XLS_COLUMNS = 512
MAX_LEGACY_XLS_CELLS = 1_000_000
MAX_LEGACY_XLS_TEXT_CHARS = 20_000_000
_OLE_COMPOUND_MAGIC = bytes.fromhex("d0cf11e0a1b11ae1")
_INVALID_SHEET_CHARACTERS = re.compile(r"[\\/*?:\[\]]")


class LegacyXlsError(ValueError):
    """The legacy workbook is unsafe, malformed, or exceeds bounded limits."""


def convert_legacy_xls_to_xlsx(source: Path, output: Path) -> Path:
    """Read BIFF values with pinned xlrd and write a sanitized XLSX copy."""
    # Lazy imports keep desktop sidecar cold-start free of Office deps until
    # an actual .xls conversion is requested.
    import xlrd
    from openpyxl import Workbook

    output = output.resolve()
    if output.suffix.lower() != ".xlsx":
        raise LegacyXlsError("legacy .xls output must be an .xlsx file")
    try:
        ensure_absent(output, "legacy .xls output already exists")
    except ValueError as exc:
        raise LegacyXlsError("legacy .xls output already exists") from exc
    source_bytes = _read_source_bytes(source)
    output.parent.mkdir(parents=True, exist_ok=True)

    diagnostics = io.StringIO()
    try:
        book = xlrd.open_workbook(
            file_contents=source_bytes,
            logfile=diagnostics,
            verbosity=0,
            use_mmap=False,
            formatting_info=False,
            on_demand=True,
            ragged_rows=True,
            ignore_workbook_corruption=False,
        )
    except Exception as exc:
        raise LegacyXlsError("legacy .xls workbook could not be parsed safely") from exc

    workbook = Workbook()
    workbook.remove(workbook.active)
    total_cells = 0
    total_text_chars = 0
    used_sheet_names: set[str] = set()
    try:
        if book.nsheets < 1 or book.nsheets > MAX_LEGACY_XLS_SHEETS:
            raise LegacyXlsError("legacy .xls workbook has an invalid sheet count")
        for sheet_index in range(book.nsheets):
            sheet = book.sheet_by_index(sheet_index)
            if sheet.nrows > MAX_LEGACY_XLS_ROWS_PER_SHEET:
                raise LegacyXlsError("legacy .xls worksheet row limit exceeded")
            if sheet.ncols > MAX_LEGACY_XLS_COLUMNS:
                raise LegacyXlsError("legacy .xls worksheet column limit exceeded")
            projected_cells = sheet.nrows * sheet.ncols
            if projected_cells > MAX_LEGACY_XLS_CELLS - total_cells:
                raise LegacyXlsError("legacy .xls workbook cell limit exceeded")
            total_cells += projected_cells

            worksheet = workbook.create_sheet(
                _safe_sheet_name(str(sheet.name), used_sheet_names, sheet_index)
            )
            for row_index in range(sheet.nrows):
                row_length = min(sheet.row_len(row_index), sheet.ncols)
                for column_index in range(row_length):
                    value = _cell_value(
                        book,
                        sheet.cell(row_index, column_index),
                        xlrd_module=xlrd,
                    )
                    if isinstance(value, str):
                        total_text_chars += len(value)
                        if total_text_chars > MAX_LEGACY_XLS_TEXT_CHARS:
                            raise LegacyXlsError("legacy .xls workbook text limit exceeded")
                    target = worksheet.cell(row=row_index + 1, column=column_index + 1)
                    target.value = value
                    if isinstance(value, str):
                        # Force strings such as =CMD(...) to remain inert text.
                        target.data_type = "s"

        with staged_path(output) as temporary:
            workbook.save(temporary)
            try:
                publish_no_clobber(
                    temporary,
                    output,
                    "legacy .xls output already exists",
                )
            except ValueError as exc:
                raise LegacyXlsError("legacy .xls output already exists") from exc
    finally:
        workbook.close()
        book.release_resources()
    return output


def _read_source_bytes(source: Path) -> bytes:
    try:
        path_metadata = source.lstat()
    except OSError as exc:
        raise LegacyXlsError("legacy .xls source is unavailable") from exc
    if not stat.S_ISREG(path_metadata.st_mode):
        raise LegacyXlsError("legacy .xls source must be a regular file")
    flags = os.O_RDONLY
    flags |= getattr(os, "O_CLOEXEC", 0)
    flags |= getattr(os, "O_NOFOLLOW", 0)
    try:
        descriptor = os.open(source, flags)
    except OSError as exc:
        raise LegacyXlsError("legacy .xls source is unavailable") from exc
    try:
        with os.fdopen(descriptor, "rb") as handle:
            before = os.fstat(handle.fileno())
            if (
                not stat.S_ISREG(before.st_mode)
                or (before.st_dev, before.st_ino)
                != (path_metadata.st_dev, path_metadata.st_ino)
            ):
                raise LegacyXlsError("legacy .xls source changed before it was read")
            if before.st_size < len(_OLE_COMPOUND_MAGIC):
                raise LegacyXlsError("legacy .xls source is truncated")
            if before.st_size > MAX_LEGACY_XLS_BYTES:
                raise LegacyXlsError("legacy .xls source exceeds 50 MiB")
            payload = handle.read(MAX_LEGACY_XLS_BYTES + 1)
            after = os.fstat(handle.fileno())
    except LegacyXlsError:
        raise
    except OSError as exc:
        raise LegacyXlsError("legacy .xls source is unavailable") from exc
    if (
        (before.st_dev, before.st_ino, before.st_size, before.st_mtime_ns, before.st_ctime_ns)
        != (after.st_dev, after.st_ino, after.st_size, after.st_mtime_ns, after.st_ctime_ns)
        or len(payload) != before.st_size
    ):
        raise LegacyXlsError("legacy .xls source changed while it was read")
    if payload[: len(_OLE_COMPOUND_MAGIC)] != _OLE_COMPOUND_MAGIC:
        raise LegacyXlsError("legacy .xls source is not an OLE BIFF workbook")
    return payload


def _cell_value(
    book: Any,
    cell: Any,
    *,
    xlrd_module: Any,
) -> str | int | float | bool | datetime | None:
    cell_type = int(cell.ctype)
    value = cell.value
    if cell_type in {xlrd_module.XL_CELL_EMPTY, xlrd_module.XL_CELL_BLANK}:
        return None
    if cell_type == xlrd_module.XL_CELL_TEXT:
        return str(value)
    if cell_type == xlrd_module.XL_CELL_NUMBER:
        number = float(value)
        if not math.isfinite(number):
            raise LegacyXlsError("legacy .xls contains a non-finite number")
        return int(number) if number.is_integer() else number
    if cell_type == xlrd_module.XL_CELL_DATE:
        try:
            converted = xlrd_module.xldate.xldate_as_datetime(value, book.datemode)
        except (OverflowError, TypeError, ValueError) as exc:
            raise LegacyXlsError("legacy .xls contains an invalid date") from exc
        if not isinstance(converted, datetime):
            raise LegacyXlsError("legacy .xls contains an invalid date")
        return converted
    if cell_type == xlrd_module.XL_CELL_BOOLEAN:
        return bool(value)
    if cell_type == xlrd_module.XL_CELL_ERROR:
        raise LegacyXlsError("legacy .xls contains a cell error")
    raise LegacyXlsError("legacy .xls contains an unsupported cell type")


def _safe_sheet_name(raw: str, used: set[str], index: int) -> str:
    candidate = _INVALID_SHEET_CHARACTERS.sub("_", raw).strip("'")[:31]
    if not candidate:
        candidate = f"Sheet{index + 1}"
    base = candidate
    suffix = 2
    while candidate.casefold() in used:
        marker = f"_{suffix}"
        candidate = base[: 31 - len(marker)] + marker
        suffix += 1
    used.add(candidate.casefold())
    return candidate
